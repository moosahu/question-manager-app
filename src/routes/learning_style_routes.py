# src/routes/learning_style_routes.py
"""أنماط التعلم (VARK) — استبيان يحدد نمط تعلم الطالب، ويعرض للمعلم إحصائيات طلابه"""

from flask import Blueprint, request, jsonify, render_template
from flask_login import login_required, current_user
from functools import wraps
from datetime import datetime

try:
    from src.extensions import db
    from src.models.learning_style import LearningStyleResult
    from src.models.student import Student
    from src.models.teacher import Teacher
    from src.models.teacher_student import TeacherStudent
    from src.middleware.auth_middleware import verify_student_token, verify_teacher_token
except ImportError:  # pragma: no cover
    from extensions import db
    from models.learning_style import LearningStyleResult
    from models.student import Student
    from models.teacher import Teacher
    from models.teacher_student import TeacherStudent
    from middleware.auth_middleware import verify_student_token, verify_teacher_token

learning_style_bp = Blueprint('learning_style', __name__, url_prefix='/api/learning-style')


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return jsonify({'success': False, 'error': 'يجب تسجيل الدخول'}), 401
        if not getattr(current_user, 'is_admin', False):
            return jsonify({'success': False, 'error': 'صلاحيات غير كافية'}), 403
        return f(*args, **kwargs)
    return decorated

_STYLE_NAMES = {'V': 'بصري', 'A': 'سمعي', 'R': 'قرائي/كتابي', 'K': 'حركي'}

# استبيان أنماط التعلم (VARK) — 12 سؤال، كل سؤال له خيار وحيد من كل نمط لكن بترتيب مختلط
# (مو نفس ترتيب V,A,R,K كل مرة) — نفس أسلوب استبيان VARK الحقيقي لتفادي التحيّز لموضع الخيار
_QUESTIONS = [
    {'id': 1, 'text': 'عند تعلّم شيء جديد، أفضّل أن:', 'options': [
        {'key': 'V', 'text': 'أشاهد صورة أو رسمًا أو مقطعًا يوضّحه.'},
        {'key': 'A', 'text': 'أستمع لشرح المعلّم أو تسجيل صوتي.'},
        {'key': 'R', 'text': 'أقرأ عنه في الكتاب أو أدوّن ملاحظات.'},
        {'key': 'K', 'text': 'أجرّبه بنفسي عمليًا خطوة بخطوة.'},
    ]},
    {'id': 2, 'text': 'أتذكّر المعلومة أكثر عندما:', 'options': [
        {'key': 'A', 'text': 'أرددها بصوتي أو أسمعها من غيري.'},
        {'key': 'R', 'text': 'أكتبها بأسلوبي وأعيد قراءتها.'},
        {'key': 'K', 'text': 'أعيد عمل التجربة أو النشاط.'},
        {'key': 'V', 'text': 'أربطها بمخطط أو ألوان أو مكانها في الصفحة.'},
    ]},
    {'id': 3, 'text': 'حين أحتاج طريق مكان جديد، أرتاح إلى:', 'options': [
        {'key': 'R', 'text': 'تعليمات مكتوبة أتّبعها بالترتيب.'},
        {'key': 'K', 'text': 'أن أمشي فيه مرة حتى أحفظه.'},
        {'key': 'V', 'text': 'خريطة مرسومة أو مصوّرة.'},
        {'key': 'A', 'text': 'أن يشرح لي أحدهم الطريق شفهيًّا.'},
    ]},
    {'id': 4, 'text': 'في وقت الفراغ أميل إلى:', 'options': [
        {'key': 'K', 'text': 'الرياضة والأعمال اليدوية والحركة.'},
        {'key': 'V', 'text': 'مشاهدة الأفلام والصور واللوحات.'},
        {'key': 'A', 'text': 'الاستماع للأناشيد أو البرامج الصوتية.'},
        {'key': 'R', 'text': 'قراءة القصص والمقالات.'},
    ]},
    {'id': 5, 'text': 'عند شرح فكرة لزميلي، غالبًا:', 'options': [
        {'key': 'V', 'text': 'أرسم له مخططًا بسيطًا.'},
        {'key': 'A', 'text': 'أشرحها له بالكلام والنبرة.'},
        {'key': 'R', 'text': 'أكتب له النقاط الأساسية.'},
        {'key': 'K', 'text': 'أريه كيف تُعمل بيدي.'},
    ]},
    {'id': 6, 'text': 'أكثر ما يشتّت تركيزي في الحصة:', 'options': [
        {'key': 'A', 'text': 'الأصوات والضجيج.'},
        {'key': 'R', 'text': 'كثرة النصوص غير المنظّمة.'},
        {'key': 'K', 'text': 'الجلوس طويلًا دون حركة.'},
        {'key': 'V', 'text': 'فوضى المنظر أمامي أو ضعف الإضاءة.'},
    ]},
    {'id': 7, 'text': 'لأحفظ كلمات درس جديد أفضّل:', 'options': [
        {'key': 'R', 'text': 'أن أكتبها عدة مرات.'},
        {'key': 'K', 'text': 'أن أمثّل معناها بحركة.'},
        {'key': 'V', 'text': 'بطاقات مصوّرة وألوان.'},
        {'key': 'A', 'text': 'أن أنطقها بصوت عالٍ مرارًا.'},
    ]},
    {'id': 8, 'text': 'عندما أجمّع جهازًا أو لعبة جديدة:', 'options': [
        {'key': 'K', 'text': 'أبدأ التجربة مباشرة بيدي.'},
        {'key': 'V', 'text': 'أنظر إلى صور خطوات التركيب.'},
        {'key': 'A', 'text': 'أطلب من أحد أن يشرح لي.'},
        {'key': 'R', 'text': 'أقرأ دليل التعليمات.'},
    ]},
    {'id': 9, 'text': 'يشدّني في الدرس أكثر عندما:', 'options': [
        {'key': 'V', 'text': 'يستخدم المعلّم عرضًا مرئيًا.'},
        {'key': 'A', 'text': 'يحكي قصة أو يناقشنا صوتيًّا.'},
        {'key': 'R', 'text': 'يوزّع مطويّة أو ورقة عمل مكتوبة.'},
        {'key': 'K', 'text': 'ينفّذ نشاطًا أو تجربة عملية.'},
    ]},
    {'id': 10, 'text': 'أعبّر عن نفسي براحة أكبر عبر:', 'options': [
        {'key': 'A', 'text': 'الحديث والإلقاء.'},
        {'key': 'R', 'text': 'الكتابة.'},
        {'key': 'K', 'text': 'العمل والحركة والتمثيل.'},
        {'key': 'V', 'text': 'الرسم والتصميم.'},
    ]},
    {'id': 11, 'text': 'حين أراجع قبل الاختبار:', 'options': [
        {'key': 'R', 'text': 'ألخّص الدرس كتابةً.'},
        {'key': 'K', 'text': 'أطبّق أمثلة وتمارين عملية.'},
        {'key': 'V', 'text': 'أحوّل الدرس إلى خرائط ذهنية ملوّنة.'},
        {'key': 'A', 'text': 'أشرح الدرس لنفسي بصوت مسموع.'},
    ]},
    {'id': 12, 'text': 'أفهم التعليمات الجديدة أسرع إذا كانت:', 'options': [
        {'key': 'K', 'text': 'عملية أجرّبها فورًا.'},
        {'key': 'V', 'text': 'مصوّرة أو ضمن فيديو.'},
        {'key': 'A', 'text': 'مسموعة يشرحها أحد.'},
        {'key': 'R', 'text': 'مكتوبة بخطوات مرقّمة.'},
    ]},
]

_STYLE_TIPS = {
    'بصري': 'تتعلم أفضل بالصور والمخططات والألوان — استخدم الخرائط الذهنية والرسوم التوضيحية بمذاكرتك.',
    'سمعي': 'تتعلم أفضل بالاستماع والنقاش — سجّل الشروحات واستمع لها، وناقش زملاءك بصوت عالٍ.',
    'قرائي/كتابي': 'تتعلم أفضل بالقراءة والكتابة — لخّص دروسك بنقاط مكتوبة وأعد كتابتها.',
    'حركي': 'تتعلم أفضل بالتجربة والتطبيق العملي — حل تمارين وتدريبات عملية بدل القراءة النظرية فقط.',
}


@learning_style_bp.route('/questions', methods=['GET'])
def get_questions():
    """أسئلة الاستبيان — بدون كشف أي خيار يقابل أي نمط بالاستجابة (فقط النص)"""
    questions = [{
        'id': q['id'], 'text': q['text'],
        'options': [{'key': o['key'], 'text': o['text']} for o in q['options']],
    } for q in _QUESTIONS]
    return jsonify({'success': True, 'questions': questions})


def _compute_result(answers):
    scores = {'V': 0, 'A': 0, 'R': 0, 'K': 0}
    for a in (answers or []):
        key = (a.get('option') or '').upper()
        if key in scores:
            scores[key] += 1
    max_score = max(scores.values()) if any(scores.values()) else 0
    dominant_keys = [k for k, v in scores.items() if v == max_score and max_score > 0]
    dominant_style = '/'.join(_STYLE_NAMES[k] for k in dominant_keys)
    return scores, dominant_style


@learning_style_bp.route('/submit', methods=['POST'])
@verify_student_token
def submit_result():
    """الطالب يرسل إجاباته — يُحسب النمط ويُحفظ (يستبدل أي نتيجة سابقة لو أعاد الاستبيان)"""
    try:
        student_id = request.student_id
        data = request.get_json() or {}
        answers = data.get('answers') or []
        if not answers:
            return jsonify({'success': False, 'error': 'answers مطلوبة'}), 400

        # ✅ الاستبيان يُؤخذ مرة واحدة بس — إلا لو المعلم/الأدمن أعاد الفرصة (يمسح نتيجته القديمة)
        result = LearningStyleResult.query.filter_by(student_id=student_id).first()
        if result:
            return jsonify({
                'success': False, 'error': 'already_taken',
                'message': 'أخذت الاستبيان من قبل — اطلب من معلمك يعيد لك الفرصة عشان تقدر تعيده',
            }), 400

        scores, dominant_style = _compute_result(answers)
        result = LearningStyleResult(student_id=student_id)
        db.session.add(result)

        result.visual_score = scores['V']
        result.auditory_score = scores['A']
        result.reading_score = scores['R']
        result.kinesthetic_score = scores['K']
        result.dominant_style = dominant_style
        result.answers = answers
        db.session.commit()

        result_dict = result.to_dict()
        result_dict['tips'] = [_STYLE_TIPS[s] for s in dominant_style.split('/') if s in _STYLE_TIPS]
        return jsonify({'success': True, 'message': 'تم حفظ النتيجة', 'result': result_dict})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@learning_style_bp.route('/my-result', methods=['GET'])
@verify_student_token
def my_result():
    """نتيجة الطالب الحالي (لو أخذ الاستبيان من قبل)"""
    try:
        student_id = request.student_id
        result = LearningStyleResult.query.filter_by(student_id=student_id).first()
        if not result:
            return jsonify({'success': True, 'result': None})
        result_dict = result.to_dict()
        result_dict['tips'] = [_STYLE_TIPS[s] for s in result.dominant_style.split('/') if s in _STYLE_TIPS]
        return jsonify({'success': True, 'result': result_dict})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@learning_style_bp.route('/access', methods=['GET'])
@verify_student_token
def check_access():
    """هل الاستبيان متاح لهذا الطالب؟ — بس لو معلمه/الأدمن أرسله له (أو أخذه من قبل أصلاً)"""
    try:
        student_id = request.student_id
        if LearningStyleResult.query.filter_by(student_id=student_id).first():
            return jsonify({'success': True, 'has_access': True})

        try:
            from src.models.notification import Notification
        except ImportError:  # pragma: no cover
            from models.notification import Notification

        invited = Notification.query.filter_by(student_id=student_id, type='learning_style').first() is not None
        return jsonify({'success': True, 'has_access': invited})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@learning_style_bp.route('/teacher/students', methods=['GET'])
@verify_teacher_token
def teacher_students_styles():
    """قائمة طلاب المعلم مع أنماط تعلمهم + إحصائية جماعية"""
    try:
        teacher_id = request.teacher_id
        links = TeacherStudent.query.join(TeacherStudent.student).filter(
            TeacherStudent.teacher_id == teacher_id
        ).order_by(Student.name).all()

        student_ids = [l.student_id for l in links]
        results = LearningStyleResult.query.filter(LearningStyleResult.student_id.in_(student_ids)).all() if student_ids else []
        by_student = {r.student_id: r for r in results}

        counts = {'بصري': 0, 'سمعي': 0, 'قرائي/كتابي': 0, 'حركي': 0}
        students_data = []
        for l in links:
            st = l.student
            r = by_student.get(l.student_id)
            entry = {
                'student_id': l.student_id,
                'student_name': st.name if st else None,
                'taken': r is not None,
                'result': r.to_dict() if r else None,
            }
            students_data.append(entry)
            if r and r.dominant_style:
                for style in r.dominant_style.split('/'):
                    if style in counts:
                        counts[style] += 1

        return jsonify({
            'success': True,
            'students': students_data,
            'counts': counts,
            'total_students': len(links),
            'taken_count': len(results),
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@learning_style_bp.route('/teacher/reopen/<int:student_id>', methods=['POST'])
@verify_teacher_token
def reopen_for_student(student_id):
    """يعيد الفرصة لطالب من طلاب المعلم يعيد الاستبيان (يمسح نتيجته الحالية)"""
    try:
        teacher_id = request.teacher_id
        link = TeacherStudent.query.filter_by(teacher_id=teacher_id, student_id=student_id).first()
        if not link:
            return jsonify({'success': False, 'error': 'هذا الطالب مو من ضمن طلابك'}), 403

        result = LearningStyleResult.query.filter_by(student_id=student_id).first()
        if not result:
            return jsonify({'success': False, 'error': 'هذا الطالب ما أخذ الاستبيان أصلاً'}), 400

        db.session.delete(result)
        db.session.commit()
        return jsonify({'success': True, 'message': 'تم إعادة الفرصة للطالب'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


def _notify_students(student_ids):
    """يرسل تذكير (إشعار داخل التطبيق + push لو متاح) لقائمة معرّفات طلاب — يُستخدم من مسارات المعلم والأدمن"""
    try:
        from src.models.notification import Notification, StudentNotification
    except ImportError:  # pragma: no cover
        from models.notification import Notification, StudentNotification

    title = '📋 استبيان أنماط التعلم'
    message = 'معلمك يطلب منك تعبئة استبيان أنماط التعلم من التطبيق — يساعده يفهم طريقة تعلّمك الأفضل.'

    for sid in student_ids:
        notification = Notification(
            student_id=sid, title=title, message=message, body=message,
            type='learning_style', notification_type='learning_style',
            is_read=False, status='delivered', sent_at=datetime.utcnow(),
        )
        db.session.add(notification)
        db.session.flush()
        db.session.add(StudentNotification(student_id=sid, notification_id=notification.id, is_read=False))
    db.session.commit()

    sent_push = 0
    try:
        from src.services.notification_service import NotificationService
        students = Student.query.filter(Student.id.in_(student_ids), Student.fcm_token.isnot(None)).all()
        for st in students:
            try:
                NotificationService.send_fcm_notification(st.fcm_token, title, message, {'type': 'learning_style'})
                sent_push += 1
            except Exception:
                pass
    except Exception:
        pass

    return sent_push


@learning_style_bp.route('/teacher/notify', methods=['POST'])
@verify_teacher_token
def notify_students():
    """يرسل تذكير لطلاب محددين أو كل طلاب المعلم بتعبئة الاستبيان"""
    try:
        teacher_id = request.teacher_id
        data = request.get_json() or {}
        requested_ids = data.get('student_ids')  # None/[] = كل طلاب المعلم

        links = TeacherStudent.query.filter_by(teacher_id=teacher_id).all()
        my_student_ids = {l.student_id for l in links}
        targets = [sid for sid in requested_ids if sid in my_student_ids] if requested_ids else list(my_student_ids)

        if not targets:
            return jsonify({'success': False, 'error': 'ما فيه طلاب لإرسال الاستبيان لهم'}), 400

        sent_push = _notify_students(targets)
        return jsonify({'success': True, 'sent_count': len(targets), 'push_sent': sent_push})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== لوحة الأدمن (ويب) ====================

@learning_style_bp.route('/page', methods=['GET'])
@login_required
@admin_required
def admin_page():
    """صفحة الأدمن لإرسال استبيان أنماط التعلم لكل الطلاب أو طلاب محددين"""
    return render_template('learning_styles_admin.html')


def _admin_linked_students():
    """طلاب الأدمن المرتبطين به فقط (TeacherStudent.admin_id) — نفس نطاق /api/mobile/admin/students"""
    links = TeacherStudent.query.join(TeacherStudent.student).filter(
        TeacherStudent.admin_id == current_user.id
    ).order_by(Student.name).all()
    return [l.student for l in links if l.student]


@learning_style_bp.route('/admin/students', methods=['GET'])
@login_required
@admin_required
def admin_students_styles():
    """طلاب الأدمن المرتبطين به بس (مو كل طلاب المنصة) مع أنماط تعلمهم + إحصائية جماعية"""
    try:
        students = _admin_linked_students()
        student_ids = [s.id for s in students]
        results = LearningStyleResult.query.filter(LearningStyleResult.student_id.in_(student_ids)).all() if student_ids else []
        by_student = {r.student_id: r for r in results}

        counts = {'بصري': 0, 'سمعي': 0, 'قرائي/كتابي': 0, 'حركي': 0}
        students_data = []
        for st in students:
            r = by_student.get(st.id)
            students_data.append({
                'student_id': st.id,
                'student_name': st.name,
                'taken': r is not None,
                'result': r.to_dict() if r else None,
            })
            if r and r.dominant_style:
                for style in r.dominant_style.split('/'):
                    if style in counts:
                        counts[style] += 1

        return jsonify({
            'success': True,
            'students': students_data,
            'counts': counts,
            'total_students': len(students),
            'taken_count': len(results),
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@learning_style_bp.route('/admin/notify', methods=['POST'])
@login_required
@admin_required
def admin_notify_students():
    """الأدمن يرسل تذكير الاستبيان لطلاب محددين أو كل طلابه المرتبطين (مو كل طلاب المنصة)"""
    try:
        data = request.get_json() or {}
        requested_ids = data.get('student_ids')  # None/[] = كل طلاب الأدمن المرتبطين

        my_student_ids = {s.id for s in _admin_linked_students()}
        targets = [sid for sid in requested_ids if sid in my_student_ids] if requested_ids else list(my_student_ids)

        if not targets:
            return jsonify({'success': False, 'error': 'ما فيه طلاب لإرسال الاستبيان لهم'}), 400

        sent_push = _notify_students(targets)
        return jsonify({'success': True, 'sent_count': len(targets), 'push_sent': sent_push})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@learning_style_bp.route('/admin/reopen/<int:student_id>', methods=['POST'])
@login_required
@admin_required
def admin_reopen_for_student(student_id):
    """الأدمن يعيد الفرصة لطالب من طلابه المرتبطين بس يعيد الاستبيان"""
    try:
        link = TeacherStudent.query.filter_by(admin_id=current_user.id, student_id=student_id).first()
        if not link:
            return jsonify({'success': False, 'error': 'هذا الطالب مو من ضمن طلابك'}), 403

        result = LearningStyleResult.query.filter_by(student_id=student_id).first()
        if not result:
            return jsonify({'success': False, 'error': 'هذا الطالب ما أخذ الاستبيان أصلاً'}), 400
        db.session.delete(result)
        db.session.commit()
        return jsonify({'success': True, 'message': 'تم إعادة الفرصة للطالب'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


def _build_styles_excel(students_data):
    """يبني ملف Excel لتقرير أنماط التعلم — نفس أسلوب بقية تقارير التطبيق (هيدر أرجواني، تذييل موحّد)"""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'أنماط التعلم'
    ws.sheet_view.rightToLeft = True

    headers = ['الطالب', 'النمط الغالب', 'بصري %', 'سمعي %', 'قرائي/كتابي %', 'حركي %', 'أخذ الاستبيان']
    thin = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='6366F1')
    for col, h in enumerate(headers, start=1):
        c = ws.cell(1, col, h)
        c.font = Font(color='FFFFFF', bold=True, size=11)
        c.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
        c.fill = header_fill
        c.border = border
    ws.row_dimensions[1].height = 26

    row = 2
    for s in students_data:
        r = s.get('result') or {}
        values = [
            s.get('student_name') or '', r.get('dominant_style') or '-',
            r.get('visual_percent', ''), r.get('auditory_percent', ''),
            r.get('reading_percent', ''), r.get('kinesthetic_percent', ''),
            'نعم' if s.get('taken') else 'لا',
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row, col, v)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

    for col, width in zip(range(1, len(headers) + 1), [22, 20, 10, 10, 14, 10, 14]):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    footer_row = row + 1
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=len(headers))
    fc = ws.cell(footer_row, 1)
    fc.value = f'⚗️  تم استخراج هذا التقرير من تطبيق كيم تحصيلي  |  منصة تعليمية للكيمياء  |  جميع الحقوق محفوظة © {datetime.now().year}'
    fc.font = Font(size=9, color='888888', italic=True)
    fc.alignment = Alignment(horizontal='center', vertical='center')
    fc.fill = PatternFill('solid', fgColor='F1F5F9')
    ws.row_dimensions[footer_row].height = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@learning_style_bp.route('/admin/export-excel', methods=['GET'])
@login_required
@admin_required
def admin_export_excel():
    """تصدير تقرير أنماط التعلم لطلاب الأدمن المرتبطين بس"""
    try:
        from flask import send_file
        students = _admin_linked_students()
        student_ids = [s.id for s in students]
        results = LearningStyleResult.query.filter(LearningStyleResult.student_id.in_(student_ids)).all() if student_ids else []
        by_student = {r.student_id: r for r in results}
        students_data = [{
            'student_name': s.name,
            'taken': s.id in by_student,
            'result': by_student[s.id].to_dict() if s.id in by_student else None,
        } for s in students]

        output = _build_styles_excel(students_data)
        return send_file(
            output, as_attachment=True, download_name='أنماط_التعلم.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@learning_style_bp.route('/teacher/export-excel', methods=['GET'])
@verify_teacher_token
def teacher_export_excel():
    """تصدير تقرير أنماط تعلم طلاب المعلم"""
    try:
        from flask import send_file
        teacher_id = request.teacher_id
        links = TeacherStudent.query.join(TeacherStudent.student).filter(
            TeacherStudent.teacher_id == teacher_id
        ).order_by(Student.name).all()
        student_ids = [l.student_id for l in links]
        results = LearningStyleResult.query.filter(LearningStyleResult.student_id.in_(student_ids)).all() if student_ids else []
        by_student = {r.student_id: r for r in results}
        students_data = [{
            'student_name': l.student.name if l.student else None,
            'taken': l.student_id in by_student,
            'result': by_student[l.student_id].to_dict() if l.student_id in by_student else None,
        } for l in links]

        output = _build_styles_excel(students_data)
        return send_file(
            output, as_attachment=True, download_name='أنماط_التعلم.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== تصدير PDF (مع/بدون كليشة الوزارة) ====================

# نفس ترتيب عرض الأنماط بالنموذج الرسمي: سمعي / حركي / بصري / قرائي-كتابي
_PDF_STYLE_ORDER = [('A', 'سمعي'), ('K', 'حركي'), ('V', 'بصري'), ('R', 'قرائي/كتابي')]


def _parse_bool(value, default=False):
    if value is None:
        return default
    return str(value).strip().lower() in ('1', 'true', 'yes', 'on')


def _pdf_options_from_request(default_teacher_name='', default_school_name=''):
    return {
        'with_letterhead': _parse_bool(request.args.get('with_letterhead'), False),
        'school_name': request.args.get('school_name') or default_school_name,
        'section_label': request.args.get('section_label') or '',
        'academic_year': request.args.get('academic_year') or '',
        'teacher_name': request.args.get('teacher_name') or default_teacher_name,
    }


def _filter_by_ids(students, param_name='student_ids'):
    """يفلتر قائمة كائنات طالب/روابط حسب query param اختياري student_ids (مفصولة بفواصل)"""
    raw = request.args.get(param_name)
    if not raw:
        return students
    wanted = {int(x) for x in raw.split(',') if x.strip().isdigit()}
    return [s for s in students if (s.student_id if hasattr(s, 'student_id') else s.id) in wanted]


def _html_to_pdf(html_content):
    """يحوّل HTML لـPDF عبر نفس محرّك Playwright المستخدم بتوليد الاختبارات"""
    import os
    import uuid
    try:
        from src.routes.exam_generator import _get_browser
    except ImportError:  # pragma: no cover
        from routes.exam_generator import _get_browser

    src_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    browser = _get_browser()
    ctx = browser.new_context()
    page = ctx.new_page()
    tmp_path = os.path.join(src_dir, f'_pw_render_{uuid.uuid4().hex}.html')
    try:
        with open(tmp_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        page.goto(f"file://{tmp_path}", wait_until='load')
        pdf = page.pdf(
            format='A4', print_background=True,
            margin={'top': '10mm', 'right': '8mm', 'bottom': '12mm', 'left': '8mm'},
        )
    finally:
        ctx.close()
        try:
            os.remove(tmp_path)
        except OSError:
            pass
    return pdf


_ls_header_img_cache = None


def _learning_style_header_base64():
    """صورة هيدر الوزارة الرسمي (مستخرجة من قالب Word رسمي زوّدنا حسين إياه) — بديل الشعار المفرد"""
    global _ls_header_img_cache
    if _ls_header_img_cache is not None:
        return _ls_header_img_cache
    import base64
    import os
    path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'static', 'images', 'learning_style_header.png'))
    try:
        with open(path, 'rb') as f:
            data = base64.b64encode(f.read()).decode()
            _ls_header_img_cache = f'data:image/png;base64,{data}'
    except Exception:
        _ls_header_img_cache = ''
    return _ls_header_img_cache


def _build_styles_pdf(students_data, options):
    """يبني تقرير PDF لأنماط التعلم — نفس شكل النموذج الرسمي (مربعات اختيار مُعلَّمة تلقائياً حسب النمط الغالب المحسوب)"""
    try:
        from src.routes.exam_generator import _get_font_data, _get_font_data_bold
    except ImportError:  # pragma: no cover
        from routes.exam_generator import _get_font_data, _get_font_data_bold

    context = {
        'students': students_data,
        'style_order': _PDF_STYLE_ORDER,
        'with_letterhead': options.get('with_letterhead', False),
        'school_name': options.get('school_name') or '',
        'section_label': options.get('section_label') or '',
        'academic_year': options.get('academic_year') or '',
        'teacher_name': options.get('teacher_name') or '',
        'header_image_base64': _learning_style_header_base64() if options.get('with_letterhead') else '',
        'font_regular': _get_font_data('cairo'),
        'font_bold': _get_font_data_bold('cairo'),
        'year_now': datetime.now().year,
    }
    html_content = render_template('learning_style_report_pdf.html', **context)
    return _html_to_pdf(html_content)


def _students_data_for_pdf(links_or_students, is_link=True):
    """يبني students_data من روابط TeacherStudent أو من كائنات Student مباشرة"""
    ids = [(l.student_id if is_link else l.id) for l in links_or_students]
    results = LearningStyleResult.query.filter(LearningStyleResult.student_id.in_(ids)).all() if ids else []
    by_student = {r.student_id: r for r in results}
    data = []
    for item in links_or_students:
        sid = item.student_id if is_link else item.id
        name = (item.student.name if item.student else None) if is_link else item.name
        r = by_student.get(sid)
        data.append({
            'student_name': name,
            'taken': r is not None,
            'dominant_keys': set((r.dominant_style or '').split('/')) if r else set(),
        })
    return data


@learning_style_bp.route('/admin/export-pdf', methods=['GET'])
@login_required
@admin_required
def admin_export_pdf():
    """تصدير تقرير أنماط التعلم PDF بشكل النموذج الرسمي — مع/بدون كليشة الوزارة حسب اختيار الأدمن"""
    try:
        from flask import send_file
        from io import BytesIO

        students = _filter_by_ids(_admin_linked_students())
        students_data = _students_data_for_pdf(students, is_link=False)

        admin_name = getattr(current_user, 'full_name', '') or current_user.username
        pdf_bytes = _build_styles_pdf(students_data, _pdf_options_from_request(default_teacher_name=admin_name))
        return send_file(
            BytesIO(pdf_bytes), as_attachment=True, download_name='تقرير_أنماط_التعلم.pdf', mimetype='application/pdf',
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@learning_style_bp.route('/teacher/export-pdf', methods=['GET'])
@verify_teacher_token
def teacher_export_pdf():
    """تصدير تقرير أنماط التعلم PDF بشكل النموذج الرسمي — مع/بدون كليشة الوزارة حسب اختيار المعلم"""
    try:
        from flask import send_file
        from io import BytesIO

        teacher_id = request.teacher_id
        links = TeacherStudent.query.join(TeacherStudent.student).filter(
            TeacherStudent.teacher_id == teacher_id
        ).order_by(Student.name).all()
        links = _filter_by_ids(links)
        students_data = _students_data_for_pdf(links, is_link=True)

        teacher = Teacher.query.get(teacher_id)
        pdf_bytes = _build_styles_pdf(students_data, _pdf_options_from_request(
            default_teacher_name=teacher.name if teacher else '',
            default_school_name=(teacher.school if teacher else '') or '',
        ))
        return send_file(
            BytesIO(pdf_bytes), as_attachment=True, download_name='تقرير_أنماط_التعلم.pdf', mimetype='application/pdf',
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

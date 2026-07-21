import os
import logging
import time
from functools import wraps
import uuid
import io # Added for reading/writing file in memory
import json
import pandas as pd # Added for reading Excel/CSV
from datetime import datetime
import random
import copy
import qrcode
import base64
import barcode
from barcode.writer import ImageWriter
from flask import (
    Blueprint, render_template, request, redirect, url_for, flash, current_app,
    send_file, jsonify # Added for sending generated files
)
from flask_login import login_required, current_user
try:
    from src.routes.api import teacher_or_admin_required, format_question
except ImportError:
    from routes.api import teacher_or_admin_required, format_question
from werkzeug.utils import secure_filename
from sqlalchemy import or_, func
from sqlalchemy.exc import IntegrityError, DBAPIError
from sqlalchemy.orm import joinedload, contains_eager
from flask_wtf import FlaskForm # إضافة استيراد FlaskForm

# Import Cloudinary
import cloudinary
import cloudinary.uploader
import cloudinary.api

try:
    from src.extensions import db
except ImportError:  # pragma: no cover
    try:
        from extensions import db
    except ImportError:
        print("Warning: Could not import db from src.extensions or extensions. Trying from main.")
        try:
            from main import db
        except ImportError:
            print("Error: Database object 'db' could not be imported.")
            raise

try:
    from src.models.question import Question, Option, MatchingPair
    from src.models.curriculum import Lesson, Unit, Course
    from src.models.activity import Activity  # استيراد نموذج النشاط
except ImportError:  # pragma: no cover
    try:
        from models.question import Question, Option, MatchingPair
        from models.curriculum import Lesson, Unit, Course
        from models.activity import Activity  # استيراد نموذج النشاط
    except ImportError:
        print("Error: Could not import models.")
        raise


class ExamHeaderSettings(db.Model):
    __tablename__ = 'exam_header_settings'
    id = db.Column(db.Integer, primary_key=True)
    country = db.Column(db.String(255), default='المملكة العربية السعودية')
    ministry = db.Column(db.String(255), default='وزارة التعليم')
    education_department = db.Column(db.String(255), default='الإدارة العامة للتعليم بالمنطقة الشرقية')
    school_name = db.Column(db.String(255), default='مدرسة عبدالرحمن بن القاسم الثانوية')
    subject = db.Column(db.String(255), default='كيمياء 4')
    time = db.Column(db.String(255), default='ثلاث ساعات')
    grade = db.Column(db.String(255), default='ثالث ثانوي')
    total_score = db.Column(db.Integer, default=30)
    checker_name = db.Column(db.String(255), nullable=True)
    reviewer_name = db.Column(db.String(255), nullable=True)
    exam_date = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def __repr__(self):
        return f"<ExamHeaderSettings {self.id}>"


class SavedExam(db.Model):
    """نموذج لحفظ الاختبارات المولدة"""
    __tablename__ = 'saved_exams'
    
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False)
    description = db.Column(db.Text, nullable=True)
    course_id = db.Column(db.Integer, nullable=True)  # بدون ForeignKey لتجنب مشاكل العلاقات
    unit_id = db.Column(db.Integer, nullable=True)
    question_ids = db.Column(db.JSON, nullable=False, default=list)
    questions_count = db.Column(db.Integer, nullable=False, default=0)
    models = db.Column(db.JSON, default=['أ'])
    settings = db.Column(db.JSON, default=dict)
    header_settings = db.Column(db.JSON, default=dict)
    exam_type = db.Column(db.String(50), nullable=True)
    semester = db.Column(db.String(50), nullable=True)
    academic_year = db.Column(db.String(50), nullable=True)
    created_by = db.Column(db.Integer, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)
    
    def get_course_name(self):
        """جلب اسم المنهج"""
        if self.course_id:
            course = Course.query.get(self.course_id)
            return course.name if course else None
        return None
    
    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'description': self.description,
            'course_id': self.course_id,
            'course_name': self.get_course_name(),
            'unit_id': self.unit_id,
            'question_ids': self.question_ids,
            'questions_count': self.questions_count,
            'models': self.models,
            'settings': self.settings,
            'header_settings': self.header_settings,
            'exam_type': self.exam_type,
            'semester': self.semester,
            'academic_year': self.academic_year,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }
    
    def __repr__(self):
        return f"<SavedExam {self.id}: {self.name}>"


# إضافة استيراد نظام الإشعارات
try:
    from src.utils.notification_system import QuestionNotifications, SystemNotifications
    notifications_available = True
except ImportError:  # pragma: no cover
    try:
        from utils.notification_system import QuestionNotifications, SystemNotifications
        notifications_available = True
    except ImportError:
        print("Warning: Could not import notification system for questions")
        QuestionNotifications = None
        SystemNotifications = None
        notifications_available = False

question_bp = Blueprint("question", __name__, template_folder="../templates/question")

# === دوال مساعدة للإشعارات ===

def notify_question_operation(operation_type, lesson_name=None, question_text=None, count=1):
    """
    دالة مساعدة موحدة لإرسال إشعارات عمليات الأسئلة
    دالة مساعدة موحدة لإرسال إشعارات عمليات الأسئلة
    """
    if not notifications_available or not QuestionNotifications:
        return
    
    try:
        if operation_type == "add":
            QuestionNotifications.notify_question_added(
                lesson_name=lesson_name or "درس غير محدد",
                question_text=question_text or "سؤال جديد",
                user_id=current_user.id
            )
        elif operation_type == "update":
            QuestionNotifications.notify_question_updated(
                lesson_name=lesson_name or "درس غير محدد",
                question_text=question_text or "سؤال محدث",
                user_id=current_user.id
            )
        elif operation_type == "delete":
            QuestionNotifications.notify_question_deleted(
                lesson_name=lesson_name or "درس غير محدد",
                user_id=current_user.id
            )
        elif operation_type == "import":
            QuestionNotifications.notify_questions_imported(
                count=count,
                lesson_name=lesson_name or "دروس متعددة",
                user_id=current_user.id
            )
            
        current_app.logger.info(f"Question {operation_type} notification sent successfully")
        
    except Exception as e:
        current_app.logger.error(f"Error sending question {operation_type} notification: {e}")

def notify_system_operation(operation_type, details=None, count=None):
    """
    دالة مساعدة موحدة لإرسال إشعارات عمليات النظام
    """
    if not notifications_available or not SystemNotifications:
        return
    
    try:
        if operation_type == "export":
            SystemNotifications.notify_data_export(
                export_type=details or "بيانات",
                count=count or 0,
                user_id=current_user.id
            )
        elif operation_type == "upload":
            SystemNotifications.notify_file_upload(
                filename=details or "ملف",
                subfolder="questions",
                user_id=current_user.id
            )
        elif operation_type == "api_usage":
            SystemNotifications.notify_api_usage(
                endpoint=details or "unknown",
                user_id=current_user.id
            )
            
        current_app.logger.info(f"System {operation_type} notification sent successfully")
        
    except Exception as e:
        current_app.logger.error(f"Error sending system {operation_type} notification: {e}")

# Allowed extensions for image uploads
ALLOWED_IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "gif"}
# Allowed extensions for question import files
ALLOWED_IMPORT_EXTENSIONS = {"xlsx", "csv"}

# Define expected columns for import template (used in import and download)
# Question Type يحدد نوع كل صف (mcq الافتراضي لو العمود غير موجود — توافق خلفي).
# باقي الأعمدة الجديدة (Correct Answer/Fill Blank .../Left|Right .../Essay Model Answer)
# كل وحدة خاصة بنوع معيّن، اختيارية دائماً بفحص الأعمدة المطلوبة.
EXPECTED_IMPORT_COLUMNS = [
    "Course Name", "Unit Name", "Lesson Name",
    "Question Type", "Question Group",
    "Question Text", "Question Image URL",
    "Option 1 Text", "Option 1 Image URL",
    "Option 2 Text", "Option 2 Image URL",
    "Option 3 Text", "Option 3 Image URL",
    "Option 4 Text", "Option 4 Image URL",
    "Correct Option Number",
    "Correct Answer",
    "Fill Blank Answer", "Fill Blank Alt Answers",
    "Left Text", "Left Image URL", "Right Text", "Right Image URL",
    "Essay Model Answer",
    "Explanation",
    "Difficulty", "Bloom Level", "Video URL", "Video Explanation", "Is Blocked"
]

# الأعمدة الإلزامية بكل الأحوال (بغض النظر عن نوع السؤال) — باقي الأعمدة اختيارية
# وتُفحص حسب نوع كل صف داخل حلقة المعالجة نفسها.
REQUIRED_IMPORT_COLUMNS = ["Course Name", "Unit Name", "Lesson Name"]

# أعمدة خاصة بكل نوع سؤال — تُستخدم لبناء قوالب التنزيل حسب الأنواع المختارة
IMPORT_COLUMNS_BY_TYPE = {
    "mcq": [
        "Question Text", "Question Image URL",
        "Option 1 Text", "Option 1 Image URL",
        "Option 2 Text", "Option 2 Image URL",
        "Option 3 Text", "Option 3 Image URL",
        "Option 4 Text", "Option 4 Image URL",
        "Correct Option Number",
    ],
    "true_false": [
        "Question Text", "Question Image URL", "Correct Answer",
    ],
    "fill_blank": [
        "Question Text", "Question Image URL",
        "Fill Blank Answer", "Fill Blank Alt Answers",
    ],
    "matching": [
        "Question Group", "Question Text",
        "Left Text", "Left Image URL", "Right Text", "Right Image URL",
    ],
    "essay": [
        "Question Text", "Question Image URL", "Essay Model Answer",
    ],
}

IMPORT_COMMON_TRAILING_COLUMNS = [
    "Explanation", "Difficulty", "Bloom Level", "Video URL", "Video Explanation", "Is Blocked"
]

IMPORT_TYPE_LABELS_AR = {
    "mcq": "اختيار من متعدد",
    "true_false": "صح وخطأ",
    "fill_blank": "إكمال فراغ",
    "matching": "مزاوجة",
    "essay": "مقالي",
}

def allowed_image_file(filename):
    return ("." in filename and
            filename.rsplit(".", 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS)

def allowed_import_file(filename):
    return ("." in filename and
            filename.rsplit(".", 1)[1].lower() in ALLOWED_IMPORT_EXTENSIONS)


# === دالة تنسيق النص للطباعة (تحويل الأسطر الجديدة) ===
from markupsafe import Markup, escape

def format_text_for_print(text):
    """
    تنسيق النص للطباعة:
    - تحويل السطور الجديدة (\n) إلى <br>
    - هذا يضمن أن كل سطر يظهر كما أدخله المعلم
    
    ملاحظة: لا نستخدم white-space: nowrap لأنه يسبب خروج النص من العمود
    """
    if not text:
        return ''
    
    # أولاً: escape أي HTML موجود في النص الأصلي (للأمان)
    safe_text = str(escape(text))
    
    # ثانياً: تحويل السطور الجديدة إلى <br>
    formatted = safe_text.replace('\n', '<br>')
    
    # إرجاع كـ Markup لمنع escape مرة أخرى في القالب
    return Markup(formatted)


# أنواع الأسئلة المدعومة حالياً بتوليد ملفات الامتحان (PDF).
# تُوسَّع تدريجياً — راجع خطة دمج الأنواع الجديدة بالامتحان قبل أي تعديل هنا.
EXAM_SUPPORTED_TYPES = ['mcq', 'true_false', 'fill_blank', 'matching', 'essay']

# أنواع تُستبعد دائماً من ورقة التظليل (OMR) — لا توجد فقاعة تظليل ممكنة لإجابة نصية حرة
OMR_EXCLUDED_TYPES = ('fill_blank', 'essay')


def get_ordered_questions(question_ids):
    """
    جلب الأسئلة مرتبة حسب الوحدة ثم الدرس (باستخدام order_num)
    """
    if not question_ids:
        return []
    
    try:
        # جلب الأسئلة مرتبة حسب ترتيب الوحدة ثم الدرس
        questions = Question.query.filter(
            Question.question_id.in_(question_ids),
            Question.question_type.in_(EXAM_SUPPORTED_TYPES)
        ).join(
            Lesson, Question.lesson_id == Lesson.id
        ).join(
            Unit, Lesson.unit_id == Unit.id
        ).order_by(
            Unit.order_num,
            Unit.id,
            Lesson.order_num,
            Lesson.id,
            Question.question_id
        ).all()
        return questions
    except Exception:
        # في حالة فشل الترتيب، نرجع الأسئلة بدون ترتيب
        return Question.query.filter(
            Question.question_id.in_(question_ids),
            Question.question_type.in_(EXAM_SUPPORTED_TYPES)
        ).all()


def get_ordered_questions_for_omr(question_ids):
    """
    مثل get_ordered_questions لكن تستبعد دائماً الأنواع غير المتوافقة مع ورقة التظليل
    (fill_blank/essay — لا توجد فقاعة تظليل ممكنة لإجابة نصية حرة).
    تُستخدم فقط بمسارات ورقة التظليل (OMR)، وليس بمسارات PDF العادي.
    """
    questions = get_ordered_questions(question_ids)
    return [q for q in questions if q.question_type not in OMR_EXCLUDED_TYPES]


def _split_omr_questions_data(questions):
    """
    يقسّم أسئلة OMR (بعد استبعاد fill_blank/essay) لثلاث مجموعات حسب النوع —
    كل نوع له مسبح ترقيم مستقل بورقة التظليل (اختر/ص-خ/مزاوجة)، بدل معاملتها
    كمسبح واحد مختلط (كان يسبب أرقام إجابات خاطئة/فاضية لصح-خطأ والمزاوجة).
    """
    mcq_data, tf_data = [], []
    matching_pairs_flat = []
    for q in questions:
        qtype = getattr(q, 'question_type', 'mcq') or 'mcq'
        if qtype == 'matching':
            for p in (q.matching_pairs or []):
                matching_pairs_flat.append({'pair_id': p.pair_id, 'right_text': p.right_text})
            continue
        q_dict = {
            'question_id': q.question_id,
            'options': [
                {'option_id': getattr(o, 'option_id', None),
                 'option_text': getattr(o, 'option_text', '') or '',
                 'is_correct': getattr(o, 'is_correct', False)}
                for o in q.options
            ],
        }
        if qtype == 'true_false':
            tf_data.append(q_dict)
        else:
            mcq_data.append(q_dict)
    return mcq_data, tf_data, matching_pairs_flat


def _resolve_matching_pairs_flat(matching_pairs_flat, matching_pair_ids):
    """
    عند توليد امتحان بأزواج مزاوجة مُجمَّعة (pooling) أو مُختارة يدوياً، الأزواج
    المستخدمة فعلياً بالامتحان لا تُستنتج من question_ids (لأن سؤال المزاوجة
    الموحّد بورقة الامتحان اصطناعي وليس صفاً حقيقياً بجدول الأسئلة) — لذا إن
    وُصلت matching_pair_ids صراحةً (من نفس استدعاء توليد الامتحان)، تُستخدم هي
    بالضبط بدل الأزواج المُستنتجة تلقائياً من أسئلة مزاوجة قد لا تُطابق الفعلي.
    """
    if not matching_pair_ids:
        return matching_pairs_flat
    pairs = MatchingPair.query.filter(MatchingPair.pair_id.in_(matching_pair_ids)).all()
    return [{'pair_id': p.pair_id, 'right_text': p.right_text} for p in pairs]


def _build_omr_answers(mcq_data, tf_data, matching_pairs_flat, shuffle_questions=True, shuffle_options=True, seed=None):
    """
    يبني قاموس answers الموحّد اللي يتوقعه قالب remark_answer_sheet.html بأوفست ثابت:
    اختر (متعدد) 1-40، ص/خ (صح-خطأ) 41-60، مزاوجة 61-70 — كل نوع بمسبح ترقيم وخلط مستقل.
    """
    letters = ['أ', 'ب', 'ج', 'د', 'هـ', 'و']
    matching_letters = ['أ', 'ب', 'ج', 'د', 'هـ', 'و', 'ز', 'ح', 'ط', 'ي']
    answers = {}

    shuffled_mcq = shuffle_exam(mcq_data, shuffle_questions=shuffle_questions, shuffle_options=shuffle_options, seed=seed)
    for q_num, q in enumerate(shuffled_mcq[:40], 1):
        for i, opt in enumerate(q.get('options', [])):
            if opt.get('is_correct'):
                answers[q_num] = letters[i] if i < len(letters) else str(i + 1)
                break

    tf_seed = (seed + 1) if seed is not None else None
    shuffled_tf = shuffle_exam(tf_data, shuffle_questions=shuffle_questions, shuffle_options=False, seed=tf_seed)
    for q_num, q in enumerate(shuffled_tf[:20], 1):
        for opt in q.get('options', []):
            if opt.get('is_correct'):
                answers[40 + q_num] = 'ص' if opt.get('option_text') == 'صح' else 'خ'
                break

    rng = random.Random(seed + 2) if seed is not None else random
    pairs = list(matching_pairs_flat)
    if shuffle_questions:
        rng.shuffle(pairs)
    pairs = pairs[:10]
    if pairs:
        order = list(range(len(pairs)))
        rng.shuffle(order)
        correct_letter_by_index = {}
        for pos, orig_idx in enumerate(order):
            letter = matching_letters[pos] if pos < len(matching_letters) else str(pos + 1)
            correct_letter_by_index[orig_idx] = letter
        for i in range(len(pairs)):
            answers[60 + i + 1] = correct_letter_by_index.get(i, '')

    return answers


def _build_answer_keys_for_models(question_ids, matching_pair_ids, models, shuffle_options=True, include_answers=True):
    """
    يبني answers (فقاعات) لكل نموذج بالاعتماد على نفس بايبلاين ترتيب/خلط ورقة
    الاختبار الفعلية (src.routes.exam_model_builder) بدل إعادة خلط مستقل —
    عشان مفتاح الريمارك يطابق بالضبط ما طُبع بورقة الاختبار لنفس النموذج
    (كان قبل هذا التصحيح مستقل تماماً بصيغة seed مختلفة وخوارزمية خلط مختلفة،
    فيؤدي لاختلاف الإجابة الصحيحة بين الورقة والمفتاح).

    ترجع: (answer_keys: {model_letter: answers_dict}, questions_count: int)
    """
    from src.routes.exam_model_builder import (
        build_exam_model, build_manual_matching_fq, compute_stable_seed, EXAM_SUPPORTED_TYPES as _EXAM_TYPES
    )

    id_to_q = {
        q.question_id: q for q in Question.query.filter(
            Question.question_id.in_(question_ids),
            Question.is_blocked == False,
            Question.question_type.in_(_EXAM_TYPES)
        ).all()
    }
    available = [id_to_q[qid] for qid in question_ids if qid in id_to_q]
    if matching_pair_ids:
        manual_matching_fq = build_manual_matching_fq(matching_pair_ids)
        if manual_matching_fq:
            available.append(manual_matching_fq)

    stable_seed = compute_stable_seed(question_ids, matching_pair_ids)
    model_letters_canonical = ['أ', 'ب', 'ج', 'د']

    letters = ['أ', 'ب', 'ج', 'د', 'هـ', 'و']
    answer_keys = {}
    questions_count = len([q for q in available if getattr(q, 'question_type', None) != 'matching']) + (
        1 if matching_pair_ids and any(getattr(q, 'question_type', None) == 'matching' for q in available) else 0
    )

    for model_letter in models:
        model_index = model_letters_canonical.index(model_letter) if model_letter in model_letters_canonical else 0
        formatted = build_exam_model(
            available, model_index, stable_seed,
            shuffle_options=shuffle_options, include_answers=include_answers
        )

        answers = {}
        mcq_items = [f for f in formatted if f.get('question_type') == 'mcq']
        for q_num, f in enumerate(mcq_items[:40], 1):
            for i, opt in enumerate(f.get('options', [])):
                if opt.get('is_correct'):
                    answers[q_num] = letters[i] if i < len(letters) else str(i + 1)
                    break

        tf_items = [f for f in formatted if f.get('question_type') == 'true_false']
        for q_num, f in enumerate(tf_items[:20], 1):
            for opt in f.get('options', []):
                if opt.get('is_correct'):
                    answers[40 + q_num] = 'ص' if opt.get('option_text') == 'صح' else 'خ'
                    break

        matching_items = [f for f in formatted if f.get('question_type') == 'matching']
        pair_idx = 0
        for f in matching_items:
            for p in (f.get('matching_pairs') or []):
                pair_idx += 1
                if pair_idx > 10:
                    break
                answers[60 + pair_idx] = p.get('correct_letter', '')

        answer_keys[model_letter] = answers

    return answer_keys, questions_count


# ============================================================
# API خاص بشاشة استخراج الاختبار (export_exam.html) — إداري فقط (login_required)
# منفصل تماماً عن /api/v1/* الذي يستهلكه تطبيق الطالب (الاختبار التفاعلي يبقى MCQ فقط دائماً).
# هذا الـ API يدعم كل الأنواع المذكورة في EXAM_SUPPORTED_TYPES.
# ============================================================

@question_bp.route("/exam-api/lessons/<int:lesson_id>/questions", methods=["GET"])
@teacher_or_admin_required
def exam_api_lesson_questions(lesson_id):
    lesson = Lesson.query.get(lesson_id)
    if not lesson:
        return jsonify({"error": "Lesson not found"}), 404
    is_bank = bool(lesson.unit.course.is_bank) if lesson.unit and lesson.unit.course else False
    questions = (
        Question.query
        .options(joinedload(Question.options))
        .filter(Question.lesson_id == lesson_id)
        .filter(Question.is_blocked == False)
        .filter(Question.is_bank == is_bank)
        .filter(Question.question_type.in_(EXAM_SUPPORTED_TYPES))
        .order_by(Question.question_id)
        .all()
    )
    return jsonify([format_question(q) for q in questions])


@question_bp.route("/exam-api/courses/<int:course_id>/questions", methods=["GET"])
@teacher_or_admin_required
def exam_api_course_questions(course_id):
    course = Course.query.get(course_id)
    if not course:
        return jsonify({"error": "Course not found"}), 404
    show_all = request.args.get('show_all', 'false').lower() == 'true'
    is_bank = bool(course.is_bank)
    query = (
        Question.query
        .join(Question.lesson)
        .join(Lesson.unit)
        .join(Unit.course)
        .options(joinedload(Question.options))
        .filter(Unit.course_id == course_id)
        .filter(Question.is_blocked == False)
        .filter(Question.is_bank == is_bank)
        .filter(Question.question_type.in_(EXAM_SUPPORTED_TYPES))
    )
    if not show_all:
        query = query.filter(Course.show_in_bot == True)
    questions = query.order_by(Question.question_id).all()
    return jsonify([format_question(q) for q in questions])


@question_bp.route("/exam-api/units/<int:unit_id>/questions", methods=["GET"])
@teacher_or_admin_required
def exam_api_unit_questions(unit_id):
    unit = Unit.query.get(unit_id)
    if not unit:
        return jsonify({"error": "Unit not found"}), 404
    show_all = request.args.get('show_all', 'false').lower() == 'true'
    is_bank = bool(unit.course.is_bank) if unit.course else False
    query = (
        Question.query
        .join(Question.lesson)
        .join(Lesson.unit)
        .join(Unit.course)
        .options(joinedload(Question.options))
        .filter(Lesson.unit_id == unit_id)
        .filter(Question.is_blocked == False)
        .filter(Question.is_bank == is_bank)
        .filter(Question.question_type.in_(EXAM_SUPPORTED_TYPES))
    )
    if not show_all:
        query = query.filter(Course.show_in_bot == True)
    questions = query.order_by(Question.question_id).all()
    return jsonify([format_question(q) for q in questions])


@question_bp.route("/exam-api/courses/<int:course_id>/units/<int:unit_id>/questions", methods=["GET"])
@teacher_or_admin_required
def exam_api_course_unit_questions(course_id, unit_id):
    course = Course.query.get(course_id)
    if not course:
        return jsonify({"error": "Course not found"}), 404
    unit = Unit.query.filter_by(id=unit_id, course_id=course_id).first()
    if not unit:
        return jsonify({"error": f"Unit {unit_id} not found in course {course_id}"}), 404
    is_bank = bool(course.is_bank)
    questions = (
        Question.query
        .join(Question.lesson)
        .options(joinedload(Question.options))
        .filter(Lesson.unit_id == unit_id)
        .filter(Question.is_blocked == False)
        .filter(Question.is_bank == is_bank)
        .filter(Question.question_type.in_(EXAM_SUPPORTED_TYPES))
        .order_by(Question.question_id)
        .all()
    )
    return jsonify([format_question(q) for q in questions])


# --- save_upload function (Modified for Cloudinary) --- #
def save_upload(file, subfolder="questions"):
    current_app.logger.debug(f"Entering save_upload for Cloudinary, subfolder: {subfolder}")
    if not file or not file.filename:
        current_app.logger.debug("No file or filename provided to save_upload.")
        return None

    current_app.logger.debug(f"Processing file: {file.filename}")

    if not allowed_image_file(file.filename):
        current_app.logger.warning(f"Image file type not allowed: {file.filename}")
        return None
    
    current_app.logger.debug(f"Image file type allowed for: {file.filename}")

    # Configure Cloudinary (should ideally be done once at app startup)
    # Ensure CLOUDINARY_URL or individual CLOUD_NAME, API_KEY, API_SECRET are set in environment
    cloud_name = os.environ.get("CLOUDINARY_CLOUD_NAME")
    api_key = os.environ.get("CLOUDINARY_API_KEY")
    api_secret = os.environ.get("CLOUDINARY_API_SECRET")

    if not all([cloud_name, api_key, api_secret]):
         current_app.logger.error("Cloudinary environment variables (CLOUD_NAME, API_KEY, API_SECRET) are missing or incomplete.")
         # Attempt to configure from CLOUDINARY_URL as a fallback
         if os.environ.get("CLOUDINARY_URL"):
             current_app.logger.info("Attempting to configure Cloudinary from CLOUDINARY_URL.")
             try:
                 cloudinary.config()
                 current_app.logger.info("Cloudinary configured from URL.")
             except Exception as config_err:
                 current_app.logger.error(f"Failed to configure Cloudinary from URL: {config_err}")
                 return None
         else:
             current_app.logger.error("CLOUDINARY_URL is also missing.")
             return None
    else:
        try:
            cloudinary.config(
                cloud_name=cloud_name,
                api_key=api_key,
                api_secret=api_secret
            )
            current_app.logger.debug("Cloudinary configured from individual variables.")
        except Exception as config_err:
            current_app.logger.error(f"Failed to configure Cloudinary from individual variables: {config_err}")
            return None

    try:
        original_filename = secure_filename(file.filename)
        # Generate a unique public_id using timestamp and UUID
        public_id = f"{subfolder}/{int(time.time())}_{uuid.uuid4().hex[:8]}_{os.path.splitext(original_filename)[0]}"
        current_app.logger.debug(f"Generated Cloudinary public_id: {public_id}")

        current_app.logger.debug(f"Attempting to upload to Cloudinary with public_id: {public_id}")
        
        # Ensure file pointer is at the beginning
        file.seek(0)
        
        upload_result = cloudinary.uploader.upload(
            file.stream, # Pass the file stream
            public_id=public_id,
            folder=subfolder, # Optional: Organize within Cloudinary folders
            resource_type="auto" # Automatically detect resource type (image/video/raw)
        )
        current_app.logger.debug("Cloudinary upload call completed.")

        if upload_result and upload_result.get("secure_url"):
            image_url = upload_result["secure_url"]
            current_app.logger.info(f"File uploaded successfully to Cloudinary: {image_url}")
            
            # === إضافة إشعار رفع الصورة ===
            if notifications_available and SystemNotifications:
                try:
                    SystemNotifications.notify_file_upload(
                        filename=original_filename,
                        subfolder=subfolder,
                        user_id=getattr(current_user, 'id', None)
                    )
                except Exception as e:
                    current_app.logger.error(f"Error sending file upload notification: {e}")
            
            return image_url
        else:
            error_message = upload_result.get("error", {}).get("message", "Unknown error") if upload_result else "No response"
            current_app.logger.error(f"Cloudinary upload failed: {error_message}")
            current_app.logger.debug(f"Cloudinary upload response: {upload_result}")
            return None

    except Exception as e:
        current_app.logger.error(f"Exception during Cloudinary upload process: {e}", exc_info=True)
        return None

# --- list_questions route (keep as is) --- #
@question_bp.route("/")
@login_required
def list_questions():
    current_app.logger.info("Entering list_questions route.")
    
    # استقبال معاملات التصفية من الطلب
    course_id = request.args.get("course_id", type=int)
    unit_id = request.args.get("unit_id", type=int)
    lesson_id = request.args.get("lesson_id", type=int)
    search_q = request.args.get("q", "").strip()
    difficulty = request.args.get("difficulty", "")
    bloom_level = request.args.get("bloom_level", "")
    question_type = request.args.get("question_type", "")
    blocked = request.args.get("blocked", "")  # "1"=محجوبة, "0"=غير محجوبة, ""=الكل
    has_video = request.args.get("has_video", "")        # "1"=فيه فيديو, "0"=ما فيه, ""=الكل
    has_explanation = request.args.get("has_explanation", "")  # "1"=فيه شرح, "0"=ما فيه, ""=الكل
    page = request.args.get("page", 1, type=int)
    per_page = 9
    bank_mode = request.args.get('bank', '0') == '1'

    # auto-detect bank mode: إذا كان المنهج المختار is_bank=True، بدّل التلقائي لوضع البنك
    if course_id and not bank_mode:
        _course_check = Course.query.get(course_id)
        if _course_check and _course_check.is_bank:
            bank_mode = True

    current_app.logger.info(f"Filter parameters: course_id={course_id}, unit_id={unit_id}, lesson_id={lesson_id}, page={page}, bank_mode={bank_mode}")

    try:
        # بناء الاستعلام الأساسي مع ضمان جلب جميع العلاقات
        query = Question.query.options(
            joinedload(Question.options),  # جلب خيارات السؤال
            joinedload(Question.lesson).joinedload(Lesson.unit).joinedload(Unit.course)
        )

        # فلتر بنك الأسئلة
        if bank_mode:
            query = query.filter(Question.is_bank == True)
        else:
            query = query.filter(Question.is_bank == False)

        # إضافة شروط التصفية بطريقة محسنة لتجنب تداخل الـ joins
        if lesson_id:
            # التصفية حسب الدرس - مباشرة بدون join إضافي
            query = query.filter(Question.lesson_id == lesson_id)
            current_app.logger.info(f"Filtering by lesson_id: {lesson_id}")
        elif unit_id:
            # التصفية حسب الوحدة - استخدام exists بدلاً من join
            query = query.filter(Question.lesson.has(Lesson.unit_id == unit_id))
            current_app.logger.info(f"Filtering by unit_id: {unit_id}")
        elif course_id:
            # التصفية حسب المنهج - استخدام exists بدلاً من join متعدد
            query = query.filter(Question.lesson.has(Lesson.unit.has(Unit.course_id == course_id)))
            current_app.logger.info(f"Filtering by course_id: {course_id}")
        
        # فلتر البحث النصي
        if search_q:
            query = query.filter(
                or_(
                    Question.question_text.ilike(f"%{search_q}%"),
                    Question.options.any(Option.option_text.ilike(f"%{search_q}%"))
                )
            )

        # فلتر الصعوبة
        if difficulty:
            query = query.filter(Question.difficulty == difficulty)

        # فلتر مستوى بلوم
        if bloom_level:
            query = query.filter(Question.bloom_level == bloom_level)

        # فلتر نوع السؤال
        if question_type:
            query = query.filter(Question.question_type == question_type)

        # فلتر المحجوبة
        if blocked == "1":
            query = query.filter(Question.is_blocked == True)
        elif blocked == "0":
            query = query.filter(Question.is_blocked == False)

        # فلتر الفيديو
        if has_video == "1":
            query = query.filter(Question.video_url != None, Question.video_url != '')
        elif has_video == "0":
            query = query.filter(or_(Question.video_url == None, Question.video_url == ''))

        # فلتر الشرح
        if has_explanation == "1":
            query = query.filter(Question.explanation != None, Question.explanation != '')
        elif has_explanation == "0":
            query = query.filter(or_(Question.explanation == None, Question.explanation == ''))

        # ترتيب النتائج: لو محدد درس معيّن، رتب حسب تسلسل الدرس (تصاعدي)
        # غير كذا (تصفح عام)، رتب الأحدث إضافة أولاً عشان الأسئلة الجديدة تظهر مباشرة بدون بحث
        if lesson_id:
            query = query.order_by(Question.lesson_id.asc(), Question.question_id.asc())
        else:
            query = query.order_by(Question.question_id.desc())

        questions_pagination = query.paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        current_app.logger.info(f"Database query successful. Found {len(questions_pagination.items)} questions on this page (total: {questions_pagination.total}).")
        
        # التحقق من جلب الخيارات بشكل صحيح
        for question in questions_pagination.items:
            options_count = len(question.options) if question.options else 0
            current_app.logger.info(f"Question {question.question_id} has {options_count} options")
        
        # جلب قوائم الدورات والوحدات والدروس للتصفية
        courses = Course.query.order_by(Course.name).all()
        units = []
        lessons = []
        
        if course_id:
            units = Unit.query.filter_by(course_id=course_id).order_by(Unit.name).all()
            if unit_id:
                lessons = Lesson.query.filter_by(unit_id=unit_id).order_by(Lesson.name).all()
        
        # إحصائيات سريعة
        base_count_query = Question.query.filter(Question.is_bank == bank_mode)
        _bq = base_count_query  # اختصار

        # إحصائيات الفيديو
        q_with_video = _bq.filter(Question.video_url != None, Question.video_url != '').count()

        # إحصائيات الشرح المفصّل (video_explanation) — يُعرض في التطبيق
        q_with_vexpl = _bq.filter(Question.video_explanation != None, Question.video_explanation != '').count()

        # إحصائيات الشرح المختصر (explanation) — يُعرض في التطبيق
        q_with_expl = _bq.filter(Question.explanation != None, Question.explanation != '').count()

        # إجمالي الدروس التي فيها أسئلة
        total_lessons = db.session.query(func.count(func.distinct(Question.lesson_id)))\
            .filter(Question.is_bank == bank_mode).scalar() or 0

        lessons_with_video = db.session.query(func.count(func.distinct(Question.lesson_id)))\
            .filter(Question.is_bank == bank_mode, Question.video_url != None, Question.video_url != '').scalar() or 0

        lessons_with_expl = db.session.query(func.count(func.distinct(Question.lesson_id)))\
            .filter(Question.is_bank == bank_mode, Question.explanation != None, Question.explanation != '').scalar() or 0

        lessons_with_vexpl = db.session.query(func.count(func.distinct(Question.lesson_id)))\
            .filter(Question.is_bank == bank_mode, Question.video_explanation != None, Question.video_explanation != '').scalar() or 0

        # الوحدات
        total_units = db.session.query(func.count(func.distinct(Lesson.unit_id)))\
            .join(Question, Question.lesson_id == Lesson.id)\
            .filter(Question.is_bank == bank_mode).scalar() or 0

        units_with_video = db.session.query(func.count(func.distinct(Lesson.unit_id)))\
            .join(Question, Question.lesson_id == Lesson.id)\
            .filter(Question.is_bank == bank_mode, Question.video_url != None, Question.video_url != '').scalar() or 0

        units_with_expl = db.session.query(func.count(func.distinct(Lesson.unit_id)))\
            .join(Question, Question.lesson_id == Lesson.id)\
            .filter(Question.is_bank == bank_mode, Question.explanation != None, Question.explanation != '').scalar() or 0

        # المناهج
        total_courses = db.session.query(func.count(func.distinct(Unit.course_id)))\
            .join(Lesson, Lesson.unit_id == Unit.id)\
            .join(Question, Question.lesson_id == Lesson.id)\
            .filter(Question.is_bank == bank_mode).scalar() or 0

        courses_with_video = db.session.query(func.count(func.distinct(Unit.course_id)))\
            .join(Lesson, Lesson.unit_id == Unit.id)\
            .join(Question, Question.lesson_id == Lesson.id)\
            .filter(Question.is_bank == bank_mode, Question.video_url != None, Question.video_url != '').scalar() or 0

        courses_with_expl = db.session.query(func.count(func.distinct(Unit.course_id)))\
            .join(Lesson, Lesson.unit_id == Unit.id)\
            .join(Question, Question.lesson_id == Lesson.id)\
            .filter(Question.is_bank == bank_mode, Question.explanation != None, Question.explanation != '').scalar() or 0

        total_regular = Question.query.filter(Question.is_bank == False).count()
        total_bank    = Question.query.filter(Question.is_bank == True).count()

        # إحصائية مراجعة التصنيف
        total_classified  = Question.query.count()
        total_verified    = Question.query.filter(Question.human_verified == True).count()
        verify_pct = round(total_verified / total_classified * 100) if total_classified else 0

        stats = {
            'total_all': _bq.count(),
            'total_blocked': _bq.filter(Question.is_blocked == True).count(),
            'total_filtered': questions_pagination.total,
            'total_regular': total_regular,
            'total_bank': total_bank,
            'total_combined': total_regular + total_bank,
            'total_verified': total_verified,
            'total_classified': total_classified,
            'verify_pct': verify_pct,
        }

        # ====== drill-down stats: منهج → وحدة → درس ======
        def _q_stats(base_q):
            total = base_q.count()
            video = base_q.filter(Question.video_url != None, Question.video_url != '').count()
            expl  = base_q.filter(Question.explanation != None, Question.explanation != '').count()
            vexpl = base_q.filter(Question.video_explanation != None, Question.video_explanation != '').count()
            return total, video, expl, vexpl

        drill_stats  = []
        drill_level  = 'course'
        drill_parent = None

        if unit_id:
            drill_level  = 'lesson'
            drill_parent = Unit.query.get(unit_id)
            for les in Lesson.query.filter_by(unit_id=unit_id).order_by(Lesson.name).all():
                bq = Question.query.filter(Question.is_bank == bank_mode, Question.lesson_id == les.id)
                total, video, expl, vexpl = _q_stats(bq)
                if total == 0:
                    continue
                drill_stats.append({'id': les.id, 'name': les.name,
                                    'total': total, 'video': video, 'expl': expl, 'vexpl': vexpl,
                                    'param': 'lesson_id'})
        elif course_id:
            drill_level  = 'unit'
            drill_parent = Course.query.get(course_id)
            for u in Unit.query.filter_by(course_id=course_id).order_by(Unit.name).all():
                bq = Question.query.filter(Question.is_bank == bank_mode,
                                           Question.lesson.has(Lesson.unit_id == u.id))
                total, video, expl, vexpl = _q_stats(bq)
                if total == 0:
                    continue
                drill_stats.append({'id': u.id, 'name': u.name,
                                    'total': total, 'video': video, 'expl': expl, 'vexpl': vexpl,
                                    'param': 'unit_id'})
        else:
            for crs in Course.query.filter_by(is_bank=bank_mode).order_by(Course.name).all():
                bq = Question.query.filter(Question.is_bank == bank_mode,
                                           Question.lesson.has(
                                               Lesson.unit.has(Unit.course_id == crs.id)))
                total, video, expl, vexpl = _q_stats(bq)
                if total == 0:
                    continue
                drill_stats.append({'id': crs.id, 'name': crs.name,
                                    'total': total, 'video': video, 'expl': expl, 'vexpl': vexpl,
                                    'param': 'course_id'})

        rendered_template = render_template(
            "question/list.html",
            questions=questions_pagination.items,
            pagination=questions_pagination,
            courses=courses,
            units=units,
            lessons=lessons,
            page=page,
            per_page=per_page,
            bank=bank_mode,
            search_q=search_q,
            difficulty=difficulty,
            bloom_level=bloom_level,
            question_type=question_type,
            blocked=blocked,
            has_video=has_video,
            has_explanation=has_explanation,
            stats=stats,
            drill_stats=drill_stats,
            drill_level=drill_level,
            drill_parent=drill_parent,
        )
        
        current_app.logger.info("Template rendering successful.")
        return rendered_template
        
    except Exception as e:
        current_app.logger.exception("Error occurred in list_questions.")
        flash("حدث خطأ غير متوقع أثناء عرض قائمة الأسئلة.", "danger")
        return redirect(url_for("index"))

# --- دوال التصدير المتقدم ---

def apply_filters(query, filters):
    """
    تطبيق التصفيات على الاستعلام
    """
    if not filters:
        return query
    
    for filter_item in filters:
        field = filter_item.get('field')
        operator = filter_item.get('operator')
        value = filter_item.get('value')
        
        if not field or not operator or not value:
            continue
            
        try:
            if field == 'course':
                if operator == 'equals':
                    query = query.join(Question.lesson).join(Lesson.unit).join(Unit.course).filter(Course.name == value)
                elif operator == 'contains':
                    query = query.join(Question.lesson).join(Lesson.unit).join(Unit.course).filter(Course.name.contains(value))
                elif operator == 'starts_with':
                    query = query.join(Question.lesson).join(Lesson.unit).join(Unit.course).filter(Course.name.startswith(value))
                elif operator == 'ends_with':
                    query = query.join(Question.lesson).join(Lessons.unit).join(Unit.course).filter(Course.name.endswith(value))
                    
            elif field == 'unit':
                if operator == 'equals':
                    query = query.join(Question.lesson).join(Lesson.unit).filter(Unit.name == value)
                elif operator == 'contains':
                    query = query.join(Question.lesson).join(Lesson.unit).filter(Unit.name.contains(value))
                elif operator == 'starts_with':
                    query = query.join(Question.lesson).join(Lesson.unit).filter(Unit.name.startswith(value))
                elif operator == 'ends_with':
                    query = query.join(Question.lesson).join(Lesson.unit).filter(Unit.name.endswith(value))
                    
            elif field == 'lesson':
                if operator == 'equals':
                    query = query.join(Question.lesson).filter(Lesson.name == value)
                elif operator == 'contains':
                    query = query.join(Question.lesson).filter(Lesson.name.contains(value))
                elif operator == 'starts_with':
                    query = query.join(Question.lesson).filter(Lesson.name.startswith(value))
                elif operator == 'ends_with':
                    query = query.join(Question.lesson).filter(Lesson.name.endswith(value))
                    
            elif field == 'question_text':
                if operator == 'equals':
                    query = query.filter(Question.question_text == value)
                elif operator == 'contains':
                    query = query.filter(Question.question_text.contains(value))
                elif operator == 'starts_with':
                    query = query.filter(Question.question_text.startswith(value))
                elif operator == 'ends_with':
                    query = query.filter(Question.question_text.endswith(value))

            elif field == 'question_type':
                if operator == 'equals':
                    query = query.filter(Question.question_type == value)

            elif field == 'created_at':
                try:
                    date_value = datetime.strptime(value, '%Y-%m-%d')
                    if operator == 'equals':
                        query = query.filter(db.func.date(Question.created_at) == date_value.date())
                    elif operator == 'greater_than':
                        query = query.filter(Question.created_at > date_value)
                    elif operator == 'less_than':
                        query = query.filter(Question.created_at < date_value)
                except ValueError:
                    current_app.logger.warning(f"Invalid date format for filter: {value}")
                    continue
                    
        except Exception as e:
            current_app.logger.error(f"Error applying filter {field} {operator} {value}: {e}")
            continue
            
    return query

def prepare_template_export_data(filters=None):
    """
    تحضير البيانات للتصدير وفقاً لتنسيق قالب الاستيراد مع إضافة ميزة الشرح
    """
    try:
        # جلب الأسئلة مع التفاصيل
        query = Question.query.options(
            joinedload(Question.options),
            joinedload(Question.lesson).joinedload(Lesson.unit).joinedload(Unit.course)
        )
        
        # تطبيق التصفيات
        if filters:
            query = apply_filters(query, filters)
        
        questions = query.all()
        
        # تحضير البيانات
        data = []
        for question in questions:
            row = {}
            
            # معلومات المنهج والوحدة والدرس
            row['Course Name'] = question.lesson.unit.course.name if question.lesson and question.lesson.unit and question.lesson.unit.course else ''
            row['Unit Name'] = question.lesson.unit.name if question.lesson and question.lesson.unit else ''
            row['Lesson Name'] = question.lesson.name if question.lesson else ''
            
            # نص السؤال وصورته
            row['Question Text'] = question.question_text or ''
            row['Question Image URL'] = question.image_url or ''
            
            # الخيارات مع صورها
            options = list(question.options)
            for i in range(1, 5):  # دعم حتى 4 خيارات
                if i <= len(options):
                    row[f'Option {i} Text'] = options[i-1].option_text or ''
                    row[f'Option {i} Image URL'] = options[i-1].image_url or ''
                else:
                    row[f'Option {i} Text'] = ''
                    row[f'Option {i} Image URL'] = ''
            
            # رقم الخيار الصحيح
            correct_option_number = ''
            for i, option in enumerate(question.options, 1):
                if option.is_correct:
                    correct_option_number = str(i)
                    break
            row['Correct Option Number'] = correct_option_number
            
            # الشرح
            row['Explanation'] = question.explanation or ''

            # حقول إضافية
            row['Difficulty']  = question.difficulty or ''
            row['Bloom Level'] = question.bloom_level or ''
            row['Video URL']          = question.video_url or question.r2_video_url or ''
            row['Video Explanation']  = question.video_explanation or ''
            row['Is Blocked']         = '1' if question.is_blocked else '0'

            data.append(row)
        
        return data
        
    except Exception as e:
        current_app.logger.error(f"Error preparing template export data: {e}")
        return []

def prepare_export_data(data_type, selected_fields, filters=None):
    """
    تحضير البيانات للتصدير بناءً على النوع والحقول المختارة والتصفيات
    """
    try:
        if data_type == 'questions':
            # جلب الأسئلة مع التفاصيل
            query = Question.query.options(
                joinedload(Question.options),
                joinedload(Question.lesson).joinedload(Lesson.unit).joinedload(Unit.course)
            )
            
            # تطبيق التصفيات
            if filters:
                query = apply_filters(query, filters)
            
            questions = query.all()
            
            # تحضير البيانات
            data = []
            for question in questions:
                row = {}
                
                if 'course' in selected_fields:
                    row['المنهج'] = question.lesson.unit.course.name if question.lesson and question.lesson.unit and question.lesson.unit.course else ''
                if 'unit' in selected_fields:
                    row['الوحدة'] = question.lesson.unit.name if question.lesson and question.lesson.unit else ''
                if 'lesson' in selected_fields:
                    row['الدرس'] = question.lesson.name if question.lesson else ''
                if 'question_text' in selected_fields:
                    row['نص السؤال'] = question.question_text or ''
                if 'question_image' in selected_fields:
                    row['صورة السؤال'] = question.image_url or ''
                if 'options' in selected_fields:
                    # فصل الخيارات في أعمدة منفصلة
                    options = list(question.options)
                    for i in range(1, 5):  # دعم حتى 4 خيارات
                        if i <= len(options):
                            row[f'الخيار {i}'] = options[i-1].option_text or ''
                        else:
                            row[f'الخيار {i}'] = ''
                if 'correct_answer' in selected_fields:
                    # إضافة رقم الإجابة الصحيحة ونص الإجابة
                    correct_option_number = ''
                    correct_option_text = ''
                    for i, option in enumerate(question.options, 1):
                        if option.is_correct:
                            correct_option_number = str(i)
                            correct_option_text = option.option_text or ''
                            break
                    row['رقم الإجابة الصحيحة'] = correct_option_number
                    row['الإجابة الصحيحة'] = correct_option_text
                if 'explanation' in selected_fields:
                    row['الشرح'] = question.explanation or ''
                if 'created_at' in selected_fields:
                    row['تاريخ الإنشاء'] = question.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(question, 'created_at') and question.created_at else ''
                
                data.append(row)
                
        elif data_type == 'curriculum':
            # جلب هيكل المنهج
            courses = Course.query.options(
                joinedload(Course.units).joinedload(Unit.lessons)
            ).all()
            
            data = []
            for course in courses:
                if 'course' in selected_fields:
                    course_row = {
                        'النوع': 'منهج',
                        'الاسم': course.name,
                        'المنهج': course.name,
                        'الوحدة': '',
                        'الدرس': '',
                        'عدد الوحدات': len(course.units),
                        'عدد الدروس': sum(len(unit.lessons) for unit in course.units),
                        'عدد الأسئلة': sum(len(lesson.questions) for unit in course.units for lesson in unit.lessons)
                    }
                    data.append(course_row)
                
                for unit in course.units:
                    if 'unit' in selected_fields:
                        unit_row = {
                            'النوع': 'وحدة',
                            'الاسم': unit.name,
                            'المنهج': course.name,
                            'الوحدة': unit.name,
                            'الدرس': '',
                            'عدد الوحدات': '',
                            'عدد الدروس': len(unit.lessons),
                            'عدد الأسئلة': sum(len(lesson.questions) for lesson in unit.lessons)
                        }
                        data.append(unit_row)
                    
                    for lesson in unit.lessons:
                        if 'lesson' in selected_fields:
                            lesson_row = {
                                'النوع': 'درس',
                                'الاسم': lesson.name,
                                'المنهج': course.name,
                                'الوحدة': unit.name,
                                'الدرس': lesson.name,
                                'عدد الوحدات': '',
                                'عدد الدروس': '',
                                'عدد الأسئلة': len(lesson.questions)
                            }
                            data.append(lesson_row)
                            
        else:  # all data
            # جلب جميع البيانات
            query = Question.query.options(
                joinedload(Question.options),
                joinedload(Question.lesson).joinedload(Lesson.unit).joinedload(Unit.course)
            )
            
            # تطبيق التصفيات
            if filters:
                query = apply_filters(query, filters)
            
            questions = query.all()
            
            data = []
            for question in questions:
                row = {}
                
                if 'course' in selected_fields:
                    row['المنهج'] = question.lesson.unit.course.name if question.lesson and question.lesson.unit and question.lesson.unit.course else ''
                if 'unit' in selected_fields:
                    row['الوحدة'] = question.lesson.unit.name if question.lesson and question.lesson.unit else ''
                if 'lesson' in selected_fields:
                    row['الدرس'] = question.lesson.name if question.lesson else ''
                if 'question_text' in selected_fields:
                    row['نص السؤال'] = question.question_text or ''
                if 'question_image' in selected_fields:
                    row['صورة السؤال'] = question.image_url or ''
                if 'options' in selected_fields:
                    # فصل الخيارات في أعمدة منفصلة
                    options = list(question.options)
                    for i in range(1, 5):  # دعم حتى 4 خيارات
                        if i <= len(options):
                            row[f'الخيار {i}'] = options[i-1].option_text or ''
                        else:
                            row[f'الخيار {i}'] = ''
                if 'correct_answer' in selected_fields:
                    # إضافة رقم الإجابة الصحيحة ونص الإجابة
                    correct_option_number = ''
                    correct_option_text = ''
                    for i, option in enumerate(question.options, 1):
                        if option.is_correct:
                            correct_option_number = str(i)
                            correct_option_text = option.option_text or ''
                            break
                    row['رقم الإجابة الصحيحة'] = correct_option_number
                    row['الإجابة الصحيحة'] = correct_option_text
                if 'explanation' in selected_fields:
                    row['الشرح'] = question.explanation or ''
                if 'created_at' in selected_fields:
                    row['تاريخ الإنشاء'] = question.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(question, 'created_at') and question.created_at else ''
                
                data.append(row)
        
        return data
        
    except Exception as e:
        current_app.logger.error(f"Error preparing export data: {e}")
        return []

@question_bp.route("/export/template_format", methods=["POST"])
@login_required
def export_template_format():
    """
    تصدير البيانات بتنسيق قالب الاستيراد مع إضافة ميزة الشرح
    """
    try:
        # استقبال التصفيات
        filter_fields = request.form.getlist('filter_field[]')
        filter_operators = request.form.getlist('filter_operator[]')
        filter_values = request.form.getlist('filter_value[]')
        
        # تحضير قائمة التصفيات
        filters = []
        for i in range(len(filter_fields)):
            if i < len(filter_operators) and i < len(filter_values):
                if filter_fields[i] and filter_operators[i] and filter_values[i]:
                    filters.append({
                        'field': filter_fields[i],
                        'operator': filter_operators[i],
                        'value': filter_values[i]
                    })
        
        # تحضير البيانات بتنسيق القالب
        data = prepare_template_export_data(filters)
        
        if not data:
            flash("لا توجد بيانات للتصدير بناءً على التصفيات المحددة.", "warning")
            return redirect(url_for("dashboard"))
        
        # إنشاء DataFrame
        df = pd.DataFrame(data)
        
        # تحضير اسم الملف
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"questions_template_format_{timestamp}.xlsx"
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Questions')
        output.seek(0)
        
        # تسجيل النشاط
        try:
            activity = Activity(
                user_id=current_user.id,
                action="تصدير البيانات",
                description=f"تم تصدير {len(data)} سؤال بتنسيق قالب الاستيراد مع الشرح",
                timestamp=datetime.now()
            )
            db.session.add(activity)
            db.session.commit()
        except Exception as e:
            current_app.logger.error(f"Error logging template export activity: {e}")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        current_app.logger.error(f"Error in export_template_format: {e}")
        flash("حدث خطأ أثناء تصدير البيانات بتنسيق القالب.", "danger")
        return redirect(url_for("dashboard"))

# إضافة مسارات API لجلب بيانات التصفية الديناميكية
@question_bp.route("/api/filter_options/<field>")
@login_required
def get_filter_options(field):
    """
    جلب خيارات التصفية الديناميكية للمناهج والوحدات والدروس
    """
    try:
        if field == "course":
            courses = Course.query.order_by(Course.name).all()
            options = [{"value": course.name, "text": course.name} for course in courses]
        elif field == "unit":
            course_name = request.args.get("course")
            if course_name:
                units = Unit.query.join(Course).filter(Course.name == course_name).order_by(Unit.name).all()
            else:
                units = Unit.query.order_by(Unit.name).all()
            options = [{"value": unit.name, "text": unit.name} for unit in units]
        elif field == "lesson":
            unit_name = request.args.get("unit")
            course_name = request.args.get("course")
            if unit_name and course_name:
                lessons = Lesson.query.join(Unit).join(Course).filter(
                    Unit.name == unit_name, Course.name == course_name
                ).order_by(Lesson.name).all()
            elif unit_name:
                lessons = Lesson.query.join(Unit).filter(Unit.name == unit_name).order_by(Lesson.name).all()
            else:
                lessons = Lesson.query.order_by(Lesson.name).all()
            options = [{"value": lesson.name, "text": lesson.name} for lesson in lessons]
        else:
            options = []
            
        return jsonify({"success": True, "options": options})
        
    except Exception as e:
        current_app.logger.error(f"Error getting filter options for {field}: {e}")
        return jsonify({"success": False, "error": str(e)})

@question_bp.route("/export/filtered_data", methods=["POST"])
@login_required
def export_filtered_data():
    """
    تصدير البيانات المفلترة مع الحقول المختارة
    """
    try:
        # استقبال البيانات من النموذج
        data_type = request.form.get('data_type', 'all')
        selected_fields = request.form.getlist('fields')
        export_format = request.form.get('format', 'xlsx')
        
        # استقبال التصفيات
        filter_fields = request.form.getlist('filter_field[]')
        filter_operators = request.form.getlist('filter_operator[]')
        filter_values = request.form.getlist('filter_value[]')
        
        # تحضير قائمة التصفيات
        filters = []
        for i in range(len(filter_fields)):
            if i < len(filter_operators) and i < len(filter_values):
                if filter_fields[i] and filter_operators[i] and filter_values[i]:
                    filters.append({
                        'field': filter_fields[i],
                        'operator': filter_operators[i],
                        'value': filter_values[i]
                    })
        
        # التحقق من وجود حقول مختارة
        if not selected_fields:
            flash("يجب اختيار حقل واحد على الأقل للتصدير.", "warning")
            return redirect(url_for("dashboard"))
        
        # تحضير البيانات
        data = prepare_export_data(data_type, selected_fields, filters)
        
        if not data:
            flash("لا توجد بيانات للتصدير بناءً على التصفيات المحددة.", "warning")
            return redirect(url_for("dashboard"))
        
        # إنشاء DataFrame
        df = pd.DataFrame(data)
        
        # تحضير اسم الملف
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename_base = f"export_{data_type}_{timestamp}"
        
        # إنشاء الملف بناءً على التنسيق المطلوب
        if export_format == 'xlsx':
            filename = f"{filename_base}.xlsx"
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='البيانات')
            output.seek(0)
            
            # تسجيل النشاط
            try:
                activity = Activity(
                    user_id=current_user.id,
                    action="تصدير البيانات",
                    description=f"تم تصدير {len(data)} عنصر من نوع {data_type} بصيغة Excel",
                    timestamp=datetime.now()
                )
                db.session.add(activity)
                db.session.commit()
            except Exception as e:
                current_app.logger.error(f"Error logging export activity: {e}")
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        elif export_format == 'csv':
            filename = f"{filename_base}.csv"
            output = io.StringIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')
            output.seek(0)
            
            # تحويل إلى BytesIO
            bytes_output = io.BytesIO()
            bytes_output.write(output.getvalue().encode('utf-8-sig'))
            bytes_output.seek(0)
            
            # تسجيل النشاط
            try:
                activity = Activity(
                    user_id=current_user.id,
                    action="تصدير البيانات",
                    description=f"تم تصدير {len(data)} عنصر من نوع {data_type} بصيغة CSV",
                    timestamp=datetime.now()
                )
                db.session.add(activity)
                db.session.commit()
            except Exception as e:
                current_app.logger.error(f"Error logging export activity: {e}")
            
            return send_file(
                bytes_output,
                mimetype='text/csv',
                as_attachment=True,
                download_name=filename
            )
            
        elif export_format == 'pdf':
            # للـ PDF، سنحتاج مكتبة إضافية مثل reportlab
            try:
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.lib import colors
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont
                
                filename = f"{filename_base}.pdf"
                output = io.BytesIO()
                
                # إنشاء مستند PDF
                doc = SimpleDocTemplate(output, pagesize=A4)
                elements = []
                
                # إضافة عنوان
                styles = getSampleStyleSheet()
                title = Paragraph(f"تقرير {data_type}", styles['Title'])
                elements.append(title)
                
                # تحويل البيانات إلى جدول
                table_data = [list(df.columns)]  # العناوين
                for _, row in df.iterrows():
                    table_data.append([str(cell) for cell in row])
                
                # إنشاء الجدول
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(table)
                doc.build(elements)
                output.seek(0)
                
                # تسجيل النشاط
                try:
                    activity = Activity(
                        user_id=current_user.id,
                        action="تصدير البيانات",
                        description=f"تم تصدير {len(data)} عنصر من نوع {data_type} بصيغة PDF",
                        timestamp=datetime.now()
                    )
                    db.session.add(activity)
                    db.session.commit()
                except Exception as e:
                    current_app.logger.error(f"Error logging export activity: {e}")
                
                return send_file(
                    output,
                    mimetype='application/pdf',
                    as_attachment=True,
                    download_name=filename
                )
                
            except ImportError:
                flash("مكتبة PDF غير متوفرة. يرجى استخدام Excel أو CSV.", "warning")
                return redirect(url_for("dashboard"))
            except Exception as e:
                current_app.logger.error(f"Error creating PDF: {e}")
                flash("حدث خطأ أثناء إنشاء ملف PDF.", "danger")
                return redirect(url_for("dashboard"))
        
        else:
            flash("تنسيق التصدير غير مدعوم.", "danger")
            return redirect(url_for("dashboard"))
            
    except Exception as e:
        current_app.logger.error(f"Error in export_filtered_data: {e}")
        flash("حدث خطأ أثناء تصدير البيانات.", "danger")
        return redirect(url_for("dashboard"))

# --- get_sorted_lessons function (keep as is) --- #
def get_sorted_lessons():
    try:
        lessons = (
            Lesson.query
            .join(Lesson.unit)
            .join(Unit.course)
            .options(
                contains_eager(Lesson.unit).contains_eager(Unit.course)
            )
            .order_by(Course.is_bank, Course.name, Unit.name, Lesson.name)
            .all()
        )
        return lessons
    except Exception as e:
        current_app.logger.exception("Error fetching sorted lessons.")
        return []

# --- add_question route (Uses modified save_upload) --- #
@question_bp.route("/add", methods=["GET", "POST"])
@login_required
def add_question():
    lessons = get_sorted_lessons()
    if not lessons:
        flash("حدث خطأ أثناء تحميل قائمة الدروس أو لا توجد دروس متاحة. الرجاء إضافة المناهج أولاً.", "warning")
        return redirect(url_for("curriculum.list_courses"))

    # إنشاء نموذج فارغ لتوفير رمز CSRF
    form = FlaskForm()

    if request.method == "POST":
        current_app.logger.info("POST request received for add_question.")
        question_text = request.form.get("text", "").strip()
        lesson_id = request.form.get("lesson_id")
        question_type = request.form.get("question_type", "mcq").strip() or "mcq"
        if question_type not in ("mcq", "true_false", "fill_blank", "matching", "essay"):
            question_type = "mcq"
        if question_type == "matching" and not question_text:
            question_text = "اربط العمود أ بما يناسبه من العمود ب"
        correct_option_index_str = request.form.get("correct_option")
        q_image_file = request.files.get("question_image")

        q_image_path = None
        if q_image_file and q_image_file.filename:
             if not allowed_image_file(q_image_file.filename):
                 flash("نوع ملف صورة السؤال غير مسموح به.", "danger")
             else:
                 # Uses the Cloudinary-compatible save_upload function
                 q_image_path = save_upload(q_image_file, subfolder="questions")
                 if q_image_path is None:
                     flash("فشل رفع صورة السؤال. تحقق من إعدادات Cloudinary والسجلات.", "danger")

        error_messages = []
        if question_type != "matching" and not question_text and not q_image_path:
            error_messages.append("يجب توفير نص للسؤال أو رفع صورة له.")
        if not lesson_id:
            error_messages.append("يجب اختيار درس.")

        options_data_from_form = []
        matching_pairs_data = []
        fill_blank_answer = None
        fill_blank_alt_answers = None
        essay_model_answer = None
        max_submitted_index = -1

        if question_type == "mcq":
            option_keys_check = [key for key in request.form if key.startswith("option_text_")]
            option_files_check = [key for key in request.files if key.startswith("option_image_")]
            if (option_keys_check or option_files_check) and correct_option_index_str is None:
                error_messages.append("يجب تحديد الإجابة الصحيحة.")

            correct_option_index = -1
            if correct_option_index_str is not None:
                try:
                    correct_option_index = int(correct_option_index_str)
                    if correct_option_index < 0:
                         error_messages.append("اختيار الإجابة الصحيحة غير صالح.")
                except ValueError:
                    error_messages.append("اختيار الإجابة الصحيحة يجب أن يكون رقمًا.")

            for key in list(request.form.keys()) + list(request.files.keys()):
                if key.startswith(("option_text_", "option_image_")):
                    try:
                        index_str = key.split("_")[-1]
                        max_submitted_index = max(max_submitted_index, int(index_str))
                    except (ValueError, IndexError):
                        continue

            for i in range(max_submitted_index + 1):
                index_str = str(i)
                option_text = request.form.get(f"option_text_{index_str}", "").strip()
                option_image_file = request.files.get(f"option_image_{index_str}")
                option_image_path = None

                if option_image_file and option_image_file.filename:
                    if not allowed_image_file(option_image_file.filename):
                        error_messages.append(f"نوع ملف صورة الخيار رقم {i+1} غير مسموح به.")
                    else:
                        # Uses the Cloudinary-compatible save_upload function
                        option_image_path = save_upload(option_image_file, subfolder="options")
                        if option_image_path is None:
                            error_messages.append(f"فشل رفع صورة الخيار رقم {i+1}. تحقق من إعدادات Cloudinary والسجلات.")

                if option_text or option_image_path:
                    is_correct = (i == correct_option_index)
                    options_data_from_form.append({
                        "index": i,
                        "option_text": option_text,
                        "image_url": option_image_path,
                        "is_correct": is_correct
                    })

            if len(options_data_from_form) < 2:
                error_messages.append("يجب إضافة خيارين صالحين على الأقل (بنص أو صورة).")

            if correct_option_index_str is not None and correct_option_index >= len(options_data_from_form):
                 error_messages.append("الخيار المحدد كصحيح غير موجود أو غير صالح.")

        elif question_type == "true_false":
            tf_answer = request.form.get("tf_answer")
            if tf_answer not in ("true", "false"):
                error_messages.append("يجب اختيار الإجابة الصحيحة (صح أو خطأ).")
            else:
                options_data_from_form = [
                    {"index": 0, "option_text": "صح", "image_url": None, "is_correct": tf_answer == "true"},
                    {"index": 1, "option_text": "خطأ", "image_url": None, "is_correct": tf_answer == "false"},
                ]

        elif question_type == "fill_blank":
            fill_blank_answer = request.form.get("fill_blank_answer", "").strip()
            if not fill_blank_answer:
                error_messages.append("يجب إدخال الإجابة الصحيحة لسؤال إكمال الفراغ.")
            alt_raw = request.form.get("fill_blank_alt_answers", "").strip()
            if alt_raw:
                fill_blank_alt_answers = [a.strip() for a in alt_raw.split(",") if a.strip()]

        elif question_type == "matching":
            max_pair_index = -1
            for key in list(request.form.keys()) + list(request.files.keys()):
                if key.startswith(("left_text_", "right_text_", "left_image_", "right_image_")):
                    try:
                        max_pair_index = max(max_pair_index, int(key.split("_")[-1]))
                    except (ValueError, IndexError):
                        continue

            for i in range(max_pair_index + 1):
                idx_str = str(i)
                left_text = request.form.get(f"left_text_{idx_str}", "").strip()
                right_text = request.form.get(f"right_text_{idx_str}", "").strip()
                left_image_file = request.files.get(f"left_image_{idx_str}")
                right_image_file = request.files.get(f"right_image_{idx_str}")
                left_image_path = None
                right_image_path = None

                if left_image_file and left_image_file.filename:
                    if not allowed_image_file(left_image_file.filename):
                        error_messages.append(f"نوع ملف صورة العمود ب رقم {i+1} غير مسموح به.")
                    else:
                        left_image_path = save_upload(left_image_file, subfolder="matching")
                if right_image_file and right_image_file.filename:
                    if not allowed_image_file(right_image_file.filename):
                        error_messages.append(f"نوع ملف صورة العمود أ رقم {i+1} غير مسموح به.")
                    else:
                        right_image_path = save_upload(right_image_file, subfolder="matching")

                if left_text or left_image_path or right_text or right_image_path:
                    matching_pairs_data.append({
                        "order_num": i,
                        "left_text": left_text or None,
                        "left_image_url": left_image_path,
                        "right_text": right_text or None,
                        "right_image_url": right_image_path,
                    })

            if len(matching_pairs_data) < 2:
                error_messages.append("يجب إضافة زوجين على الأقل لسؤال المزاوجة.")

        elif question_type == "essay":
            essay_model_answer = request.form.get("essay_model_answer", "").strip() or None

        if error_messages:
            for error in error_messages:
                flash(error, "danger")
            form_data = request.form.to_dict()
            if question_type == "mcq":
                repop_options = []
                for i in range(max_submitted_index + 1):
                     idx_str = str(i)
                     opt_text = request.form.get(f"option_text_{idx_str}", "")
                     processed_opt = next((opt for opt in options_data_from_form if opt["index"] == i), None)
                     img_url = processed_opt["image_url"] if processed_opt else None
                     repop_options.append({"option_text": opt_text, "image_url": img_url})
                form_data["options_repop"] = repop_options
                form_data["correct_option_repop"] = correct_option_index_str
            form_data["question_type"] = question_type
            form_data["question_image_url_repop"] = q_image_path
            return render_template("question/form.html", title="إضافة سؤال جديد", lessons=lessons, question=form_data, submit_text="إضافة سؤال", form=form)

        try:
            # فحص التكرار — نص + درس + إجابة صحيحة (خاص بـ MCQ فقط، لأن نفس النص ممكن يكون سؤالين مختلفين)
            if question_text and question_type == "mcq":
                correct_answer_text = next(
                    (o["option_text"] for o in options_data_from_form if o["is_correct"]), None
                )
                existing_qs = Question.query.filter_by(
                    question_text=question_text,
                    lesson_id=lesson_id
                ).all()
                is_duplicate = any(
                    Option.query.filter_by(
                        question_id=q.question_id,
                        is_correct=True,
                        option_text=correct_answer_text
                    ).first()
                    for q in existing_qs
                )
                if is_duplicate:
                    flash("هذا السؤال موجود بالفعل في هذا الدرس.", "warning")
                    return redirect(url_for("question.add_question"))

            # هل الدرس تابع لمنهج بنك؟
            lesson_obj = Lesson.query.get(lesson_id)
            auto_is_bank = False
            if lesson_obj and lesson_obj.unit and lesson_obj.unit.course:
                auto_is_bank = bool(lesson_obj.unit.course.is_bank)

            new_question = Question(
                question_text=question_text if question_text else None,
                lesson_id=lesson_id,
                image_url=q_image_path,
                explanation=request.form.get("explanation", "").strip() or None,
                video_explanation=request.form.get("video_explanation", "").strip() or None,
                explanation_image_path=None,
                is_blocked=(request.form.get("is_blocked") == "1"),  # معالجة حقل منع السؤال
                is_bank=auto_is_bank,
                question_type=question_type,
                fill_blank_answer=fill_blank_answer,
                fill_blank_alt_answers=fill_blank_alt_answers,
                essay_model_answer=essay_model_answer,
            )
            db.session.add(new_question)
            db.session.flush()
            current_app.logger.info(f"New question added (pending commit) with ID: {new_question.question_id}")

            for opt_data in options_data_from_form:
                # --- Logic to set option_text to image_url if option_text is empty and image_url exists ---
                option_text_to_save = opt_data["option_text"]
                if not option_text_to_save and opt_data["image_url"]:
                    option_text_to_save = opt_data["image_url"] # Set option_text to the image_url
                elif not option_text_to_save: # If option_text is still empty (and no image_url or image_url was not used)
                    option_text_to_save = None

                option = Option(
                    option_text=option_text_to_save,
                    image_url=opt_data["image_url"],
                    is_correct=opt_data["is_correct"],
                    question_id=new_question.question_id
                )
                db.session.add(option)

            for pair_data in matching_pairs_data:
                pair = MatchingPair(
                    left_text=pair_data["left_text"],
                    left_image_url=pair_data["left_image_url"],
                    right_text=pair_data["right_text"],
                    right_image_url=pair_data["right_image_url"],
                    order_num=pair_data["order_num"],
                    question_id=new_question.question_id
                )
                db.session.add(pair)

            db.session.commit()
            current_app.logger.info("Transaction committed successfully. Question and options saved.")
            flash("تمت إضافة السؤال بنجاح!", "success")
            
            # تسجيل النشاط بعد إضافة السؤال بنجاح
            try:
                # الحصول على معلومات الدرس والوحدة والدورة
                lesson = Lesson.query.get(lesson_id)
                lesson_name = lesson.name if lesson else None
                unit_name = lesson.unit.name if lesson and lesson.unit else None
                course_name = lesson.unit.course.name if lesson and lesson.unit and lesson.unit.course else None
                
                # تسجيل النشاط
                Activity.log_activity(
                    action_type="create",
                    entity_type="question",
                    entity_id=new_question.question_id,
                    description=f"تمت إضافة سؤال جديد في الدرس: {lesson_name}",
                    lesson_name=lesson_name,
                    unit_name=unit_name,
                    course_name=course_name,
                    user_id=current_user.id if current_user.is_authenticated else None
                )
            except Exception as activity_error:
                current_app.logger.error(f"Error logging activity: {activity_error}")
                # لا نريد أن يؤثر خطأ تسجيل النشاط على تدفق العملية الأساسية

            return redirect(url_for("question.list_questions", lesson_id=lesson_id))

        except (IntegrityError, DBAPIError) as db_err:
            db.session.rollback()
            current_app.logger.exception("Database error during question creation.")
            flash("حدث خطأ في قاعدة البيانات أثناء إضافة السؤال.", "danger")
            return render_template("question/form.html", title="إضافة سؤال جديد", lessons=lessons, question=request.form, submit_text="إضافة سؤال", form=form)
        
        except Exception as e:
            db.session.rollback()
            current_app.logger.exception("Unexpected error during question creation.")
            flash("حدث خطأ غير متوقع أثناء إضافة السؤال.", "danger")
            return render_template("question/form.html", title="إضافة سؤال جديد", lessons=lessons, question=request.form, submit_text="إضافة سؤال", form=form)

    # GET request
    return render_template("question/form.html", title="إضافة سؤال جديد", lessons=lessons, submit_text="إضافة سؤال", form=form)

# --- START: Import Questions Route (Modified for Cloudinary) --- #
@question_bp.route("/import", methods=["GET", "POST"])
@login_required
def import_questions():
    """Import questions from Excel or CSV file."""
    current_app.logger.info("Entering import_questions route.")
    lessons = get_sorted_lessons()
    if not lessons:
        flash("حدث خطأ أثناء تحميل قائمة الدروس أو لا توجد دروس متاحة. الرجاء إضافة المناهج أولاً.", "warning")
        return redirect(url_for("curriculum.list_courses"))

    # إنشاء نموذج فارغ لتوفير رمز CSRF
    form = FlaskForm()

    if request.method == "POST":
        current_app.logger.info("POST request received for import_questions.")
        
        # Debug logging for request data
        current_app.logger.debug(f"Form data: {request.form}")
        current_app.logger.debug(f"Files: {request.files}")
        
        lesson_id  = request.form.get("lesson_id")
        course_id  = request.form.get("course_id_temp")   # المنهج المحدد من الـ dropdown
        
        file = request.files.get("question_file")
        if not file or not file.filename:
            flash("الرجاء اختيار ملف للاستيراد.", "danger")
            current_app.logger.warning("No file provided in import form.")
            return render_template("question/import_questions.html", lessons=lessons, selected_lesson_id=lesson_id, form=form)
        
        if not allowed_import_file(file.filename):
            flash("نوع الملف غير مسموح به. يرجى استخدام ملف Excel (.xlsx) أو CSV (.csv).", "danger")
            current_app.logger.warning(f"File type not allowed: {file.filename}")
            return render_template("question/import_questions.html", lessons=lessons, selected_lesson_id=lesson_id, form=form)
        
        # Process the file
        try:
            # Read the file into a pandas DataFrame
            if file.filename.endswith('.xlsx'):
                df = pd.read_excel(file)
            else:  # CSV
                df = pd.read_csv(file)
            
            current_app.logger.info(f"File read successfully. Shape: {df.shape}")
            current_app.logger.debug(f"Columns in file: {df.columns.tolist()}")
            
            # الأعمدة الإلزامية بغض النظر عن النوع — باقي الأعمدة اختيارية وتُفحص
            # حسب نوع كل صف (Question Type) داخل حلقة المعالجة نفسها.
            missing_columns = [col for col in REQUIRED_IMPORT_COLUMNS if col not in df.columns]
            if missing_columns:
                flash(f"الملف يفتقد إلى الأعمدة التالية: {', '.join(missing_columns)}", "danger")
                current_app.logger.warning(f"Missing columns in import file: {missing_columns}")
                return render_template("question/import_questions.html", lessons=lessons, selected_lesson_id=lesson_id, form=form)

            # Process each row
            imported_count = 0
            skipped_count = 0
            error_details = []
            matching_groups = {}  # (lesson_id, group_key) -> [{"index":, "row":, "import_is_bank":}, ...]

            for index, row in df.iterrows():
                try:
                    # Extract course, unit, and lesson names
                    course_name = row["Course Name"] if pd.notna(row.get("Course Name")) else None
                    unit_name = row["Unit Name"] if pd.notna(row.get("Unit Name")) else None
                    lesson_name = row["Lesson Name"] if pd.notna(row.get("Lesson Name")) else None

                    # Validate course, unit, and lesson names
                    if not course_name or not unit_name or not lesson_name:
                        error_details.append(f"صف {index+2}: يجب توفير اسم المنهج والوحدة والدرس.")
                        continue

                    # Find the lesson by course, unit, and lesson names
                    # لو المستخدم اختار منهج محدد من الـ dropdown → نفلتر بـ course_id عشان نتجنب تضارب الأسماء
                    lesson_q = Lesson.query.join(Unit).join(Course).filter(
                        Unit.name == unit_name,
                        Lesson.name == lesson_name
                    )
                    if course_id:
                        lesson_q = lesson_q.filter(Course.id == int(course_id))
                    else:
                        lesson_q = lesson_q.filter(Course.name == course_name)
                    lesson = lesson_q.first()

                    if not lesson:
                        error_details.append(f"صف {index+2}: لم يتم العثور على الدرس '{lesson_name}' في الوحدة '{unit_name}' في المنهج '{course_name}'.")
                        continue

                    current_lesson_id = lesson.id
                    # هل الدرس تابع لمنهج بنك؟
                    import_is_bank = False
                    if lesson.unit and lesson.unit.course:
                        import_is_bank = bool(lesson.unit.course.is_bank)

                    # نوع السؤال — mcq افتراضياً لو العمود غير موجود/فاضي (توافق خلفي مع ملفات قديمة)
                    raw_type = row.get("Question Type")
                    question_type = str(raw_type).strip().lower() if pd.notna(raw_type) else "mcq"
                    if question_type not in ("mcq", "true_false", "fill_blank", "matching", "essay"):
                        question_type = "mcq"

                    # المزاوجة تُجمَّع أولاً (كل صف = زوج)، تُعالج كمجموعات بعد الحلقة
                    if question_type == "matching":
                        raw_group = row.get("Question Group")
                        group_key = str(raw_group).strip() if pd.notna(raw_group) else None
                        if not group_key:
                            error_details.append(f"صف {index+2}: سؤال المزاوجة يحتاج قيمة بعمود Question Group.")
                            continue
                        matching_groups.setdefault((current_lesson_id, group_key), []).append({
                            "index": index, "row": row, "import_is_bank": import_is_bank,
                        })
                        continue

                    # الأنواع غير-matching: نص/صورة السؤال إلزامية للجميع
                    question_text = row["Question Text"] if pd.notna(row.get("Question Text")) else None
                    question_image_url = row["Question Image URL"] if pd.notna(row.get("Question Image URL")) else None

                    if not question_text and not question_image_url:
                        error_details.append(f"صف {index+2}: يجب توفير نص السؤال أو صورة له.")
                        continue

                    options_data = []
                    fill_blank_answer = None
                    fill_blank_alt_answers = None
                    essay_model_answer = None

                    if question_type == "mcq":
                        valid_options_count = 0
                        correct_option_number = None

                        if pd.notna(row.get("Correct Option Number")):
                            try:
                                correct_option_number = int(row["Correct Option Number"])
                                if correct_option_number < 1 or correct_option_number > 4:
                                    error_details.append(f"صف {index+2}: رقم الإجابة الصحيحة يجب أن يكون بين 1 و 4.")
                                    continue
                            except (ValueError, TypeError):
                                error_details.append(f"صف {index+2}: رقم الإجابة الصحيحة يجب أن يكون رقمًا صحيحًا.")
                                continue
                        else:
                            error_details.append(f"صف {index+2}: يجب تحديد رقم الإجابة الصحيحة.")
                            continue

                        for i in range(1, 5):
                            option_text = row[f"Option {i} Text"] if pd.notna(row.get(f"Option {i} Text")) else None
                            option_image_url = row[f"Option {i} Image URL"] if pd.notna(row.get(f"Option {i} Image URL")) else None

                            if option_text or option_image_url:
                                valid_options_count += 1
                                options_data.append({
                                    "option_text": option_text,
                                    "image_url": option_image_url,
                                    "is_correct": (i == correct_option_number)
                                })

                        if valid_options_count < 2:
                            error_details.append(f"صف {index+2}: يجب توفير خيارين صالحين على الأقل.")
                            continue

                        if correct_option_number > valid_options_count:
                            error_details.append(f"صف {index+2}: رقم الإجابة الصحيحة يشير إلى خيار غير موجود.")
                            continue

                    elif question_type == "true_false":
                        raw_answer = row.get("Correct Answer")
                        answer_str = str(raw_answer).strip().lower() if pd.notna(raw_answer) else ""
                        if answer_str in ("true", "صح", "1", "yes", "t"):
                            tf_correct = True
                        elif answer_str in ("false", "خطأ", "خطا", "0", "no", "f"):
                            tf_correct = False
                        else:
                            error_details.append(f"صف {index+2}: قيمة Correct Answer يجب أن تكون true/false أو صح/خطأ.")
                            continue
                        options_data = [
                            {"option_text": "صح", "image_url": None, "is_correct": tf_correct},
                            {"option_text": "خطأ", "image_url": None, "is_correct": not tf_correct},
                        ]

                    elif question_type == "fill_blank":
                        raw_fb = row.get("Fill Blank Answer")
                        fill_blank_answer = str(raw_fb).strip() if pd.notna(raw_fb) else None
                        if not fill_blank_answer:
                            error_details.append(f"صف {index+2}: يجب إدخال الإجابة الصحيحة لسؤال إكمال الفراغ (Fill Blank Answer).")
                            continue
                        raw_alt = row.get("Fill Blank Alt Answers")
                        if pd.notna(raw_alt) and str(raw_alt).strip():
                            fill_blank_alt_answers = [a.strip() for a in str(raw_alt).split(",") if a.strip()]

                    elif question_type == "essay":
                        raw_essay = row.get("Essay Model Answer")
                        essay_model_answer = str(raw_essay).strip() if pd.notna(raw_essay) else None

                    # Extract optional fields
                    explanation  = row["Explanation"]  if pd.notna(row.get("Explanation"))  else None
                    difficulty   = row["Difficulty"]   if pd.notna(row.get("Difficulty"))   else None
                    bloom_level  = row["Bloom Level"]  if pd.notna(row.get("Bloom Level"))  else None
                    video_url         = row["Video URL"]         if pd.notna(row.get("Video URL"))         else None
                    video_explanation = row["Video Explanation"] if pd.notna(row.get("Video Explanation")) else None
                    is_blocked        = str(row.get("Is Blocked", "0")).strip() == "1"

                    # فحص التكرار — نص + درس + إجابة صحيحة (mcq فقط، مطابقة سلوك الإضافة اليدوية)
                    if question_type == "mcq" and question_text:
                        correct_opt_text = next(
                            (o["option_text"] for o in options_data if o["is_correct"]), None
                        )
                        existing_qs = Question.query.filter_by(
                            question_text=question_text,
                            lesson_id=current_lesson_id
                        ).all()
                        is_duplicate = any(
                            Option.query.filter_by(
                                question_id=q.question_id,
                                is_correct=True,
                                option_text=correct_opt_text
                            ).first()
                            for q in existing_qs
                        )
                        if is_duplicate:
                            skipped_count += 1
                            continue

                    # Create question
                    new_question = Question(
                        question_text=question_text,
                        lesson_id=current_lesson_id,
                        image_url=question_image_url,
                        explanation=explanation,
                        difficulty=difficulty,
                        bloom_level=bloom_level,
                        video_url=video_url,
                        video_explanation=video_explanation,
                        is_blocked=is_blocked,
                        is_bank=import_is_bank,
                        question_type=question_type,
                        fill_blank_answer=fill_blank_answer,
                        fill_blank_alt_answers=fill_blank_alt_answers,
                        essay_model_answer=essay_model_answer,
                    )
                    db.session.add(new_question)
                    db.session.flush()  # Get the question ID

                    # Create options (mcq/true_false فقط — باقي الأنواع بدون Option rows)
                    for opt_data in options_data:
                        option = Option(
                            option_text=opt_data["option_text"],
                            image_url=opt_data["image_url"],
                            is_correct=opt_data["is_correct"],
                            question_id=new_question.question_id
                        )
                        db.session.add(option)

                    imported_count += 1

                except Exception as row_error:
                    error_details.append(f"صف {index+2}: {str(row_error)}")
                    current_app.logger.exception(f"Error processing row {index+2}: {row_error}")

            # ── معالجة مجموعات المزاوجة (كل مجموعة = سؤال واحد بعدة أزواج) ──────
            for (group_lesson_id, group_key), rows_in_group in matching_groups.items():
                first_index = rows_in_group[0]["index"]
                try:
                    pairs_data = []
                    group_question_text = None
                    group_question_image_url = None
                    group_explanation = None
                    group_difficulty = None
                    group_bloom_level = None
                    group_video_url = None
                    group_video_explanation = None
                    group_is_blocked = False
                    group_import_is_bank = rows_in_group[0]["import_is_bank"]

                    for order_num, entry in enumerate(rows_in_group):
                        row = entry["row"]
                        left_text = row["Left Text"] if pd.notna(row.get("Left Text")) else None
                        left_image_url = row["Left Image URL"] if pd.notna(row.get("Left Image URL")) else None
                        right_text = row["Right Text"] if pd.notna(row.get("Right Text")) else None
                        right_image_url = row["Right Image URL"] if pd.notna(row.get("Right Image URL")) else None

                        if left_text or left_image_url or right_text or right_image_url:
                            pairs_data.append({
                                "order_num": order_num,
                                "left_text": left_text, "left_image_url": left_image_url,
                                "right_text": right_text, "right_image_url": right_image_url,
                            })

                        if group_question_text is None and pd.notna(row.get("Question Text")):
                            group_question_text = row["Question Text"]
                        if group_question_image_url is None and pd.notna(row.get("Question Image URL")):
                            group_question_image_url = row["Question Image URL"]
                        if group_explanation is None and pd.notna(row.get("Explanation")):
                            group_explanation = row["Explanation"]
                        if group_difficulty is None and pd.notna(row.get("Difficulty")):
                            group_difficulty = row["Difficulty"]
                        if group_bloom_level is None and pd.notna(row.get("Bloom Level")):
                            group_bloom_level = row["Bloom Level"]
                        if group_video_url is None and pd.notna(row.get("Video URL")):
                            group_video_url = row["Video URL"]
                        if group_video_explanation is None and pd.notna(row.get("Video Explanation")):
                            group_video_explanation = row["Video Explanation"]
                        if str(row.get("Is Blocked", "0")).strip() == "1":
                            group_is_blocked = True

                    if len(pairs_data) < 2:
                        error_details.append(f"صف {first_index+2}: سؤال المزاوجة (Question Group: {group_key}) يحتاج زوجين على الأقل.")
                        continue

                    if not group_question_text:
                        group_question_text = "اربط العمود أ بما يناسبه من العمود ب"

                    new_question = Question(
                        question_text=group_question_text,
                        lesson_id=group_lesson_id,
                        image_url=group_question_image_url,
                        explanation=group_explanation,
                        difficulty=group_difficulty,
                        bloom_level=group_bloom_level,
                        video_url=group_video_url,
                        video_explanation=group_video_explanation,
                        is_blocked=group_is_blocked,
                        is_bank=group_import_is_bank,
                        question_type="matching",
                    )
                    db.session.add(new_question)
                    db.session.flush()

                    for pair_data in pairs_data:
                        pair = MatchingPair(
                            left_text=pair_data["left_text"],
                            left_image_url=pair_data["left_image_url"],
                            right_text=pair_data["right_text"],
                            right_image_url=pair_data["right_image_url"],
                            order_num=pair_data["order_num"],
                            question_id=new_question.question_id,
                        )
                        db.session.add(pair)

                    imported_count += 1
                except Exception as group_error:
                    error_details.append(f"صف {first_index+2}: {str(group_error)}")
                    current_app.logger.exception(f"Error processing matching group {group_key}: {group_error}")
            
            # Commit all changes if there were any successful imports
            if imported_count > 0:
                db.session.commit()
                current_app.logger.info(f"Successfully imported {imported_count} questions.")
                msg = f"تم استيراد {imported_count} سؤال بنجاح!"
                if skipped_count > 0:
                    msg += f" (تم تخطي {skipped_count} مكرر موجود مسبقاً)"
                flash(msg, "success")
            elif skipped_count > 0:
                flash(f"لم يُضَف أي سؤال — كل الأسئلة ({skipped_count}) موجودة مسبقاً.", "warning")
                
                # تسجيل النشاط بعد استيراد الأسئلة بنجاح
                try:
                    # الحصول على معلومات الدرس والوحدة والدورة
                    lesson = Lesson.query.get(lesson_id)
                    lesson_name = lesson.name if lesson else None
                    unit_name = lesson.unit.name if lesson and lesson.unit else None
                    course_name = lesson.unit.course.name if lesson and lesson.unit and lesson.unit.course else None
                    
                    # تسجيل النشاط
                    Activity.log_activity(
                        action_type="import",
                        entity_type="question",
                        entity_id=None,  # لا يوجد معرف محدد لأننا استوردنا عدة أسئلة
                        description=f"تم استيراد {imported_count} سؤال إلى الدرس: {lesson_name}",
                        lesson_name=lesson_name,
                        unit_name=unit_name,
                        course_name=course_name,
                        user_id=current_user.id if current_user.is_authenticated else None
                    )
                except Exception as activity_error:
                    current_app.logger.error(f"Error logging activity: {activity_error}")
                    # لا نريد أن يؤثر خطأ تسجيل النشاط على تدفق العملية الأساسية
            
            # Show errors if any
            if error_details:
                error_summary = f"تم استيراد {imported_count} سؤال، مع {len(error_details)} أخطاء:"
                for i, error in enumerate(error_details[:5]):  # Show first 5 errors
                    flash(error, "warning")
                if len(error_details) > 5:
                    flash(f"... و {len(error_details) - 5} أخطاء أخرى.", "warning")
                
                flash(error_summary, "danger")
                current_app.logger.error(f"Import errors occurred: {error_details}")
                return render_template("question/import_questions.html", lessons=lessons, selected_lesson_id=lesson_id, form=form)
            else:
                if imported_count > 0:
                     return redirect(url_for("question.list_questions", lesson_id=lesson_id))
                else:
                     flash("لم يتم العثور على أسئلة صالحة للاستيراد في الملف.", "warning")
                     return render_template("question/import_questions.html", lessons=lessons, selected_lesson_id=lesson_id, form=form)

        except Exception as e:
            current_app.logger.exception(f"Error processing import file: {e}")
            flash(f"حدث خطأ أثناء معالجة ملف الاستيراد: {str(e)}", "danger")
            return render_template("question/import_questions.html", lessons=lessons, selected_lesson_id=lesson_id, form=form)
    
    # GET request - show the form
    return render_template("question/import_questions.html", lessons=lessons, form=form)

def _import_sample_row_for_type(qtype):
    """صف عيّنة واقعي (أو أكثر لو النوع يحتاج عدة صفوف كالمزاوجة) لنوع سؤال معيّن."""
    base = {
        "Course Name": "مثال: كيمياء 1",
        "Unit Name": "مثال: الوحدة الأولى",
        "Lesson Name": "مثال: الدرس الأول",
        "Question Type": qtype,
    }
    if qtype == "mcq":
        return [{
            **base,
            "Question Text": "ما هي الصيغة الكيميائية للماء؟",
            "Option 1 Text": "H₂O", "Option 2 Text": "CO₂",
            "Option 3 Text": "NaCl", "Option 4 Text": "O₂",
            "Correct Option Number": 1,
            "Explanation": "الماء يتكون من ذرتين من الهيدروجين وذرة واحدة من الأكسجين",
            "Difficulty": "medium", "Bloom Level": "remember", "Is Blocked": 0,
        }]
    if qtype == "true_false":
        return [{
            **base,
            "Question Text": "الأوزون غاز عديم اللون.",
            "Correct Answer": "صح",
            "Difficulty": "easy", "Bloom Level": "remember", "Is Blocked": 0,
        }]
    if qtype == "fill_blank":
        return [{
            **base,
            "Question Text": "الصيغة الكيميائية للماء هي ......",
            "Fill Blank Answer": "H2O", "Fill Blank Alt Answers": "H₂O, ماء",
            "Difficulty": "medium", "Bloom Level": "remember", "Is Blocked": 0,
        }]
    if qtype == "matching":
        return [
            {**base, "Question Group": "مثال-1", "Question Text": "اربط العمود أ بما يناسبه من العمود ب",
             "Left Text": "الأوزون", "Right Text": "O3", "Difficulty": "medium", "Bloom Level": "understand", "Is Blocked": 0},
            {**base, "Question Group": "مثال-1",
             "Left Text": "الماء", "Right Text": "H2O"},
        ]
    if qtype == "essay":
        return [{
            **base,
            "Question Text": "اشرح كيف تتكون طبقة الأوزون.",
            "Essay Model Answer": "تتكون طبقة الأوزون عندما تتحلل جزيئات الأكسجين بفعل الأشعة فوق البنفسجية...",
            "Difficulty": "hard", "Bloom Level": "understand", "Is Blocked": 0,
        }]
    return []


@question_bp.route("/import/template")
@login_required
def download_import_template():
    try:
        requested_types = [t.strip().lower() for t in request.args.get('types', 'mcq').split(',') if t.strip()]
        selected_types = [t for t in requested_types if t in IMPORT_COLUMNS_BY_TYPE] or ["mcq"]
        file_format = request.args.get('format', 'xlsx').strip().lower()
        if file_format not in ('xlsx', 'csv'):
            file_format = 'xlsx'

        # الأعمدة = الثابتة + عمود النوع/المجموعة + اتحاد أعمدة الأنواع المختارة + الأعمدة العامة بالنهاية
        columns = ["Course Name", "Unit Name", "Lesson Name", "Question Type", "Question Group"]
        for t in selected_types:
            for col in IMPORT_COLUMNS_BY_TYPE[t]:
                if col not in columns:
                    columns.append(col)
        columns += IMPORT_COMMON_TRAILING_COLUMNS

        df = pd.DataFrame(columns=columns)
        sample_rows = []
        for t in selected_types:
            sample_rows.extend(_import_sample_row_for_type(t))
        if sample_rows:
            df = pd.concat([df, pd.DataFrame(sample_rows)], ignore_index=True)
        df = df.fillna("")

        output = io.BytesIO()
        type_labels = "-".join(IMPORT_TYPE_LABELS_AR.get(t, t) for t in selected_types)

        if file_format == 'csv':
            output.write(df.to_csv(index=False).encode('utf-8-sig'))
            output.seek(0)
            return send_file(
                output,
                mimetype='text/csv',
                as_attachment=True,
                download_name=f'question_import_template_{"_".join(selected_types)}.csv'
            )

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Questions')

            # Auto-adjust column widths
            worksheet = writer.sheets['Questions']
            from openpyxl.utils import get_column_letter
            for i, col in enumerate(df.columns):
                col_max = df[col].astype(str).map(len).max()
                max_width = max(col_max if pd.notna(col_max) else 0, len(col)) + 2
                worksheet.column_dimensions[get_column_letter(i + 1)].width = max_width

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'question_import_template_{"_".join(selected_types)}.xlsx'
        )

    except Exception as e:
        current_app.logger.exception("Error generating import template")
        flash("حدث خطأ أثناء إنشاء قالب الاستيراد.", "danger")
        return redirect(url_for("question.import_questions"))

# --- edit_question route (Uses modified save_upload) --- #
@question_bp.route("/edit/<int:question_id>", methods=["GET", "POST"])
@login_required
def edit_question(question_id):
    # إنشاء نموذج فارغ لتوفير رمز CSRF
    form = FlaskForm()
    
    question = Question.query.options(joinedload(Question.options)).get_or_404(question_id)
    lessons = get_sorted_lessons()
    
    if request.method == "POST":
        current_app.logger.info(f"POST request received for edit_question ID: {question_id}")
        question_text = request.form.get("text", "").strip()
        lesson_id = request.form.get("lesson_id")
        # نوع السؤال لا يتغير بعد الإنشاء — نعتمد على القيمة المحفوظة دائماً
        question_type = question.question_type or "mcq"
        if question_type == "matching" and not question_text:
            question_text = "اربط العمود أ بما يناسبه من العمود ب"
        correct_option_index_str = request.form.get("correct_option")
        q_image_file = request.files.get("question_image")
        delete_question_image = request.form.get("delete_question_image") == "1"
        explanation = request.form.get("explanation", "").strip()

        q_image_path = question.image_url
        if delete_question_image:
            q_image_path = None
        elif q_image_file and q_image_file.filename:
            if not allowed_image_file(q_image_file.filename):
                flash("نوع ملف صورة السؤال غير مسموح به.", "danger")
            else:
                # Uses the Cloudinary-compatible save_upload function
                new_q_image_path = save_upload(q_image_file, subfolder="questions")
                if new_q_image_path is None:
                    flash("فشل رفع صورة السؤال. تحقق من إعدادات Cloudinary والسجلات.", "danger")
                else:
                    q_image_path = new_q_image_path

        error_messages = []
        if question_type != "matching" and not question_text and not q_image_path:
            error_messages.append("يجب توفير نص للسؤال أو رفع صورة له.")
        if not lesson_id:
            error_messages.append("يجب اختيار درس.")

        options_data_from_form = []
        matching_pairs_data = []
        fill_blank_answer = question.fill_blank_answer
        fill_blank_alt_answers = question.fill_blank_alt_answers
        essay_model_answer = question.essay_model_answer
        max_submitted_index = -1

        if question_type == "mcq":
            option_keys_check = [key for key in request.form if key.startswith("option_text_")]
            option_files_check = [key for key in request.files if key.startswith("option_image_")]
            if (option_keys_check or option_files_check) and correct_option_index_str is None:
                error_messages.append("يجب تحديد الإجابة الصحيحة.")

            correct_option_index = -1
            if correct_option_index_str is not None:
                try:
                    correct_option_index = int(correct_option_index_str)
                    if correct_option_index < 0:
                         error_messages.append("اختيار الإجابة الصحيحة غير صالح.")
                except ValueError:
                    error_messages.append("اختيار الإجابة الصحيحة يجب أن يكون رقمًا.")

            for key in list(request.form.keys()) + list(request.files.keys()):
                if key.startswith(("option_text_", "option_image_")):
                    try:
                        index_str = key.split("_")[-1]
                        max_submitted_index = max(max_submitted_index, int(index_str))
                    except (ValueError, IndexError):
                        continue

            for i in range(max_submitted_index + 1):
                index_str = str(i)
                option_text = request.form.get(f"option_text_{index_str}", "").strip()
                option_image_file = request.files.get(f"option_image_{index_str}")
                delete_option_image = request.form.get(f"delete_option_image_{index_str}") == "1"
                option_id = request.form.get(f"option_id_{index_str}")

                # Find existing option if we have an option_id
                existing_option = None
                if option_id:
                    try:
                        option_id = int(option_id)
                        existing_option = next((opt for opt in question.options if opt.option_id == option_id), None)
                    except ValueError:
                        pass

                option_image_path = existing_option.image_url if existing_option else None

                if delete_option_image:
                    option_image_path = None
                elif option_image_file and option_image_file.filename:
                    if not allowed_image_file(option_image_file.filename):
                        error_messages.append(f"نوع ملف صورة الخيار رقم {i+1} غير مسموح به.")
                    else:
                        # Uses the Cloudinary-compatible save_upload function
                        new_option_image_path = save_upload(option_image_file, subfolder="options")
                        if new_option_image_path is None:
                            error_messages.append(f"فشل رفع صورة الخيار رقم {i+1}. تحقق من إعدادات Cloudinary والسجلات.")
                        else:
                            option_image_path = new_option_image_path

                if option_text or option_image_path:
                    is_correct = (i == correct_option_index)
                    options_data_from_form.append({
                        "index": i,
                        "option_id": option_id,
                        "option_text": option_text,
                        "image_url": option_image_path,
                        "is_correct": is_correct
                    })

            if len(options_data_from_form) < 2:
                error_messages.append("يجب إضافة خيارين صالحين على الأقل (بنص أو صورة).")

            if correct_option_index_str is not None and correct_option_index >= len(options_data_from_form):
                 error_messages.append("الخيار المحدد كصحيح غير موجود أو غير صالح.")

        elif question_type == "true_false":
            tf_answer = request.form.get("tf_answer")
            if tf_answer not in ("true", "false"):
                error_messages.append("يجب اختيار الإجابة الصحيحة (صح أو خطأ).")
            else:
                existing_opts_sorted = sorted(question.options, key=lambda o: o.option_id)
                opt0_id = existing_opts_sorted[0].option_id if len(existing_opts_sorted) > 0 else None
                opt1_id = existing_opts_sorted[1].option_id if len(existing_opts_sorted) > 1 else None
                options_data_from_form = [
                    {"index": 0, "option_id": opt0_id, "option_text": "صح", "image_url": None, "is_correct": tf_answer == "true"},
                    {"index": 1, "option_id": opt1_id, "option_text": "خطأ", "image_url": None, "is_correct": tf_answer == "false"},
                ]

        elif question_type == "fill_blank":
            fill_blank_answer = request.form.get("fill_blank_answer", "").strip()
            if not fill_blank_answer:
                error_messages.append("يجب إدخال الإجابة الصحيحة لسؤال إكمال الفراغ.")
            alt_raw = request.form.get("fill_blank_alt_answers", "").strip()
            fill_blank_alt_answers = [a.strip() for a in alt_raw.split(",") if a.strip()] if alt_raw else None

        elif question_type == "matching":
            max_pair_index = -1
            for key in list(request.form.keys()) + list(request.files.keys()):
                if key.startswith(("left_text_", "right_text_", "left_image_", "right_image_")):
                    try:
                        max_pair_index = max(max_pair_index, int(key.split("_")[-1]))
                    except (ValueError, IndexError):
                        continue

            for i in range(max_pair_index + 1):
                idx_str = str(i)
                pair_id = request.form.get(f"pair_id_{idx_str}")
                existing_pair = None
                if pair_id:
                    try:
                        pair_id = int(pair_id)
                        existing_pair = next((p for p in question.matching_pairs if p.pair_id == pair_id), None)
                    except ValueError:
                        pair_id = None

                left_text = request.form.get(f"left_text_{idx_str}", "").strip()
                right_text = request.form.get(f"right_text_{idx_str}", "").strip()
                left_image_file = request.files.get(f"left_image_{idx_str}")
                right_image_file = request.files.get(f"right_image_{idx_str}")
                left_image_path = existing_pair.left_image_url if existing_pair else None
                right_image_path = existing_pair.right_image_url if existing_pair else None

                if left_image_file and left_image_file.filename:
                    if not allowed_image_file(left_image_file.filename):
                        error_messages.append(f"نوع ملف صورة العمود ب رقم {i+1} غير مسموح به.")
                    else:
                        left_image_path = save_upload(left_image_file, subfolder="matching")
                if right_image_file and right_image_file.filename:
                    if not allowed_image_file(right_image_file.filename):
                        error_messages.append(f"نوع ملف صورة العمود أ رقم {i+1} غير مسموح به.")
                    else:
                        right_image_path = save_upload(right_image_file, subfolder="matching")

                if left_text or left_image_path or right_text or right_image_path:
                    matching_pairs_data.append({
                        "pair_id": pair_id,
                        "order_num": i,
                        "left_text": left_text or None,
                        "left_image_url": left_image_path,
                        "right_text": right_text or None,
                        "right_image_url": right_image_path,
                    })

            if len(matching_pairs_data) < 2:
                error_messages.append("يجب إضافة زوجين على الأقل لسؤال المزاوجة.")

        elif question_type == "essay":
            essay_model_answer = request.form.get("essay_model_answer", "").strip() or None

        if error_messages:
            for error in error_messages:
                flash(error, "danger")
            form_data = request.form.to_dict()
            if question_type == "mcq":
                repop_options = []
                for i in range(max_submitted_index + 1):
                     idx_str = str(i)
                     opt_text = request.form.get(f"option_text_{idx_str}", "")
                     opt_id = request.form.get(f"option_id_{idx_str}")
                     processed_opt = next((opt for opt in options_data_from_form if opt["index"] == i), None)
                     img_url = processed_opt["image_url"] if processed_opt else None
                     repop_options.append({"option_id": opt_id, "option_text": opt_text, "image_url": img_url})
                form_data["options_repop"] = repop_options
                form_data["correct_option_repop"] = correct_option_index_str
            form_data["question_type"] = question_type
            form_data["question_image_url_repop"] = q_image_path
            return render_template("question/form.html", title="تعديل السؤال", lessons=lessons, question=form_data, submit_text="تحديث السؤال", form=form)

        try:
            # Update question
            question.question_text = question_text if question_text else None
            question.lesson_id = lesson_id
            question.image_url = q_image_path
            question.explanation = explanation or None
            question.video_explanation = request.form.get("video_explanation", "").strip() or None
            question.is_blocked = (request.form.get("is_blocked") == "1")  # معالجة حقل منع السؤال
            # حفظ رابط الفيديو إذا تم تعديله يدوياً
            video_url_form = request.form.get("video_url", "").strip()
            if video_url_form:
                question.video_url = video_url_form
                question.video_status = 'ready'
            elif 'video_url' in request.form and not video_url_form:
                question.video_url = None
                question.video_status = 'none'
            r2_url_form = request.form.get("r2_video_url", "").strip()
            if r2_url_form:
                question.r2_video_url = r2_url_form
                question.video_status = 'ready'
            elif 'r2_video_url' in request.form and not r2_url_form:
                question.r2_video_url = None

            # حقول الأنواع الجديدة
            question.fill_blank_answer = fill_blank_answer
            question.fill_blank_alt_answers = fill_blank_alt_answers
            question.essay_model_answer = essay_model_answer

            if question_type in ("mcq", "true_false"):
                # Track existing options to determine which to delete
                existing_option_ids = {opt.option_id for opt in question.options}
                updated_option_ids = set()

                # Update or create options
                for opt_data in options_data_from_form:
                    option_id = opt_data.get("option_id")

                    # --- Logic to set option_text to image_url if option_text is empty and image_url exists ---
                    option_text_to_save = opt_data["option_text"]
                    if not option_text_to_save and opt_data["image_url"]:
                        option_text_to_save = opt_data["image_url"] # Set option_text to the image_url
                    elif not option_text_to_save: # If option_text is still empty (and no image_url or image_url was not used)
                        option_text_to_save = None

                    if option_id:
                        # Update existing option
                        try:
                            option_id = int(option_id)
                            option = next((opt for opt in question.options if opt.option_id == option_id), None)
                            if option:
                                option.option_text = option_text_to_save
                                option.image_url = opt_data["image_url"]
                                option.is_correct = opt_data["is_correct"]
                                updated_option_ids.add(option_id)
                        except (ValueError, TypeError):
                            # If option_id is not a valid integer, create a new option
                            option = Option(
                                option_text=option_text_to_save,
                                image_url=opt_data["image_url"],
                                is_correct=opt_data["is_correct"],
                                question_id=question.question_id
                            )
                            db.session.add(option)
                    else:
                        # Create new option
                        option = Option(
                            option_text=option_text_to_save,
                            image_url=opt_data["image_url"],
                            is_correct=opt_data["is_correct"],
                            question_id=question.question_id
                        )
                        db.session.add(option)

                # Delete options that were not updated or created
                options_to_delete = existing_option_ids - updated_option_ids
                if options_to_delete:
                    Option.query.filter(Option.option_id.in_(options_to_delete)).delete(synchronize_session=False)

            elif question_type == "matching":
                existing_pair_ids = {p.pair_id for p in question.matching_pairs}
                updated_pair_ids = set()

                for pair_data in matching_pairs_data:
                    pair_id = pair_data.get("pair_id")
                    if pair_id:
                        pair = next((p for p in question.matching_pairs if p.pair_id == pair_id), None)
                        if pair:
                            pair.left_text = pair_data["left_text"]
                            pair.left_image_url = pair_data["left_image_url"]
                            pair.right_text = pair_data["right_text"]
                            pair.right_image_url = pair_data["right_image_url"]
                            pair.order_num = pair_data["order_num"]
                            updated_pair_ids.add(pair_id)
                            continue
                    pair = MatchingPair(
                        left_text=pair_data["left_text"],
                        left_image_url=pair_data["left_image_url"],
                        right_text=pair_data["right_text"],
                        right_image_url=pair_data["right_image_url"],
                        order_num=pair_data["order_num"],
                        question_id=question.question_id
                    )
                    db.session.add(pair)

                pairs_to_delete = existing_pair_ids - updated_pair_ids
                if pairs_to_delete:
                    MatchingPair.query.filter(MatchingPair.pair_id.in_(pairs_to_delete)).delete(synchronize_session=False)

            db.session.commit()
            current_app.logger.info(f"Question ID {question_id} updated successfully.")
            flash("تم تحديث السؤال بنجاح!", "success")
            
            # تسجيل النشاط بعد تعديل السؤال بنجاح
            try:
                # الحصول على معلومات الدرس والوحدة والدورة
                lesson = Lesson.query.get(lesson_id)
                lesson_name = lesson.name if lesson else None
                unit_name = lesson.unit.name if lesson and lesson.unit else None
                course_name = lesson.unit.course.name if lesson and lesson.unit and lesson.unit.course else None
                
                # تسجيل النشاط
                Activity.log_activity(
                    action_type="update",
                    entity_type="question",
                    entity_id=question.question_id,
                    description=f"تم تعديل سؤال في الدرس: {lesson_name}",
                    lesson_name=lesson_name,
                    unit_name=unit_name,
                    course_name=course_name,
                    user_id=current_user.id if current_user.is_authenticated else None
                )
            except Exception as activity_error:
                current_app.logger.error(f"Error logging activity: {activity_error}")
                # لا نريد أن يؤثر خطأ تسجيل النشاط على تدفق العملية الأساسية

            return redirect(url_for("question.list_questions", lesson_id=question.lesson_id))

        except (IntegrityError, DBAPIError) as db_err:
            db.session.rollback()
            current_app.logger.exception(f"Database error during question update for ID {question_id}.")
            flash("حدث خطأ في قاعدة البيانات أثناء تحديث السؤال.", "danger")
            return render_template("question/form.html", title="تعديل السؤال", lessons=lessons, question=question, submit_text="تحديث السؤال", form=form)
        
        except Exception as e:
            db.session.rollback()
            current_app.logger.exception(f"Unexpected error during question update for ID {question_id}.")
            flash("حدث خطأ غير متوقع أثناء تحديث السؤال.", "danger")
            return render_template("question/form.html", title="تعديل السؤال", lessons=lessons, question=question, submit_text="تحديث السؤال", form=form)

    # GET request - show the form with existing data
    return render_template("question/form.html", title="تعديل السؤال", lessons=lessons, question=question, submit_text="تحديث السؤال", form=form)

# --- delete_question route (keep as is) --- #
@question_bp.route("/delete/<int:question_id>", methods=["POST"])
@login_required
def delete_question(question_id):
    question = Question.query.get_or_404(question_id)
    
    try:
        # Get lesson info before deleting for activity logging
        lesson_id = question.lesson_id
        lesson = Lesson.query.get(lesson_id) if lesson_id else None
        lesson_name = lesson.name if lesson else None
        unit_name = lesson.unit.name if lesson and lesson.unit else None
        course_name = lesson.unit.course.name if lesson and lesson.unit and lesson.unit.course else None
        
        # Delete associated options first (should happen automatically with cascade, but being explicit)
        Option.query.filter_by(question_id=question_id).delete()
        
        # Delete the question
        db.session.delete(question)
        db.session.commit()
        
        flash("تم حذف السؤال بنجاح!", "success")
        
        # تسجيل النشاط بعد حذف السؤال بنجاح
        try:
            # تسجيل النشاط
            Activity.log_activity(
                action_type="delete",
                entity_type="question",
                entity_id=question_id,
                description=f"تم حذف سؤال من الدرس: {lesson_name}",
                lesson_name=lesson_name,
                unit_name=unit_name,
                course_name=course_name,
                user_id=current_user.id if current_user.is_authenticated else None
            )
        except Exception as activity_error:
            current_app.logger.error(f"Error logging activity: {activity_error}")
            # لا نريد أن يؤثر خطأ تسجيل النشاط على تدفق العملية الأساسية
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception(f"Error deleting question ID {question_id}.")
        flash("حدث خطأ أثناء محاولة حذف السؤال.", "danger")
    
    return redirect(request.referrer or url_for("question.list_questions"))


# ===== Route لصفحة الاختبار التفاعلي =====
@question_bp.route('/quiz')
@login_required
def quiz():
    """
    صفحة الاختبار التفاعلي
    """
    return render_template('quiz.html')


# ===== Route لصفحة استخراج وتوليد نماذج الاختبار =====
@question_bp.route('/export-exam')
@login_required
def export_exam():
    """
    صفحة استخراج وتوليد نماذج الاختبار
    """
    return render_template('question/export_exam.html')


@question_bp.route('/download-exam-word', methods=['POST'])
@login_required
def download_exam_word():
    """
    تحميل نموذج الاختبار كملف Word
    
    Request JSON:
    {
        "question_ids": [1, 2, 3, ...],
        "include_answers": true/false,
        "exam_title": "عنوان الاختبار",
        "course_name": "اسم المنهج",
        "unit_name": "اسم الوحدة",
        "lesson_name": "اسم الدرس"
    }
    """
    try:
        import sys
        import os
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from exam_generator import generate_exam
        from models.exam_header_settings import ExamHeaderSettings
        
        data = request.get_json()
        question_ids = data.get("question_ids", [])
        include_answers = data.get("include_answers", False)
        exam_title = data.get("exam_title", "نموذج الاختبار")
        course_name = data.get("course_name", "")
        unit_name = data.get("unit_name", "")
        lesson_name = data.get("lesson_name", "")
        output_format = data.get("output_format", "word").lower()  # أضفنا هذا
        
        # استخراج الإعدادات المحفوظة
        header_settings = ExamHeaderSettings.query.first()
        if header_settings:
            country = header_settings.country or ""
            ministry = header_settings.ministry or ""
            education_department = header_settings.education_department or ""
            school_name = header_settings.school_name or ""
            subject = header_settings.subject or ""
            exam_time = header_settings.time or ""
            grade = header_settings.grade or ""
            total_score = header_settings.total_score or 30
            checker_name = header_settings.checker_name or ""
            reviewer_name = header_settings.reviewer_name or ""
            exam_date = header_settings.exam_date or ""
        else:
            # قيم افتراضية
            country = "المملكة العربية السعودية"
            ministry = "وزارة التعليم"
            education_department = "الإدارة العامة للتعليم"
            school_name = ""
            subject = ""
            exam_time = ""
            grade = ""
            total_score = 30
            checker_name = ""
            reviewer_name = ""
            exam_date = ""
        
        if not question_ids:
            return jsonify({
                'success': False,
                'error': 'لم يتم تحديد أسئلة'
            }), 400
        
        # الحصول على الأسئلة من قاعدة البيانات (MCQ فقط — تصدير Word لا يدعم بقية الأنواع بعد)
        questions = Question.query.filter(
            Question.question_id.in_(question_ids),
            Question.question_type == 'mcq'
        ).all()

        if not questions:
            return jsonify({
                'success': False,
                'error': 'لم يتم العثور على الأسئلة المحددة'
            }), 404
        
        # تنسيق الأسئلة مع تنسيق النص للطباعة
        formatted_questions = []
        for question in questions:
            formatted_q = {
                'question_id': question.question_id,
                'question_text': format_text_for_print(question.question_text),
                'options': [
                    {
                        'option_id': opt.option_id,
                        'option_text': format_text_for_print(opt.option_text),
                        'is_correct': opt.is_correct
                    }
                    for opt in sorted(question.options, key=lambda o: o.option_id)
                ],
                'correct_option_id': next(
                    (opt.option_id for opt in question.options if opt.is_correct),
                    None
                )
            }
            formatted_questions.append(formatted_q)
        
        # توليد ملف Word أو PDF باستخدام النظام الموحد
        file_bytes = generate_exam(
            formatted_questions,
            exam_title=exam_title,
            output_format=output_format,  # استخدم الصيغة المطلوبة
            show_answers=include_answers,
            country=country,
            ministry=ministry,
            education_department=education_department,
            school_name=school_name,
            subject=subject,
            time=exam_time,
            grade=grade,
            total_score=total_score,
            checker_name=checker_name,
            reviewer_name=reviewer_name,
            exam_date=exam_date
        )
        
        # إرسال الملف بالصيغة الصحيحة
        if output_format == 'pdf':
            return send_file(
                io.BytesIO(file_bytes),
                mimetype='application/pdf',
                as_attachment=True,
                download_name=f"exam_{int(time.time())}.pdf"
            )
        else:  # word
            return send_file(
                io.BytesIO(file_bytes),
                mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                as_attachment=True,
                download_name=f"exam_{int(time.time())}.docx"
            )
        
    except ImportError as ie:
        current_app.logger.error(f"Import error: {ie}")
        return jsonify({
            'success': False,
            'error': 'وحدة توليد الاختبارات غير متاحة'
        }), 500
    except Exception as e:
        current_app.logger.exception(f"Error downloading exam word: {e}")
        return jsonify({
            'success': False,
            'error': f'خطأ في توليد الملف: {str(e)}'
        }), 500


@question_bp.route('/header-settings')
@login_required
def header_settings():
    """عرض صفحة إعدادات الكليشة"""
    return render_template('question/exam_header_settings.html')


@question_bp.route('/save-header-settings', methods=['POST'])
@login_required
def save_header_settings():
    """حفظ إعدادات الكليشة في قاعدة البيانات"""
    try:
        data = request.get_json()
        
        # البحث عن الإعدادات الموجودة أو إنشاء جديدة
        settings = ExamHeaderSettings.query.first()
        if not settings:
            settings = ExamHeaderSettings()
        
        # تحديث البيانات
        settings.country = data.get('country', 'المملكة العربية السعودية')
        settings.ministry = data.get('ministry', 'وزارة التعليم')
        settings.education_department = data.get('education_department', 'الإدارة العامة للتعليم بالمنطقة الشرقية')
        settings.school_name = data.get('school_name', 'مدرسة عبدالرحمن بن القاسم الثانوية')
        settings.subject = data.get('subject', 'كيمياء 4')
        settings.time = data.get('time', 'ثلاث ساعات')
        settings.grade = data.get('grade', 'ثالث ثانوي')
        settings.total_score = data.get('total_score', 30)
        settings.checker_name = data.get('checker_name', '')
        settings.reviewer_name = data.get('reviewer_name', '')
        settings.exam_date = data.get('exam_date', '')
        
        db.session.add(settings)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم حفظ الإعدادات بنجاح'
        }), 200
        
    except Exception as e:
        current_app.logger.exception(f"Error saving header settings: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@question_bp.route('/get-header-settings', methods=['GET'])
@login_required
def get_header_settings():
    """الحصول على إعدادات الكليشة المحفوظة"""
    try:
        settings = ExamHeaderSettings.query.first()
        
        if settings:
            return jsonify({
                'success': True,
                'settings': {
                    'country': settings.country,
                    'ministry': settings.ministry,
                    'education_department': settings.education_department,
                    'school_name': settings.school_name,
                    'subject': settings.subject,
                    'time': settings.time,
                    'grade': settings.grade,
                    'total_score': settings.total_score,
                    'checker_name': settings.checker_name,
                    'reviewer_name': settings.reviewer_name,
                    'exam_date': settings.exam_date
                }
            }), 200
        else:
            return jsonify({
                'success': True,
                'settings': None
            }), 200
            
    except Exception as e:
        current_app.logger.exception(f"Error getting header settings: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@question_bp.route('/export-exam-pdf', methods=['POST'])
@login_required
def export_exam_pdf():
    """استخراج ملف PDF للاختبار مع جلب البيانات بشكل صحيح"""
    try:
        # محاولة استيراد المولد
        try:
            from src.routes.exam_generator import ExamGenerator
        except ImportError:
            from exam_generator import ExamGenerator
            
        # محاولة استيراد نموذج الإعدادات من المسار الصحيح (كما في دالة الوورد)
        try:
            from src.models.exam_header_settings import ExamHeaderSettings as SettingsModel
        except ImportError:
            try:
                from models.exam_header_settings import ExamHeaderSettings as SettingsModel
            except ImportError:
                # في حال الفشل نستخدم الكلاس المعرف محلياً
                SettingsModel = ExamHeaderSettings
        
        data = request.get_json()
        question_ids = data.get('question_ids', [])
        include_answers = data.get('include_answers', False)
        exam_title = data.get('exam_title', 'نموذج الاختبار')
        
        # جلب الأسئلة
        questions = get_ordered_questions(question_ids)
        
        if not questions:
            return jsonify({'error': 'لا توجد أسئلة محددة'}), 400
        
        # تحويل الأسئلة مع تنسيق النص للطباعة
        questions_data = []
        for q in questions:
            q_dict = {
                'id': q.question_id, 
                'question_text': format_text_for_print(q.question_text), 
                'points': 1, 
                'options': []
            }
            for opt in q.options:
                q_dict['options'].append({
                    'option_text': format_text_for_print(opt.option_text), 
                    'is_correct': opt.is_correct
                })
            questions_data.append(q_dict)
        
        # جلب الإعدادات من قاعدة البيانات
        # نستخدم SettingsModel الذي تم استيراده لضمان التطابق مع دالة الوورد
        header_settings = db.session.query(SettingsModel).first()
        
        settings_dict = {}
        if header_settings:
            # استخدام or "" لضمان عدم تمرير None
            settings_dict = {
                'country': header_settings.country or "المملكة العربية السعودية",
                'ministry': header_settings.ministry or "وزارة التعليم",
                'education_department': header_settings.education_department or "",
                'school_name': header_settings.school_name or "",
                'subject': header_settings.subject or "",
                'time': header_settings.time or "",
                'grade': header_settings.grade or "",
                'total_score': header_settings.total_score or 30,
                'checker_name': header_settings.checker_name or "",
                'reviewer_name': header_settings.reviewer_name or "",
                'exam_date': header_settings.exam_date or ""
            }
        
        # تمرير الإعدادات إلى المنشئ وإلى دالة التوليد
        generator = ExamGenerator(header_settings=settings_dict)
        
        pdf_bytes = generator.generate_pdf(
            questions_data, 
            exam_title, 
            include_answers,
            **settings_dict  # فك القاموس تمريره كـ kwargs
        )
        
        # ── حفظ في تاريخ الاختبارات ──────────────────────────────
        try:
            from src.models.generated_exam import GeneratedExam as _GE
            # استخراج course_id من أول سؤال (عبر lesson → unit → course)
            first_q = Question.query.get(question_ids[0]) if question_ids else None
            _course_id = None
            _unit_id   = None
            _lesson_id = None
            _course_name = None
            if first_q and first_q.lesson_id:
                from src.models.curriculum import Lesson as _Lesson, Unit as _Unit, Course as _Course
                _lesson = _Lesson.query.get(first_q.lesson_id)
                if _lesson:
                    _lesson_id   = _lesson.id
                    _unit        = _Unit.query.get(_lesson.unit_id)
                    if _unit:
                        _unit_id     = _unit.id
                        _course      = _Course.query.get(_unit.course_id)
                        if _course:
                            _course_id   = _course.id
                            _course_name = _course.name

            _ge = _GE(
                user_id=current_user.id,
                course_id=_course_id or 0,
                unit_id=_unit_id,
                lesson_id=_lesson_id,
                question_count=len(question_ids),
                include_answers=include_answers,
                shuffle_questions=False,
                shuffle_options=False,
                course_name=_course_name,
            )
            _ge.header = settings_dict
            db.session.add(_ge)
            db.session.commit()
        except Exception:
            pass

        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"exam_{int(time.time())}.pdf"
        )

    except Exception as e:
        current_app.logger.exception(f"Error exporting PDF: {e}")
        return jsonify({'error': str(e)}), 500


@question_bp.route('/preview-exam-paper', methods=['POST'])
@login_required
def preview_exam_paper():
    """عرض معاينة الاختبار في المتصفح باستخدام نفس القالب الموحد"""
    try:
        # استيراد المولد فقط (لأنه غير موجود في هذا الملف)
        try:
            from src.routes.exam_generator import ExamGenerator
        except ImportError:
            from exam_generator import ExamGenerator
            
        # ملاحظة: لا نقم باستيراد ExamHeaderSettings هنا 
        # لأننا سنستخدم الكلاس المعرف في السطر 48 من هذا الملف مباشرة.

        # استقبال البيانات من الفورم
        question_ids_str = request.form.get('question_ids', '')
        if not question_ids_str:
             return "لا توجد أسئلة محددة", 400
             
        question_ids = [int(x) for x in question_ids_str.split(',')]
        include_answers = request.form.get('include_answers') == 'true'
        
        # ... (باقي الكود كما هو تماماً) ...
        
        questions = get_ordered_questions(question_ids)
        
        # تحويل الأسئلة مع تنسيق النص للطباعة
        questions_data = []
        for q in questions:
            q_dict = {
                'id': q.question_id, 
                'question_text': format_text_for_print(q.question_text), 
                'points': 1, 
                'options': []
            }
            for opt in q.options:
                q_dict['options'].append({
                    'option_text': format_text_for_print(opt.option_text), 
                    'is_correct': opt.is_correct, 
                    'option_id': opt.option_id
                })
            q_dict['correct_option_id'] = next((o.option_id for o in q.options if o.is_correct), None)
            questions_data.append(q_dict)

        header_settings = ExamHeaderSettings.query.first()
        settings_dict = {
            'country': header_settings.country or "المملكة العربية السعودية",
            'ministry': header_settings.ministry or "وزارة التعليم",
            'education_department': header_settings.education_department or "",
            'school_name': header_settings.school_name or "",
            'subject': header_settings.subject or "",
            'time': header_settings.time or "",
            'grade': header_settings.grade or "",
            'total_score': header_settings.total_score or 30,
            'checker_name': header_settings.checker_name or "",
            'reviewer_name': header_settings.reviewer_name or "",
            'exam_date': header_settings.exam_date or ""
        }
        
        generator = ExamGenerator(header_settings=settings_dict)
        # توليد HTML فقط للعرض في المتصفح
        html_content = generator.generate_html(questions_data, "نموذج الاختبار", include_answers, **settings_dict)
        
        return html_content
        
    except Exception as e:
        return f"Error: {str(e)}", 500


# ============================================================
# كود توليد النماذج المتعددة مع الباركود
# ============================================================

def generate_qr_code(data_dict):
    """
    توليد باركود QR يحتوي على معلومات الاختبار
    
    Args:
        data_dict: قاموس يحتوي على معلومات الاختبار
        
    Returns:
        Base64 string للصورة
    """
    # تحويل البيانات لنص
    qr_text = f"""المادة: {data_dict.get('subject', '')}
النموذج: {data_dict.get('model', '')}
عدد الأسئلة: {data_dict.get('questions_count', '')}
التاريخ: {data_dict.get('date', '')}""".strip()
    
    # إنشاء الباركود
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=3,
        border=2,
    )
    qr.add_data(qr_text)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # تحويل لـ Base64
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    
    return f"data:image/png;base64,{img_base64}"


def shuffle_exam(questions, shuffle_questions=True, shuffle_options=True, seed=None, saved_options_order=None):
    """
    خلط ترتيب الأسئلة والخيارات
    
    Args:
        questions: قائمة الأسئلة
        shuffle_questions: خلط ترتيب الأسئلة
        shuffle_options: خلط ترتيب الخيارات
        seed: بذرة للعشوائية (لإعادة الإنتاج)
        saved_options_order: ترتيب الخيارات المحفوظ لكل سؤال (dict: question_id -> [option_ids])
        
    Returns:
        قائمة الأسئلة المخلوطة مع تتبع الإجابات الصحيحة
    """
    if seed is not None:
        random.seed(seed)
    
    # نسخة عميقة لتجنب تعديل الأصل
    shuffled = copy.deepcopy(questions)
    
    # خلط ترتيب الأسئلة
    if shuffle_questions:
        random.shuffle(shuffled)
    
    # تحويل مفاتيح saved_options_order لـ strings للمقارنة
    saved_order_normalized = {}
    if saved_options_order:
        for k, v in saved_options_order.items():
            saved_order_normalized[str(k)] = [str(x) for x in v]
    
    # خلط ترتيب الخيارات لكل سؤال
    letters = ['أ', 'ب', 'ج', 'د', 'هـ', 'و']
    
    for question in shuffled:
        options = question.get('options', [])
        if not options:
            continue
            
        question_id = str(question.get('question_id', ''))
        
        # التحقق من وجود ترتيب محفوظ لهذا السؤال
        if saved_order_normalized and question_id in saved_order_normalized:
            saved_order = saved_order_normalized[question_id]
            # إنشاء قاموس للخيارات حسب الـ ID
            options_by_id = {str(opt.get('option_id', '')): opt for opt in options}
            
            # ترتيب الخيارات حسب الترتيب المحفوظ
            ordered_options = []
            for opt_id in saved_order:
                if opt_id in options_by_id:
                    ordered_options.append(options_by_id[opt_id])
            
            # إذا تم ترتيب جميع الخيارات بنجاح، نستخدم الترتيب المحفوظ
            if len(ordered_options) == len(options):
                options = ordered_options
            elif shuffle_options:
                # إذا فشل الترتيب المحفوظ، نخلط عشوائياً
                random.shuffle(options)
        elif shuffle_options:
            # لا يوجد ترتيب محفوظ، نخلط عشوائياً
            random.shuffle(options)
        
        # تحديث الإجابة الصحيحة بعد الترتيب
        for i, opt in enumerate(options):
            if opt.get('is_correct'):
                question['correct_answer_index'] = i
                question['correct_answer_letter'] = letters[i] if i < len(letters) else str(i+1)
                break
        
        question['options'] = options
    
    return shuffled


@question_bp.route('/generate-multi-models', methods=['POST'])
@login_required
def generate_multi_models():
    """
    توليد نماذج متعددة من الاختبار بترتيب مختلف + باركود
    """
    try:
        # استيراد WeasyPrint
        from weasyprint import HTML as WeasyHTML
        
        data = request.get_json()
        current_app.logger.info(f"Received data for multi-models: {data}")
        
        question_ids = data.get('question_ids', [])
        models = data.get('models', ['أ'])  # النماذج المطلوبة
        include_answers = data.get('include_answers', False)
        include_answer_sheet = data.get('include_answer_sheet', False)
        include_barcode = data.get('include_barcode', True)
        shuffle_options = data.get('shuffle_options', True)
        
        if not question_ids:
            return jsonify({'error': 'لم يتم تحديد أسئلة'}), 400
        
        current_app.logger.info(f"Processing {len(question_ids)} questions for models: {models}")
        
        # جلب الأسئلة من قاعدة البيانات
        questions = get_ordered_questions(question_ids)
        
        if not questions:
            return jsonify({'error': 'لم يتم العثور على الأسئلة'}), 404
        
        current_app.logger.info(f"Found {len(questions)} questions in database")
        
        # تحويل الأسئلة لقاموس مع تنسيق النص للطباعة
        questions_data = []
        for q in questions:
            q_dict = {
                'question_id': q.question_id,
                'question_text': format_text_for_print(q.question_text or ''),
                'image_url': getattr(q, 'image_url', None) or '',
                'options': []
            }
            for opt in q.options:
                q_dict['options'].append({
                    'option_id': getattr(opt, 'option_id', None),
                    'option_text': format_text_for_print(getattr(opt, 'option_text', '') or ''),
                    'image_url': getattr(opt, 'image_url', None) or '',
                    'is_correct': getattr(opt, 'is_correct', False)
                })
            questions_data.append(q_dict)
        
        current_app.logger.info(f"Converted {len(questions_data)} questions to dict")
        
        # جلب إعدادات الهيدر
        settings = ExamHeaderSettings.query.first()
        header_settings = {
            'country': settings.country if settings else 'المملكة العربية السعودية',
            'ministry': settings.ministry if settings else 'وزارة التعليم',
            'education_department': settings.education_department if settings else '',
            'school_name': settings.school_name if settings else '',
            'subject': settings.subject if settings else '',
            'time': settings.time if settings else '',
            'grade': settings.grade if settings else '',
            'total_score': settings.total_score if settings else 30,
            'checker_name': settings.checker_name if settings else '',
            'reviewer_name': settings.reviewer_name if settings else ''
        }
        
        # تحويل الشعار لـ base64 لتجنب مشاكل WeasyPrint مع timeout
        logo_base64 = None
        try:
            logo_path = os.path.join(current_app.static_folder, 'images', 'logo.png')
            if os.path.exists(logo_path):
                with open(logo_path, 'rb') as f:
                    logo_data = f.read()
                    logo_base64 = f"data:image/png;base64,{base64.b64encode(logo_data).decode('utf-8')}"
        except Exception as logo_err:
            current_app.logger.warning(f"Could not load logo: {logo_err}")
        
        header_settings['logo_base64'] = logo_base64
        header_settings['logo_url'] = url_for('static', filename='images/logo.png', _external=True)
        
        # توليد HTML لكل نموذج
        all_models_html = []
        answer_keys = {}  # مفاتيح الإجابات لكل نموذج
        
        for i, model_letter in enumerate(models):
            # 🔧 seed يعتمد على: محتوى الأسئلة + النموذج + أرقام أولية ضخمة متباعدة
            # استخدام أرقام أولية ضخمة جداً لضمان اختلاف كبير بين النماذج
            question_ids_str = ''.join(str(q['question_id']) for q in questions_data)
            random_offset = [15485863, 32452843, 49979687, 67867967][i % 4]  # أرقام أولية ضخمة
            seed = (sum(q['question_id'] for q in questions_data) * (ord(model_letter[0]) + 1) + i * 7919 + random_offset) % (2**31)
            shuffled_questions = shuffle_exam(
                questions_data, 
                shuffle_questions=True, 
                shuffle_options=shuffle_options,
                seed=seed
            )
            
            # حفظ مفتاح الإجابات
            letters = ['أ', 'ب', 'ج', 'د', 'هـ', 'و']
            answer_keys[model_letter] = []
            for j, q in enumerate(shuffled_questions):
                correct_letter = q.get('correct_answer_letter', '')
                if not correct_letter:
                    # البحث عن الإجابة الصحيحة
                    for k, opt in enumerate(q['options']):
                        if opt.get('is_correct'):
                            correct_letter = letters[k] if k < len(letters) else str(k+1)
                            break
                answer_keys[model_letter].append({
                    'question_num': j + 1,
                    'answer': correct_letter
                })
            
            # توليد الباركود
            qr_code_data = None
            if include_barcode:
                qr_code_data = generate_qr_code({
                    'subject': header_settings['subject'],
                    'model': model_letter,
                    'questions_count': len(shuffled_questions),
                    'date': datetime.now().strftime('%Y-%m-%d')
                })
            
            # توليد HTML للنموذج
            model_html = render_template(
                'question/exam_paper_layout_with_barcode.html',
                questions=shuffled_questions,
                model_letter=model_letter,
                qr_code=qr_code_data,
                show_answers=include_answers,
                exam_title=f"نموذج الاختبار - {model_letter}",
                **header_settings
            )
            
            all_models_html.append(model_html)
        
        # إضافة بطاقة التصحيح إذا مطلوبة
        if include_answer_sheet:
            answer_sheet_html = render_template(
                'question/answer_sheet_template.html',
                answer_keys=answer_keys,
                models=models,
                questions_count=len(questions_data),
                **header_settings
            )
            all_models_html.append(answer_sheet_html)
        
        # دمج كل النماذج في HTML واحد - مبسط للأداء
        combined_html = """<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
<meta charset="UTF-8">
<style>
@page { size: A4; margin: 15mm; }
.page-break { page-break-after: always; }
body { font-family: Arial, Tahoma, sans-serif; font-size: 12px; }
</style>
</head>
<body>
"""
        
        for i, model_html in enumerate(all_models_html):
            combined_html += model_html
            if i < len(all_models_html) - 1:
                combined_html += '<div class="page-break"></div>'
        
        combined_html += "</body></html>"
        
        current_app.logger.info(f"Combined HTML length: {len(combined_html)} chars")
        
        # تحويل لـ PDF مع timeout handling
        pdf_buffer = io.BytesIO()
        try:
            # استخدام base_url محلي لتجنب HTTP requests
            WeasyHTML(string=combined_html, base_url=None).write_pdf(pdf_buffer)
        except Exception as pdf_err:
            current_app.logger.error(f"WeasyPrint PDF generation failed: {pdf_err}")
            # Fallback: إرجاع HTML بدلاً من PDF
            return combined_html, 200, {'Content-Type': 'text/html; charset=utf-8'}
        
        pdf_buffer.seek(0)
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'exam_models_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        )
        
    except Exception as e:
        current_app.logger.exception(f"Error generating multi models: {e}")
        return jsonify({'error': str(e)}), 500


@question_bp.route('/preview-multi-models', methods=['POST'])
@login_required
def preview_multi_models():
    """
    معاينة النماذج المتعددة كـ HTML في المتصفح (للطباعة)
    هذا يتجنب مشاكل WeasyPrint timeout وتم تحديثه لضمان اختلاف الإجابات بين النماذج
    """
    try:
        data = request.get_json()
        
        question_ids = data.get('question_ids', [])
        models = data.get('models', ['أ'])
        include_answers = data.get('include_answers', False)
        include_answer_sheet = data.get('include_answer_sheet', False)
        include_barcode = data.get('include_barcode', True)
        font_size = data.get('font_size', 14)  # حجم الخط الافتراضي 14px
        image_size = data.get('image_size', 100)  # حجم الصور الافتراضي 100%
        
        # ترتيب الخيارات المحفوظ (من اختبار محفوظ سابقاً)
        saved_options_order = data.get('saved_options_order', {})
        
        # إعدادات التنسيق المتقدمة
        columns = data.get('columns', 2)  # عدد الأعمدة الافتراضي 2
        spacing = data.get('spacing', 'normal')  # المسافة الافتراضية متوسطة
        options_layout = data.get('options_layout', 'vertical')  # تنسيق الخيارات الافتراضي عمودي
        
        # إعدادات تنسيق الكليشة
        header_format = {
            'header_size': data.get('header_size', 'medium'),
            'show_logo': data.get('show_logo', True),
            'logo_size': data.get('logo_size', 'medium'),
            'qr_size': data.get('qr_size', 'medium'),
            'show_grades_table': data.get('show_grades_table', True),
            'show_extra_grade_field': data.get('show_extra_grade_field', False),
            'show_student_name': data.get('show_student_name', True),
            'show_student_class': data.get('show_student_class', True),
            'show_student_seat_no': data.get('show_student_seat_no', True),
            'show_student_signature': data.get('show_student_signature', False),
            'name_line_length': data.get('name_line_length', 'medium'),
            'exam_type': data.get('exam_type', 'نهاية'),
            'semester': data.get('semester', 'الأول'),
            'academic_year': data.get('academic_year', '1447هـ')
        }
        
        # التعديل الأساسي: نضمن خلط الخيارات دائماً عند تعدد النماذج لكسر نمط الإجابات
        shuffle_options = True if len(models) > 1 else data.get('shuffle_options', True)
        
        if not question_ids:
            return jsonify({'error': 'لم يتم تحديد أسئلة'}), 400
        
        # جلب الأسئلة
        questions = get_ordered_questions(question_ids)
        
        if not questions:
            return jsonify({'error': 'لم يتم العثور على الأسئلة'}), 404
        
        # تحويل الأسئلة لقاموس (كما في كودك الأصلي)
        # مع تنسيق النص للطباعة (تحويل \n إلى <br> ومنع كسر الأسطر)
        questions_data = []
        for q in questions:
            q_dict = {
                'question_id': q.question_id,
                'question_text': format_text_for_print(getattr(q, 'question_text', '') or ''),
                'image_url': getattr(q, 'image_url', None) or '',
                'options': []
            }
            for opt in q.options:
                q_dict['options'].append({
                    'option_id': getattr(opt, 'option_id', None),
                    'option_text': format_text_for_print(getattr(opt, 'option_text', '') or ''),
                    'image_url': getattr(opt, 'image_url', None) or '',
                    'is_correct': getattr(opt, 'is_correct', False)
                })
            questions_data.append(q_dict)
        
        # جلب إعدادات الهيدر (كما في كودك الأصلي)
        settings = ExamHeaderSettings.query.first()
        header_settings = {
            'country': settings.country if settings else 'المملكة العربية السعودية',
            'ministry': settings.ministry if settings else 'وزارة التعليم',
            'education_department': settings.education_department if settings else '',
            'school_name': settings.school_name if settings else '',
            'subject': settings.subject if settings else '',
            'time': settings.time if settings else '',
            'grade': settings.grade if settings else '',
            'total_score': settings.total_score if settings else 30,
            'checker_name': settings.checker_name if settings else '',
            'reviewer_name': settings.reviewer_name if settings else ''
        }
        
        # تحويل الشعار لـ base64 (كما في كودك الأصلي)
        logo_base64 = None
        try:
            logo_path = os.path.join(current_app.static_folder, 'images', 'logo.png')
            if os.path.exists(logo_path):
                with open(logo_path, 'rb') as f:
                    logo_data = f.read()
                    logo_base64 = f"data:image/png;base64,{base64.b64encode(logo_data).decode('utf-8')}"
        except Exception as logo_err:
            current_app.logger.warning(f"Could not load logo: {logo_err}")
        
        header_settings['logo_base64'] = logo_base64
        
        # توليد HTML لكل نموذج
        all_models_html = []
        answer_keys = {}
        letters = ['أ', 'ب', 'ج', 'د', 'هـ', 'و']
        
        for idx, model_letter in enumerate(models):
            # 🔧 seed يعتمد على: محتوى الأسئلة + النموذج + رقم عشوائي كبير
            question_ids_str = ''.join(str(q['question_id']) for q in questions_data)
            random_offset = [15485863, 32452843, 49979687, 67867967][idx % 4]
            seed = (sum(q['question_id'] for q in questions_data) * (ord(model_letter[0]) + 1) + idx * 7919 + random_offset) % (2**31)
            
            # التحقق من وجود ترتيب محفوظ (للنموذج الأول فقط)
            has_saved_order = saved_options_order and len(saved_options_order) > 0 and idx == 0
            
            current_app.logger.info(f"Model {model_letter}: has_saved_order={has_saved_order}, saved_options_order keys={list(saved_options_order.keys())[:5] if saved_options_order else 'None'}")
            
            shuffled_questions = shuffle_exam(
                questions_data,
                shuffle_questions=not has_saved_order,  # لا تخلط الأسئلة إذا فيه ترتيب محفوظ
                shuffle_options=shuffle_options if not has_saved_order else True,  # استخدم الترتيب المحفوظ
                seed=seed if not has_saved_order else None,
                saved_options_order=saved_options_order if has_saved_order else None
            )
            
            # بناء مفتاح الإجابات من الأسئلة المخلوطة لهذا النموذج تحديداً
            model_answers = []
            for q in shuffled_questions:
                correct_letter = ''
                # البحث عن الإجابة الصحيحة بعد الخلط (هذا يضمن اختلاف الحرف بين النماذج)
                for i, opt in enumerate(q.get('options', [])):
                    if opt.get('is_correct'):
                        correct_letter = letters[i] if i < len(letters) else str(i+1)
                        break
                model_answers.append({'answer': correct_letter})
            answer_keys[model_letter] = model_answers
            
            # توليد QR code (كما في كودك الأصلي)
            qr_code_data = None
            if include_barcode:
                qr_data = {
                    'subject': header_settings.get('subject', ''),
                    'model': model_letter,
                    'questions': len(questions_data),
                    'date': datetime.now().strftime('%Y-%m-%d')
                }
                qr_code_data = generate_qr_code(qr_data)
            
            # توليد HTML للنموذج (كما في كودك الأصلي)
            model_html = render_template(
                'question/exam_paper_layout_with_barcode.html',
                questions=shuffled_questions,
                model_letter=model_letter,
                qr_code=qr_code_data,
                show_answers=include_answers,
                exam_title=f"نموذج الاختبار - {model_letter}",
                font_size=font_size,
                image_size=image_size,
                columns=columns,
                spacing=spacing,
                options_layout=options_layout,
                header_format=header_format,
                **header_settings
            )
            
            all_models_html.append(model_html)
        
        # إضافة بطاقة التصحيح (كما في كودك الأصلي)
        if include_answer_sheet:
            answer_sheet_html = render_template(
                'question/answer_sheet_template.html',
                answer_keys=answer_keys,
                models=models,
                questions_count=len(questions_data),
                **header_settings
            )
            all_models_html.append(answer_sheet_html)
        
        # دمج كل النماذج في HTML واحد للطباعة (كما في كودك الأصلي)
        combined_html = """<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
<meta charset="UTF-8">
<title>النماذج المتعددة</title>
<style>
@media print {
    .page-break { page-break-after: always; }
    .no-print { display: none; }
}
@page { size: A4; margin: 15mm; }
.print-btn {
    position: fixed;
    top: 10px;
    left: 10px;
    padding: 15px 30px;
    background: #4CAF50;
    color: white;
    border: none;
    border-radius: 5px;
    cursor: pointer;
    font-size: 16px;
    z-index: 9999;
}
.print-btn:hover { background: #45a049; }

/* === تنسيق الأسطر لمنع كسر المعادلات الكيميائية === */
.line-block {
    display: inline-block;
    white-space: nowrap;
}
</style>
</head>
<body>
<button class="print-btn no-print" onclick="window.print()">🖨️ طباعة النماذج</button>
"""
        
        for i, model_html in enumerate(all_models_html):
            combined_html += model_html
            if i < len(all_models_html) - 1:
                combined_html += '<div class="page-break"></div>'
        
        # إضافة سكربت لحفظ ترتيب الخيارات في النافذة الأصلية
        # نبني ترتيب الخيارات من النموذج الأول فقط
        first_model_order = {}
        if all_models_html and questions_data:
            # نستخدم الأسئلة المخلوطة من أول نموذج
            first_shuffled = shuffle_exam(
                questions_data,
                shuffle_questions=True if not saved_options_order else False,
                shuffle_options=shuffle_options,
                seed=(sum(q['question_id'] for q in questions_data) * (ord(models[0][0]) + 1) + 15485863) % (2**31) if not saved_options_order else None,
                saved_options_order=saved_options_order if saved_options_order else None
            )
            for q in first_shuffled:
                qid = str(q.get('question_id', ''))
                options = q.get('options', [])
                first_model_order[qid] = [opt.get('option_id') for opt in options]
        
        # إضافة السكربت لإرسال الترتيب للنافذة الأصلية
        options_order_json = json.dumps(first_model_order, ensure_ascii=False)
        combined_html += f'''
<script>
// إرسال ترتيب الخيارات للنافذة الأصلية
if (window.opener && !window.opener.closed) {{
    window.opener.shuffledOptionsOrder = {options_order_json};
    console.log('تم حفظ ترتيب الخيارات:', window.opener.shuffledOptionsOrder);
}}
</script>
'''
        
        combined_html += "</body></html>"
        
        return combined_html, 200, {'Content-Type': 'text/html; charset=utf-8'}
        
    except Exception as e:
        current_app.logger.exception(f"Error previewing multi models: {e}")
        return jsonify({'error': str(e)}), 500


# 1. دالة المعاينة: قراءة ملف الإكسل وتمرير البيانات للواجهة
@question_bp.route('/preview-students', methods=['POST'])
@teacher_or_admin_required
def preview_students():
    # تعريف القائمة في البداية لتجنب خطأ التعريف "not defined"
    final_students = [] 
    try:
        file = request.files.get('student_file')
        if not file:
            return jsonify({'error': 'الرجاء اختيار ملف إكسل'}), 400
        
        df = pd.read_excel(file)
        # تنظيف أسماء الأعمدة من المسافات
        df.columns = [str(c).strip() for c in df.columns]
        
        for _, row in df.iterrows():
            # البحث عن البيانات بمسميات مرنة (عربي/إنجليزي) لضمان القراءة
            name = row.get('الاسم') or row.get('اسم الطالب') or row.get('Name') or '..........'
            academic_id = row.get('الرقم الأكاديمي') or row.get('الرقم الاكاديمي') or row.get('Academic ID') or '..........'
            section = row.get('الشعبة') or row.get('الشعبه') or row.get('Section') or '....'
            
            # استخدام مفاتيح إنجليزية لتطابق قالب HTML (student.name)
            final_students.append({
                'name': str(name),
                'academic_id': str(academic_id),
                'section': str(section)
            })
        
        return jsonify({'success': True, 'students': final_students})
    except Exception as e:
        current_app.logger.error(f"Error in preview_students: {str(e)}")
        return jsonify({'error': f'خطأ في معالجة الملف: {str(e)}'}), 500

# ============================================================
# دالة توليد باركود الرقم الأكاديمي
# ============================================================
def generate_student_barcode(academic_id):
    """
    توليد باركود Code39 للرقم الأكاديمي
    
    Args:
        academic_id: الرقم الأكاديمي للطالب (string أو int)
        
    Returns:
        Base64 string للصورة
    """
    try:
        # تحويل الرقم لنص وإزالة المسافات
        academic_id_str = str(academic_id).strip()
        
        if not academic_id_str or academic_id_str == 'None':
            return None
        
        # إنشاء باركود Code39
        code39 = barcode.get_barcode_class('code39')
        
        # إعدادات الباركود
        writer = ImageWriter()
        writer.set_options({
            'module_width': 0.5,
            'module_height': 18,
            'quiet_zone': 4,
            'font_size': 0,
            'text_distance': 18,
            'write_text': True
        })
        
        # إنشاء الباركود
        barcode_obj = code39(academic_id_str, writer=writer, add_checksum=False)
        
        # حفظ في الذاكرة
        buffer = io.BytesIO()
        barcode_obj.write(buffer)
        buffer.seek(0)
        
        # تحويل لـ Base64
        img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
        
        return f"data:image/png;base64,{img_base64}"
        
    except Exception as e:
        current_app.logger.error(f"Error generating barcode for {academic_id}: {e}")
        return None


# 2. دالة الطباعة: توليد أوراق Remark بناءً على الكليشة والأسماء المرفوعة
@question_bp.route('/print-remark-sheets', methods=['POST'])
@login_required
def print_remark_sheets():
    """طباعة أوراق إجابة ريمارك مع الباركود"""
    try:
        data = request.get_json()
        students_list = data.get('students', [])
        exam_type = data.get('exam_type', 'نهاية')
        semester = data.get('semester', 'الأول')
        academic_year = data.get('academic_year', '1447هـ')
        
        # جلب إعدادات الكليشة من قاعدة البيانات
        settings = ExamHeaderSettings.query.first()
        
        header_context = {
            'country': settings.country if settings else 'المملكة العربية السعودية',
            'ministry': settings.ministry if settings else 'وزارة التعليم',
            'education_department': settings.education_department if settings else '',
            'school_name': settings.school_name if settings else '',
            'subject': settings.subject if settings else '',
            'grade': settings.grade if settings else '',
            'total_score': settings.total_score if settings else 30,
            'exam_type': exam_type,
            'semester': semester,
            'academic_year': academic_year,
            'logo_base64': ''
        }

        # تحويل الشعار لـ Base64
        try:
            logo_path = os.path.join(current_app.static_folder, 'images', 'logo.png')
            if os.path.exists(logo_path):
                with open(logo_path, 'rb') as f:
                    header_context['logo_base64'] = f"data:image/png;base64,{base64.b64encode(f.read()).decode('utf-8')}"
        except Exception as e:
            current_app.logger.warning(f"Could not load logo: {e}")

        all_html = ""
        for student in students_list:
            # توليد الباركود للرقم الأكاديمي
            academic_id = student.get('academic_id', '')
            if academic_id:
                student['barcode'] = generate_student_barcode(academic_id)
            else:
                student['barcode'] = None
            
            # توليد HTML للطالب
            all_html += render_template('question/remark_answer_sheet.html', 
                                       student=student, 
                                       **header_context)
            all_html += '<div style="page-break-after: always;"></div>'

        return jsonify({'success': True, 'html_content': all_html})
        
    except Exception as e:
        current_app.logger.error(f"Error in print_remark_sheets: {str(e)}")
        return jsonify({'error': f'فشل تجهيز الطباعة: {str(e)}'}), 500


# ==================== Route لاستخراج مفتاح إجابة OMR ====================
@question_bp.route('/generate-omr-answer-key', methods=['POST'])
@login_required
def generate_omr_answer_key():
    """استخراج مفتاح إجابة OMR من الأسئلة المختارة"""
    try:
        data = request.get_json()
        question_ids = data.get('question_ids', [])
        matching_pair_ids = data.get('matching_pair_ids', [])
        model_letter = data.get('model_letter', 'أ')
        exam_type = data.get('exam_type', 'نهاية')
        semester = data.get('semester', 'الأول')
        academic_year = data.get('academic_year', '1447هـ')

        if not question_ids:
            return jsonify({'error': 'لم يتم تحديد أسئلة'}), 400

        # جلب الأسئلة مع الخيارات (تستبعد fill_blank/essay غير المتوافقة مع التظليل)
        questions = get_ordered_questions_for_omr(question_ids)

        if not questions:
            return jsonify({'error': 'لم يتم العثور على الأسئلة'}), 404

        # استخراج الإجابات الصحيحة — مفصولة حسب النوع (متعدد/صح-خطأ/مزاوجة)
        mcq_data, tf_data, matching_pairs_flat = _split_omr_questions_data(questions)
        matching_pairs_flat = _resolve_matching_pairs_flat(matching_pairs_flat, matching_pair_ids)
        answers = _build_omr_answers(mcq_data, tf_data, matching_pairs_flat,
                                      shuffle_questions=False, shuffle_options=False, seed=None)

        # جلب إعدادات الكليشة
        header_settings_record = ExamHeaderSettings.query.first()
        header_settings = {
            'country': 'المملكة العربية السعودية',
            'ministry': 'وزارة التعليم',
            'education_department': '',
            'school_name': '',
            'subject': '',
            'time': '',
            'grade': '',
            'total_score': 30,
            'logo_base64': ''
        }
        if header_settings_record:
            header_settings.update({
                'country': getattr(header_settings_record, 'country', 'المملكة العربية السعودية'),
                'ministry': getattr(header_settings_record, 'ministry', 'وزارة التعليم'),
                'education_department': getattr(header_settings_record, 'education_department', ''),
                'school_name': getattr(header_settings_record, 'school_name', ''),
                'subject': getattr(header_settings_record, 'subject', ''),
                'time': getattr(header_settings_record, 'time', ''),
                'grade': getattr(header_settings_record, 'grade', ''),
                'total_score': getattr(header_settings_record, 'total_score', 30)
            })
        
        # تحويل الشعار لـ Base64
        try:
            logo_path = os.path.join(current_app.static_folder, 'images', 'logo.png')
            if os.path.exists(logo_path):
                with open(logo_path, 'rb') as f:
                    header_settings['logo_base64'] = f"data:image/png;base64,{base64.b64encode(f.read()).decode('utf-8')}"
        except Exception as e:
            current_app.logger.warning(f"Could not load logo: {e}")
        
        # توليد HTML لمفتاح الإجابة
        answer_key_html = render_template(
            'question/remark_answer_sheet.html',
            student={'name': '🔑 مفتاح الإجابة', 'academic_id': '---', 'section': '---', 'barcode': None},
            is_answer_key=True,
            answers=answers,
            model_letter=model_letter,
            exam_type=exam_type,
            semester=semester,
            academic_year=academic_year,
            questions_count=len(questions),
            **header_settings
        )
        
        # إضافة زر الطباعة - بدون header ثابت لتجنب انزياح الأوراق
        full_html = f"""<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
<meta charset="UTF-8">
<title>🔑 مفتاح إجابة OMR - النموذج {model_letter}</title>
<style>
@media print {{
    .no-print {{ display: none !important; }}
}}
@media screen {{
    .screen-only-header {{
        background: #c41e3a;
        color: white;
        padding: 15px 20px;
        text-align: center;
        margin-bottom: 20px;
        border-radius: 8px;
    }}
    .screen-only-header button {{
        background: white;
        color: #c41e3a;
        border: none;
        padding: 10px 25px;
        border-radius: 5px;
        cursor: pointer;
        font-weight: bold;
        font-size: 1.1rem;
        margin-bottom: 10px;
    }}
    .screen-only-header button:hover {{
        background: #f0f0f0;
    }}
}}
body {{
    margin: 0;
    padding: 0;
}}
</style>
</head>
<body>
<div class="screen-only-header no-print">
    <button onclick="window.print()">🖨️ طباعة مفتاح الإجابة</button>
    <div style="margin-top: 10px;">🔑 مفتاح إجابة OMR - النموذج {model_letter} - عدد الأسئلة: {len(questions)}</div>
</div>
{answer_key_html}
</body>
</html>"""
        
        return jsonify({'success': True, 'html': full_html})
        
    except Exception as e:
        current_app.logger.exception(f"Error generating OMR answer key: {e}")
        return jsonify({'error': str(e)}), 500


@question_bp.route('/print-remark-sheets-multi-models', methods=['POST'])
@login_required
def print_remark_sheets_multi_models():
    """طباعة أوراق إجابة ريمارك للنماذج المتعددة"""
    try:
        data = request.get_json()
        students_list = data.get('students', [])
        models = data.get('models', ['أ'])  # النماذج المطلوبة
        question_ids = data.get('question_ids', [])
        matching_pair_ids = data.get('matching_pair_ids', [])
        shuffle_options = data.get('shuffle_options', True)
        exam_type = data.get('exam_type', 'نهاية')
        semester = data.get('semester', 'الأول')
        academic_year = data.get('academic_year', '1447هـ')

        if not students_list:
            return jsonify({'error': 'لم يتم تحديد قائمة الطلاب'}), 400

        if not question_ids:
            return jsonify({'error': 'لم يتم تحديد الأسئلة'}), 400

        # جلب الأسئلة (تستبعد fill_blank/essay غير المتوافقة مع التظليل) — للعدّ فقط
        questions = get_ordered_questions_for_omr(question_ids)

        if not questions:
            return jsonify({'error': 'لم يتم العثور على الأسئلة'}), 404

        # جلب إعدادات الكليشة
        settings = ExamHeaderSettings.query.first()
        header_context = {
            'country': settings.country if settings else 'المملكة العربية السعودية',
            'ministry': settings.ministry if settings else 'وزارة التعليم',
            'education_department': settings.education_department if settings else '',
            'school_name': settings.school_name if settings else '',
            'subject': settings.subject if settings else '',
            'grade': settings.grade if settings else '',
            'total_score': settings.total_score if settings else 30,
            'exam_type': exam_type,
            'semester': semester,
            'academic_year': academic_year,
            'logo_base64': ''
        }

        # تحويل الشعار لـ Base64
        try:
            logo_path = os.path.join(current_app.static_folder, 'images', 'logo.png')
            if os.path.exists(logo_path):
                with open(logo_path, 'rb') as f:
                    header_context['logo_base64'] = f"data:image/png;base64,{base64.b64encode(f.read()).decode('utf-8')}"
        except Exception as e:
            current_app.logger.warning(f"Could not load logo: {e}")

        # توليد مفاتيح الإجابة لكل نموذج — عبر نفس بايبلاين ترتيب/خلط ورقة الاختبار
        # الفعلية (exam_model_builder) عشان تطابق المطبوع بالضبط، مو خلط مستقل
        answer_keys, _ = _build_answer_keys_for_models(
            question_ids, matching_pair_ids, models, shuffle_options=shuffle_options
        )

        # توزيع النماذج على الطلاب بالتساوي
        students_per_model = len(students_list) // len(models)
        remainder = len(students_list) % len(models)
        
        all_html = ""
        student_idx = 0
        
        for model_idx, model_letter in enumerate(models):
            # عدد الطلاب لهذا النموذج
            count = students_per_model + (1 if model_idx < remainder else 0)
            
            for _ in range(count):
                if student_idx >= len(students_list):
                    break
                    
                student = students_list[student_idx]
                student_idx += 1
                
                # توليد الباركود للرقم الأكاديمي
                academic_id = student.get('academic_id', '')
                if academic_id:
                    student['barcode'] = generate_student_barcode(academic_id)
                else:
                    student['barcode'] = None
                
                # إضافة معلومات النموذج
                student['model_letter'] = model_letter
                
                # توليد HTML للطالب مع مفتاح إجابات النموذج
                all_html += render_template(
                    'question/remark_answer_sheet.html', 
                    student=student,
                    model_letter=model_letter,
                    is_answer_key=False,
                    answers=None,  # لا نعرض الإجابات في ورقة الطالب
                    questions_count=len(questions),
                    **header_context
                )
                all_html += '<div style="page-break-after: always;"></div>'

        # إضافة مفاتيح الإجابة لكل نموذج في النهاية
        for model_letter in models:
            all_html += render_template(
                'question/remark_answer_sheet.html',
                student={
                    'name': f'🔑 مفتاح الإجابة - نموذج {model_letter}',
                    'academic_id': '---',
                    'section': '---',
                    'barcode': None,
                    'model_letter': model_letter
                },
                is_answer_key=True,
                answers=answer_keys[model_letter],
                model_letter=model_letter,
                questions_count=len(questions),
                **header_context
            )
            all_html += '<div style="page-break-after: always;"></div>'

        return jsonify({'success': True, 'html_content': all_html})
        
    except Exception as e:
        current_app.logger.error(f"Error in print_remark_sheets_multi_models: {str(e)}")
        return jsonify({'error': f'فشل تجهيز الطباعة: {str(e)}'}), 500


@question_bp.route('/print-blank-remark-sheets', methods=['POST'])
@login_required
def print_blank_remark_sheets():
    """طباعة نماذج ريمارك فارغة للنماذج المتعددة"""
    try:
        data = request.get_json()
        models = data.get('models', ['أ'])
        count_per_model = data.get('count_per_model', 10)
        question_ids = data.get('question_ids', [])
        exam_type = data.get('exam_type', 'نهاية')
        semester = data.get('semester', 'الأول')
        academic_year = data.get('academic_year', '1447هـ')
        
        if not question_ids:
            return jsonify({'error': 'لم يتم تحديد الأسئلة'}), 400
        
        # جلب الأسئلة لمعرفة العدد (تستبعد fill_blank/essay غير المتوافقة مع التظليل)
        questions = get_ordered_questions_for_omr(question_ids)
        questions_count = len(questions)
        
        # جلب إعدادات الكليشة
        settings = ExamHeaderSettings.query.first()
        header_context = {
            'country': settings.country if settings else 'المملكة العربية السعودية',
            'ministry': settings.ministry if settings else 'وزارة التعليم',
            'education_department': settings.education_department if settings else '',
            'school_name': settings.school_name if settings else '',
            'subject': settings.subject if settings else '',
            'grade': settings.grade if settings else '',
            'total_score': settings.total_score if settings else 30,
            'exam_type': exam_type,
            'semester': semester,
            'academic_year': academic_year,
            'logo_base64': ''
        }

        # تحويل الشعار لـ Base64
        try:
            logo_path = os.path.join(current_app.static_folder, 'images', 'logo.png')
            if os.path.exists(logo_path):
                with open(logo_path, 'rb') as f:
                    header_context['logo_base64'] = f"data:image/png;base64,{base64.b64encode(f.read()).decode('utf-8')}"
        except Exception as e:
            current_app.logger.warning(f"Could not load logo: {e}")
        
        all_html = ""
        
        # توليد نماذج فارغة لكل نموذج
        for model_letter in models:
            for i in range(count_per_model):
                blank_student = {
                    'name': '..........................................',
                    'academic_id': '..........................................',
                    'section': '.....',
                    'barcode': None,
                    'model_letter': model_letter
                }
                
                all_html += render_template(
                    'question/remark_answer_sheet.html',
                    student=blank_student,
                    model_letter=model_letter,
                    is_answer_key=False,
                    answers=None,
                    questions_count=questions_count,
                    **header_context
                )
                all_html += '<div style="page-break-after: always;"></div>'
        
        return jsonify({'success': True, 'html_content': all_html})
        
    except Exception as e:
        current_app.logger.error(f"Error in print_blank_remark_sheets: {str(e)}")
        return jsonify({'error': f'فشل تجهيز الطباعة: {str(e)}'}), 500


@question_bp.route('/generate-all-models-answer-keys', methods=['POST'])
@login_required
def generate_all_models_answer_keys():
    """استخراج مفاتيح إجابة OMR لكل النماذج"""
    try:
        data = request.get_json()
        question_ids = data.get('question_ids', [])
        matching_pair_ids = data.get('matching_pair_ids', [])
        models = data.get('models', ['أ'])
        shuffle_options = data.get('shuffle_options', True)
        exam_type = data.get('exam_type', 'نهاية')
        semester = data.get('semester', 'الأول')
        academic_year = data.get('academic_year', '1447هـ')

        if not question_ids:
            return jsonify({'error': 'لم يتم تحديد أسئلة'}), 400

        # جلب الأسئلة (تستبعد fill_blank/essay غير المتوافقة مع التظليل)
        questions = get_ordered_questions_for_omr(question_ids)

        if not questions:
            return jsonify({'error': 'لم يتم العثور على الأسئلة'}), 404

        # جلب إعدادات الكليشة
        header_settings_record = ExamHeaderSettings.query.first()
        header_settings = {
            'country': 'المملكة العربية السعودية',
            'ministry': 'وزارة التعليم',
            'education_department': '',
            'school_name': '',
            'subject': '',
            'time': '',
            'grade': '',
            'total_score': 30,
            'logo_base64': ''
        }
        if header_settings_record:
            header_settings.update({
                'country': getattr(header_settings_record, 'country', 'المملكة العربية السعودية'),
                'ministry': getattr(header_settings_record, 'ministry', 'وزارة التعليم'),
                'education_department': getattr(header_settings_record, 'education_department', ''),
                'school_name': getattr(header_settings_record, 'school_name', ''),
                'subject': getattr(header_settings_record, 'subject', ''),
                'time': getattr(header_settings_record, 'time', ''),
                'grade': getattr(header_settings_record, 'grade', ''),
                'total_score': getattr(header_settings_record, 'total_score', 30)
            })

        # تحويل الشعار لـ Base64
        try:
            logo_path = os.path.join(current_app.static_folder, 'images', 'logo.png')
            if os.path.exists(logo_path):
                with open(logo_path, 'rb') as f:
                    header_settings['logo_base64'] = f"data:image/png;base64,{base64.b64encode(f.read()).decode('utf-8')}"
        except Exception as e:
            current_app.logger.warning(f"Could not load logo: {e}")

        # توليد مفاتيح الإجابة لكل نموذج — عبر نفس بايبلاين ترتيب/خلط ورقة الاختبار
        # الفعلية (exam_model_builder) عشان تطابق المطبوع بالضبط، مو خلط مستقل
        answer_keys, _ = _build_answer_keys_for_models(
            question_ids, matching_pair_ids, models, shuffle_options=shuffle_options
        )
        all_keys_html = ""

        for idx, model_letter in enumerate(models):
            answers = answer_keys[model_letter]

            # توليد HTML لمفتاح الإجابة
            answer_key_html = render_template(
                'question/remark_answer_sheet.html',
                student={'name': f'🔑 مفتاح الإجابة - نموذج {model_letter}', 'academic_id': '---', 'section': '---', 'barcode': None},
                is_answer_key=True,
                answers=answers,
                model_letter=model_letter,
                exam_type=exam_type,
                semester=semester,
                academic_year=academic_year,
                questions_count=len(questions),
                **header_settings
            )
            all_keys_html += answer_key_html
            if idx < len(models) - 1:
                all_keys_html += '<div style="page-break-after: always;"></div>'
        
        # إضافة زر الطباعة - بدون header ثابت لتجنب انزياح الأوراق
        full_html = f"""<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
<meta charset="UTF-8">
<title>🔑 مفاتيح إجابة OMR - النماذج {', '.join(models)}</title>
<style>
@media print {{
    .no-print {{ display: none !important; }}
}}
@media screen {{
    .screen-only-header {{
        background: #9c27b0;
        color: white;
        padding: 15px 20px;
        text-align: center;
        margin-bottom: 20px;
        border-radius: 8px;
    }}
    .screen-only-header button {{
        background: white;
        color: #9c27b0;
        border: none;
        padding: 10px 25px;
        border-radius: 5px;
        cursor: pointer;
        font-weight: bold;
        font-size: 1.1rem;
        margin-bottom: 10px;
    }}
    .screen-only-header button:hover {{
        background: #f0f0f0;
    }}
}}
body {{
    margin: 0;
    padding: 0;
}}
</style>
</head>
<body>
<div class="screen-only-header no-print">
    <button onclick="window.print()">🖨️ طباعة مفاتيح الإجابة</button>
    <div style="margin-top: 10px;">🔑 مفاتيح إجابة OMR - النماذج: {', '.join(models)} - عدد الأسئلة: {len(questions)}</div>
</div>
{all_keys_html}
</body>
</html>"""
        
        return jsonify({'success': True, 'html': full_html})
        
    except Exception as e:
        current_app.logger.exception(f"Error generating all models answer keys: {e}")
        return jsonify({'error': str(e)}), 500


# =====================================================
# ===== ورقة التظليل Remark — PDF Endpoints =====
# =====================================================

def _remark_header_context(exam_type='نهاية', semester='الأول', academic_year='1447هـ'):
    """دالة مساعدة: تجيب إعدادات الكليشة + الشعار Base64"""
    settings = ExamHeaderSettings.query.first()
    ctx = {
        'country':              settings.country             if settings else 'المملكة العربية السعودية',
        'ministry':             settings.ministry            if settings else 'وزارة التعليم',
        'education_department': settings.education_department if settings else '',
        'school_name':          settings.school_name         if settings else '',
        'subject':              settings.subject             if settings else '',
        'grade':                settings.grade               if settings else '',
        'total_score':          settings.total_score         if settings else 30,
        'exam_type':            exam_type,
        'semester':             semester,
        'academic_year':        academic_year,
        'logo_base64':          '',
    }
    try:
        logo_path = os.path.join(current_app.static_folder, 'images', 'logo.png')
        if os.path.exists(logo_path):
            with open(logo_path, 'rb') as f:
                ctx['logo_base64'] = f"data:image/png;base64,{base64.b64encode(f.read()).decode('utf-8')}"
    except Exception:
        pass
    return ctx


def _remark_auth_required(f):
    """Decorator يقبل أدمن (session) أو معلم (JWT) لـ remark endpoints"""
    @wraps(f)
    def decorated(*args, **kwargs):
        import jwt as pyjwt
        if current_user.is_authenticated:
            return f(*args, **kwargs)
        auth = request.headers.get('Authorization', '')
        if auth.startswith('Bearer '):
            try:
                data = pyjwt.decode(
                    auth[7:],
                    current_app.config['JWT_SECRET_KEY'],
                    algorithms=[current_app.config.get('JWT_ALGORITHM', 'HS256')]
                )
                if data.get('user_type') == 'teacher':
                    return f(*args, **kwargs)
            except Exception:
                pass
        try:
            from src.models.teacher import Teacher
        except ImportError:
            from models.teacher import Teacher
        tok = request.headers.get('X-Session-Token')
        if tok and Teacher.query.filter_by(session_token=tok, is_active=True).first():
            return f(*args, **kwargs)
        return jsonify({'success': False, 'error': 'يرجى تسجيل الدخول'}), 401
    return decorated


def _q_get_teacher_id():
    """يرجع teacher_id إذا المرسِل معلم، None إذا أدمن — لاستخدام داخل question.py"""
    import jwt as pyjwt
    auth = request.headers.get('Authorization', '')
    if auth.startswith('Bearer '):
        try:
            data = pyjwt.decode(
                auth[7:],
                current_app.config['JWT_SECRET_KEY'],
                algorithms=[current_app.config.get('JWT_ALGORITHM', 'HS256')]
            )
            if data.get('user_type') == 'teacher':
                return data.get('teacher_id')
        except Exception:
            pass
    try:
        from src.models.teacher import Teacher
    except ImportError:
        from models.teacher import Teacher
    tok = request.headers.get('X-Session-Token')
    if tok:
        t = Teacher.query.filter_by(session_token=tok, is_active=True).first()
        if t:
            return t.id
    return None


def _q_check_remark_quota(teacher_id):
    """تحقق من كوتا ورقة التظليل للمعلم — يرجع (allowed, remaining, limit, msg)"""
    try:
        from src.models.teacher_feature import TeacherFeatureOverride, FEATURE_DEFAULTS
        from src.models.teacher_export_log import TeacherExportLog
    except ImportError:
        from models.teacher_feature import TeacherFeatureOverride, FEATURE_DEFAULTS
        from models.teacher_export_log import TeacherExportLog

    try:
        from src.models.ai_setting import AISetting
        def _gs(key): return AISetting.get_setting(key)
    except Exception:
        def _gs(key): return None

    from datetime import timezone, timedelta as _td

    # هل الميزة مفعّلة؟
    ov = TeacherFeatureOverride.get_override(teacher_id, 'remark_export_enabled')
    if ov:
        enabled = ov.value == 'true'
    else:
        gv = _gs('remark_export_enabled')
        enabled = (gv or FEATURE_DEFAULTS.get('remark_export_enabled', 'false')) == 'true'
    if not enabled:
        return False, 0, 0, 'هذه الميزة غير مفعّلة لحسابك'

    # الحد اليومي
    oq = TeacherFeatureOverride.get_override(teacher_id, 'quota_remark_export')
    if oq:
        limit = int(oq.value)
    else:
        gq = _gs('quota_remark_export')
        limit = int(gq) if gq else int(FEATURE_DEFAULTS.get('quota_remark_export', '2'))

    # الاستخدام اليوم
    sa_tz = timezone(_td(hours=3))
    today_start = datetime.now(sa_tz).replace(
        hour=0, minute=0, second=0, microsecond=0
    ).astimezone(timezone.utc).replace(tzinfo=None)

    used = db.session.query(TeacherExportLog).filter(
        TeacherExportLog.teacher_id == teacher_id,
        TeacherExportLog.export_type == 'remark',
        TeacherExportLog.created_at >= today_start,
    ).count()

    remaining = max(0, limit - used)
    if remaining == 0:
        return False, 0, limit, f'تجاوزت الحد اليومي ({limit} أوراق/يوم)'
    return True, remaining, limit, ''


def _remark_pdf_wrapper(sheets_html):
    """يلف صفحات ورقة التظليل في HTML واحد صحيح مع CSS مرة واحدة فقط"""
    # نولّد صفحة واحدة بوضع standalone=True للحصول على الـ CSS
    # ثم نستخدم standalone=False لباقي الصفحات
    return (
        '<!DOCTYPE html><html dir="rtl" lang="ar">'
        '<head><meta charset="UTF-8">'
        '<style>'
        '@page{size:A4 landscape;margin:0;}'
        '*{margin:0;padding:0;box-sizing:border-box;-webkit-print-color-adjust:exact;print-color-adjust:exact;}'
        'html,body{width:100%;background:white;font-family:"Noto Sans Arabic","DejaVu Sans",Arial,sans-serif;font-size:12px;-webkit-print-color-adjust:exact;print-color-adjust:exact;}'
        '.sheet-container{width:285mm;height:182mm;border:2px solid #000;margin:5mm auto;display:flex;flex-direction:column;overflow:hidden;page-break-inside:avoid;break-inside:avoid;}'
        '.sheet-container:not(:first-of-type){page-break-before:always;break-before:page;}'
        '.sheet-top{display:flex;flex-direction:row;flex-shrink:0;}'
        '.top-right{width:47%;border-left:2px solid #000;border-bottom:2px solid #000;padding:8px;display:flex;flex-direction:column;}'
        '.top-left{width:53%;border-bottom:2px solid #000;padding:8px;display:flex;flex-direction:column;}'
        '.logo-section{display:flex;align-items:flex-start;gap:10px;margin-bottom:6px;}'
        '.logo-img{width:65px;height:auto;}'
        '.ministry-info{flex:1;text-align:right;font-size:11px;line-height:1.4;}'
        '.ministry-info .title{font-weight:bold;color:#000;font-size:12px;}'
        '.school-name{font-weight:bold;color:#000;font-size:12px;margin-top:2px;}'
        '.exam-info-box{background:white;border:1px solid #000;border-radius:4px;padding:6px;margin-top:5px;font-size:12px;}'
        '.dor-section{display:flex;gap:15px;align-items:center;margin-top:5px;}'
        '.dor-option{display:flex;align-items:center;gap:4px;font-weight:bold;}'
        '.checkbox{width:16px;height:16px;border:2px solid #000;display:inline-block;background:white;}'
        '.instructions-box{background:white;border:2px solid #000;border-radius:5px;padding:6px;margin-top:6px;display:flex;justify-content:space-between;align-items:flex-start;gap:6px;flex:1;}'
        '.instructions-content{flex:1;}'
        '.instructions-title{font-weight:bold;color:#000;margin-bottom:3px;font-size:12px;}'
        '.instructions-list{font-size:11px;line-height:1.5;list-style:disc;padding-right:14px;color:#333;font-weight:bold;}'
        '.shading-example{display:flex;flex-direction:column;gap:6px;padding:3px;min-width:80px;}'
        '.shading-item{display:flex;align-items:center;gap:5px;font-size:11px;font-weight:bold;}'
        '.bubble-correct{width:20px;height:20px;background:#000;border-radius:50%;}'
        '.bubble-wrong{width:20px;height:20px;border:2px solid #000;border-radius:50%;text-align:center;line-height:16px;color:#000;font-weight:bold;}'
        '.student-section{display:flex;flex-direction:row;gap:6px;margin-bottom:6px;}'
        '.student-name-box{flex:1;background:white;border:2px solid #000;border-radius:5px;padding:5px;display:flex;flex-direction:column;justify-content:center;}'
        '.student-name-label{font-size:10px;color:#555;margin-bottom:2px;font-weight:bold;}'
        '.student-name{font-size:15px;font-weight:bold;color:#000;}'
        '.barcode-box{display:flex;flex-direction:column;align-items:center;justify-content:center;border:2px solid #000;padding:4px;border-radius:5px;background:white;width:185px;}'
        '.barcode-img{width:168px;height:55px;}'
        '.barcode-id-box{display:flex;flex-direction:column;align-items:center;justify-content:center;border:1px solid #000;padding:4px 8px;border-radius:4px;background:white;width:125px;}'
        '.status-options{display:flex;gap:8px;justify-content:center;margin-bottom:2px;font-size:10px;font-weight:bold;}'
        '.academic-id-label{font-size:9px;color:#666;margin-bottom:2px;text-align:center;}'
        '.academic-id-number{font-size:17px;font-weight:bold;letter-spacing:2px;}'
        '.info-grid{display:flex;flex-direction:row;gap:4px;margin:4px 0;}'
        '.info-item{flex:1;text-align:center;border:1px solid #000;padding:3px;background:white;}'
        '.info-label{font-size:10px;color:#333;font-weight:bold;}'
        '.info-value{font-size:11px;font-weight:bold;}'
        '.score-table-section{margin-top:4px;border:2px solid #000;}'
        '.score-table{width:100%;border-collapse:collapse;}'
        '.score-table th{background:white;border:1px solid #000;padding:2px;text-align:center;font-size:10px;font-weight:bold;}'
        '.score-table td{border:1px solid #000;padding:2px;text-align:center;height:26px;vertical-align:middle;}'
        '.score-bubble{width:13px;height:13px;border:2px solid #000;border-radius:50%;display:inline-block;background:white;}'
        '.score-num-header{font-size:11px;font-weight:bold;display:block;}'
        '.model-section{display:flex;align-items:center;gap:12px;margin-top:6px;padding:5px;background:white;border:1px solid #000;justify-content:center;}'
        '.model-bubbles{display:flex;flex-direction:column;align-items:center;gap:2px;}'
        '.model-bubble{width:17px;height:17px;border:2px solid #000;border-radius:50%;display:inline-block;background:white;}'
        '.model-bubble.selected{background:#000;}'
        '.questions-section{display:flex;flex-direction:row;flex:1;border-top:2px solid #000;overflow:hidden;}'
        '.question-column{border-left:2px solid #000;padding:4px 5px;display:flex;flex-direction:column;overflow:hidden;}'
        '.question-column:last-child{border-left:none;}'
        '.col-mcq1{width:22.5%;}.col-mcq2{width:22.5%;}.col-tf{width:16%;}.col-match{width:39%;}'
        '.column-title{text-align:center;padding:3px;font-weight:bold;font-size:11px;margin-bottom:3px;border-radius:3px;color:#000;background:white;border:2px solid #000;}'
        '.column-title.mcq{}.column-title.tf{}.column-title.matching{}'
        '.questions-grid{display:flex;flex-direction:row;gap:0 6px;flex:1;overflow:hidden;}'
        '.questions-grid>div{flex:1;overflow:hidden;display:flex;flex-direction:column;}'
        '.questions-grid.matching-grid{flex-direction:column;}'
        '.questions-grid.matching-grid>div{flex:1;display:flex;flex-direction:column;}'
        '.rows-wrapper{flex:1;display:flex;flex-direction:column;justify-content:space-evenly;}'
        '.question-row{display:flex;align-items:center;justify-content:center;padding:1px 0;gap:3px;}'
        '.question-num{width:19px;font-weight:bold;font-size:12px;text-align:center;}'
        '.answer-bubble{width:16px;height:16px;border:2px solid #000;border-radius:50%;display:inline-block;background:white;}'
        '.answer-bubble.filled{background:#000;}'
        '.column-headers{margin-bottom:4px;padding-bottom:3px;border-bottom:2px solid #333;display:flex;flex-direction:column;align-items:center;margin-right:22px;}'
        '.bubbles-labels-ar{display:flex;gap:4px;font-size:11px;font-weight:bold;color:#000;margin-bottom:1px;}'
        '.bubbles-labels-ar span{width:16px;text-align:center;display:inline-block;}'
        '.bubbles-labels-en{display:flex;gap:4px;font-size:10px;font-weight:bold;color:#666;margin-bottom:1px;}'
        '.bubbles-labels-en span{width:16px;text-align:center;display:inline-block;}'
        '.bubbles-row-only{display:flex;gap:4px;}'
        '.matching-bubble{width:16px;height:16px;border:2px solid #000;border-radius:50%;display:inline-block;background:white;}'
        '.matching-bubble.filled{background:#000;}'
        '.matching-labels-ar{display:flex;gap:4px;font-size:11px;font-weight:bold;color:#000;margin-bottom:1px;}'
        '.matching-labels-ar span{width:16px;text-align:center;display:inline-block;}'
        '.matching-labels-en{display:flex;gap:4px;font-size:10px;font-weight:bold;color:#666;margin-bottom:1px;}'
        '.matching-labels-en span{width:16px;text-align:center;display:inline-block;}'
        '.matching-row-only{display:flex;gap:4px;}'
        '</style></head><body>'
        + sheets_html
        + '</body></html>'
    )


def _remark_pdf_wrapper_colored(sheets_html):
    """يلف صفحات ورقة التظليل الملونة (CSS Grid) في HTML واحد — للاستخدام مع WebView فقط"""
    return (
        '<!DOCTYPE html><html dir="rtl" lang="ar">'
        '<head><meta charset="UTF-8">'
        '<style>'
        '@page{size:A4 landscape;margin:0;}'
        '*{margin:0;padding:0;box-sizing:border-box;-webkit-print-color-adjust:exact;print-color-adjust:exact;}'
        'html,body{width:100%;background:white;font-family:"Segoe UI","Traditional Arabic",Arial,sans-serif;font-size:12px;-webkit-print-color-adjust:exact;print-color-adjust:exact;}'
        '.sheet-container{width:285mm;height:190mm;display:grid;grid-template-columns:1fr 1.1fr;grid-template-rows:auto 1fr;border:2px solid #000;margin:5mm auto;page-break-inside:avoid;break-inside:avoid;}'
        '.sheet-container:not(:first-of-type){page-break-before:always;break-before:page;}'
        '.top-right{border-left:2px solid #000;border-bottom:2px solid #000;padding:8px;display:flex;flex-direction:column;}'
        '.logo-section{display:flex;align-items:flex-start;gap:12px;margin-bottom:6px;}'
        '.logo-img{width:70px;height:auto;}'
        '.ministry-info{text-align:right;font-size:12px;line-height:1.4;flex:1;}'
        '.ministry-info .title{font-weight:bold;color:#1a5c4c;font-size:13px;}'
        '.school-name{font-weight:bold;color:#c41e3a;font-size:13px;margin-top:2px;}'
        '.exam-info-box{background:#f0f7f0;border:1px solid #1a5c4c;border-radius:5px;padding:8px;margin-top:6px;font-size:12px;}'
        '.dor-section{display:flex;gap:20px;align-items:center;margin-top:6px;}'
        '.dor-option{display:flex;align-items:center;gap:6px;font-weight:bold;}'
        '.checkbox{width:18px;height:18px;border:2px solid #000;display:inline-block;background:white;}'
        '.instructions-box{background:#fff8e1;border:2px solid #f9a825;border-radius:6px;padding:8px;margin-top:8px;display:flex;justify-content:space-between;align-items:flex-start;gap:8px;flex:1;}'
        '.instructions-content{flex:1;}'
        '.instructions-title{font-weight:bold;color:#e65100;margin-bottom:4px;font-size:13px;}'
        '.instructions-list{font-size:11px;line-height:1.5;list-style:disc;padding-right:16px;color:#333;font-weight:bold;}'
        '.shading-example{display:flex;flex-direction:column;gap:8px;padding:4px;min-width:90px;}'
        '.shading-item{display:flex;align-items:center;gap:6px;font-size:12px;font-weight:bold;}'
        '.bubble-correct{width:22px;height:22px;background:#000;border-radius:50%;}'
        '.bubble-wrong{width:22px;height:22px;border:2px solid #000;border-radius:50%;display:flex;align-items:center;justify-content:center;color:#c41e3a;font-size:16px;font-weight:bold;}'
        '.top-left{border-bottom:2px solid #000;padding:8px;display:flex;flex-direction:column;}'
        '.student-section{display:grid;grid-template-columns:1fr auto auto;gap:8px;margin-bottom:8px;}'
        '.student-name-box{background:#e3f2fd;border:2px solid #1976d2;border-radius:6px;padding:6px;display:flex;flex-direction:column;justify-content:center;}'
        '.student-name-label{font-size:11px;color:#555;margin-bottom:3px;font-weight:bold;}'
        '.student-name{font-size:16px;font-weight:bold;color:#1565c0;}'
        '.barcode-box{display:flex;flex-direction:column;align-items:center;justify-content:center;border:2px solid #4caf50;padding:5px;border-radius:6px;background:#f1f8f4;min-width:180px;}'
        '.barcode-img{width:170px;height:60px;}'
        '.barcode-id-box{display:flex;flex-direction:column;align-items:center;justify-content:center;border:1px solid #ccc;padding:5px 10px;border-radius:5px;background:#fafafa;min-width:150px;}'
        '.status-options{display:flex;gap:10px;justify-content:center;margin-bottom:3px;font-size:11px;font-weight:bold;}'
        '.academic-id-label{font-size:9px;color:#666;margin-bottom:2px;}'
        '.academic-id-number{font-size:20px;font-weight:bold;letter-spacing:4px;}'
        '.info-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:5px;margin:5px 0;}'
        '.info-item{text-align:center;border:1px solid #000;padding:4px;background:#fafafa;}'
        '.info-label{font-size:11px;color:#333;font-weight:bold;}'
        '.info-value{font-size:12px;font-weight:bold;}'
        '.score-table-section{margin-top:4px;border:2px solid #000;}'
        '.score-table{width:100%;border-collapse:collapse;}'
        '.score-table th{background:#ddd;border:1px solid #000;padding:3px;text-align:center;font-size:11px;font-weight:bold;}'
        '.score-table td{border:1px solid #000;padding:2px;text-align:center;height:30px;vertical-align:middle;}'
        '.score-bubble{width:16px;height:16px;border:2px solid #000;border-radius:50%;display:inline-block;background:white;}'
        '.score-num-header{font-size:14px;font-weight:bold;display:block;}'
        '.model-section{display:flex;align-items:center;gap:15px;margin-top:8px;padding:6px;background:#f5f5f5;border:1px solid #999;justify-content:center;}'
        '.model-bubbles{display:flex;flex-direction:column;align-items:center;gap:2px;}'
        '.model-bubble{width:18px;height:18px;border:2px solid #000;border-radius:50%;display:inline-block;background:white;}'
        '.model-bubble.selected{background:#000;}'
        '.questions-section{grid-column:1 / -1;display:grid;grid-template-columns:0.9fr 0.9fr 0.65fr 1.55fr;border-top:2px solid #000;}'
        '.question-column{border-left:2px solid #000;padding:8px 10px;display:flex;flex-direction:column;}'
        '.question-column:last-child{border-left:none;}'
        '.column-title{text-align:center;padding:5px;font-weight:bold;font-size:13px;margin-bottom:6px;border-radius:4px;color:white;border:1px solid #000;}'
        '.column-title.mcq{background:#1a5c4c;}.column-title.tf{background:#e65100;}.column-title.matching{background:#7b1fa2;}'
        '.questions-grid{display:grid;grid-template-columns:1fr 1fr;gap:0 15px;}'
        '.questions-grid>div:first-child{border-left:1px solid #ddd;padding-left:8px;}'
        '.questions-grid>div:last-child{border-right:1px solid #ddd;padding-right:8px;}'
        '.questions-grid.matching-grid{grid-template-columns:1fr;}'
        '.questions-grid.matching-grid>div{border-left:none!important;border-right:none!important;padding-left:0!important;padding-right:0!important;}'
        '.question-row{display:flex;align-items:center;justify-content:center;padding:2px 0;gap:5px;}'
        '.question-num{width:22px;font-weight:bold;font-size:12px;text-align:center;}'
        '.answer-bubble{width:16px;height:16px;border:2px solid #000;border-radius:50%;display:inline-block;background:white;}'
        '.answer-bubble.filled{background:#000;}'
        '.column-headers{margin-bottom:8px;padding-bottom:5px;border-bottom:2px solid #333;display:flex;flex-direction:column;align-items:center;margin-right:27px;}'
        '.bubbles-labels-ar{display:flex;gap:4px;font-size:11px;font-weight:bold;color:#000;margin-bottom:2px;}'
        '.bubbles-labels-ar span{width:16px;text-align:center;display:inline-block;}'
        '.bubbles-labels-en{display:flex;gap:4px;font-size:10px;font-weight:bold;color:#666;margin-bottom:2px;}'
        '.bubbles-labels-en span{width:16px;text-align:center;display:inline-block;}'
        '.bubbles-row-only{display:flex;gap:4px;}'
        '.matching-bubble{width:16px;height:16px;border:2px solid #000;border-radius:50%;display:inline-block;background:white;}'
        '.matching-bubble.filled{background:#000;}'
        '.matching-labels-ar{display:flex;gap:4px;font-size:11px;font-weight:bold;color:#000;margin-bottom:2px;}'
        '.matching-labels-ar span{width:16px;text-align:center;display:inline-block;}'
        '.matching-labels-en{display:flex;gap:4px;font-size:10px;font-weight:bold;color:#666;margin-bottom:2px;}'
        '.matching-labels-en span{width:16px;text-align:center;display:inline-block;}'
        '.matching-row-only{display:flex;gap:4px;}'
        '</style></head><body>'
        + sheets_html
        + '</body></html>'
    )


def _html_to_pdf_response(html_str, filename):
    """دالة مساعدة: تحوّل HTML إلى PDF وترجعه كـ send_file"""
    from weasyprint import HTML as WeasyHTML
    buf = io.BytesIO()
    try:
        WeasyHTML(string=html_str, base_url=None).write_pdf(buf)
    except Exception as e:
        current_app.logger.error(f"WeasyPrint error: {e}")
        return html_str, 200, {'Content-Type': 'text/html; charset=utf-8'}
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=filename)


@question_bp.route('/remark-blank-pdf', methods=['POST'])
@login_required
def remark_blank_pdf():
    """ورقة تظليل فارغة → PDF"""
    try:
        data             = request.get_json()
        models           = data.get('models', ['أ'])
        count_per_model  = int(data.get('count_per_model', 10))
        question_ids     = data.get('question_ids', [])
        exam_type        = data.get('exam_type', 'نهاية')
        semester         = data.get('semester', 'الأول')
        academic_year    = data.get('academic_year', '1447هـ')

        if not question_ids:
            return jsonify({'error': 'لم يتم تحديد أسئلة'}), 400

        questions_count = len(get_ordered_questions_for_omr(question_ids))
        ctx = _remark_header_context(exam_type, semester, academic_year)

        all_html = ""
        for model_letter in models:
            for _ in range(count_per_model):
                all_html += render_template(
                    'question/remark_answer_sheet.html',
                    standalone=False,
                    student={
                        'name':         '..........................................',
                        'academic_id':  '..........................................',
                        'section':      '.....',
                        'barcode':      None,
                        'model_letter': model_letter,
                        'seat_no':      '',
                    },
                    model_letter=model_letter,
                    is_answer_key=False,
                    answers=None,
                    questions_count=questions_count,
                    **ctx
                )

        fname = f'remark_blank_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        return _html_to_pdf_response(_remark_pdf_wrapper(all_html), fname)

    except Exception as e:
        current_app.logger.exception(f"remark_blank_pdf error: {e}")
        return jsonify({'error': str(e)}), 500


@question_bp.route('/remark-answer-key-pdf', methods=['POST'])
@login_required
def remark_answer_key_pdf():
    """مفاتيح إجابة ريمارك → PDF (نموذج واحد أو كل النماذج)"""
    try:
        data          = request.get_json()
        question_ids  = data.get('question_ids', [])
        matching_pair_ids = data.get('matching_pair_ids', [])
        models        = data.get('models', ['أ'])
        shuffle_opts  = data.get('shuffle_options', True)
        exam_type     = data.get('exam_type', 'نهاية')
        semester      = data.get('semester', 'الأول')
        academic_year = data.get('academic_year', '1447هـ')

        if not question_ids:
            return jsonify({'error': 'لم يتم تحديد أسئلة'}), 400

        questions = get_ordered_questions_for_omr(question_ids)
        if not questions:
            return jsonify({'error': 'لم يتم العثور على الأسئلة'}), 404

        answer_keys, _ = _build_answer_keys_for_models(
            question_ids, matching_pair_ids, models, shuffle_options=shuffle_opts
        )

        ctx = _remark_header_context(exam_type, semester, academic_year)
        all_html = ""

        for idx, model_letter in enumerate(models):
            answers = answer_keys[model_letter]

            all_html += render_template(
                'question/remark_answer_sheet.html',
                standalone=False,
                student={'name': f'مفتاح الإجابة - نموذج {model_letter}', 'academic_id': '---', 'section': '---', 'barcode': None},
                is_answer_key=True,
                answers=answers,
                model_letter=model_letter,
                questions_count=len(questions),
                **ctx
            )

        fname = f'remark_answer_key_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        return _html_to_pdf_response(_remark_pdf_wrapper(all_html), fname)

    except Exception as e:
        current_app.logger.exception(f"remark_answer_key_pdf error: {e}")
        return jsonify({'error': str(e)}), 500


@question_bp.route('/remark-students-pdf', methods=['POST'])
@login_required
def remark_students_pdf():
    """
    ورقة تظليل بأسماء الطلاب → PDF
    يقبل مصدرين:
      - students: [{name, academic_id, section, seat_no}]  ← من Excel
      - student_ids: [1, 2, 3]                              ← من DB
    """
    try:
        data          = request.get_json()
        question_ids  = data.get('question_ids', [])
        matching_pair_ids = data.get('matching_pair_ids', [])
        models        = data.get('models', ['أ'])
        shuffle_opts  = data.get('shuffle_options', True)
        exam_type     = data.get('exam_type', 'نهاية')
        semester      = data.get('semester', 'الأول')
        academic_year = data.get('academic_year', '1447هـ')

        if not question_ids:
            return jsonify({'error': 'لم يتم تحديد أسئلة'}), 400

        # ── جلب الطلاب ──────────────────────────────────────────────
        students_list = data.get('students', [])

        if not students_list:
            # من قاعدة البيانات
            student_ids = data.get('student_ids', [])
            if not student_ids:
                return jsonify({'error': 'لم يتم تحديد طلاب'}), 400
            try:
                from src.models.student import Student as StudentModel
            except ImportError:
                from models.student import Student as StudentModel
            db_students = StudentModel.query.filter(StudentModel.id.in_(student_ids)).all()
            for idx, s in enumerate(db_students):
                students_list.append({
                    'name':        s.name,
                    'academic_id': str(s.id),
                    'section':     s.grade or '',
                    'seat_no':     str(idx + 1),
                })

        if not students_list:
            return jsonify({'error': 'لا يوجد طلاب'}), 400

        # ── جلب الأسئلة (تستبعد fill_blank/essay غير المتوافقة مع التظليل) ──────
        questions = get_ordered_questions_for_omr(question_ids)
        if not questions:
            return jsonify({'error': 'لم يتم العثور على الأسئلة'}), 404

        ctx = _remark_header_context(exam_type, semester, academic_year)

        # ── بناء مفاتيح الإجابة لكل نموذج — عبر نفس بايبلاين ورقة الاختبار الفعلية ──
        answer_keys, _ = _build_answer_keys_for_models(
            question_ids, matching_pair_ids, models, shuffle_options=shuffle_opts
        )

        # ── توزيع الطلاب على النماذج ────────────────────────────────
        per_model   = len(students_list) // len(models)
        remainder   = len(students_list) % len(models)
        all_html    = ""
        student_idx = 0

        for m_idx, model_letter in enumerate(models):
            count = per_model + (1 if m_idx < remainder else 0)
            for _ in range(count):
                if student_idx >= len(students_list):
                    break
                s = dict(students_list[student_idx])
                student_idx += 1
                # رقم الجلوس: من البيانات أو تسلسلي
                if not s.get('seat_no'):
                    s['seat_no'] = str(student_idx)
                # باركود
                acad_id = s.get('academic_id', '')
                s['barcode'] = generate_student_barcode(acad_id) if acad_id else None
                s['model_letter'] = model_letter

                all_html += render_template(
                    'question/remark_answer_sheet.html',
                    standalone=False,
                    student=s,
                    model_letter=model_letter,
                    is_answer_key=False,
                    answers=None,
                    questions_count=len(questions),
                    **ctx
                )

        # ── مفاتيح الإجابة في النهاية ────────────────────────────────
        for model_letter in models:
            all_html += render_template(
                'question/remark_answer_sheet.html',
                standalone=False,
                student={'name': f'مفتاح الإجابة - نموذج {model_letter}', 'academic_id': '---', 'section': '---', 'barcode': None},
                is_answer_key=True,
                answers=answer_keys[model_letter],
                model_letter=model_letter,
                questions_count=len(questions),
                **ctx
            )

        fname = f'remark_students_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        return _html_to_pdf_response(_remark_pdf_wrapper(all_html), fname)

    except Exception as e:
        current_app.logger.exception(f"remark_students_pdf error: {e}")
        return jsonify({'error': str(e)}), 500


@question_bp.route('/remark-excel-template', methods=['GET'])
@login_required
def remark_excel_template():
    """تحميل نموذج Excel فارغ لرفع أسماء الطلاب"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = 'قائمة الطلاب'
        headers = ['الاسم', 'الرقم الأكاديمي', 'الشعبة', 'رقم الجلوس']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color='FFFFFF')
            cell.fill      = PatternFill('solid', fgColor='2563EB')
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[chr(64 + col)].width = 22
        # صفوف مثال
        ws.append(['أحمد محمد علي', '202501001', 'أ', '1'])
        ws.append(['سارة خالد المطيري', '202501002', 'ب', '2'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name='نموذج_قائمة_الطلاب.xlsx')
    except Exception as e:
        current_app.logger.exception(f"remark_excel_template error: {e}")
        return jsonify({'error': str(e)}), 500


# =====================================================
# ===== ورقة تظليل Remark — HTML (للتحويل على الجهاز) =====
# =====================================================

@question_bp.route('/remark-blank-html', methods=['POST'])
@_remark_auth_required
def remark_blank_html():
    """ورقة تظليل فارغة → HTML JSON (يتحوّل لـ PDF على جهاز المستخدم)"""
    try:
        _tid = _q_get_teacher_id()
        if _tid:
            _ok, _rem, _lim, _msg = _q_check_remark_quota(_tid)
            if not _ok:
                return jsonify({'success': False, 'error': _msg,
                                'quota_exceeded': True, 'limit': _lim, 'remaining': 0}), 403
        data            = request.get_json()
        models          = data.get('models', ['أ'])
        count_per_model = int(data.get('count_per_model', 10))
        question_ids    = data.get('question_ids', [])
        exam_type       = data.get('exam_type', 'نهاية')
        semester        = data.get('semester', 'الأول')
        academic_year   = data.get('academic_year', '1447هـ')
        style           = data.get('style', 'clean')

        if not question_ids:
            return jsonify({'error': 'لم يتم تحديد أسئلة'}), 400

        questions_count = len(get_ordered_questions_for_omr(question_ids))
        ctx = _remark_header_context(exam_type, semester, academic_year)
        template_name = 'question/remark_answer_sheet_colored.html' if style == 'colored' else 'question/remark_answer_sheet.html'

        all_html = ""
        for model_letter in models:
            for _ in range(count_per_model):
                all_html += render_template(
                    template_name,
                    standalone=False,
                    student={'name': '..........................................',
                             'academic_id': '..........................................',
                             'section': '.....', 'barcode': None, 'seat_no': ''},
                    model_letter=model_letter,
                    is_answer_key=False,
                    answers=None,
                    questions_count=questions_count,
                    **ctx
                )
        wrapper = _remark_pdf_wrapper_colored if style == 'colored' else _remark_pdf_wrapper
        if _tid:
            try:
                from src.models.teacher_export_log import TeacherExportLog
                db.session.add(TeacherExportLog(teacher_id=_tid, export_type='remark'))
                db.session.commit()
            except Exception:
                pass
        return jsonify({'html': wrapper(all_html)})
    except Exception as e:
        current_app.logger.exception(f"remark_blank_html error: {e}")
        return jsonify({'error': str(e)}), 500


@question_bp.route('/remark-answer-key-html', methods=['POST'])
@_remark_auth_required
def remark_answer_key_html():
    """مفاتيح إجابة ريمارك → HTML JSON"""
    try:
        _tid = _q_get_teacher_id()
        if _tid:
            _ok, _rem, _lim, _msg = _q_check_remark_quota(_tid)
            if not _ok:
                return jsonify({'success': False, 'error': _msg,
                                'quota_exceeded': True, 'limit': _lim, 'remaining': 0}), 403
        data          = request.get_json()
        question_ids  = data.get('question_ids', [])
        matching_pair_ids = data.get('matching_pair_ids', [])
        models        = data.get('models', ['أ'])
        shuffle_opts  = data.get('shuffle_options', True)
        exam_type     = data.get('exam_type', 'نهاية')
        semester      = data.get('semester', 'الأول')
        academic_year = data.get('academic_year', '1447هـ')
        style         = data.get('style', 'clean')

        if not question_ids:
            return jsonify({'error': 'لم يتم تحديد أسئلة'}), 400

        questions = get_ordered_questions_for_omr(question_ids)
        if not questions:
            return jsonify({'error': 'لم يتم العثور على الأسئلة'}), 404

        answer_keys, _ = _build_answer_keys_for_models(
            question_ids, matching_pair_ids, models, shuffle_options=shuffle_opts
        )

        ctx = _remark_header_context(exam_type, semester, academic_year)
        template_name = 'question/remark_answer_sheet_colored.html' if style == 'colored' else 'question/remark_answer_sheet.html'
        all_html = ""

        for idx, model_letter in enumerate(models):
            answers = answer_keys[model_letter]

            all_html += render_template(
                template_name,
                standalone=False,
                student={'name': f'مفتاح الإجابة - نموذج {model_letter}',
                         'academic_id': '---', 'section': '---', 'barcode': None},
                is_answer_key=True,
                answers=answers,
                model_letter=model_letter,
                questions_count=len(questions),
                **ctx
            )
        wrapper = _remark_pdf_wrapper_colored if style == 'colored' else _remark_pdf_wrapper
        if _tid:
            try:
                from src.models.teacher_export_log import TeacherExportLog
                db.session.add(TeacherExportLog(teacher_id=_tid, export_type='remark'))
                db.session.commit()
            except Exception:
                pass
        return jsonify({'html': wrapper(all_html)})
    except Exception as e:
        current_app.logger.exception(f"remark_answer_key_html error: {e}")
        return jsonify({'error': str(e)}), 500


@question_bp.route('/remark-students-html', methods=['POST'])
@_remark_auth_required
def remark_students_html():
    """ورقة تظليل بأسماء الطلاب → HTML JSON"""
    try:
        _tid = _q_get_teacher_id()
        if _tid:
            _ok, _rem, _lim, _msg = _q_check_remark_quota(_tid)
            if not _ok:
                return jsonify({'success': False, 'error': _msg,
                                'quota_exceeded': True, 'limit': _lim, 'remaining': 0}), 403
        data          = request.get_json()
        question_ids  = data.get('question_ids', [])
        matching_pair_ids = data.get('matching_pair_ids', [])
        models        = data.get('models', ['أ'])
        shuffle_opts  = data.get('shuffle_options', True)
        exam_type     = data.get('exam_type', 'نهاية')
        semester      = data.get('semester', 'الأول')
        academic_year = data.get('academic_year', '1447هـ')
        students_raw  = data.get('students', [])
        student_ids   = data.get('student_ids', [])
        style         = data.get('style', 'clean')

        if not question_ids:
            return jsonify({'error': 'لم يتم تحديد أسئلة'}), 400

        questions = get_ordered_questions_for_omr(question_ids)
        if not questions:
            return jsonify({'error': 'لم يتم العثور على الأسئلة'}), 404

        # بناء قائمة الطلاب
        if students_raw:
            students_list = [dict(s) for s in students_raw]
        elif student_ids:
            try:
                from src.models.student import Student as StudentModel
            except ImportError:
                from models.student import Student as StudentModel
            db_students = StudentModel.query.filter(StudentModel.id.in_(student_ids)).all()
            students_list = [{'name': s.name,
                              'academic_id': str(s.id),
                              'section': s.grade or '',
                              'seat_no': str(idx + 1)} for idx, s in enumerate(db_students)]
        else:
            return jsonify({'error': 'لم يتم تحديد طلاب'}), 400

        ctx = _remark_header_context(exam_type, semester, academic_year)
        template_name = 'question/remark_answer_sheet_colored.html' if style == 'colored' else 'question/remark_answer_sheet.html'

        # بناء مفاتيح الإجابة لكل نموذج — عبر نفس بايبلاين ورقة الاختبار الفعلية
        answer_keys, _ = _build_answer_keys_for_models(
            question_ids, matching_pair_ids, models, shuffle_options=shuffle_opts
        )

        all_html = ""
        student_idx = 0
        for model_letter in models:
            while student_idx < len(students_list):
                if student_idx >= len(students_list):
                    break
                s = dict(students_list[student_idx])
                student_idx += 1
                if not s.get('seat_no'):
                    s['seat_no'] = str(student_idx)
                acad_id = s.get('academic_id', '')
                s['barcode'] = generate_student_barcode(acad_id) if acad_id else None
                all_html += render_template(
                    template_name,
                    standalone=False,
                    student=s,
                    model_letter=model_letter,
                    is_answer_key=False,
                    answers=None,
                    questions_count=len(questions),
                    **ctx
                )

        for model_letter in models:
            all_html += render_template(
                template_name,
                standalone=False,
                student={'name': f'مفتاح الإجابة - نموذج {model_letter}',
                         'academic_id': '---', 'section': '---', 'barcode': None},
                is_answer_key=True,
                answers=answer_keys[model_letter],
                model_letter=model_letter,
                questions_count=len(questions),
                **ctx
            )

        wrapper = _remark_pdf_wrapper_colored if style == 'colored' else _remark_pdf_wrapper
        if _tid:
            try:
                from src.models.teacher_export_log import TeacherExportLog
                db.session.add(TeacherExportLog(teacher_id=_tid, export_type='remark'))
                db.session.commit()
            except Exception:
                pass
        return jsonify({'html': wrapper(all_html)})
    except Exception as e:
        current_app.logger.exception(f"remark_students_html error: {e}")
        return jsonify({'error': str(e)}), 500


# =====================================================
# ===== صفحة تصنيف الأسئلة بالذكاء الاصطناعي =====
# =====================================================

@question_bp.route('/classify')
@login_required
def classify_questions():
    """
    صفحة تصنيف الأسئلة بالذكاء الاصطناعي
    تعرض إحصائيات التصنيف وتتيح التصنيف التلقائي والتعديل اليدوي
    """
    return render_template('classify_questions.html')


@question_bp.route('/review-classifications')
@login_required
def review_classifications():
    """صفحة مراجعة تصنيف الذكاء الاصطناعي — سؤال واحد في كل مرة"""
    course_id  = request.args.get('course_id',  type=int)
    unit_id    = request.args.get('unit_id',    type=int)
    lesson_id  = request.args.get('lesson_id',  type=int)
    show_all   = request.args.get('show_all', '0') == '1'

    query = Question.query.options(
        joinedload(Question.options),
        joinedload(Question.lesson).joinedload(Lesson.unit).joinedload(Unit.course)
    )

    if not show_all:
        query = query.filter(Question.human_verified == False)

    if lesson_id:
        query = query.filter(Question.lesson_id == lesson_id)
    elif unit_id:
        query = query.filter(Question.lesson.has(Lesson.unit_id == unit_id))
    elif course_id:
        query = query.filter(Question.lesson.has(Lesson.unit.has(Unit.course_id == course_id)))

    query = query.order_by(Question.lesson_id.asc(), Question.question_id.asc())

    total_pending   = Question.query.filter(Question.human_verified == False).count()
    total_verified  = Question.query.filter(Question.human_verified == True).count()
    total_all       = Question.query.count()

    questions = query.all()
    courses   = Course.query.order_by(Course.name).all()

    return render_template(
        'question/review_classifications.html',
        questions=questions,
        courses=courses,
        total_pending=total_pending,
        total_verified=total_verified,
        total_all=total_all,
        show_all=show_all,
        course_id=course_id,
        unit_id=unit_id,
        lesson_id=lesson_id,
    )


@question_bp.route('/verify-classification', methods=['POST'])
@login_required
def verify_classification():
    """
    مراجعة تصنيف الذكاء الاصطناعي من قِبل المعلم.
    action=approve → يثبّت التصنيف الحالي ويضع human_verified=True
    action=correct → يحدّث difficulty/bloom_level ويضع human_verified=True
    """
    data = request.get_json(silent=True) or {}
    question_id = data.get('question_id') or request.form.get('question_id', type=int)
    action = data.get('action') or request.form.get('action', '')

    if not question_id:
        return jsonify({'success': False, 'error': 'question_id مطلوب'}), 400

    question = Question.query.get(question_id)
    if not question:
        return jsonify({'success': False, 'error': 'السؤال غير موجود'}), 404

    if action == 'correct':
        difficulty  = data.get('difficulty')  or request.form.get('difficulty', '')
        bloom_level = data.get('bloom_level') or request.form.get('bloom_level', '')
        valid_diff  = {'easy', 'medium', 'hard'}
        valid_bloom = {'remember', 'understand', 'apply', 'analyze', 'evaluate', 'create'}
        if difficulty not in valid_diff or bloom_level not in valid_bloom:
            return jsonify({'success': False, 'error': 'قيم غير صالحة'}), 400
        question.difficulty  = difficulty
        question.bloom_level = bloom_level

    question.human_verified = True
    db.session.commit()

    return jsonify({
        'success': True,
        'question_id': question_id,
        'difficulty':  question.difficulty,
        'bloom_level': question.bloom_level,
        'human_verified': True
    })


# =====================================================
# ===== APIs حفظ واسترجاع الاختبارات =====
# =====================================================

@question_bp.route('/saved-exams', methods=['GET'])
@login_required
def get_saved_exams():
    """جلب قائمة الاختبارات المحفوظة"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        course_id = request.args.get('course_id', type=int)
        search = request.args.get('search', '')
        
        query = SavedExam.query.filter_by(is_active=True)
        
        if course_id:
            query = query.filter_by(course_id=course_id)
        
        if search:
            query = query.filter(SavedExam.name.ilike(f'%{search}%'))
        
        query = query.order_by(SavedExam.created_at.desc())
        
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        
        return jsonify({
            'success': True,
            'exams': [exam.to_dict() for exam in pagination.items],
            'total': pagination.total,
            'pages': pagination.pages,
            'current_page': page
        })
        
    except Exception as e:
        current_app.logger.error(f"Error fetching saved exams: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@question_bp.route('/saved-exams', methods=['POST'])
@login_required
def save_exam():
    """حفظ اختبار جديد"""
    try:
        data = request.get_json()
        
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'success': False, 'error': 'اسم الاختبار مطلوب'}), 400
        
        question_ids = data.get('question_ids', [])
        if not question_ids:
            return jsonify({'success': False, 'error': 'يجب اختيار أسئلة للحفظ'}), 400
        
        # الحصول على ترتيب الخيارات لكل سؤال (إذا كان موجوداً)
        questions_with_order = data.get('questions_with_order', [])
        options_order = {}
        for q in questions_with_order:
            if q.get('options_order'):
                options_order[str(q['question_id'])] = q['options_order']
        
        # إنشاء الاختبار الجديد
        new_exam = SavedExam(
            name=name,
            description=data.get('description', ''),
            course_id=data.get('course_id'),
            unit_id=data.get('unit_id'),
            question_ids=question_ids,
            questions_count=len(question_ids),
            models=data.get('models', ['أ']),
            settings={
                'shuffle_questions': data.get('shuffle_questions', True),
                'shuffle_options': data.get('shuffle_options', True),
                'font_size': data.get('font_size', 14),
                'image_size': data.get('image_size', 100),
                'columns': data.get('columns', 2),
                'spacing': data.get('spacing', 'normal'),
                'options_layout': data.get('options_layout', 'vertical'),
                'options_order': options_order  # ترتيب الخيارات المخلوط لكل سؤال
            },
            header_settings=data.get('header_settings', {}),
            exam_type=data.get('exam_type', ''),
            semester=data.get('semester', ''),
            academic_year=data.get('academic_year', ''),
            created_by=current_user.id if hasattr(current_user, 'id') else None
        )
        
        db.session.add(new_exam)
        db.session.commit()
        
        current_app.logger.info(f"Exam saved: {new_exam.id} - {new_exam.name}")
        
        return jsonify({
            'success': True,
            'message': 'تم حفظ الاختبار بنجاح',
            'exam': new_exam.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error saving exam: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@question_bp.route('/saved-exams/<int:exam_id>', methods=['GET'])
@login_required
def get_saved_exam(exam_id):
    """جلب تفاصيل اختبار محفوظ"""
    try:
        exam = SavedExam.query.filter_by(id=exam_id, is_active=True).first()
        
        if not exam:
            return jsonify({'success': False, 'error': 'الاختبار غير موجود'}), 404
        
        return jsonify({
            'success': True,
            'exam': exam.to_dict()
        })
        
    except Exception as e:
        current_app.logger.error(f"Error fetching exam {exam_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@question_bp.route('/saved-exams/<int:exam_id>', methods=['PUT'])
@login_required
def update_saved_exam(exam_id):
    """تحديث اختبار محفوظ"""
    try:
        exam = SavedExam.query.filter_by(id=exam_id, is_active=True).first()
        
        if not exam:
            return jsonify({'success': False, 'error': 'الاختبار غير موجود'}), 404
        
        data = request.get_json()
        
        if 'name' in data:
            exam.name = data['name'].strip()
        if 'description' in data:
            exam.description = data['description']
        if 'question_ids' in data:
            exam.question_ids = data['question_ids']
            exam.questions_count = len(data['question_ids'])
        if 'models' in data:
            exam.models = data['models']
        if 'settings' in data:
            exam.settings = data['settings']
        if 'header_settings' in data:
            exam.header_settings = data['header_settings']
        if 'exam_type' in data:
            exam.exam_type = data['exam_type']
        if 'semester' in data:
            exam.semester = data['semester']
        if 'academic_year' in data:
            exam.academic_year = data['academic_year']
        
        exam.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم تحديث الاختبار بنجاح',
            'exam': exam.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating exam {exam_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@question_bp.route('/saved-exams/<int:exam_id>', methods=['DELETE'])
@login_required
def delete_saved_exam(exam_id):
    """حذف اختبار محفوظ (حذف ناعم)"""
    try:
        exam = SavedExam.query.filter_by(id=exam_id, is_active=True).first()
        
        if not exam:
            return jsonify({'success': False, 'error': 'الاختبار غير موجود'}), 404
        
        # حذف ناعم
        exam.is_active = False
        exam.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم حذف الاختبار بنجاح'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting exam {exam_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@question_bp.route('/saved-exams/<int:exam_id>/load', methods=['POST'])
@login_required
def load_saved_exam(exam_id):
    """تحميل اختبار محفوظ مع جلب الأسئلة بنفس الترتيب المحفوظ"""
    try:
        exam = SavedExam.query.filter_by(id=exam_id, is_active=True).first()
        
        if not exam:
            return jsonify({'success': False, 'error': 'الاختبار غير موجود'}), 404
        
        # جلب الأسئلة المحفوظة بنفس الترتيب الأصلي
        saved_question_ids = exam.question_ids or []
        
        # جلب ترتيب الخيارات المحفوظ (إذا وجد)
        settings = exam.settings or {}
        options_order = settings.get('options_order', {})
        
        # جلب جميع الأسئلة دفعة واحدة
        questions_dict = {}
        if saved_question_ids:
            questions = Question.query.filter(
                Question.question_id.in_(saved_question_ids)
            ).options(
                joinedload(Question.options),
                joinedload(Question.lesson).joinedload(Lesson.unit)
            ).all()
            
            # تحويل لقاموس للوصول السريع
            for q in questions:
                questions_dict[q.question_id] = q
        
        # بناء قائمة الأسئلة بنفس الترتيب المحفوظ
        questions_data = []
        for qid in saved_question_ids:
            q = questions_dict.get(qid)
            if q:  # السؤال موجود
                # جلب الخيارات
                options_list = []
                for opt in q.options:
                    options_list.append({
                        'option_id': getattr(opt, 'option_id', None),
                        'option_text': getattr(opt, 'option_text', '') or '',
                        'image_url': getattr(opt, 'image_url', None) or '',
                        'is_correct': getattr(opt, 'is_correct', False)
                    })
                
                # إذا كان هناك ترتيب محفوظ للخيارات، نرتبها حسبه
                saved_order = options_order.get(str(qid))
                if saved_order and len(saved_order) == len(options_list):
                    # إنشاء قاموس للخيارات حسب الـ ID
                    options_by_id = {opt['option_id']: opt for opt in options_list}
                    # ترتيب الخيارات حسب الترتيب المحفوظ
                    ordered_options = []
                    for opt_id in saved_order:
                        if opt_id in options_by_id:
                            ordered_options.append(options_by_id[opt_id])
                    # إذا تم ترتيب جميع الخيارات بنجاح
                    if len(ordered_options) == len(options_list):
                        options_list = ordered_options
                
                q_dict = {
                    'question_id': q.question_id,
                    'question_text': q.question_text or '',
                    'image_url': getattr(q, 'image_url', None) or '',
                    'difficulty': getattr(q, 'difficulty', 'medium'),
                    'bloom_level': getattr(q, 'bloom_level', 'remember'),
                    'unit': q.lesson.unit.name if q.lesson and q.lesson.unit else 'غير محدد',
                    'lesson': q.lesson.name if q.lesson else 'غير محدد',
                    'options': options_list
                }
                questions_data.append(q_dict)
        
        return jsonify({
            'success': True,
            'exam': exam.to_dict(),
            'questions': questions_data,
            'options_order': options_order  # إرسال ترتيب الخيارات للـ Frontend
        })
        
    except Exception as e:
        current_app.logger.error(f"Error loading exam {exam_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# --- Export Courses/Units/Lessons list route --- #
@question_bp.route("/export/courses_units_lessons")
@login_required
def export_courses_units_lessons():
    """
    تصدير قائمة بجميع المناهج والوحدات والدروس المتاحة في النظام
    """
    try:
        # جلب جميع الدروس مع علاقاتها
        lessons = Lesson.query.join(Unit).join(Course).order_by(
            Course.name,
            Unit.order_num,
            Lesson.order_num
        ).all()
        
        # إنشاء قائمة بالبيانات
        data = []
        for lesson in lessons:
            data.append({
                "Course Name": lesson.unit.course.name,
                "Unit Name": lesson.unit.name,
                "Lesson Name": lesson.name
            })
        
        # إنشاء DataFrame
        df = pd.DataFrame(data)
        
        # إزالة التكرارات (اختياري - يمكن إبقاء التكرارات لتوضيح جميع الدروس)
        # df = df.drop_duplicates()
        
        # إنشاء BytesIO object
        output = io.BytesIO()
        
        # كتابة DataFrame إلى BytesIO
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Courses_Units_Lessons')
            
            # ضبط عرض الأعمدة
            worksheet = writer.sheets['Courses_Units_Lessons']
            for i, col in enumerate(df.columns):
                max_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                col_letter = chr(65 + i) if i < 26 else chr(65 + i // 26 - 1) + chr(65 + i % 26)
                worksheet.column_dimensions[col_letter].width = max_width
        
        output.seek(0)
        
        # إرسال الملف
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='courses_units_lessons.xlsx'
        )
        
    except Exception as e:
        current_app.logger.exception(f"Error exporting courses/units/lessons list: {e}")
        flash(f"حدث خطأ أثناء تصدير قائمة المناهج والوحدات والدروس: {str(e)}", "danger")
        return redirect(url_for("question.import_questions"))


# ============================================================
# صفحة توليد الشرح بالجملة (درس / وحدة)
# ============================================================

@question_bp.route('/generate-explanations')
def generate_explanations_page():
    """صفحة توليد الشرح بالذكاء الاصطناعي للدروس والوحدات"""
    try:
        courses = Course.query.order_by(Course.name).all()
    except Exception:
        courses = []
    return render_template('question/generate_explanations.html', courses=courses)

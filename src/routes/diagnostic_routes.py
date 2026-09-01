# src/routes/diagnostic_routes.py
"""
API للاختبارات التشخيصية
- متوافق مع Flutter
- توليد اختبارات قبلية/بعدية
- استخراج PDF
"""

from flask import Blueprint, request, jsonify, send_file, render_template, current_app
from flask_login import login_required, current_user
from datetime import datetime, timezone, timedelta
from functools import wraps
import io
import os
import base64

try:
    from src.extensions import db
    from src.models.diagnostic_test import DiagnosticTest, DiagnosticResult, DiagnosticComparison
    from src.services.diagnostic_service import diagnostic_service
    from src.models.curriculum import Lesson, Unit, Course
    from src.models.student import Student
    from src.utils.field_encryption import make_email_hash
except ImportError:  # pragma: no cover
    from extensions import db
    from models.diagnostic_test import DiagnosticTest, DiagnosticResult, DiagnosticComparison
    from services.diagnostic_service import diagnostic_service
    from models.curriculum import Lesson, Unit, Course
    from models.student import Student
    from utils.field_encryption import make_email_hash

# ✅ استيراد موديل الإشعارات لحفظها في قاعدة البيانات
try:
    from src.models.notification import Notification
except ImportError:  # pragma: no cover
    try:
        from models.notification import Notification
    except:
        Notification = None
        print("⚠️ Notification model غير متوفر - الإشعارات لن تُحفظ في قاعدة البيانات")

# ✅ استيراد StudentNotification (جدول الربط بين الطالب والإشعار)
try:
    from src.models.student_notification import StudentNotification
except ImportError:  # pragma: no cover
    try:
        from models.student_notification import StudentNotification
    except:
        try:
            from src.models.notification import StudentNotification
        except:
            try:
                from models.notification import StudentNotification
            except:
                StudentNotification = None
                print("⚠️ StudentNotification model غير متوفر")

# ✅ خدمة الإشعارات
try:
    from src.services.notification_service import NotificationService
except ImportError:  # pragma: no cover
    try:
        from services.notification_service import NotificationService
    except:
        NotificationService = None
        print("⚠️ NotificationService غير متوفر")

diagnostic_bp = Blueprint('diagnostic', __name__, url_prefix='/api/diagnostic')


def _save_notification_to_db(student_id, title, message, notification_type='reminder', data=None):
    """
    حفظ الإشعار في قاعدة البيانات ليظهر في صفحة الإشعارات
    ✅ يحفظ في جدولين: Notification + StudentNotification (نفس طريقة إشعارات الأدمن)
    """
    try:
        if Notification is None:
            print(f"⚠️ Notification model غير متوفر - لن يتم حفظ الإشعار في DB")
            return False
        
        # 1️⃣ إنشاء سجل في جدول Notification
        notification = Notification(
            student_id=student_id,
            title=title,
            message=message,
            body=message,  # بعض الأماكن تقرأ body بدل message
            notification_type=notification_type,
            type=notification_type,
            is_read=False,
            status='delivered',
            sent_at=datetime.utcnow(),
        )
        
        if data:
            notification.data = data
        
        db.session.add(notification)
        db.session.flush()  # للحصول على notification.id
        
        # 2️⃣ إنشاء سجل في جدول StudentNotification (الربط بين الطالب والإشعار)
        if StudentNotification is not None:
            student_notif = StudentNotification(
                student_id=student_id,
                notification_id=notification.id,
                is_read=False,
            )
            db.session.add(student_notif)
            print(f"  ✅ تم إنشاء Notification #{notification.id} + StudentNotification للطالب {student_id}")
        else:
            print(f"  ✅ تم إنشاء Notification #{notification.id} للطالب {student_id} (بدون StudentNotification)")
        
        return True
    except Exception as e:
        print(f"❌ خطأ في حفظ الإشعار في DB: {e}")
        import traceback
        traceback.print_exc()
        return False



def convert_saudi_to_utc(dt_string):
    """تحويل وقت من السعودية (UTC+3) إلى UTC"""
    try:
        # إزالة Z أو timezone info
        dt_string = dt_string.replace('Z', '').replace('+00:00', '')
        
        # parse التاريخ
        dt = datetime.fromisoformat(dt_string)
        
        # السعودية = UTC + 3، نطرح 3 ساعات
        dt_utc = dt - timedelta(hours=3)
        
        print(f"⏰ Converted {dt_string} (Saudi) → {dt_utc} (UTC)")
        return dt_utc
    except Exception as e:
        print(f"⚠️ Error converting timezone: {e}")
        # fallback - استخدم الوقت كما هو
        return datetime.fromisoformat(dt_string.replace('Z', ''))


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return jsonify({'success': False, 'error': 'يجب تسجيل الدخول'}), 401
        if not getattr(current_user, 'is_admin', False):
            return jsonify({'success': False, 'error': 'صلاحيات غير كافية'}), 403
        return f(*args, **kwargs)
    return decorated


# ==========================================
# توليد الاختبارات
# ==========================================

@diagnostic_bp.route('/generate', methods=['POST'])
@login_required
@admin_required
def generate_test():
    """
    توليد اختبار تشخيصي جديد
    
    POST /api/diagnostic/generate
    {
        "lesson_id": 1,  // أو unit_id أو course_id
        "test_type": "pre_test",  // أو post_test
        "questions_count": 5,
        "difficulty_distribution": {"easy": 2, "medium": 2, "hard": 1}
    }
    """
    try:
        data = request.get_json() or {}
        
        lesson_id = data.get('lesson_id')
        unit_id = data.get('unit_id')
        course_id = data.get('course_id')
        test_type = data.get('test_type', 'pre_test')
        questions_count = data.get('questions_count', 5)
        difficulty_dist = data.get('difficulty_distribution', {'easy': 2, 'medium': 2, 'hard': 1})
        force_ai = data.get('force_ai', False)  # توليد كل الأسئلة بـ AI
        
        if not any([lesson_id, unit_id, course_id]):
            return jsonify({'success': False, 'error': 'يجب تحديد درس أو وحدة أو منهج'}), 400
        
        # توليد الاختبار
        result = diagnostic_service.generate_test(
            lesson_id=lesson_id,
            unit_id=unit_id,
            course_id=course_id,
            test_type=test_type,
            questions_count=questions_count,
            difficulty_dist=difficulty_dist,
            force_ai=force_ai
        )
        
        if not result.get('success'):
            return jsonify(result), 400
        
        # جلب الأسماء من السياق
        context = result.get('context', {})
        lesson_name = context.get('name') if context.get('type') == 'lesson' else None
        unit_name = context.get('name') if context.get('type') == 'unit' else context.get('unit_name')
        course_name = context.get('course_name') or (context.get('name') if context.get('type') == 'course' else None)
        
        # حفظ في قاعدة البيانات
        test = DiagnosticTest(
            title=result['title'],
            description=result['description'],
            test_type=test_type,
            lesson_id=lesson_id,
            unit_id=unit_id,
            course_id=course_id,
            lesson_name=lesson_name,
            unit_name=unit_name,
            course_name=course_name,
            questions_count=result['questions_count'],
            questions_data=result['questions'],
            difficulty_distribution=difficulty_dist,
            ai_generated=result.get('ai_generated', False),
            created_by=current_user.id if hasattr(current_user, 'id') else None
        )
        
        db.session.add(test)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'تم إنشاء الاختبار {"القبلي" if test_type == "pre_test" else "البعدي"} بنجاح',
            'test': test.to_dict(include_questions=True)
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@diagnostic_bp.route('/generate-pair', methods=['POST'])
@login_required
@admin_required
def generate_test_pair():
    """
    توليد زوج اختبارات (قبلي + بعدي) معاً
    
    POST /api/diagnostic/generate-pair
    {
        "lesson_id": 1,
        "questions_count": 5
    }
    """
    try:
        data = request.get_json() or {}
        
        lesson_id = data.get('lesson_id')
        unit_id = data.get('unit_id')
        course_id = data.get('course_id')
        questions_count = data.get('questions_count', 5)

        if not any([lesson_id, unit_id, course_id]):
            return jsonify({'success': False, 'error': 'يجب تحديد درس أو وحدة أو منهج'}), 400

        # توليد القبلي
        pre_result = diagnostic_service.generate_test(
            lesson_id=lesson_id,
            unit_id=unit_id,
            course_id=course_id,
            test_type='pre_test',
            questions_count=questions_count
        )
        
        if not pre_result.get('success'):
            return jsonify(pre_result), 400
        
        # جلب الأسماء
        context = pre_result.get('context', {})
        lesson_name = context.get('name') if context.get('type') == 'lesson' else None
        unit_name = context.get('name') if context.get('type') == 'unit' else context.get('unit_name')
        course_name = context.get('course_name') or (context.get('name') if context.get('type') == 'course' else None)

        # حفظ القبلي
        pre_test = DiagnosticTest(
            title=pre_result['title'],
            description=pre_result['description'],
            test_type='pre_test',
            lesson_id=lesson_id,
            unit_id=unit_id,
            course_id=course_id,
            lesson_name=lesson_name,
            unit_name=unit_name,
            course_name=course_name,
            questions_count=pre_result['questions_count'],
            questions_data=pre_result['questions'],
            ai_generated=pre_result.get('ai_generated', False),
            created_by=current_user.id if hasattr(current_user, 'id') else None
        )
        db.session.add(pre_test)
        db.session.flush()  # للحصول على ID

        # توليد البعدي (نفس الأسئلة أو مختلفة)
        post_result = diagnostic_service.generate_test(
            lesson_id=lesson_id,
            unit_id=unit_id,
            course_id=course_id,
            test_type='post_test',
            questions_count=questions_count
        )

        # حفظ البعدي
        post_test = DiagnosticTest(
            title=post_result['title'],
            description=post_result['description'],
            test_type='post_test',
            lesson_id=lesson_id,
            unit_id=unit_id,
            course_id=course_id,
            lesson_name=lesson_name,
            unit_name=unit_name,
            course_name=course_name,
            questions_count=post_result.get('questions_count', questions_count),
            questions_data=post_result.get('questions', pre_result['questions']),
            ai_generated=post_result.get('ai_generated', False),
            paired_test_id=pre_test.id,
            created_by=current_user.id if hasattr(current_user, 'id') else None
        )
        db.session.add(post_test)
        
        # ربط القبلي بالبعدي
        pre_test.paired_test_id = post_test.id
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم إنشاء الاختبارين القبلي والبعدي بنجاح',
            'pre_test': pre_test.to_dict(),
            'post_test': post_test.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==========================================
# جلب الاختبارات
# ==========================================

@diagnostic_bp.route('/tests', methods=['GET'])
@login_required
def get_tests():
    """جلب قائمة الاختبارات"""
    try:
        lesson_id = request.args.get('lesson_id', type=int)
        unit_id = request.args.get('unit_id', type=int)
        test_type = request.args.get('test_type')
        
        query = DiagnosticTest.query.filter_by(is_active=True)
        
        if lesson_id:
            query = query.filter_by(lesson_id=lesson_id)
        if unit_id:
            query = query.filter_by(unit_id=unit_id)
        if test_type:
            query = query.filter_by(test_type=test_type)
        
        tests = query.order_by(DiagnosticTest.created_at.desc()).all()
        
        return jsonify({
            'success': True,
            'tests': [t.to_dict() for t in tests],
            'count': len(tests)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@diagnostic_bp.route('/tests/<int:test_id>', methods=['GET'])
@login_required
def get_test(test_id):
    """جلب تفاصيل اختبار مع الأسئلة"""
    try:
        test = DiagnosticTest.query.filter_by(id=test_id, is_active=True).first()
        
        if not test:
            return jsonify({'success': False, 'error': 'الاختبار غير موجود'}), 404
        
        return jsonify({
            'success': True,
            'test': test.to_dict(include_questions=True)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==========================================
# استخراج PDF
# ==========================================

@diagnostic_bp.route('/tests/<int:test_id>/pdf', methods=['GET'])
@login_required
def export_pdf(test_id):
    """
    استخراج ورقة اختبار PDF بتنسيق احترافي
    
    GET /api/diagnostic/tests/1/pdf?include_answers=true&columns=2&layout=grid
    
    Parameters:
        include_answers: إظهار الإجابات (true/false)
        columns: عدد الأعمدة (1, 2, 3) - افتراضي: 2
        layout: تنسيق الخيارات (vertical, horizontal, grid) - افتراضي: grid
    """
    try:
        test = DiagnosticTest.query.filter_by(id=test_id, is_active=True).first()
        
        if not test:
            return jsonify({'success': False, 'error': 'الاختبار غير موجود'}), 404
        
        # قراءة المعاملات
        include_answers = request.args.get('include_answers', 'false').lower() == 'true'
        columns = int(request.args.get('columns', 2))
        options_layout = request.args.get('layout', 'grid')
        font_family = request.args.get('font_family', 'cairo')

        # التحقق من القيم
        if columns not in [1, 2, 3]:
            columns = 2
        if options_layout not in ['vertical', 'horizontal', 'grid']:
            options_layout = 'grid'
        if font_family not in ['traditional', 'cairo', 'tajawal', 'amiri', 'tahoma', 'scheherazade', 'noto']:
            font_family = 'cairo'
        
        # جلب إعدادات الكليشة
        header_settings = {}
        try:
            from src.models.exam_header_settings import ExamHeaderSettings
            header = ExamHeaderSettings.query.first()
            if header:
                header_settings = {
                    'country': header.country or "المملكة العربية السعودية",
                    'ministry': header.ministry or "وزارة التعليم",
                    'school_name': header.school_name or "",
                    'subject': header.subject or "كيمياء",
                    'grade': header.grade or ""
                }
            
            # جلب الشعار من الملف
            import os
            import base64
            logo_path = os.path.join(current_app.root_path, 'static', 'images', 'logo.png')
            if os.path.exists(logo_path):
                with open(logo_path, 'rb') as f:
                    logo_base64 = base64.b64encode(f.read()).decode('utf-8')
                    header_settings['logo_base64'] = f"data:image/png;base64,{logo_base64}"
        except Exception as e:
            print(f"⚠️ Error loading header settings: {e}")
        
        # بناء HTML بالتنسيق المطلوب
        html_content = generate_diagnostic_html(
            test,
            include_answers,
            header_settings,
            columns=columns,
            options_layout=options_layout,
            font_family=font_family,
        )
        
        # تحويل إلى PDF باستخدام WeasyPrint
        try:
            from weasyprint import HTML, CSS
            pdf_bytes = HTML(string=html_content).write_pdf()
        except Exception as e:
            print(f"❌ WeasyPrint Error: {e}")
            return jsonify({'success': False, 'error': f'خطأ في توليد PDF: {str(e)}'}), 500
        
        if not pdf_bytes:
            return jsonify({'success': False, 'error': 'فشل توليد PDF'}), 500
        
        # إرسال الملف
        filename = f"diagnostic_test_{test_id}{'_answers' if include_answers else ''}.pdf"
        
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ PDF Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def generate_diagnostic_html(test, include_answers=False, header_settings=None, columns=2, options_layout='grid', font_family='cairo'):
    """
    توليد HTML للاختبار التشخيصي بتنسيق احترافي
    
    Args:
        test: كائن الاختبار
        include_answers: إظهار الإجابات
        header_settings: إعدادات الكليشة
        columns: عدد الأعمدة (1، 2، 3)
        options_layout: تنسيق الخيارات (vertical، horizontal، grid)
    """
    
    questions = test.questions_data or []
    
    # تحديد عرض الأعمدة
    if columns == 1:
        column_width = "100%"
        questions_per_column = len(questions)
    elif columns == 2:
        column_width = "48%"
        questions_per_column = (len(questions) + 1) // 2
    else:  # 3 columns
        column_width = "31%"
        questions_per_column = (len(questions) + 2) // 3
    
    # تضمين الخط المختار بـ base64
    from src.routes.exam_generator import _font_base64, _FONT_FILES
    _FONT_CSS_NAMES = {
        'traditional':  "'Traditional Arabic', Arial, sans-serif",
        'cairo':        "'Cairo', Arial, sans-serif",
        'tajawal':      "'Tajawal', Arial, sans-serif",
        'amiri':        "'Amiri', Arial, sans-serif",
        'tahoma':       "Tahoma, Arial, sans-serif",
        'scheherazade': "'Scheherazade New', Arial, sans-serif",
        'noto':         "'Noto Naskh Arabic', Arial, sans-serif",
    }
    font_css_name = _FONT_CSS_NAMES.get(font_family, _FONT_CSS_NAMES['cairo'])
    font_filename = _FONT_FILES.get(font_family, _FONT_FILES['cairo'])[0]
    font_b64 = _font_base64(font_filename)
    font_face_css = f"@font-face {{ font-family: '{font_css_name.split(',')[0].strip(chr(39))}'; src: url('{font_b64}'); }}" if font_b64 else ''

    # CSS للتنسيق الاحترافي
    css = f"""
    <style>
        {font_face_css}
        @page {{
            size: A4;
            margin: 1cm 1cm 1.5cm 1cm;
            @bottom-center {{
                content: "صفحة " counter(page) " من " counter(pages);
                font-family: {font_css_name};
                font-size: 7pt;
                color: #94a3b8;
            }}
            @bottom-left {{
                content: "تم إنشاؤه بواسطة تطبيق كيم تحصيلي";
                font-family: {font_css_name};
                font-size: 7pt;
                color: #94a3b8;
            }}
        }}
        * {{
            box-sizing: border-box;
        }}
        body {{
            font-family: {font_css_name};
            direction: rtl;
            text-align: right;
            font-size: 12px;
            line-height: 1.4;
            margin: 0;
            padding: 10px;
        }}
        
        /* === الكليشة === */
        .header-table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 10px;
        }}
        .header-table td {{
            padding: 3px 8px;
            vertical-align: top;
        }}
        .header-right, .header-left {{
            width: 35%;
            font-size: 11px;
        }}
        .header-center {{
            width: 30%;
            text-align: center;
        }}
        .header-center img {{
            max-width: 60px;
            max-height: 60px;
        }}
        
        /* === عنوان الاختبار === */
        .exam-title {{
            text-align: center;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 10px;
            border-radius: 8px;
            margin: 10px 0;
            font-size: 16px;
            font-weight: bold;
        }}
        
        /* === معلومات الطالب === */
        .student-info {{
            display: flex;
            justify-content: space-between;
            border: 1px solid #333;
            padding: 8px 15px;
            margin: 10px 0;
            background: #f9f9f9;
            border-radius: 5px;
        }}
        .student-info span {{
            font-size: 12px;
        }}
        
        /* === الأعمدة === */
        .questions-table {{
            width: 100%;
            border-collapse: separate;
            border-spacing: 8px 0;
        }}
        .questions-table td {{
            width: {column_width};
            vertical-align: top;
            padding: 0;
        }}
        
        /* === السؤال === */
        .question {{
            margin-bottom: 12px;
            padding: 8px;
            border: 1px solid #e0e0e0;
            border-radius: 6px;
            background: #fafafa;
            page-break-inside: avoid;
        }}
        .question-text {{
            font-weight: bold;
            font-size: 12px;
            margin-bottom: 8px;
            color: #333;
            line-height: 1.5;
        }}
        
        /* === الخيارات - شبكة === */
        .options-grid {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 6px 20px;
        }}
        /* ترتيب الخيارات: أ يمين، ب يسار */
        .options-grid .option:nth-child(1) {{ order: 1; }}
        .options-grid .option:nth-child(2) {{ order: 2; }}
        .options-grid .option:nth-child(3) {{ order: 3; }}
        .options-grid .option:nth-child(4) {{ order: 4; }}
        
        /* === الخيارات - أفقي === */
        .options-horizontal {{
            display: flex;
            flex-wrap: wrap;
            gap: 10px;
        }}
        .options-horizontal .option {{
            flex: 1;
            min-width: 45%;
        }}
        
        /* === الخيارات - عمودي === */
        .options-vertical {{
            display: flex;
            flex-direction: column;
            gap: 4px;
        }}
        
        /* === الخيار === */
        .option {{
            padding: 5px 8px;
            border: 1px solid #ddd;
            border-radius: 4px;
            font-size: 12px;
            background: white;
            text-align: right;
        }}
        .option.correct {{
            background: #d4edda;
            border-color: #28a745;
            font-weight: bold;
        }}
        .option-letter {{
            display: inline-block;
            width: 22px;
            height: 22px;
            line-height: 22px;
            text-align: center;
            background: #667eea;
            color: white;
            border-radius: 50%;
            font-size: 11px;
            margin-left: 8px;
            font-weight: bold;
        }}
        
        /* === نموذج الإجابة === */
        .answer-key {{
            page-break-before: always;
            margin-top: 20px;
        }}
        .answer-key h2 {{
            text-align: center;
            background: #28a745;
            color: white;
            padding: 10px;
            border-radius: 8px;
            margin-bottom: 15px;
        }}
        .answer-grid {{
            display: grid;
            grid-template-columns: repeat(5, 1fr);
            gap: 8px;
            max-width: 500px;
            margin: 0 auto;
        }}
        .answer-box {{
            border: 2px solid #333;
            padding: 8px;
            text-align: center;
            border-radius: 5px;
            background: #f5f5f5;
        }}
        .answer-box .q-num {{
            font-weight: bold;
            color: #667eea;
        }}
        .answer-box .q-ans {{
            font-size: 14px;
            font-weight: bold;
            color: #28a745;
        }}
        
        @media print {{
            body {{
                -webkit-print-color-adjust: exact;
                print-color-adjust: exact;
            }}
        }}
    </style>
    """
    
    # === الكليشة ===
    header_html = ""
    if header_settings:
        # الشعار
        logo_base64 = header_settings.get('logo_base64', '')
        if logo_base64:
            logo_html = f'<img src="{logo_base64}" style="max-width: 60px; max-height: 60px;">'
        else:
            logo_html = ''
        
        header_html = f"""
        <table class="header-table">
            <tr>
                <td class="header-right">
                    <div style="font-weight: bold;">{header_settings.get('country', 'المملكة العربية السعودية')}</div>
                    <div>{header_settings.get('ministry', 'وزارة التعليم')}</div>
                    <div>{header_settings.get('school_name', '')}</div>
                </td>
                <td class="header-center">
                    {logo_html}
                </td>
                <td class="header-left" style="text-align: left;">
                    <div>المادة: {header_settings.get('subject', 'كيمياء')}</div>
                    <div>الصف: {header_settings.get('grade', '')}</div>
                    <div>الزمن: {test.time_limit_minutes} دقيقة</div>
                </td>
            </tr>
        </table>
        """
    
    # === عنوان الاختبار ===
    test_type_ar = 'قبلي' if test.test_type == 'pre_test' else 'بعدي'
    title_html = f"""
    <div class="exam-title">
        {test.title or f'اختبار تشخيصي {test_type_ar}'}
        <br>
        <span style="font-size: 12px; font-weight: normal;">عدد الأسئلة: {len(questions)}</span>
    </div>
    """
    
    # === معلومات الطالب ===
    student_html = """
    <div class="student-info">
        <span>الاسم: ________________________________</span>
        <span>الصف: ____________</span>
        <span>التاريخ: ____/____/______</span>
    </div>
    """
    
    # === تحديد class الخيارات ===
    options_class = f"options-{options_layout}"
    
    # === بناء الأسئلة ===
    # === الأحرف العربية للخيارات ===
    arabic_letters = ['أ', 'ب', 'ج', 'د', 'هـ', 'و']
    
    def build_question_html(q, num):
        q_text = q.get('text', '')
        options_list = q.get('options', [])
        
        # بناء الخيارات بترتيب صحيح للشبكة RTL
        # الترتيب المطلوب في HTML: [ب، أ، د، ج] ليظهر [أ، ب] في الصف الأول و[ج، د] في الثاني
        if options_layout == 'grid' and len(options_list) >= 4:
            # إعادة ترتيب: نضع ب قبل أ، د قبل ج
            reordered = [
                options_list[1] if len(options_list) > 1 else None,  # ب
                options_list[0] if len(options_list) > 0 else None,  # أ
                options_list[3] if len(options_list) > 3 else None,  # د
                options_list[2] if len(options_list) > 2 else None,  # ج
            ]
            reordered = [o for o in reordered if o]  # إزالة None
            letter_order = ['ب', 'أ', 'د', 'ج']
        else:
            reordered = options_list
            letter_order = arabic_letters
        
        options_html = ""
        for idx, opt in enumerate(reordered):
            if options_layout == 'grid' and len(options_list) >= 4:
                letter = letter_order[idx] if idx < len(letter_order) else str(idx + 1)
            else:
                letter = arabic_letters[idx] if idx < len(arabic_letters) else str(idx + 1)
            
            text = opt.get('text', '')
            is_correct = opt.get('is_correct', False)
            
            correct_class = ' correct' if include_answers and is_correct else ''
            check_mark = ' ✓' if include_answers and is_correct else ''
            
            options_html += f'''
            <div class="option{correct_class}">{letter}- {text}{check_mark}</div>
            '''
        
        return f'''
        <div class="question">
            <div class="question-text">س{num}: {q_text}</div>
            <div class="{options_class}">{options_html}</div>
        </div>
        '''
    
    # === توزيع الأسئلة ===
    if columns == 1:
        questions_html = '<div>'
        for i, q in enumerate(questions, 1):
            questions_html += build_question_html(q, i)
        questions_html += '</div>'
    else:
        # توزيع على جدول: كل صف يحمل (columns) أسئلة متجاورة
        questions_html = '<table class="questions-table">'
        for row_start in range(0, len(questions), columns):
            questions_html += '<tr>'
            for col in range(columns):
                idx = row_start + col
                if idx < len(questions):
                    questions_html += f'<td>{build_question_html(questions[idx], idx + 1)}</td>'
                else:
                    questions_html += '<td></td>'
            questions_html += '</tr>'
        questions_html += '</table>'
    
    # === نموذج الإجابة ===
    answer_key_html = ""
    if include_answers:
        answers_grid = ""
        for i, q in enumerate(questions, 1):
            correct = next((o for o in q.get('options', []) if o.get('is_correct')), None)
            if correct:
                answers_grid += f'''
                <div class="answer-box">
                    <div class="q-num">س{i}</div>
                    <div class="q-ans">{correct.get('letter', '')}</div>
                </div>
                '''
        
        answer_key_html = f"""
        <div class="answer-key">
            <h2>🔑 نموذج الإجابة</h2>
            <div class="answer-grid">
                {answers_grid}
            </div>
        </div>
        """
    
    # === HTML النهائي ===
    html = f"""
    <!DOCTYPE html>
    <html dir="rtl" lang="ar">
    <head>
        <meta charset="UTF-8">
        {css}
    </head>
    <body>
        {header_html}
        {title_html}
        {student_html}
        {questions_html}
        {answer_key_html}
    </body>
    </html>
    """
    
    return html
    
    # Answer Key (if needed)
    answer_key_html = ""
    if include_answers:
        rows = ""
        for i, q in enumerate(questions, 1):
            correct = next((o for o in q.get('options', []) if o.get('is_correct')), None)
            if correct:
                rows += f"<tr><td>س{i}</td><td>{correct.get('letter', '')}</td></tr>"
        
        answer_key_html = f"""
        <div class="answer-key">
            <h2>نموذج الإجابة</h2>
            <table class="answer-table">
                <tr><th>السؤال</th><th>الإجابة</th></tr>
                {rows}
            </table>
        </div>
        """
    
    # Final HTML
    html = f"""
    <!DOCTYPE html>
    <html dir="rtl" lang="ar">
    <head>
        <meta charset="UTF-8">
        {css}
    </head>
    <body>
        {header_html}
        {title_html}
        {student_html}
        {questions_html}
        {answer_key_html}
    </body>
    </html>
    """
    
    return html


# ==========================================
# حل الاختبار (للطالب)
# ==========================================

def _get_shuffled_questions(questions_data, seed):
    """
    يرجّع نسخة من الأسئلة (وخياراتها) بترتيب عشوائي يعتمد على seed ثابت (result.id).
    نفس الـ seed يرجع نفس الترتيب بالضبط دائماً — يُستخدم وقت عرض الاختبار للطالب
    وأيضاً وقت التصحيح، فيبقى التصحيح صحيح دون الحاجة لتخزين أي بيانات إضافية.
    """
    import random
    import copy
    rng = random.Random(seed)
    questions = copy.deepcopy(questions_data)
    rng.shuffle(questions)
    for q in questions:
        options = q.get('options')
        if isinstance(options, list):
            rng.shuffle(options)
    return questions


@diagnostic_bp.route('/tests/<int:test_id>/start', methods=['POST'])
def start_test(test_id):
    """
    بدء اختبار (للطالب من التطبيق)
    
    POST /api/diagnostic/tests/1/start
    Headers: Authorization: Bearer <token>
    Body: {"student_id": 5}  // أو من التوكن
    """
    try:
        data = request.get_json() or {}
        
        # ✅ استخرج student_id من session cookie (كما في assigned tests)
        student_id = None
        
        # 1. جرّب من session cookie
        for cookie_name, cookie_value in request.cookies.items():
            if cookie_name.startswith('student_session_'):
                username = cookie_name.replace('student_session_', '')
                student = Student.query.filter_by(username=username).first()
                if student:
                    student_id = student.id
                    print(f"✅ Got student_id from session cookie: {student_id} (username: {username})")
                    break
        
        # 2. جرّب من body
        if not student_id:
            student_id = data.get('student_id')
        
        # 3. جرّب من current_user
        if not student_id and current_user.is_authenticated:
            student_id = current_user.id
        
        if not student_id:
            return jsonify({'success': False, 'error': 'يجب تحديد الطالب'}), 400
        
        test = DiagnosticTest.query.filter_by(id=test_id, is_active=True).first()
        if not test:
            return jsonify({'success': False, 'error': 'الاختبار غير موجود'}), 404
        
        # التحقق من عدم وجود نتيجة سابقة مكتملة
        existing = DiagnosticResult.query.filter_by(
            diagnostic_test_id=test_id,
            student_id=student_id,
            status='completed'
        ).first()
        
        if existing:
            return jsonify({
                'success': False,
                'error': 'لقد أكملت هذا الاختبار مسبقاً',
                'result': existing.to_dict()
            }), 400
        
        # إنشاء أو تحديث نتيجة
        result = DiagnosticResult.query.filter_by(
            diagnostic_test_id=test_id,
            student_id=student_id
        ).first()

        device_id = request.headers.get('X-Device-ID') or data.get('device_id')

        # ✅ منع فتح نفس الاختبار من جهازين بنفس الوقت (نفس الحساب على أكثر من جهاز)
        if result and result.status == 'in_progress' and result.device_id and device_id \
                and result.device_id != device_id:
            return jsonify({
                'success': False,
                'error': 'هذا الاختبار مفتوح حالياً من جهاز آخر بنفس حسابك'
            }), 409

        if not result:
            result = DiagnosticResult(
                diagnostic_test_id=test_id,
                student_id=student_id,
                total_questions=test.questions_count
            )
            db.session.add(result)

        if device_id:
            result.device_id = device_id
        result.started_at = datetime.utcnow()
        result.status = 'in_progress'
        db.session.commit()

        # ✅ جلب الأسئلة بترتيب عشوائي (أسئلة + خيارات) يختلف لكل طالب، لمنع الغش
        # الترتيب مبني على result.id كـ seed ثابت، فيرجع نفسه بالضبط وقت التصحيح بدون تخزين أي شي إضافي
        questions = _get_shuffled_questions(test.questions_data or [], result.id)

        return jsonify({
            'success': True,
            'message': 'تم بدء الاختبار',
            'result_id': result.id,
            'questions': questions,
            'time_limit_minutes': test.time_limit_minutes
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@diagnostic_bp.route('/results/<int:result_id>/submit', methods=['POST'])
def submit_test(result_id):
    """
    تسليم إجابات الاختبار
    
    POST /api/diagnostic/results/1/submit
    {
        "answers": [
            {"question_id": 1, "selected_option_id": 3, "time_spent": 30},
            ...
        ]
    }
    """
    try:
        data = request.get_json() or {}
        answers = data.get('answers', [])
        print(f"📝 Submit - result_id: {result_id}, answers: {len(answers)}")
        
        result = DiagnosticResult.query.get(result_id)
        if not result:
            return jsonify({'success': False, 'error': 'النتيجة غير موجودة'}), 404
        
        if result.status == 'completed':
            return jsonify({'success': False, 'error': 'تم تسليم الاختبار مسبقاً'}), 400
        
        test = result.test
        # ✅ نفس الترتيب العشوائي المبني على result.id اللي شافه الطالب بالضبط وقت الحل
        questions = _get_shuffled_questions(test.questions_data or [], result.id)

        # تصحيح الإجابات
        corrected = []
        correct_count = 0
        
        # ✅ استخدم index-based correction
        for i, ans in enumerate(answers):
            selected_answer = ans.get('selected_answer')  # index من Flutter
            
            # استخدام index للوصول للسؤال
            if i >= len(questions):
                continue
            
            question = questions[i]
            options = question.get('options', [])
            
            # البحث عن الإجابة الصحيحة بالـ index
            correct_index = None
            for opt_idx, opt in enumerate(options):
                if opt.get('is_correct'):
                    correct_index = opt_idx
                    break
            
            is_correct = (selected_answer == correct_index) if selected_answer is not None else False
            
            if is_correct:
                correct_count += 1
            
            corrected.append({
                'question_id': i,
                'question_text': question.get('text', question.get('question_text', '')),
                'selected_answer': selected_answer,
                'correct_answer': correct_index,
                'is_correct': is_correct,
                'time_spent': ans.get('time_spent', 0),
                # ✅ اسم الدرس من السؤال نفسه، وإلا اسم الدرس/الوحدة/المقرر المرتبط بالاختبار نفسه
                'topic': question.get('lesson_name') or test.lesson_name or test.unit_name or test.course_name or ''
            })
        
        # تحديث النتيجة
        result.answers = corrected
        result.score = correct_count
        result.total_questions = len(corrected)
        result.correct_answers = correct_count
        result.percentage = (correct_count / len(corrected) * 100) if corrected else 0
        result.score_percentage = result.percentage
        result.passed = result.score_percentage >= test.passing_score
        result.completed_at = datetime.utcnow()
        summed_time = sum(a.get('time_spent', 0) for a in corrected)
        result.time_spent_seconds = data.get('time_spent_seconds') or summed_time
        result.status = 'completed'
        
        # تحليل النتيجة
        context = {'name': test.lesson_name or test.unit_name or 'عام'}
        analysis = diagnostic_service.analyze_result(result, context, test.test_type)
        
        result.ai_analysis = analysis.get('analysis', '')
        
        # تحديد نقاط القوة والضعف
        weak = [a['topic'] for a in corrected if not a['is_correct'] and a.get('topic')]
        strong = [a['topic'] for a in corrected if a['is_correct'] and a.get('topic')]
        result.weak_topics = list(set(weak))
        result.strong_topics = list(set(strong))

        # ✅ مؤشرات غش (بدون عمود DB جديد — تُخزَّن كعنصر _meta داخل answers نفسها)
        left_app_count = int(data.get('left_app_count') or 0)
        screenshot_count = int(data.get('screenshot_count') or 0)
        if left_app_count or screenshot_count:
            corrected.append({
                '_meta': True,
                'left_app_count': left_app_count,
                'screenshot_count': screenshot_count,
            })
            result.answers = corrected

        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'تم تسليم الاختبار بنجاح',
            'result': result.to_dict(),
            'analysis': analysis
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Submit Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def _topic_fallback(stored_topic, test):
    """اسم الموضوع/الدرس، وإذا كان فارغاً (نتائج قديمة) يرجع لاسم الاختبار نفسه بدل 'عام'"""
    if stored_topic:
        return stored_topic
    if test:
        return test.lesson_name or test.unit_name or test.course_name or 'عام'
    return 'عام'


def _get_student_section(student_id):
    """يرجع شعبة الطالب (أول ربط له فيه شعبة) أو نص فاضي"""
    try:
        from src.models.teacher_student import TeacherStudent
        link = TeacherStudent.query.filter(
            TeacherStudent.student_id == student_id,
            TeacherStudent.section.isnot(None),
            TeacherStudent.section != ''
        ).first()
        return link.section if link else ''
    except Exception:
        return ''


def _get_historical_avg(student_id, exclude_result_id=None):
    """متوسط أداء الطالب بباقي الاختبارات التشخيصية المكتملة (لرصد الشذوذ)"""
    q = DiagnosticResult.query.filter_by(student_id=student_id, status='completed')
    if exclude_result_id:
        q = q.filter(DiagnosticResult.id != exclude_result_id)
    others = q.all()
    if not others:
        return None
    return sum(o.percentage or 0 for o in others) / len(others)


def _get_student_answer_signature(result):
    """يرجع {نص السؤال: نص الإجابة المختارة} لإجابات الطالب الخاطئة فقط (لمقارنة التطابق بين طلاب)"""
    sig = {}
    for a in (result.answers or []):
        if not isinstance(a, dict) or a.get('_meta') or a.get('is_correct'):
            continue
        qtext = a.get('question_text')
        if not qtext:
            continue
        _, sel_text = _resolve_question_options(result, a)
        sig[qtext] = sel_text or '(لم يُجب)'
    return sig


@diagnostic_bp.route('/tests/<int:test_id>/collusion-check', methods=['GET'])
@login_required
@admin_required
def get_collusion_check(test_id):
    """رصد تشابه مشبوه بالإجابات الخاطئة بين الطلاب (مؤشر تبادل/غش) لنفس الاختبار"""
    try:
        test = DiagnosticTest.query.get(test_id)
        if not test:
            return jsonify({'success': False, 'error': 'الاختبار غير موجود'}), 404

        results = DiagnosticResult.query.filter_by(diagnostic_test_id=test_id, status='completed').all()

        entries = []
        for r in results:
            sig = _get_student_answer_signature(r)
            if len(sig) < 2:
                continue
            student = Student.query.get(r.student_id) if r.student_id else None
            entries.append({
                'result_id': r.id,
                'student_name': student.name if student else f'طالب #{r.student_id}',
                'section': _get_student_section(r.student_id) if r.student_id else '',
                'sig': sig,
            })

        pairs = []
        for i in range(len(entries)):
            for j in range(i + 1, len(entries)):
                a, b = entries[i], entries[j]
                common_qs = set(a['sig'].keys()) & set(b['sig'].keys())
                if len(common_qs) < 2:
                    continue
                matches = sum(1 for q in common_qs if a['sig'][q] == b['sig'][q])
                similarity = matches / len(common_qs)
                if similarity >= 0.7 and matches >= 2:
                    pairs.append({
                        'student_a': a['student_name'],
                        'student_b': b['student_name'],
                        'section_a': a['section'],
                        'section_b': b['section'],
                        'matching_wrong_answers': matches,
                        'compared_questions': len(common_qs),
                        'similarity_percent': round(similarity * 100, 1),
                    })

        pairs.sort(key=lambda p: p['similarity_percent'], reverse=True)

        return jsonify({
            'success': True,
            'test_title': test.title,
            'suspicious_pairs': pairs,
            'count': len(pairs),
        })
    except Exception as e:
        print(f"❌ Error checking collusion: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@diagnostic_bp.route('/results/<int:result_id>/detail', methods=['GET'])
def get_result_detail(result_id):
    """تفاصيل نتيجة طالب: كل سؤال، إجابته، الإجابة الصحيحة، ونقاط القوة/الضعف"""
    try:
        result = DiagnosticResult.query.get(result_id)
        if not result:
            return jsonify({'success': False, 'error': 'النتيجة غير موجودة'}), 404

        test = result.test
        canonical_data = (test.questions_data or []) if test else []
        # ✅ نفس الترتيب العشوائي اللي شافه الطالب وقت الحل (نفس result.id = نفس seed)
        shuffled_data = _get_shuffled_questions(canonical_data, result.id) if test else []
        stored_answers = result.answers or []

        questions = []
        topic_breakdown = {}
        cheat_flags = {'left_app_count': 0, 'screenshot_count': 0}

        for a in stored_answers:
            if a.get('_meta'):
                cheat_flags['left_app_count'] = a.get('left_app_count', 0)
                cheat_flags['screenshot_count'] = a.get('screenshot_count', 0)
                continue
            idx = a.get('question_id')
            q = {}
            if isinstance(idx, int):
                # جرّب الترتيب العشوائي أولاً، وإذا نص السؤال ما تطابق (نتيجة قديمة من قبل هذه الميزة)
                # ارجع للترتيب الأصلي حتى تبقى النتائج القديمة صحيحة
                if 0 <= idx < len(shuffled_data) and shuffled_data[idx].get('text') == a.get('question_text'):
                    q = shuffled_data[idx]
                elif 0 <= idx < len(canonical_data):
                    q = canonical_data[idx]
            options = q.get('options', [])

            def _opt_text(opt_idx):
                if opt_idx is None or not (0 <= opt_idx < len(options)):
                    return None
                return options[opt_idx].get('text', '')

            topic = _topic_fallback(a.get('topic'), test)
            bucket = topic_breakdown.setdefault(topic, {'correct': 0, 'total': 0})
            bucket['total'] += 1
            if a.get('is_correct'):
                bucket['correct'] += 1

            questions.append({
                'question_text': a.get('question_text', ''),
                'topic': topic,
                'options': [o.get('text', '') for o in options],
                'selected_index': a.get('selected_answer'),
                'selected_text': _opt_text(a.get('selected_answer')),
                'correct_index': a.get('correct_answer'),
                'correct_text': _opt_text(a.get('correct_answer')),
                'is_correct': a.get('is_correct', False),
            })

        weak_topics = [t for t, b in topic_breakdown.items() if b['correct'] < b['total']]
        strong_topics = [t for t, b in topic_breakdown.items() if b['correct'] == b['total'] and b['total'] > 0]

        result_dict = result.to_dict()
        if test:
            result_dict['test_title'] = test.title
            result_dict['test_type'] = test.test_type
        student = Student.query.get(result.student_id) if result.student_id else None
        if student:
            result_dict['student_name'] = student.name
            result_dict['section'] = _get_student_section(student.id)

        return jsonify({
            'success': True,
            'result': result_dict,
            'questions': questions,
            'stats': {
                'weak_topics': weak_topics,
                'strong_topics': strong_topics,
                'topic_breakdown': topic_breakdown,
                'left_app_count': cheat_flags['left_app_count'],
                'screenshot_count': cheat_flags['screenshot_count'],
            }
        })
    except Exception as e:
        print(f"❌ Error getting result detail: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@diagnostic_bp.route('/tests/<int:test_id>/assignment-status', methods=['GET'])
def get_assignment_status(test_id):
    """حالة كل طالب معيّن له الاختبار: لم يبدأ / قيد التنفيذ / مكتمل"""
    try:
        test = DiagnosticTest.query.get(test_id)
        if not test:
            return jsonify({'success': False, 'error': 'الاختبار غير موجود'}), 404

        assigned_ids = test.assigned_students or []

        results_by_student = {}
        for r in DiagnosticResult.query.filter_by(diagnostic_test_id=test_id).all():
            try:
                sid = int(r.student_id)
            except (TypeError, ValueError):
                continue
            # لو فيه أكثر من محاولة، خذ الأحدث/الأكمل
            existing = results_by_student.get(sid)
            if not existing or (r.status == 'completed' and existing.status != 'completed'):
                results_by_student[sid] = r

        data = []
        for sid in assigned_ids:
            student = Student.query.get(sid)
            r = results_by_student.get(sid)
            if r and r.status == 'completed':
                status = 'completed'
            elif r:
                status = 'in_progress'
            else:
                status = 'not_started'

            data.append({
                'student_id': sid,
                'student_name': student.name if student else f'طالب #{sid}',
                'section': _get_student_section(sid),
                'status': status,
                'percentage': r.percentage if r else None,
                'completed_at': ((r.completed_at.isoformat() + 'Z') if (r and r.completed_at and status == 'completed') else None),
            })

        counts = {
            'not_started': sum(1 for d in data if d['status'] == 'not_started'),
            'in_progress': sum(1 for d in data if d['status'] == 'in_progress'),
            'completed': sum(1 for d in data if d['status'] == 'completed'),
        }

        return jsonify({
            'success': True,
            'test_title': test.title,
            'students': data,
            'counts': counts,
        })
    except Exception as e:
        print(f"❌ Error getting assignment status: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@diagnostic_bp.route('/tests/<int:test_id>/reopen/<int:student_id>', methods=['POST'])
@login_required
@admin_required
def reopen_test_for_student(test_id, student_id):
    """يمسح محاولة طالب معيّن على اختبار (مكتملة أو قيد التنفيذ) عشان ياخذ فرصة ثانية،
    دون الحاجة لإعادة تعيين الاختبار لكل الطلاب من جديد"""
    try:
        test = DiagnosticTest.query.get(test_id)
        if not test:
            return jsonify({'success': False, 'error': 'الاختبار غير موجود'}), 404

        result = DiagnosticResult.query.filter_by(
            diagnostic_test_id=test_id, student_id=str(student_id)
        ).first()
        if not result:
            return jsonify({'success': False, 'error': 'ما فيه محاولة سابقة لهذا الطالب'}), 404

        db.session.delete(result)
        db.session.commit()

        return jsonify({'success': True, 'message': 'تم فتح الاختبار للطالب من جديد'})
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error reopening test: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


def _resolve_question_options(result, a):
    """يرجع (options, نص إجابة الطالب) لعنصر إجابة واحد، مع مراعاة الترتيب العشوائي لكل طالب"""
    idx = a.get('question_id')
    test = result.test
    if not test or not isinstance(idx, int):
        return [], None
    canonical = test.questions_data or []
    shuffled = _get_shuffled_questions(canonical, result.id)
    q = {}
    if 0 <= idx < len(shuffled) and shuffled[idx].get('text') == a.get('question_text'):
        q = shuffled[idx]
    elif 0 <= idx < len(canonical):
        q = canonical[idx]
    options = q.get('options', [])
    sel = a.get('selected_answer')
    sel_text = options[sel].get('text', '') if (isinstance(sel, int) and 0 <= sel < len(options)) else None
    return options, sel_text


def _compute_item_analysis(test):
    """تحليل كل سؤال بالاختبار عبر كل الطلاب: نسبة الصح، وأكثر مشتت (إجابة خاطئة) تم اختياره"""
    results = DiagnosticResult.query.filter_by(diagnostic_test_id=test.id, status='completed').all()

    stats = {}
    order = []
    for r in results:
        for a in (r.answers or []):
            if not isinstance(a, dict) or a.get('_meta'):
                continue
            qtext = a.get('question_text') or ''
            if not qtext:
                continue
            if qtext not in stats:
                stats[qtext] = {'correct': 0, 'total': 0, 'wrong_counts': {}}
                order.append(qtext)
            s = stats[qtext]
            s['total'] += 1
            if a.get('is_correct'):
                s['correct'] += 1
            else:
                _, sel_text = _resolve_question_options(r, a)
                key = sel_text if sel_text else '(لم يُجب)'
                s['wrong_counts'][key] = s['wrong_counts'].get(key, 0) + 1

    items = []
    for qtext in order:
        s = stats[qtext]
        accuracy = round(s['correct'] / s['total'] * 100, 1) if s['total'] else 0
        top_distractor = max(s['wrong_counts'].items(), key=lambda kv: kv[1]) if s['wrong_counts'] else None
        items.append({
            'question_text': qtext,
            'total_answers': s['total'],
            'correct_count': s['correct'],
            'accuracy': accuracy,
            'top_distractor': top_distractor[0] if top_distractor else None,
            'top_distractor_count': top_distractor[1] if top_distractor else 0,
            'distractor_breakdown': s['wrong_counts'],
        })
    items.sort(key=lambda it: it['accuracy'])
    return items


@diagnostic_bp.route('/tests/<int:test_id>/item-analysis', methods=['GET'])
@login_required
@admin_required
def get_item_analysis(test_id):
    """تحليل كل سؤال بالاختبار: أضعف الأسئلة وأكثر مشتت (إجابة خاطئة) تم اختياره"""
    try:
        test = DiagnosticTest.query.get(test_id)
        if not test:
            return jsonify({'success': False, 'error': 'الاختبار غير موجود'}), 404

        items = _compute_item_analysis(test)
        return jsonify({'success': True, 'test_title': test.title, 'items': items})
    except Exception as e:
        print(f"❌ Error getting item analysis: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@diagnostic_bp.route('/tests/<int:test_id>/export-excel', methods=['GET'])
@login_required
@admin_required
def export_results_excel(test_id):
    """تصدير نتائج اختبار تشخيصي كملف Excel"""
    try:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from flask import send_file

        test = DiagnosticTest.query.get(test_id)
        if not test:
            return jsonify({'success': False, 'error': 'الاختبار غير موجود'}), 404

        results = DiagnosticResult.query.filter_by(diagnostic_test_id=test_id)\
            .order_by(DiagnosticResult.percentage.desc()).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'النتائج'
        ws.sheet_view.rightToLeft = True

        header_fill = PatternFill('solid', fgColor='0D9488')
        thin = Side(style='thin', color='CBD5E1')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ['الطالب', 'الشعبة', 'الحالة', 'الدرجة', 'من', 'النسبة', 'الوقت المستغرق (د)', 'تاريخ الإكمال']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, value=h)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
            cell.border = border
        ws.row_dimensions[1].height = 26

        status_ar = {'completed': 'مكتمل', 'in_progress': 'قيد التنفيذ'}
        for r, res in enumerate(results, 2):
            student = Student.query.get(res.student_id) if res.student_id else None
            row = [
                (student.name if student else f'طالب #{res.student_id}'),
                _get_student_section(res.student_id) if res.student_id else '',
                status_ar.get(res.status, res.status),
                res.correct_answers or 0,
                res.total_questions or 0,
                round(res.percentage or 0, 1),
                round((res.time_spent_seconds or 0) / 60, 1),
                res.completed_at.strftime('%Y-%m-%d %H:%M') if (res.completed_at and res.status == 'completed') else '',
            ]
            for col, val in enumerate(row, 1):
                cell = ws.cell(r, col, value=val)
                cell.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
                cell.border = border

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

        # صف المصدر (نفس تذييل بقية تقارير التطبيق)
        footer_row1 = len(results) + 3
        ws.merge_cells(start_row=footer_row1, start_column=1, end_row=footer_row1, end_column=len(headers))
        fc1 = ws.cell(footer_row1, 1)
        fc1.value = f'⚗️  تم استخراج هذا التقرير من تطبيق كيم تحصيلي  |  منصة تعليمية للكيمياء  |  جميع الحقوق محفوظة © {datetime.now().year}'
        fc1.font = Font(size=9, color='888888', italic=True)
        fc1.alignment = Alignment(horizontal='center', vertical='center')
        fc1.fill = PatternFill('solid', fgColor='F1F5F9')
        ws.row_dimensions[footer_row1].height = 18

        # ✅ ورقة ثانية: تحليل كل سؤال (أضعف الأسئلة + أكثر مشتت تم اختياره)
        items = _compute_item_analysis(test)
        if items:
            ws2 = wb.create_sheet('تحليل الأسئلة')
            ws2.sheet_view.rightToLeft = True
            headers2 = ['السؤال', 'عدد الإجابات', 'صح', 'الدقة %', 'أكثر مشتت تم اختياره', 'عدد من اختاره']
            for col, h in enumerate(headers2, 1):
                cell = ws2.cell(1, col, value=h)
                cell.font = Font(bold=True, color='FFFFFF', size=11)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2, wrap_text=True)
                cell.border = border
            ws2.row_dimensions[1].height = 30

            for r, it in enumerate(items, 2):
                row = [
                    it['question_text'],
                    it['total_answers'],
                    it['correct_count'],
                    it['accuracy'],
                    it['top_distractor'] or '-',
                    it['top_distractor_count'],
                ]
                for col, val in enumerate(row, 1):
                    cell = ws2.cell(r, col, value=val)
                    cell.alignment = Alignment(horizontal='center', vertical='center',
                                               readingOrder=2, wrap_text=True)
                    cell.border = border

            ws2.column_dimensions['A'].width = 45
            for col in range(2, len(headers2) + 1):
                ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

            footer_row2 = len(items) + 3
            ws2.merge_cells(start_row=footer_row2, start_column=1, end_row=footer_row2, end_column=len(headers2))
            fc2 = ws2.cell(footer_row2, 1)
            fc2.value = f'⚗️  تم استخراج هذا التقرير من تطبيق كيم تحصيلي  |  منصة تعليمية للكيمياء  |  جميع الحقوق محفوظة © {datetime.now().year}'
            fc2.font = Font(size=9, color='888888', italic=True)
            fc2.alignment = Alignment(horizontal='center', vertical='center')
            fc2.fill = PatternFill('solid', fgColor='F1F5F9')
            ws2.row_dimensions[footer_row2].height = 18
            ws2.column_dimensions['E'].width = 30

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        safe_title = (test.title or 'اختبار').replace('/', '-')
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'نتائج_{safe_title}.xlsx',
        )
    except Exception as e:
        print(f"❌ Error exporting results excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@diagnostic_bp.route('/sections-report', methods=['GET'])
@login_required
@admin_required
def get_sections_report():
    """تقرير أداء شامل عبر الشعب — متوسط كل شعبة وأضعف المواضيع عبر كل الاختبارات التشخيصية مجتمعة"""
    try:
        results = DiagnosticResult.query.filter_by(status='completed').all()

        section_stats = {}
        topic_stats = {}
        cheat_count = 0

        for r in results:
            section = (_get_student_section(r.student_id) if r.student_id else '') or 'بدون شعبة'
            s = section_stats.setdefault(section, {'count': 0, 'sum_pct': 0.0})
            s['count'] += 1
            s['sum_pct'] += (r.percentage or 0)

            for a in (r.answers or []):
                if not isinstance(a, dict):
                    continue
                if a.get('_meta'):
                    if a.get('left_app_count') or a.get('screenshot_count'):
                        cheat_count += 1
                    continue
                topic = _topic_fallback(a.get('topic'), r.test)
                t = topic_stats.setdefault(topic, {'correct': 0, 'total': 0})
                t['total'] += 1
                if a.get('is_correct'):
                    t['correct'] += 1

        sections = [{
            'section': k,
            'count': v['count'],
            'avg_percentage': round(v['sum_pct'] / v['count'], 1) if v['count'] else 0,
        } for k, v in section_stats.items()]
        sections.sort(key=lambda s: s['avg_percentage'])

        topics = [{
            'topic': k,
            'correct': v['correct'],
            'total': v['total'],
            'accuracy': round(v['correct'] / v['total'] * 100, 1) if v['total'] else 0,
        } for k, v in topic_stats.items()]
        topics.sort(key=lambda t: t['accuracy'])

        return jsonify({
            'success': True,
            'sections': sections,
            'weak_topics': topics[:10],
            'total_results': len(results),
            'flagged_attempts': cheat_count,
        })
    except Exception as e:
        print(f"❌ Error getting sections report: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@diagnostic_bp.route('/tests/<int:test_id>/comparison', methods=['GET'])
def get_test_comparison(test_id):
    """مقارنة نتائج القبلي/البعدي لكل طالب أكمل الاثنين (يعتمد على paired_test_id)"""
    try:
        test = DiagnosticTest.query.get(test_id)
        if not test:
            return jsonify({'success': False, 'error': 'الاختبار غير موجود'}), 404

        paired_id = test.paired_test_id
        if not paired_id:
            return jsonify({'success': False, 'error': 'هذا الاختبار غير مرتبط باختبار قبلي/بعدي'}), 400

        if test.test_type == 'pre_test':
            pre_id, post_id = test_id, paired_id
        else:
            pre_id, post_id = paired_id, test_id

        pre_test = DiagnosticTest.query.get(pre_id)
        post_test = DiagnosticTest.query.get(post_id)
        if not pre_test or not post_test:
            return jsonify({'success': False, 'error': 'الاختبار المرتبط غير موجود'}), 404

        pre_results = {
            r.student_id: r for r in
            DiagnosticResult.query.filter_by(diagnostic_test_id=pre_id, status='completed').all()
        }
        post_results = {
            r.student_id: r for r in
            DiagnosticResult.query.filter_by(diagnostic_test_id=post_id, status='completed').all()
        }
        common_ids = set(pre_results.keys()) & set(post_results.keys())

        students = []
        for sid in common_ids:
            pre_r = pre_results[sid]
            post_r = post_results[sid]
            student = Student.query.get(sid) if sid else None
            improvement = (post_r.percentage or 0) - (pre_r.percentage or 0)
            students.append({
                'student_id': sid,
                'student_name': student.name if student else f'طالب #{sid}',
                'section': _get_student_section(sid) if sid else '',
                'pre_percentage': round(pre_r.percentage or 0, 1),
                'post_percentage': round(post_r.percentage or 0, 1),
                'improvement': round(improvement, 1),
            })

        students.sort(key=lambda s: s['improvement'], reverse=True)
        avg_improvement = round(sum(s['improvement'] for s in students) / len(students), 1) if students else 0

        return jsonify({
            'success': True,
            'pre_test_title': pre_test.title,
            'post_test_title': post_test.title,
            'students': students,
            'avg_improvement': avg_improvement,
            'count': len(students),
        })
    except Exception as e:
        print(f"❌ Error getting comparison: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==========================================
# المقارنة بين القبلي والبعدي
# ==========================================

@diagnostic_bp.route('/compare', methods=['POST'])
def compare_tests():
    """
    مقارنة نتائج القبلي والبعدي
    
    POST /api/diagnostic/compare
    {
        "pre_test_id": 1,
        "post_test_id": 2,
        "student_id": 5
    }
    """
    try:
        data = request.get_json() or {}
        
        pre_test_id = data.get('pre_test_id')
        post_test_id = data.get('post_test_id')
        student_id = data.get('student_id')
        
        if not all([pre_test_id, post_test_id, student_id]):
            return jsonify({'success': False, 'error': 'بيانات ناقصة'}), 400
        
        # جلب النتائج
        pre_result = DiagnosticResult.query.filter_by(
            test_id=pre_test_id,
            student_id=student_id,
            status='completed'
        ).first()
        
        post_result = DiagnosticResult.query.filter_by(
            test_id=post_test_id,
            student_id=student_id,
            status='completed'
        ).first()
        
        if not pre_result:
            return jsonify({'success': False, 'error': 'لم تكمل الاختبار القبلي'}), 404
        
        if not post_result:
            return jsonify({'success': False, 'error': 'لم تكمل الاختبار البعدي'}), 404
        
        # المقارنة
        pre_test = DiagnosticTest.query.get(pre_test_id)
        context = {'name': pre_test.lesson_name or 'عام'}
        
        comparison_data = diagnostic_service.compare_results(pre_result, post_result, context)
        
        # حفظ المقارنة
        comparison = DiagnosticComparison(
            student_id=student_id,
            pre_test_id=pre_test_id,
            post_test_id=post_test_id,
            pre_result_id=pre_result.id,
            post_result_id=post_result.id,
            pre_score=comparison_data['pre_score'],
            post_score=comparison_data['post_score'],
            improvement=comparison_data['improvement'],
            effectiveness=comparison_data['effectiveness'],
            ai_analysis=comparison_data['analysis']
        )
        db.session.add(comparison)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'comparison': comparison.to_dict(),
            'pre_result': pre_result.to_dict(),
            'post_result': post_result.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==========================================
# إحصائيات
# ==========================================

@diagnostic_bp.route('/stats', methods=['GET'])
@login_required
def get_stats():
    """إحصائيات الاختبارات التشخيصية"""
    try:
        from sqlalchemy import func
        
        total = DiagnosticTest.query.filter_by(is_active=True).count()
        pre_count = DiagnosticTest.query.filter_by(is_active=True, test_type='pre_test').count()
        post_count = DiagnosticTest.query.filter_by(is_active=True, test_type='post_test').count()
        results_count = DiagnosticResult.query.filter_by(status='completed').count()
        
        avg_improvement = db.session.query(
            func.avg(DiagnosticComparison.improvement)
        ).scalar() or 0
        
        return jsonify({
            'success': True,
            'stats': {
                'total_tests': total,
                'pre_tests': pre_count,
                'post_tests': post_count,
                'total_results': results_count,
                'avg_improvement': round(avg_improvement, 1)
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@diagnostic_bp.route('/student/<int:student_id>/history', methods=['GET'])
def get_student_history(student_id):
    """سجل اختبارات الطالب"""
    try:
        results = DiagnosticResult.query.filter_by(
            student_id=student_id,
            status='completed'
        ).order_by(DiagnosticResult.completed_at.desc()).all()
        
        comparisons = DiagnosticComparison.query.filter_by(
            student_id=student_id
        ).order_by(DiagnosticComparison.created_at.desc()).all()
        
        return jsonify({
            'success': True,
            'results': [r.to_dict() for r in results],
            'comparisons': [c.to_dict() for c in comparisons]
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==========================================
# حذف
# ==========================================

@diagnostic_bp.route('/tests/<int:test_id>', methods=['PUT'])
@login_required
@admin_required
def update_test(test_id):
    """تعديل اسم الاختبار"""
    try:
        test = DiagnosticTest.query.filter_by(id=test_id, is_active=True).first()
        if not test:
            return jsonify({'success': False, 'error': 'الاختبار غير موجود'}), 404

        data = request.get_json() or {}
        title = (data.get('title') or '').strip()
        if not title:
            return jsonify({'success': False, 'error': 'اسم الاختبار مطلوب'}), 400

        test.title = title
        db.session.commit()

        return jsonify({'success': True, 'message': 'تم تعديل الاسم بنجاح', 'test': test.to_dict()})
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error updating test: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@diagnostic_bp.route('/tests/<int:test_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_test(test_id):
    """حذف اختبار"""
    try:
        test = DiagnosticTest.query.filter_by(id=test_id, is_active=True).first()
        
        if not test:
            return jsonify({'success': False, 'error': 'الاختبار غير موجود'}), 404
        
        test.is_active = False
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'تم حذف الاختبار'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==========================================
# صفحة الأدمن
# ==========================================

@diagnostic_bp.route('/admin', methods=['GET'])
@login_required
@admin_required
def admin_page():
    """صفحة إدارة الاختبارات التشخيصية"""
    return render_template('diagnostic/admin.html')


# =====================================================
# ✅ Routes الجدولة والإسناد (مضافة - جديدة)
# =====================================================

@diagnostic_bp.route('/scheduled', methods=['GET'])
@login_required
@admin_required
def get_scheduled_tests():
    """جلب الاختبارات المجدولة"""
    try:
        tests = DiagnosticTest.query.filter_by(
            is_scheduled=True,
            is_active=True
        ).order_by(DiagnosticTest.scheduled_start.desc()).all()
        
        # تحديث حالة كل اختبار
        for test in tests:
            if hasattr(test, 'update_schedule_status'):
                test.update_schedule_status()
        db.session.commit()
        
        return jsonify({
            'scheduled_tests': [test.to_dict() for test in tests]
        }), 200
        
    except Exception as e:
        print(f"❌ Error getting scheduled tests: {e}")
        return jsonify({'error': str(e)}), 500


@diagnostic_bp.route('/assign', methods=['POST'])
@login_required
@admin_required
def assign_test():
    """إسناد اختبار لطلاب مع جدولة - محدث لدعم الإرسال المتعدد والصفوف الدراسية"""
    try:
        data = request.get_json()
        test_id = data.get('test_id')
        student_ids = data.get('student_ids')
        grade = data.get('grade')  # ✅ جديد: الصف الدراسي
        sections = data.get('sections')  # ✅ جديد: فلترة شعبة واحدة أو أكثر (مع my_students)
        scheduled_start = data.get('scheduled_start')
        scheduled_end = data.get('scheduled_end')
        time_limit = data.get('time_limit_minutes', 30)
        send_notification = data.get('send_notification', True)
        # ✅ إضافة بدلاً من الاستبدال (checkbox بالموقع). التطبيق ما يرسل هذا الحقل، فنخليه يضيف
        # افتراضياً (منعاً لضياع طلاب أُرسل لهم الاختبار سابقاً عند إعادة الإرسال لشعبة ثانية)
        append_students = data.get('append_students', True)
        # ✅ هل نُشعر حتى من أكمل الاختبار مسبقاً، أو الطلاب الجدد بس (الافتراضي)
        notify_all = data.get('notify_all', False)
        
        test = DiagnosticTest.query.get(test_id)
        if not test:
            return jsonify({'error': 'Test not found'}), 404
        
        # تحضير قائمة الطلاب
        student_ids_list = []

        if grade:
            # طلاب صف دراسي محدد (أولوية عليا)
            students = Student.query.filter_by(is_active=True, grade=grade).all()
            student_ids_list = [s.id for s in students]
            print(f"✅ تم اختيار {len(student_ids_list)} طالب من الصف {grade}")
        elif student_ids == 'all':
            # جميع الطلاب النشطين
            students = Student.query.filter_by(is_active=True).all()
            student_ids_list = [s.id for s in students]
            print(f"✅ تم اختيار جميع الطلاب: {len(student_ids_list)} طالب")
        elif student_ids == 'my_students':
            # طلابي — الطلاب المرتبطين بالأدمن الحالي
            try:
                from src.models.teacher_student import TeacherStudent
                links = TeacherStudent.query.filter_by(admin_id=current_user.id).all()
                if sections:
                    links = [lnk for lnk in links if (lnk.section or '') in sections]
                student_ids_list = [lnk.student_id for lnk in links]
                print(f"✅ طلابي (أدمن {current_user.id}): {len(student_ids_list)} طالب")
            except Exception as e:
                print(f"⚠️ خطأ في جلب طلابي: {e}")
                student_ids_list = []
        else:
            # طلاب محددين
            student_ids_list = student_ids if isinstance(student_ids, list) else [student_ids]
            print(f"✅ تم اختيار {len(student_ids_list)} طالب محدد")
        
        # ✅ جديد: دعم الإضافة بدلاً من الاستبدال
        if append_students and test.assigned_students:
            # إضافة الطلاب الجدد للقائمة الحالية
            existing_ids = set(test.assigned_students)
            new_ids = set(student_ids_list)
            student_ids_list = list(existing_ids.union(new_ids))
            print(f"✅ تم إضافة طلاب جدد. الإجمالي: {len(student_ids_list)}")
        else:
            # استبدال القائمة بالكامل
            print(f"✅ تم استبدال قائمة الطلاب. العدد: {len(student_ids_list)}")
        
        # تحديث الاختبار
        test.is_scheduled = True
        test.scheduled_start = convert_saudi_to_utc(scheduled_start) if scheduled_start else None
        test.scheduled_end = convert_saudi_to_utc(scheduled_end) if scheduled_end else None
        test.assigned_students = student_ids_list
        test.time_limit_minutes = time_limit
        test.schedule_status = 'pending'
        
        if hasattr(test, 'update_schedule_status'):
            test.update_schedule_status()
        
        db.session.commit()

        # ✅ لا نُشعر مين خلّص الاختبار فعلاً إلا لو طلب notify_all صراحة (يصير مزعج وهو أصلاً ما يقدر يفتحه مرة ثانية)
        if notify_all:
            notify_ids = student_ids_list
        else:
            already_completed_ids = {
                int(r.student_id) for r in DiagnosticResult.query.filter_by(
                    diagnostic_test_id=test.id, status='completed'
                ).all() if r.student_id and str(r.student_id).isdigit()
            }
            notify_ids = [sid for sid in student_ids_list if sid not in already_completed_ids]

        # إرسال إشعارات
        if send_notification and NotificationService:
            try:
                students = Student.query.filter(
                    Student.id.in_(notify_ids),
                    Student.fcm_token.isnot(None)
                ).all()
                
                # ✅ استخدام NotificationService
                success_count = 0
                
                # ✅ تنسيق رسالة FCM بشكل مقروء
                try:
                    from datetime import datetime as dt_parse
                    start_dt = dt_parse.fromisoformat(scheduled_start.replace('Z', ''))
                    end_dt = dt_parse.fromisoformat(scheduled_end.replace('Z', ''))
                    
                    def _fmt_t(dt):
                        h = dt.hour % 12 or 12
                        p = 'م' if dt.hour >= 12 else 'ص'
                        return f'{h}:{dt.minute:02d} {p}'
                    
                    if start_dt.date() == end_dt.date():
                        fcm_body = f'{test.title}\n📅 {start_dt.day}/{start_dt.month} | 🕐 {_fmt_t(start_dt)} - {_fmt_t(end_dt)}'
                    else:
                        fcm_body = f'{test.title}\n🟢 {start_dt.day}/{start_dt.month} {_fmt_t(start_dt)}\n🔴 {end_dt.day}/{end_dt.month} {_fmt_t(end_dt)}'
                except:
                    fcm_body = f'{test.title}'
                
                for student in students:
                    if student.fcm_token:
                        result = NotificationService.send_fcm_notification(
                            student.fcm_token,
                            '📝 اختبار تشخيصي جديد',
                            fcm_body,
                            {
                                'type': 'diagnostic_test',
                                'test_id': str(test.id),
                                'scheduled_start': scheduled_start,
                                'scheduled_end': scheduled_end,
                                'time_limit_minutes': str(test.time_limit_minutes)
                            }
                        )
                        if result:
                            success_count += 1
                
                # ✅ حفظ الإشعار في قاعدة البيانات لكل الطلاب المعينين (ليظهر في صفحة الإشعارات)
                db_save_count = 0
                notification_title = '📝 اختبار تشخيصي جديد'
                
                # ✅ تنسيق الوقت بشكل مقروء
                try:
                    from datetime import datetime as dt_parse
                    start_dt = dt_parse.fromisoformat(scheduled_start.replace('Z', ''))
                    end_dt = dt_parse.fromisoformat(scheduled_end.replace('Z', ''))
                    
                    def _format_period(hour):
                        return 'مساءً' if hour >= 12 else 'صباحاً'
                    
                    def _format_time_12h(dt):
                        h = dt.hour % 12 or 12
                        return f'{h}:{dt.minute:02d}'
                    
                    start_time = _format_time_12h(start_dt)
                    start_p = _format_period(start_dt.hour)
                    end_time = _format_time_12h(end_dt)
                    end_p = _format_period(end_dt.hour)
                    
                    start_date_str = f'{start_dt.day}/{start_dt.month}/{start_dt.year}'
                    end_date_str = f'{end_dt.day}/{end_dt.month}/{end_dt.year}'
                    
                    # إذا نفس اليوم
                    if start_dt.date() == end_dt.date():
                        notification_message = (
                            f'تم تعيين اختبار: {test.title}\n\n'
                            f'📅 التاريخ: {start_date_str}\n'
                            f'🕐 البداية: {start_time} {start_p}\n'
                            f'🕐 النهاية: {end_time} {end_p}\n'
                            f'⏱ المدة: {test.time_limit_minutes} دقيقة'
                        )
                    else:
                        notification_message = (
                            f'تم تعيين اختبار: {test.title}\n\n'
                            f'🟢 البداية: {start_date_str} - {start_time} {start_p}\n'
                            f'🔴 النهاية: {end_date_str} - {end_time} {end_p}\n'
                            f'⏱ المدة: {test.time_limit_minutes} دقيقة'
                        )
                except:
                    notification_message = f'تم تعيين اختبار: {test.title}'
                
                notification_data = {
                    'type': 'diagnostic_test',
                    'test_id': str(test.id),
                }
                
                for sid in notify_ids:
                    if _save_notification_to_db(
                        student_id=sid,
                        title=notification_title,
                        message=notification_message,
                        notification_type='reminder',
                        data=notification_data
                    ):
                        db_save_count += 1
                
                db.session.commit()
                print(f"✅ تم إرسال {success_count}/{len(students)} إشعار FCM")
                print(f"✅ تم حفظ {db_save_count}/{len(notify_ids)} إشعار في قاعدة البيانات "
                      f"(استُثني {len(student_ids_list) - len(notify_ids)} أكملوا الاختبار مسبقاً)")
                
                if success_count > 0:
                    test.notification_sent = True
                    test.notification_sent_at = datetime.utcnow()
                    db.session.commit()
                
            except Exception as e:
                print(f"⚠️ Error sending notifications: {e}")
        
        return jsonify({
            'success': True,
            'message': 'Test assigned successfully',
            'test': test.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error assigning test: {e}")
        return jsonify({'error': str(e)}), 500


@diagnostic_bp.route('/teacher/assign', methods=['POST'])
def teacher_assign_test():
    """تعيين اختبار تشخيصي لطلاب المعلم المرتبطين به — يستخدم JWT Token"""
    from src.middleware.auth_middleware import verify_teacher_token as _vtt
    from src.models.teacher_student import TeacherStudent

    # التحقق من توكن المعلم يدوياً (لأننا لا نستخدم decorator هنا)
    import jwt as _jwt
    token = None
    auth = request.headers.get('Authorization', '')
    if auth.startswith('Bearer '):
        token = auth.split(' ', 1)[1]
    if not token:
        return jsonify({'success': False, 'error': 'رمز المصادقة مطلوب'}), 401
    try:
        from flask import current_app
        data = _jwt.decode(token, current_app.config['JWT_SECRET_KEY'],
                           algorithms=[current_app.config['JWT_ALGORITHM']])
        if data.get('user_type') != 'teacher':
            return jsonify({'success': False, 'error': 'مخصص للمعلمين فقط'}), 403
        teacher_id = data.get('teacher_id')
    except _jwt.ExpiredSignatureError:
        return jsonify({'success': False, 'error': 'انتهت صلاحية الرمز'}), 401
    except _jwt.InvalidTokenError as e:
        return jsonify({'success': False, 'error': str(e)}), 401

    try:
        body = request.get_json() or {}
        test_id          = body.get('test_id')
        time_limit       = body.get('time_limit_minutes', 30)
        send_notif       = body.get('send_notification', True)
        scheduled_start  = body.get('scheduled_start')
        scheduled_end    = body.get('scheduled_end')
        sections         = body.get('sections')  # ✅ جديد: فلترة شعبة واحدة أو أكثر
        notify_all       = body.get('notify_all', False)  # ✅ إشعار حتى من أكمل الاختبار مسبقاً

        test = DiagnosticTest.query.get(test_id)
        if not test:
            return jsonify({'success': False, 'error': 'الاختبار غير موجود'}), 404

        # جلب طلاب المعلم
        links = TeacherStudent.query.filter_by(teacher_id=teacher_id).all()
        if sections:
            links = [lnk for lnk in links if (lnk.section or '') in sections]
        student_ids_list = [lnk.student_id for lnk in links]

        if not student_ids_list:
            return jsonify({'success': False, 'error': 'لا يوجد طلاب مرتبطون بك بعد ضمن الشعبة المحددة' if sections else 'لا يوجد طلاب مرتبطون بك بعد'}), 400

        # ✅ نضيف للقائمة الحالية بدل الاستبدال — منعاً لضياع طلاب أُرسل لهم الاختبار سابقاً
        # عند إعادة الإرسال لشعبة ثانية لاحقاً
        if not body.get('replace_students', False) and test.assigned_students:
            student_ids_list = list(set(test.assigned_students) | set(student_ids_list))

        test.is_scheduled     = True
        test.assigned_students = student_ids_list
        test.time_limit_minutes = time_limit
        test.schedule_status  = 'pending'
        if scheduled_start:
            test.scheduled_start = convert_saudi_to_utc(scheduled_start)
        if scheduled_end:
            test.scheduled_end = convert_saudi_to_utc(scheduled_end)
        if hasattr(test, 'update_schedule_status'):
            test.update_schedule_status()

        db.session.commit()

        # ✅ لا نُشعر مين خلّص الاختبار فعلاً إلا لو طلب notify_all صراحة
        if notify_all:
            notify_ids = student_ids_list
        else:
            already_completed_ids = {
                int(r.student_id) for r in DiagnosticResult.query.filter_by(
                    diagnostic_test_id=test.id, status='completed'
                ).all() if r.student_id and str(r.student_id).isdigit()
            }
            notify_ids = [sid for sid in student_ids_list if sid not in already_completed_ids]

        # إشعارات FCM
        if send_notif:
            try:
                students = Student.query.filter(Student.id.in_(notify_ids), Student.is_active == True).all()
                for s in students:
                    if getattr(s, 'fcm_token', None):
                        try:
                            from src.services.fcm_service import send_fcm_notification
                            send_fcm_notification(
                                token=s.fcm_token,
                                title='اختبار تشخيصي جديد',
                                body=f'تم تعيين اختبار: {test.title}',
                                data={'type': 'diagnostic_test', 'test_id': str(test.id)},
                            )
                        except Exception:
                            pass
            except Exception as e:
                print(f"⚠️ FCM error: {e}")

        return jsonify({
            'success': True,
            'message': f'تم تعيين الاختبار لـ {len(student_ids_list)} طالب',
            'students_count': len(student_ids_list),
        })

    except Exception as e:
        db.session.rollback()
        print(f"❌ teacher_assign_test error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@diagnostic_bp.route('/student/assigned', methods=['GET', 'POST'])
def get_student_assigned_tests():
    """اختبارات الطالب المخصصة (للتطبيق والويب)"""
    try:
        # جرب جميع الطرق للحصول على student_id
        student_id = None
        
        # 1. من query parameter (GET)
        student_id = request.args.get('student_id', type=int)
        
        # 2. من body (POST)
        if not student_id and request.method == 'POST':
            data = request.get_json() or {}
            student_id = data.get('student_id')
            if student_id:
                student_id = int(student_id)
        
        # 3. من headers
        if not student_id:
            header_id = request.headers.get('X-Student-ID') or request.headers.get('Student-ID')
            if header_id:
                try:
                    student_id = int(header_id)
                except:
                    pass
        
        # 4. من cookies (للتطبيق Flutter)
        if not student_id:
            # جرب cookie مباشر
            cookie_id = request.cookies.get('student_id')
            if cookie_id:
                try:
                    student_id = int(cookie_id)
                    print(f"📱 Got student_id from cookie: {student_id}")
                except:
                    pass
            
            # جرب استخراج من session cookie
            if not student_id:
                for cookie_name, cookie_value in request.cookies.items():
                    # ابحث عن pattern: student_session_{username}
                    if cookie_name.startswith('student_session_'):
                        print(f"🔍 Found session cookie: {cookie_name}")
                        # جرب query الـ database
                        try:
                            username = cookie_name.replace('student_session_', '')
                            student = Student.query.filter_by(username=username).first()
                            if not student:
                                student = Student.query.filter_by(email_hash=make_email_hash(username)).first()
                            if student:
                                student_id = student.id
                                print(f"✅ Got student_id from session cookie: {student_id} (username: {username})")
                                break
                        except Exception as e:
                            print(f"⚠️ Error extracting from session cookie: {e}")
                            pass
        
        # 5. من session (Flask session)
        if not student_id:
            from flask import session
            session_id = session.get('student_id') or session.get('user_id')
            if session_id:
                try:
                    student_id = int(session_id)
                    print(f"📱 Got student_id from session: {student_id}")
                except:
                    pass
        
        # 6. من current_user (للويب)
        if not student_id:
            try:
                if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated:
                    student_id = current_user.id
                    pass  # student_id from current_user
            except:
                pass
        
        # Debug info removed for security
        
        if not student_id:
            return jsonify({
                'success': False,
                'error': 'student_id required',
                'message_ar': 'يرجى تحديث التطبيق أو إرسال student_id',
                'hint': 'Send student_id as: ?student_id=7 (query parameter)',
                'fix_flutter': 'في diagnostic_service.dart أضف: ?student_id=$studentId في الـ URL',
                'example_url': '/api/diagnostic/student/assigned?student_id=7',
                'debug': {
                    'cookies_received': list(request.cookies.keys()),
                    'all_headers': {k: v for k, v in request.headers.items() if k.lower() not in ['cookie', 'authorization']}
                }
            }), 400
        
        print(f"✅ Getting assigned tests for student {student_id}")
        
        tests = DiagnosticTest.query.filter(
            DiagnosticTest.is_scheduled == True,
            DiagnosticTest.is_active == True
        ).all()
        
        print(f"📋 Found {len(tests)} scheduled tests")
        
        # فلتر الاختبارات حسب الطالب
        assigned_tests = []
        for test in tests:
            print(f"🔍 Test {test.id}: assigned_students={test.assigned_students}")
            
            # تحقق من الإسناد
            is_assigned = False
            
            if not test.assigned_students or test.assigned_students == 'all':
                is_assigned = True
            elif test.assigned_students:
                if isinstance(test.assigned_students, list):
                    is_assigned = student_id in test.assigned_students
                elif isinstance(test.assigned_students, str):
                    import json
                    try:
                        students_list = json.loads(test.assigned_students)
                        is_assigned = student_id in students_list
                    except:
                        is_assigned = str(student_id) in test.assigned_students or (student_id == int(test.assigned_students) if test.assigned_students.isdigit() else False)
            
            print(f"  → Assigned: {is_assigned}")
            
            if is_assigned:
                is_available = True
                if test.scheduled_start and test.scheduled_end:
                    now = datetime.utcnow()
                    
                    # إزالة timezone للمقارنة الصحيحة
                    start = test.scheduled_start.replace(tzinfo=None) if test.scheduled_start.tzinfo else test.scheduled_start
                    end = test.scheduled_end.replace(tzinfo=None) if test.scheduled_end.tzinfo else test.scheduled_end
                    
                    is_available = start <= now <= end
                    
                    print(f"  → Available (time): {is_available}")
                    print(f"     Now (UTC): {now}")
                    print(f"     Start: {start}")
                    print(f"     End: {end}")
                
                if is_available:
                    assigned_tests.append(test)
        
        print(f"✅ Returning {len(assigned_tests)} assigned tests")
        
        return jsonify({
            'success': True,
            'assigned_tests': [test.to_dict() for test in assigned_tests]
        }), 200
        
    except Exception as e:
        print(f"❌ Error getting student tests: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@diagnostic_bp.route('/tests/<int:test_id>/cancel-schedule', methods=['POST'])
@login_required
@admin_required
def cancel_schedule(test_id):
    """إلغاء جدولة اختبار"""
    try:
        test = DiagnosticTest.query.get(test_id)
        if not test:
            return jsonify({'error': 'Test not found'}), 404
        
        test.is_scheduled = False
        test.schedule_status = 'cancelled'
        db.session.commit()
        
        return jsonify({
            'message': 'Schedule cancelled successfully',
            'test': test.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error cancelling schedule: {e}")
        return jsonify({'error': str(e)}), 500


@diagnostic_bp.route('/tests/<int:test_id>/send-notification', methods=['POST'])
@login_required
@admin_required
def resend_notification(test_id):
    """إعادة إرسال إشعار"""
    try:
        test = DiagnosticTest.query.get(test_id)
        if not test:
            return jsonify({'error': 'Test not found'}), 404
        
        if not test.is_scheduled or not test.assigned_students:
            return jsonify({'error': 'Test not scheduled or no students assigned'}), 400
        
        try:
            if not NotificationService:
                return jsonify({'error': 'Notification service not available'}), 500
            
            students = Student.query.filter(
                Student.id.in_(test.assigned_students),
                Student.fcm_token.isnot(None)
            ).all()
            
            # ✅ استخدام NotificationService
            success_count = 0
            for student in students:
                if student.fcm_token:
                    result = NotificationService.send_fcm_notification(
                        student.fcm_token,
                        '📝 اختبار تشخيصي',
                        f'{test.title}',
                        {
                            'type': 'diagnostic_test',
                            'test_id': str(test.id)
                        }
                    )
                    if result:
                        success_count += 1
            
            # ✅ حفظ الإشعار في قاعدة البيانات لكل الطلاب المعينين
            db_save_count = 0
            for sid in test.assigned_students:
                if _save_notification_to_db(
                    student_id=sid,
                    title='📝 تذكير: اختبار تشخيصي',
                    message=f'{test.title}',
                    notification_type='reminder',
                    data={'type': 'diagnostic_test', 'test_id': str(test.id)}
                ):
                    db_save_count += 1
            
            test.notification_sent = True
            test.notification_sent_at = datetime.utcnow()
            db.session.commit()
            print(f"✅ تم حفظ {db_save_count} إشعار في قاعدة البيانات (إعادة إرسال)")
            
            return jsonify({
                'message': f'Notifications sent to {success_count} students',
                'success_count': success_count
            }), 200
            
        except Exception as e:
            print(f"❌ Error sending notifications: {e}")
            return jsonify({'error': str(e)}), 500
        
    except Exception as e:
        print(f"❌ Error in resend_notification: {e}")
        return jsonify({'error': str(e)}), 500


# =====================================================
# ✅ Routes مساعدة للصفحة (مضافة - جديدة)
# =====================================================

@diagnostic_bp.route('/lessons', methods=['GET'])
def get_lessons():
    """جلب قائمة الدروس"""
    try:
        lessons = Lesson.query.filter_by(is_active=True).all()
        return jsonify({
            'lessons': [{'id': l.id, 'name': l.name, 'unit_id': l.unit_id} for l in lessons]
        }), 200
    except Exception as e:
        print(f"❌ Error getting lessons: {e}")
        return jsonify({'error': str(e)}), 500


@diagnostic_bp.route('/students', methods=['GET'])
def get_students():
    """جلب قائمة الطلاب"""
    try:
        students = Student.query.filter_by(is_active=True).all()
        return jsonify({
            'students': [{'id': s.id, 'name': s.name, 'grade': getattr(s, 'grade', None)} for s in students]
        }), 200
    except Exception as e:
        print(f"❌ Error getting students: {e}")
        return jsonify({'error': str(e)}), 500


@diagnostic_bp.route('/grades', methods=['GET'])
def get_grades():
    """جلب قائمة الصفوف الدراسية المتوفرة"""
    try:
        # جلب جميع الصفوف الفريدة من جدول الطلاب
        grades_query = db.session.query(Student.grade).filter(
            Student.is_active == True,
            Student.grade.isnot(None),
            Student.grade != ''
        ).distinct().all()
        
        grades = [g[0] for g in grades_query if g[0]]
        
        # حساب عدد الطلاب لكل صف
        grades_with_count = []
        for grade in grades:
            count = Student.query.filter_by(is_active=True, grade=grade).count()
            grades_with_count.append({
                'grade': grade,
                'count': count
            })
        
        return jsonify({
            'success': True,
            'grades': grades_with_count
        }), 200
    except Exception as e:
        print(f"❌ Error getting grades: {e}")
        return jsonify({'error': str(e)}), 500


# ✅ تحسين route الإحصائيات (معدل)
@diagnostic_bp.route('/stats', methods=['GET'])
@login_required
@admin_required
def get_diagnostic_stats():
    """إحصائيات الاختبارات التشخيصية"""
    try:
        total_tests = DiagnosticTest.query.filter_by(is_active=True).count()
        
        stats = {
            'total_tests': total_tests,
            'pre_tests': DiagnosticTest.query.filter_by(
                test_type='pre_test', 
                is_active=True
            ).count(),
            'post_tests': DiagnosticTest.query.filter_by(
                test_type='post_test',
                is_active=True
            ).count(),
            'scheduled_tests': 0  # قيمة افتراضية
        }
        
        # إحصائيات الجدولة (إذا كانت متوفرة)
        try:
            stats['scheduled_tests'] = DiagnosticTest.query.filter_by(
                is_scheduled=True,
                is_active=True
            ).count()
        except:
            pass
        
        return jsonify(stats), 200
        
    except Exception as e:
        print(f"❌ Error getting stats: {e}")
        return jsonify({
            'total_tests': 0,
            'pre_tests': 0,
            'post_tests': 0,
            'scheduled_tests': 0
        }), 200


print("🧪 Diagnostic Tests System with Scheduling - Loaded successfully!")


@diagnostic_bp.route('/results', methods=['GET'])
def get_all_results():
    """جلب جميع النتائج"""
    try:
        # جلب آخر 50 نتيجة
        results = DiagnosticResult.query\
            .order_by(DiagnosticResult.completed_at.desc())\
            .limit(50)\
            .all()
        
        results_data = []
        for r in results:
            try:
                result_dict = r.to_dict()
                # إضافة معلومات الاختبار
                if r.test:
                    result_dict['test_title'] = r.test.title
                    result_dict['test_type'] = r.test.test_type
                
                # جلب اسم الطالب + الشعبة
                if r.student_id:
                    student = Student.query.get(r.student_id)
                    if student:
                        result_dict['student_name'] = student.name
                        result_dict['section'] = _get_student_section(student.id)

                # ✅ مؤشر غش سريع بدون فتح التفاصيل
                meta = next((a for a in (r.answers or []) if isinstance(a, dict) and a.get('_meta')), None)
                if meta:
                    result_dict['left_app_count'] = meta.get('left_app_count', 0)
                    result_dict['screenshot_count'] = meta.get('screenshot_count', 0)

                # ✅ إكمال سريع بشكل غير طبيعي (أقل من ٨ ثواني لكل سؤال بالمتوسط)
                if r.status == 'completed' and r.time_spent_seconds and r.total_questions:
                    avg_per_q = r.time_spent_seconds / r.total_questions
                    if avg_per_q < 8:
                        result_dict['too_fast'] = True
                        result_dict['avg_seconds_per_question'] = round(avg_per_q, 1)

                # ✅ شذوذ مقارنة بمتوسط أداء الطالب بباقي الاختبارات التشخيصية
                if r.status == 'completed' and r.student_id:
                    hist_avg = _get_historical_avg(r.student_id, exclude_result_id=r.id)
                    if hist_avg is not None and (r.percentage or 0) - hist_avg > 30:
                        result_dict['anomalous_vs_history'] = True
                        result_dict['historical_avg'] = round(hist_avg, 1)

                results_data.append(result_dict)
            except Exception as e:
                print(f"⚠️ Error processing result {r.id}: {e}")
                continue
        
        return jsonify({
            'success': True,
            'results': results_data,
            'count': len(results_data)
        })
    except Exception as e:
        print(f"❌ Error getting results: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


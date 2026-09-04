# src/routes/academic_calendar_routes.py
"""مسرد إعداد الدروس — تقويم دراسي (تواريخ + إجازات + توزيع دروس تلقائي)"""

from flask import Blueprint, request, jsonify, render_template
from flask_login import login_required, current_user
from functools import wraps
from datetime import datetime, timedelta
from hijridate import Gregorian

try:
    from src.extensions import db
    from src.models.academic_calendar import AcademicCalendar
    from src.models.curriculum import Lesson, Unit, Course
    from src.models.teacher import Teacher
    from src.services.lesson_prep_service import lesson_prep_service
except ImportError:  # pragma: no cover
    from extensions import db
    from models.academic_calendar import AcademicCalendar
    from models.curriculum import Lesson, Unit, Course
    from models.teacher import Teacher
    from services.lesson_prep_service import lesson_prep_service

academic_calendar_bp = Blueprint('academic_calendar', __name__, url_prefix='/api/academic-calendar')

# حقول اليوم اللي يُسمح للمعلم يعدّلها (تعبئة محتوى بس) — الباقي (الإجازات/الحصص الأسبوعية/الهيكل) أدمن فقط
_TEACHER_EDITABLE_DAY_FIELDS = {
    'lesson_name', 'lesson_id', 'unit_id', 'unit_name',  # اختيار الدرس من قائمة المنهج بدل كتابة حرة
    'homework', 'notes', 'solved_problems', 'section', 'period_number',
}


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return jsonify({'success': False, 'error': 'يجب تسجيل الدخول'}), 401
        if not getattr(current_user, 'is_admin', False):
            return jsonify({'success': False, 'error': 'صلاحيات غير كافية'}), 403
        return f(*args, **kwargs)
    return decorated


def admin_or_teacher_required(f):
    """للشاشات اللي يشوفها الأدمن والمعلم (عرض + تعبئة الدرس) — التمييز الفعلي بين الاثنين يصير داخل الدالة"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return jsonify({'success': False, 'error': 'يجب تسجيل الدخول'}), 401
        if not (getattr(current_user, 'is_admin', False) or isinstance(current_user, Teacher)):
            return jsonify({'success': False, 'error': 'صلاحيات غير كافية'}), 403
        return f(*args, **kwargs)
    return decorated


_ARABIC_WEEKDAY = {
    6: 'الأحد',      # Python weekday(): Monday=0 ... Sunday=6
    0: 'الاثنين',
    1: 'الثلاثاء',
    2: 'الأربعاء',
    3: 'الخميس',
}
_ARABIC_WEEK_ORDINALS = [
    'الأول', 'الثاني', 'الثالث', 'الرابع', 'الخامس', 'السادس', 'السابع', 'الثامن', 'التاسع', 'العاشر',
    'الحادي عشر', 'الثاني عشر', 'الثالث عشر', 'الرابع عشر', 'الخامس عشر', 'السادس عشر',
    'السابع عشر', 'الثامن عشر', 'التاسع عشر', 'العشرون', 'الحادي والعشرون', 'الثاني والعشرون',
]


def _week_label(n):
    """رقم -> اسم أسبوع عربي (١، ٢، ٣...)؛ يرجع الرقم نفسه لو تعدّى القائمة الجاهزة"""
    return _ARABIC_WEEK_ORDINALS[n - 1] if 1 <= n <= len(_ARABIC_WEEK_ORDINALS) else str(n)


# ترتيب أيام الأسبوع الدراسي (الأحد أولاً) بترميز weekday() بايثون
_WEEKDAY_ORDER = [6, 0, 1, 2, 3]  # أحد، اثنين، ثلاثاء، أربعاء، خميس


def _auto_pick_weekdays(count):
    """يختار عدد أيام موزّع بالتساوي على الأسبوع الدراسي (أحد-خميس) بدل ما يحدده المستخدم يدوياً —
    مفيد لما يكون عندك عدة شعب لنفس المقرر بأيام مختلفة وما تعرف مسبقاً أي أيام بالضبط"""
    n = max(1, min(5, int(count)))
    if n == 5:
        return list(_WEEKDAY_ORDER)
    if n == 1:
        return [_WEEKDAY_ORDER[2]]  # الثلاثاء (منتصف الأسبوع)
    step = (len(_WEEKDAY_ORDER) - 1) / (n - 1)
    indices = sorted({round(i * step) for i in range(n)})
    return [_WEEKDAY_ORDER[i] for i in indices]


def _generate_weeks_from_range(start_date_str, end_date_str, holidays, class_weekdays=None):
    """يبني هيكل الأسابيع/الأيام تلقائياً من تاريخ بداية/نهاية ميلادي، مستثنياً الجمعة/السبت،
    ومحوّلاً كل تاريخ للهجري للعرض، ومعلّماً أيام الإجازات المحددة.
    class_weekdays: مجموعة أرقام weekday() (الاثنين=0..الأحد=6) للأيام اللي فيها حصة لهذا المقرر —
    None يعني كل أيام الأسبوع الدراسي (الأحد-الخميس) فيها حصة (السلوك الافتراضي القديم)"""
    start = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    if end < start:
        raise ValueError('تاريخ النهاية قبل تاريخ البداية')

    # حوّل نطاقات الإجازات/الاختبارات/المراجعة لتواريخ فعلية — كلها تُستثنى من توزيع الدروس
    # type: holiday (إجازة) | exam (اختبارات نظرية) | practical_exam (اختبارات عملية) | review (مراجعة)
    _DEFAULT_LABELS = {
        'holiday': 'إجازة',
        'exam': 'اختبارات نهاية الفصل الدراسي',
        'practical_exam': 'اختبارات عملية',
        'review': 'مراجعة',
    }
    holiday_ranges = []
    for h in (holidays or []):
        try:
            h_start = datetime.strptime(h['start_date'], '%Y-%m-%d').date()
            h_end = datetime.strptime(h.get('end_date') or h['start_date'], '%Y-%m-%d').date()
            h_type = h.get('type') or 'holiday'
            default_label = _DEFAULT_LABELS.get(h_type, 'إجازة')
            holiday_ranges.append((h_start, h_end, h.get('label') or default_label, h_type))
        except Exception:
            continue

    def _holiday_label_for(d):
        for h_start, h_end, label, h_type in holiday_ranges:
            if h_start <= d <= h_end:
                return label, h_type
        return None, None

    # مرحلة 1: جمّع الأيام بمجموعات كل 5 أيام (بغض النظر عن كونها إجازة) بدون ترقيم بعد
    raw_weeks = []
    current_week_days = []
    d = start
    while d <= end:
        wd = d.weekday()
        if wd in _ARABIC_WEEKDAY:  # يستثني الجمعة (4) والسبت (5) تلقائياً
            hijri = Gregorian(d.year, d.month, d.day).to_hijri()
            holiday_label, day_type = _holiday_label_for(d)
            if day_type is None and class_weekdays is not None and wd not in class_weekdays:
                day_type = 'no_class'  # يوم دراسي عادي لكن ما فيه حصة لهذا المقرر
            current_week_days.append({
                'day_name': _ARABIC_WEEKDAY[wd],
                'hijri_date': f'{hijri.day}/{hijri.month}',
                'is_holiday': holiday_label is not None,
                'holiday_label': holiday_label,
                'day_type': day_type or 'study',  # study | holiday | exam | no_class
            })
            if len(current_week_days) == 5:
                raw_weeks.append(current_week_days)
                current_week_days = []
        d += timedelta(days=1)
    if current_week_days:  # آخر أسبوع ناقص (أقل من 5 أيام)
        raw_weeks.append(current_week_days)

    # مرحلة 2: رقّم الأسابيع بالتسلسل — أسبوع كامل إجازة (كل أيامه is_holiday) ما يُحسب
    # ضمن عدّاد "الأسبوع الدراسي" (نفس منطق المسرد الرسمي)، يُعرض كبطاقة إجازة بدون رقم
    weeks = []
    week_num = 0
    for days in raw_weeks:
        is_full_holiday = all(day['is_holiday'] for day in days)
        if is_full_holiday:
            labels = {day['holiday_label'] for day in days if day['holiday_label']}
            weeks.append({
                'week_number': None,
                'week_label': None,
                'is_full_holiday_week': True,
                'holiday_summary': ' / '.join(sorted(labels)) if labels else 'إجازة',
                'days': days,
            })
        else:
            week_num += 1
            weeks.append({'week_number': week_num, 'week_label': _week_label(week_num), 'days': days})

    return weeks


def _auto_fill_lessons(course_id, weeks):
    """يوزّع دروس المقرر (بالترتيب: وحدة فوحدة، درس فدرس) على الأيام غير العطلة تسلسلياً"""
    units = Unit.query.filter_by(course_id=course_id).order_by(Unit.order_num).all()
    lesson_queue = []
    for u in units:
        lessons = Lesson.query.filter_by(unit_id=u.id).order_by(Lesson.order_num).all()
        for l in lessons:
            lesson_queue.append({
                'unit_id': u.id, 'unit_name': u.name,
                'lesson_id': l.id, 'lesson_name': l.name,
            })

    idx = 0
    for week in weeks:
        for day in week.get('days', []):
            if day.get('is_holiday') or day.get('day_type') == 'no_class':
                day['unit_id'] = None
                day['unit_name'] = None
                day['lesson_id'] = None
                day['lesson_name'] = None
            elif idx < len(lesson_queue):
                item = lesson_queue[idx]
                day['unit_id'] = item['unit_id']
                day['unit_name'] = item['unit_name']
                day['lesson_id'] = item['lesson_id']
                day['lesson_name'] = item['lesson_name']
                idx += 1
            else:
                day['unit_id'] = None
                day['unit_name'] = None
                day['lesson_id'] = None
                day['lesson_name'] = None
            day.setdefault('period_number', 1)
            day.setdefault('section', '')
            day.setdefault('solved_problems', '')
            day.setdefault('homework', '')
            day.setdefault('notes', '')
            day.setdefault('holiday_label', None)
            day.setdefault('day_type', 'exam' if day.get('is_holiday') and 'اختبار' in (day.get('holiday_label') or '') else ('holiday' if day.get('is_holiday') else 'study'))
    return weeks, len(lesson_queue), idx


@academic_calendar_bp.route('/page', methods=['GET'])
@login_required
@admin_required
def calendar_page():
    """صفحة مسرد إعداد الدروس"""
    return render_template('academic_calendar.html')


@academic_calendar_bp.route('/course/<int:course_id>/lessons', methods=['GET'])
@login_required
@admin_or_teacher_required
def course_lessons(course_id):
    """دروس مقرر معيّن مرتّبة (وحدة فدرس) — تُستخدم بقائمة اختيار الدرس بدل الكتابة الحرة"""
    try:
        units = Unit.query.filter_by(course_id=course_id).order_by(Unit.order_num).all()
        lessons = []
        for u in units:
            for l in Lesson.query.filter_by(unit_id=u.id).order_by(Lesson.order_num).all():
                lessons.append({
                    'unit_id': u.id, 'unit_name': u.name,
                    'lesson_id': l.id, 'lesson_name': l.name,
                })
        return jsonify({'success': True, 'lessons': lessons})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@academic_calendar_bp.route('/list', methods=['GET'])
@login_required
@admin_or_teacher_required
def list_calendars():
    """قائمة كل التقاويم المحفوظة (بدون weeks_data الكاملة لتخفيف الحمل)"""
    try:
        calendars = AcademicCalendar.query.order_by(AcademicCalendar.updated_at.desc()).all()
        return jsonify({
            'success': True,
            'calendars': [{
                'id': c.id,
                'course_id': c.course_id,
                'course_name': c.course.name if c.course else None,
                'semester_number': c.semester_number,
                'academic_year_label': c.academic_year_label,
                'section': c.section,
                'weeks_count': len([w for w in (c.weeks_data or []) if w.get('week_number') is not None]),
            } for c in calendars],
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@academic_calendar_bp.route('/<int:calendar_id>', methods=['GET'])
@login_required
@admin_or_teacher_required
def get_calendar(calendar_id):
    try:
        cal = AcademicCalendar.query.get(calendar_id)
        if not cal:
            return jsonify({'success': False, 'error': 'التقويم غير موجود'}), 404
        return jsonify({'success': True, 'calendar': cal.to_dict()})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@academic_calendar_bp.route('/<int:calendar_id>/export-excel', methods=['GET'])
@login_required
@admin_or_teacher_required
def export_calendar_excel(calendar_id):
    """تصدير التقويم كملف Excel — متاح للأدمن والمعلم"""
    try:
        cal = AcademicCalendar.query.get(calendar_id)
        if not cal:
            return jsonify({'success': False, 'error': 'التقويم غير موجود'}), 404

        from io import BytesIO
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from flask import send_file

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'مسرد إعداد الدروس'
        ws.sheet_view.rightToLeft = True

        headers = ['الأسبوع', 'اليوم والتاريخ', 'الحصة', 'الشعبة', 'الوحدة', 'الدرس',
                   'المسائل المحلولة', 'الواجب', 'ملاحظات']
        thin = Side(style='thin', color='CBD5E1')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        header_fill = PatternFill('solid', fgColor='0D9488')
        for col, h in enumerate(headers, start=1):
            c = ws.cell(1, col, h)
            c.font = Font(color='FFFFFF', bold=True, size=11)
            c.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
            c.fill = header_fill
            c.border = border
        ws.row_dimensions[1].height = 26

        _EXAM_TYPE_LABELS = {'holiday': 'إجازة', 'review': 'مراجعة', 'practical_exam': 'اختبارات عملية', 'exam': 'اختبارات نهاية الفصل'}
        row_num = 2
        for week in (cal.weeks_data or []):
            week_header = f"الأسبوع {week['week_label']}" if week.get('week_number') is not None else (week.get('holiday_summary') or 'إجازة')
            for i, day in enumerate(week.get('days', [])):
                date_str = f"{day.get('day_name', '')} {day.get('hijri_date', '')}هـ"
                if day.get('is_holiday'):
                    label = day.get('holiday_label') or _EXAM_TYPE_LABELS.get(day.get('day_type'), 'إجازة')
                    ws.cell(row_num, 1, week_header if i == 0 else '')
                    ws.cell(row_num, 2, date_str)
                    ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=len(headers))
                    lc = ws.cell(row_num, 3, label)
                    lc.alignment = Alignment(horizontal='center', vertical='center')
                    lc.font = Font(bold=True, color='9A3412' if day.get('day_type') != 'exam' else '9D174D')
                elif day.get('day_type') == 'no_class':
                    ws.cell(row_num, 1, week_header if i == 0 else '')
                    ws.cell(row_num, 2, date_str)
                    ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=len(headers))
                    ws.cell(row_num, 3, 'لا يوجد حصة لهذا المقرر').alignment = Alignment(horizontal='center', vertical='center')
                else:
                    values = [
                        week_header if i == 0 else '', date_str, day.get('period_number') or '', day.get('section') or '',
                        day.get('unit_name') or '', day.get('lesson_name') or '', day.get('solved_problems') or '',
                        day.get('homework') or '', day.get('notes') or '',
                    ]
                    for col, v in enumerate(values, start=1):
                        ws.cell(row_num, col, v)
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row_num, col)
                    cell.border = border
                    if col <= 2:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                row_num += 1

        for col, width in zip(range(1, len(headers) + 1), [10, 14, 8, 10, 18, 26, 20, 18, 20]):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # صف المصدر (نفس تذييل بقية تقارير التطبيق)
        footer_row = row_num + 1
        ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=len(headers))
        fc = ws.cell(footer_row, 1)
        fc.value = f'⚗️  تم استخراج هذا التقرير من تطبيق كيم تحصيلي  |  منصة تعليمية للكيمياء  |  جميع الحقوق محفوظة © {datetime.now().year}'
        fc.font = Font(size=9, color='888888', italic=True)
        fc.alignment = Alignment(horizontal='center', vertical='center')
        fc.fill = PatternFill('solid', fgColor='F1F5F9')
        ws.row_dimensions[footer_row].height = 18

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        course_name = (cal.course.name if cal.course else 'مقرر').replace('/', '-')
        section_part = f'_شعبة{cal.section}' if cal.section else ''
        filename = f'مسرد_{course_name}{section_part}_فصل{cal.semester_number}.xlsx'
        return send_file(
            output, as_attachment=True, download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        print(f"❌ Error exporting calendar to excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@academic_calendar_bp.route('/extract-from-pdf', methods=['POST'])
@login_required
@admin_required
def extract_from_pdf():
    """يرفع PDF تقويم دراسي رسمي (مثل مسرد الوزارة) ويستخرج منه بداية/نهاية الفصل + قائمة الإجازات
    بالذكاء الاصطناعي — بدون حفظ، يرجّع البيانات للمراجعة والتعديل قبل التأكيد عبر /setup"""
    try:
        if 'pdf' not in request.files:
            return jsonify({'success': False, 'error': 'لم يتم إرفاق ملف PDF'}), 400
        pdf_file = request.files['pdf']
        if not pdf_file.filename.lower().endswith('.pdf'):
            return jsonify({'success': False, 'error': 'الملف يجب أن يكون PDF'}), 400

        pdf_bytes = pdf_file.read()
        import fitz
        import gc
        images = []
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        for page_num in range(len(doc)):
            page = doc[page_num]
            pix = page.get_pixmap(matrix=fitz.Matrix(1.5, 1.5))
            images.append(pix.tobytes("jpeg"))
            del pix
        doc.close()
        gc.collect()

        if not images:
            return jsonify({'success': False, 'error': 'تعذّر قراءة الملف'}), 400

        prompt = """حلّل صور تقويم دراسي رسمي مرفقة (جدول أسابيع/أيام دراسية). أحتاج منك بس:
1. تاريخ بداية الفصل الدراسي (أول يوم دراسي فعلي) بالميلادي
2. تاريخ نهاية الفصل الدراسي (آخر يوم دراسي، عادة آخر يوم اختبارات) بالميلادي
3. قائمة كل فترات التوقف عن الدروس العادية (إجازات رسمية، اختبارات، مراجعة، إجازات إضافية) بتواريخها الميلادية

لكل فترة حدد "type":
- "holiday" لأي إجازة عادية
- "review" لأسبوع/أيام المراجعة
- "practical_exam" للاختبارات العملية
- "exam" لاختبارات نهاية الفصل

لو التاريخ مذكور بالهجري بس، حوّله للميلادي بالاستناد لأي تاريخ ميلادي مرافق مذكور بالجدول (زي "20 - 28 نوفمبر 2026م")، أو احسبه بنفسك. لو ما تقدر تحدد سنة ميلادية بثقة لفترة معينة، اذكرها بحقل "warnings" بدل ما تخمّن.

أجب بصيغة JSON فقط بدون أي نص إضافي، بالضبط بهذا الشكل:
```json
{
  "start_date": "2026-08-23",
  "end_date": "2027-01-14",
  "holidays": [
    {"start_date": "2026-11-20", "end_date": "2026-11-28", "type": "holiday", "label": "إجازة الخريف"},
    {"start_date": "2027-01-10", "end_date": "2027-01-14", "type": "exam", "label": "اختبارات نهاية الفصل الدراسي الأول"}
  ],
  "warnings": []
}
```"""
        text, _usage = lesson_prep_service._call_ai(
            prompt, label='academic_calendar_extract', images=images,
            plan_id=None, teacher_id=None, operation_type='calendar_extract',
        )
        extracted = lesson_prep_service._extract_json(text)
        if not extracted or not extracted.get('start_date') or not extracted.get('end_date'):
            return jsonify({'success': False, 'error': 'تعذّر استخراج البيانات من الملف بثقة — جرّب إدخالها يدوياً أو راجع وضوح الصور بالملف'}), 400

        return jsonify({'success': True, 'extracted': extracted})
    except Exception as e:
        print(f"❌ Error extracting calendar from PDF: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@academic_calendar_bp.route('/setup', methods=['POST'])
@login_required
@admin_required
def setup_calendar():
    """إنشاء تقويم جديد: يستقبل هيكل الأسابيع/الأيام (تواريخ + إجازات) ويوزّع دروس المنهج عليه تلقائياً"""
    try:
        data = request.get_json() or {}
        course_id = data.get('course_id')
        semester_number = data.get('semester_number')
        academic_year_label = (data.get('academic_year_label') or '').strip()
        section = (data.get('section') or '').strip() or None

        if not course_id or not semester_number or not academic_year_label:
            return jsonify({'success': False, 'error': 'course_id و semester_number و academic_year_label مطلوبة'}), 400

        course = Course.query.get(course_id)
        if not course:
            return jsonify({'success': False, 'error': 'المقرر غير موجود'}), 404

        # ✅ طريقتان لبناء الأسابيع: (أ) تاريخ بداية/نهاية + إجازات (تلقائي، موصى به)
        #    (ب) هيكل weeks جاهز يدوياً (الطريقة القديمة، تبقى مدعومة)
        raw_holidays = data.get('holidays') or []
        raw_class_weekdays = data.get('class_weekdays')
        weekly_periods_count = data.get('weekly_periods_count')
        # لو حدد عدد الحصص بس (مو أيام معيّنة) — النظام يختار الأيام بنفسه موزّعة بالتساوي
        # (مفيد لما يكون عندك عدة شعب لنفس المقرر بأيام مختلفة وما تعرف مسبقاً أي أيام بالضبط)
        if not raw_class_weekdays and weekly_periods_count:
            raw_class_weekdays = _auto_pick_weekdays(weekly_periods_count)
        if data.get('start_date') and data.get('end_date'):
            # class_weekdays: أرقام weekday() (الاثنين=0..الأحد=6) للأيام اللي فيها حصة لهذا المقرر
            # ما تُرسل = كل أيام الأسبوع الدراسي فيها حصة (الافتراضي)
            class_weekdays = set(raw_class_weekdays) if raw_class_weekdays else None
            try:
                weeks = _generate_weeks_from_range(
                    data['start_date'], data['end_date'], raw_holidays, class_weekdays)
            except ValueError as ve:
                return jsonify({'success': False, 'error': str(ve)}), 400
        else:
            weeks = data.get('weeks') or []

        if not weeks:
            return jsonify({'success': False, 'error': 'ما فيه أيام دراسية بالنطاق المحدد'}), 400

        existing = AcademicCalendar.query.filter_by(
            course_id=course_id, semester_number=semester_number,
            academic_year_label=academic_year_label, section=section,
        ).first()
        if existing:
            return jsonify({'success': False, 'error': 'يوجد تقويم محفوظ مسبقاً بنفس المقرر والشعبة والفصل والعام — عدّله من شاشة التعديل بدل إنشاء واحد جديد'}), 400

        weeks, total_lessons, used_lessons = _auto_fill_lessons(course_id, weeks)

        cal = AcademicCalendar(
            course_id=course_id,
            semester_number=semester_number,
            academic_year_label=academic_year_label,
            section=section,
            weeks_data=weeks,
            start_date=datetime.strptime(data['start_date'], '%Y-%m-%d').date() if data.get('start_date') else None,
            end_date=datetime.strptime(data['end_date'], '%Y-%m-%d').date() if data.get('end_date') else None,
            class_weekdays=list(raw_class_weekdays) if raw_class_weekdays else None,
            holidays=raw_holidays or None,
        )
        db.session.add(cal)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'تم إنشاء التقويم وتوزيع الدروس تلقائياً',
            'calendar': cal.to_dict(),
            'total_lessons': total_lessons,
            'used_lessons': used_lessons,
        })
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error setting up academic calendar: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@academic_calendar_bp.route('/<int:calendar_id>/regenerate', methods=['POST'])
@login_required
@admin_required
def regenerate_calendar(calendar_id):
    """إعادة توزيع الدروس تلقائياً من جديد (لو تغيّر المنهج) — يمسح أي تعديل يدوي سابق"""
    try:
        cal = AcademicCalendar.query.get(calendar_id)
        if not cal:
            return jsonify({'success': False, 'error': 'التقويم غير موجود'}), 404

        weeks, total_lessons, used_lessons = _auto_fill_lessons(cal.course_id, cal.weeks_data or [])
        cal.weeks_data = weeks
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'تم إعادة التوزيع',
            'calendar': cal.to_dict(),
            'total_lessons': total_lessons,
            'used_lessons': used_lessons,
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@academic_calendar_bp.route('/<int:calendar_id>/holidays', methods=['POST'])
@login_required
@admin_required
def add_holiday(calendar_id):
    """إضافة إجازة/اختبارات/مراجعة على تقويم موجود — يعيد بناء الأسابيع كاملة (يمسح أي تعديل يدوي سابق)"""
    try:
        cal = AcademicCalendar.query.get(calendar_id)
        if not cal:
            return jsonify({'success': False, 'error': 'التقويم غير موجود'}), 404
        if not cal.start_date or not cal.end_date:
            return jsonify({'success': False, 'error': 'هذا التقويم أُنشئ بطريقة قديمة ما تخزّن فيها تواريخ البداية/النهاية — احذفه وأنشئ واحد جديد عشان تقدر تضيف إجازات لاحقاً'}), 400

        data = request.get_json() or {}
        start_date = data.get('start_date')
        end_date = data.get('end_date') or start_date
        h_type = data.get('type') or 'holiday'
        label = (data.get('label') or '').strip()
        if not start_date:
            return jsonify({'success': False, 'error': 'start_date مطلوب'}), 400

        holidays = list(cal.holidays or [])
        holidays.append({'start_date': start_date, 'end_date': end_date, 'type': h_type, 'label': label})

        class_weekdays = set(cal.class_weekdays) if cal.class_weekdays else None
        weeks = _generate_weeks_from_range(cal.start_date.isoformat(), cal.end_date.isoformat(), holidays, class_weekdays)
        weeks, total_lessons, used_lessons = _auto_fill_lessons(cal.course_id, weeks)

        cal.holidays = holidays
        cal.weeks_data = weeks
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'تمت إضافة الإجازة وإعادة بناء التقويم',
            'calendar': cal.to_dict(),
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@academic_calendar_bp.route('/<int:calendar_id>/holidays/<int:index>', methods=['DELETE'])
@login_required
@admin_required
def remove_holiday(calendar_id, index):
    """حذف إجازة مضافة سابقاً (بالترتيب الظاهر بالقائمة) — يعيد بناء الأسابيع كاملة"""
    try:
        cal = AcademicCalendar.query.get(calendar_id)
        if not cal:
            return jsonify({'success': False, 'error': 'التقويم غير موجود'}), 404
        if not cal.start_date or not cal.end_date:
            return jsonify({'success': False, 'error': 'هذا التقويم ما يدعم التعديل — احذفه وأنشئ واحد جديد'}), 400

        holidays = list(cal.holidays or [])
        if index < 0 or index >= len(holidays):
            return jsonify({'success': False, 'error': 'إجازة غير موجودة'}), 404
        holidays.pop(index)

        class_weekdays = set(cal.class_weekdays) if cal.class_weekdays else None
        weeks = _generate_weeks_from_range(cal.start_date.isoformat(), cal.end_date.isoformat(), holidays, class_weekdays)
        weeks, total_lessons, used_lessons = _auto_fill_lessons(cal.course_id, weeks)

        cal.holidays = holidays
        cal.weeks_data = weeks
        db.session.commit()

        return jsonify({'success': True, 'message': 'تم الحذف وإعادة بناء التقويم', 'calendar': cal.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@academic_calendar_bp.route('/<int:calendar_id>/day', methods=['PUT'])
@login_required
@admin_or_teacher_required
def update_day(calendar_id):
    """تعديل خانة يوم واحد يدوياً (موضوع الدرس/الواجب/الملاحظات) — الأدمن يعدّل أي حقل،
    المعلم يعدّل حقول التعبئة بس (_TEACHER_EDITABLE_DAY_FIELDS)"""
    try:
        cal = AcademicCalendar.query.get(calendar_id)
        if not cal:
            return jsonify({'success': False, 'error': 'التقويم غير موجود'}), 404

        is_admin = getattr(current_user, 'is_admin', False)
        data = request.get_json() or {}
        if not is_admin:
            data = {k: v for k, v in data.items() if k in _TEACHER_EDITABLE_DAY_FIELDS or k in ('week_number', 'day_name')}
        week_number = data.get('week_number')
        day_name = data.get('day_name')
        if week_number is None or not day_name:
            return jsonify({'success': False, 'error': 'week_number و day_name مطلوبة'}), 400

        weeks = cal.weeks_data or []
        found = False
        for week in weeks:
            if week.get('week_number') != week_number:
                continue
            for day in week.get('days', []):
                if day.get('day_name') != day_name:
                    continue
                if 'lesson_id' in data:
                    day['lesson_id'] = data.get('lesson_id')
                if 'lesson_name' in data:
                    day['lesson_name'] = data.get('lesson_name')
                if 'unit_id' in data:
                    day['unit_id'] = data.get('unit_id')
                if 'unit_name' in data:
                    day['unit_name'] = data.get('unit_name')
                if 'period_number' in data:
                    day['period_number'] = data.get('period_number')
                if 'section' in data:
                    day['section'] = data.get('section')
                if 'solved_problems' in data:
                    day['solved_problems'] = data.get('solved_problems')
                if 'homework' in data:
                    day['homework'] = data.get('homework')
                if 'notes' in data:
                    day['notes'] = data.get('notes')
                if 'is_holiday' in data:
                    day['is_holiday'] = data.get('is_holiday')
                if 'holiday_label' in data:
                    day['holiday_label'] = data.get('holiday_label')
                if 'day_type' in data:
                    day['day_type'] = data.get('day_type')  # study | holiday | exam
                found = True
                break
            if found:
                break

        if not found:
            return jsonify({'success': False, 'error': 'اليوم غير موجود بهذا التقويم'}), 404

        cal.weeks_data = weeks
        db.session.commit()
        return jsonify({'success': True, 'message': 'تم الحفظ'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@academic_calendar_bp.route('/<int:calendar_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_calendar(calendar_id):
    try:
        cal = AcademicCalendar.query.get(calendar_id)
        if not cal:
            return jsonify({'success': False, 'error': 'التقويم غير موجود'}), 404
        db.session.delete(cal)
        db.session.commit()
        return jsonify({'success': True, 'message': 'تم الحذف'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

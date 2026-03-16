# ==================== Backend - Flask Reports Routes (Updated for StudentResult) ====================
# routes/reports.py
"""
نظام التقارير المتكامل - محدث لاستخدام StudentResult
✅ متوافق مع students.py
✅ يستخدم نظام التوقيت الذكي
✅ APIs للتطبيق وصفحات للويب
✅ متوافق مع StudentResult model
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from functools import wraps
from src.extensions import db
from src.models.student import Student
from src.models.student_result import StudentResult  # ✅ تم التحديث
from src.models.curriculum import Course
from datetime import datetime, timedelta
import pytz
import io

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')


# ==================== نظام التوقيت (مطابق لـ students.py) ====================
def get_utc_time():
    """الحصول على الوقت الحالي بتوقيت UTC"""
    return datetime.now(pytz.utc)


def convert_utc_to_timezone(utc_dt, timezone_str='Asia/Riyadh'):
    """تحويل وقت UTC إلى أي منطقة زمنية"""
    if utc_dt is None:
        return None
    
    try:
        if utc_dt.tzinfo is None:
            utc_dt = pytz.utc.localize(utc_dt)
        
        target_tz = pytz.timezone(timezone_str)
        return utc_dt.astimezone(target_tz)
    except Exception as e:
        print(f"⚠️ خطأ في تحويل التوقيت: {e}")
        return utc_dt


def get_user_timezone_from_request():
    """الحصول على المنطقة الزمنية من الـ request"""
    timezone = request.headers.get('X-Timezone', 'Asia/Riyadh')
    
    try:
        pytz.timezone(timezone)
        return timezone
    except:
        return 'Asia/Riyadh'


# ==================== Middleware ====================
def admin_required(f):
    """التحقق من صلاحيات الأدمن"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            if request.is_json:
                return jsonify({'success': False, 'error': 'Unauthorized'}), 403
            flash('ليس لديك صلاحية الوصول لهذه الصفحة', 'danger')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function


# ==================== 1. تقرير الطلاب الشامل (API) ====================
@reports_bp.route('/api/students-performance', methods=['GET'])
@login_required
@admin_required
def api_students_performance_report():
    """
    API: تقرير شامل عن أداء جميع الطلاب
    يستخدم من التطبيق (Flutter)
    """
    try:
        # الحصول على timezone المستخدم
        user_timezone = get_user_timezone_from_request()
        
        # جلب جميع الطلاب
        students = Student.query.all()
        report = []
        
        for student in students:
            # ✅ جلب نتائج الطالب من StudentResult
            results = StudentResult.query.filter_by(student_id=student.id).order_by(StudentResult.created_at.desc()).all()
            
            if not results:
                # طالب بدون اختبارات
                created_at_local = convert_utc_to_timezone(student.created_at, user_timezone) if student.created_at else None
                
                report.append({
                    'student_id': student.id,
                    'student_name': student.name,
                    'email': student.email or '',
                    'phone': student.phone or '',
                    'grade': student.grade or '',
                    'total_quizzes': 0,
                    'avg_score': 0,
                    'total_time_spent': 0,
                    'last_activity': None,
                    'best_score': 0,
                    'worst_score': 0,
                    'improvement_trend': 0,
                    'active_status': 'غير نشط',
                    'performance_level': 'لا توجد بيانات',
                    'created_at': created_at_local.isoformat() if created_at_local else None,
                    'strengths': [],
                    'weaknesses': [],
                    'recent_scores': [],
                })
                continue
            
            # حساب الإحصائيات
            total_quizzes = len(results)
            scores = [r.score_percentage for r in results]
            avg_score = sum(scores) / len(scores) if scores else 0
            total_time = sum(r.time_spent or 0 for r in results)
            best_score = max(scores) if scores else 0
            worst_score = min(scores) if scores else 0
            
            # تحويل آخر نشاط للتوقيت المحلي
            last_activity_utc = results[0].created_at
            last_activity_local = convert_utc_to_timezone(last_activity_utc, user_timezone)
            
            # حساب اتجاه التحسن
            improvement = calculate_improvement_trend(results)
            
            # تحديد حالة النشاط
            active_status = get_activity_status(last_activity_utc)
            
            # تحديد مستوى الأداء
            performance_level = get_performance_level(avg_score)
            
            # ✅ تحليل نقاط القوة والضعف (محدث)
            strengths, weaknesses = analyze_topics_from_student_results(student.id)
            
            # تحويل تاريخ التسجيل
            created_at_local = convert_utc_to_timezone(student.created_at, user_timezone) if student.created_at else None
            
            report.append({
                'student_id': student.id,
                'student_name': student.name,
                'email': student.email or '',
                'phone': student.phone or '',
                'grade': student.grade or '',
                'total_quizzes': total_quizzes,
                'avg_score': round(avg_score, 2),
                'total_time_spent': total_time,
                'avg_time_per_quiz': round(total_time / total_quizzes, 1) if total_quizzes > 0 else 0,
                'last_activity': last_activity_local.isoformat() if last_activity_local else None,
                'best_score': round(best_score, 2),
                'worst_score': round(worst_score, 2),
                'improvement_trend': round(improvement, 2),
                'active_status': active_status,
                'performance_level': performance_level,
                'strengths': strengths[:3],
                'weaknesses': weaknesses[:3],
                'recent_scores': scores[:10],
                'created_at': created_at_local.isoformat() if created_at_local else None,
            })
        
        # حساب الإحصائيات العامة
        total_students = len(students)
        active_students = sum(1 for s in report if s['active_status'] in ['نشط جداً', 'نشط'])
        avg_all_score = sum(s['avg_score'] for s in report) / len(report) if report else 0
        total_quizzes_all = sum(s['total_quizzes'] for s in report)
        
        # توزيع مستويات الأداء
        performance_distribution = {
            'excellent': sum(1 for s in report if s['avg_score'] >= 90),
            'very_good': sum(1 for s in report if 80 <= s['avg_score'] < 90),
            'good': sum(1 for s in report if 70 <= s['avg_score'] < 80),
            'acceptable': sum(1 for s in report if 60 <= s['avg_score'] < 70),
            'weak': sum(1 for s in report if 0 < s['avg_score'] < 60),
            'no_data': sum(1 for s in report if s['avg_score'] == 0),
        }
        
        return jsonify({
            'success': True,
            'report': report,
            'summary': {
                'total_students': total_students,
                'active_students': active_students,
                'inactive_students': total_students - active_students,
                'avg_all_score': round(avg_all_score, 2),
                'total_quizzes_all': total_quizzes_all,
                'performance_distribution': performance_distribution,
            },
            'timezone': user_timezone,
        })
        
    except Exception as e:
        print(f'❌ Error in students performance report: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==================== 2. تقرير الطلاب المتميزين (API) ====================
@reports_bp.route('/api/top-performers', methods=['GET'])
@login_required
@admin_required
def api_top_performers():
    """API: تقرير الطلاب المتميزين"""
    try:
        limit = request.args.get('limit', 10, type=int)
        user_timezone = get_user_timezone_from_request()
        
        # جلب جميع الطلاب مع نتائجهم
        students = Student.query.all()
        students_with_scores = []
        
        for student in students:
            # ✅ استخدام StudentResult
            results = StudentResult.query.filter_by(student_id=student.id).all()
            
            if results:
                avg_score = sum(r.score_percentage for r in results) / len(results)
                total_quizzes = len(results)
                last_activity = convert_utc_to_timezone(results[0].created_at, user_timezone) if results else None
                
                students_with_scores.append({
                    'student_id': student.id,
                    'student_name': student.name,
                    'email': student.email or '',
                    'avg_score': round(avg_score, 2),
                    'total_quizzes': total_quizzes,
                    'best_score': round(max(r.score_percentage for r in results), 2),
                    'last_activity': last_activity.isoformat() if last_activity else None,
                    'performance_level': get_performance_level(avg_score),
                })
        
        # ترتيب حسب المعدل
        students_with_scores.sort(key=lambda x: x['avg_score'], reverse=True)
        top_performers = students_with_scores[:limit]
        
        # إضافة الترتيب والشارات
        for idx, student in enumerate(top_performers, 1):
            student['rank'] = idx
            student['badge'] = get_badge(idx)
        
        return jsonify({
            'success': True,
            'top_performers': top_performers,
            'total_count': len(students_with_scores),
            'timezone': user_timezone,
        })
        
    except Exception as e:
        print(f'❌ Error in top performers: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 3. تقرير الطلاب المحتاجين للمساعدة (API) ====================
@reports_bp.route('/api/need-help', methods=['GET'])
@login_required
@admin_required
def api_students_need_help():
    """API: تقرير الطلاب المحتاجين للمساعدة"""
    try:
        user_timezone = get_user_timezone_from_request()
        students = Student.query.filter_by(is_active=True).all()
        need_help = []
        
        current_time_utc = get_utc_time()
        
        for student in students:
            # ✅ استخدام StudentResult
            results = StudentResult.query.filter_by(student_id=student.id).order_by(StudentResult.created_at.desc()).all()
            
            if not results:
                # طالب بدون اختبارات منذ التسجيل
                if student.created_at:
                    student_created = student.created_at
                    if student_created.tzinfo is None:
                        student_created = pytz.utc.localize(student_created)
                    
                    days_since_registration = (current_time_utc - student_created).days
                    if days_since_registration > 7:
                        need_help.append({
                            'student_id': student.id,
                            'student_name': student.name,
                            'email': student.email or '',
                            'issue': 'لم يبدأ الاختبارات بعد',
                            'severity': 'high',
                            'avg_score': 0,
                            'total_quizzes': 0,
                            'last_activity': None,
                            'recommendation': 'تواصل مع الطالب لتشجيعه على البدء',
                            'days_inactive': days_since_registration
                        })
                continue
            
            avg_score = sum(r.score_percentage for r in results) / len(results)
            last_activity_utc = results[0].created_at
            last_activity_local = convert_utc_to_timezone(last_activity_utc, user_timezone)
            
            # التأكد من timezone
            if last_activity_utc.tzinfo is None:
                last_activity_utc = pytz.utc.localize(last_activity_utc)
            
            days_since_activity = (current_time_utc - last_activity_utc).days
            
            issue = None
            severity = 'low'
            recommendation = ''
            
            # 1. أداء ضعيف جداً
            if avg_score < 50:
                issue = f'أداء ضعيف جداً - المعدل {avg_score:.1f}%'
                severity = 'high'
                recommendation = 'يحتاج لمتابعة مكثفة ومراجعة المناهج الأساسية'
            
            # 2. أداء ضعيف
            elif avg_score < 60:
                issue = f'أداء ضعيف - المعدل {avg_score:.1f}%'
                severity = 'medium'
                recommendation = 'مراجعة المواضيع الصعبة والتدرب أكثر'
            
            # 3. تراجع مستمر
            elif len(results) >= 3:
                recent_scores = [r.score_percentage for r in results[:3]]
                if all(recent_scores[i] > recent_scores[i+1] for i in range(len(recent_scores)-1)):
                    issue = 'تراجع مستمر في الأداء'
                    severity = 'medium'
                    recommendation = 'مراجعة الأسباب والمواضيع الصعبة'
            
            # 4. غير نشط
            elif days_since_activity > 14:
                issue = f'لم يختبر منذ {days_since_activity} يوم'
                severity = 'high' if days_since_activity > 30 else 'medium'
                recommendation = 'إرسال تذكير للطالب'
            
            if issue:
                need_help.append({
                    'student_id': student.id,
                    'student_name': student.name,
                    'email': student.email or '',
                    'phone': student.phone or '',
                    'issue': issue,
                    'severity': severity,
                    'avg_score': round(avg_score, 2),
                    'total_quizzes': len(results),
                    'last_activity': last_activity_local.isoformat() if last_activity_local else None,
                    'recommendation': recommendation,
                    'days_inactive': days_since_activity
                })
        
        # ترتيب حسب الأولوية
        severity_order = {'high': 0, 'medium': 1, 'low': 2}
        need_help.sort(key=lambda x: (severity_order[x['severity']], -x['days_inactive']))
        
        return jsonify({
            'success': True,
            'students_need_help': need_help,
            'total': len(need_help),
            'high_priority': sum(1 for s in need_help if s['severity'] == 'high'),
            'medium_priority': sum(1 for s in need_help if s['severity'] == 'medium'),
            'low_priority': sum(1 for s in need_help if s['severity'] == 'low'),
        })
        
    except Exception as e:
        print(f'❌ Error in students need help: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 4. تقرير المناهج (API) ====================
@reports_bp.route('/api/courses-analysis', methods=['GET'])
@login_required
@admin_required
def api_courses_analysis():
    """
    ✅ API: تحليل أداء المناهج (محدث للعمل مع StudentResult)
    ملاحظة: هذا الـ route يعتمد على course_id في StudentResult
    """
    try:
        courses = Course.query.all()
        analysis = []
        
        for course in courses:
            # ✅ جلب النتائج مباشرة من StudentResult حسب course_id
            course_results = StudentResult.query.filter_by(course_id=course.id).all()
            
            if not course_results:
                continue
            
            avg_score = sum(r.score_percentage for r in course_results) / len(course_results)
            total_attempts = len(course_results)
            unique_students = len(set(r.student_id for r in course_results))
            
            # تحديد مستوى الصعوبة
            if avg_score >= 80:
                difficulty = 'سهل'
                difficulty_color = 'green'
            elif avg_score >= 60:
                difficulty = 'متوسط'
                difficulty_color = 'orange'
            else:
                difficulty = 'صعب'
                difficulty_color = 'red'
            
            # حساب عدد الاختبارات الفريدة (بناءً على quiz_name)
            unique_quizzes = len(set(r.quiz_name for r in course_results if r.quiz_name))
            
            analysis.append({
                'course_id': course.id,
                'course_name': course.name,
                'total_quizzes': unique_quizzes,
                'total_attempts': total_attempts,
                'unique_students': unique_students,
                'avg_score': round(avg_score, 2),
                'difficulty': difficulty,
                'difficulty_color': difficulty_color,
                'popularity_rank': total_attempts,
            })
        
        # ترتيب حسب الشعبية
        analysis.sort(key=lambda x: x['popularity_rank'], reverse=True)
        for idx, item in enumerate(analysis, 1):
            item['popularity_rank'] = idx
        
        difficulty_analysis = sorted(analysis, key=lambda x: x['avg_score'])
        
        return jsonify({
            'success': True,
            'courses_analysis': analysis,
            'most_popular': analysis[:5] if len(analysis) >= 5 else analysis,
            'most_difficult': difficulty_analysis[:5] if len(difficulty_analysis) >= 5 else difficulty_analysis,
            'easiest': difficulty_analysis[-5:] if len(difficulty_analysis) >= 5 else difficulty_analysis,
        })
        
    except Exception as e:
        print(f'❌ Error in courses analysis: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 5. تقرير النشاط (API) ====================
@reports_bp.route('/api/activity', methods=['GET'])
@login_required
@admin_required
def api_activity_report():
    """API: تقرير النشاط اليومي/الأسبوعي/الشهري"""
    try:
        period = request.args.get('period', 'week')
        user_timezone = get_user_timezone_from_request()
        
        # تحديد الفترة الزمنية بتوقيت UTC
        current_time_utc = get_utc_time()
        if period == 'day':
            start_date_utc = current_time_utc - timedelta(days=1)
        elif period == 'week':
            start_date_utc = current_time_utc - timedelta(weeks=1)
        else:  # month
            start_date_utc = current_time_utc - timedelta(days=30)
        
        # ✅ استخدام StudentResult
        results = StudentResult.query.filter(StudentResult.created_at >= start_date_utc).all()
        
        # تجميع البيانات حسب التاريخ (بالتوقيت المحلي للمستخدم)
        daily_stats = {}
        for result in results:
            # تحويل للتوقيت المحلي
            local_dt = convert_utc_to_timezone(result.created_at, user_timezone)
            date_key = local_dt.date().isoformat()
            
            if date_key not in daily_stats:
                daily_stats[date_key] = {
                    'date': date_key,
                    'total_quizzes': 0,
                    'total_students': set(),
                    'avg_score': [],
                    'total_time': 0,
                }
            
            daily_stats[date_key]['total_quizzes'] += 1
            daily_stats[date_key]['total_students'].add(result.student_id)
            daily_stats[date_key]['avg_score'].append(result.score_percentage)
            daily_stats[date_key]['total_time'] += result.time_spent or 0
        
        activity_data = []
        for date_key, stats in sorted(daily_stats.items()):
            activity_data.append({
                'date': date_key,
                'total_quizzes': stats['total_quizzes'],
                'unique_students': len(stats['total_students']),
                'avg_score': round(sum(stats['avg_score']) / len(stats['avg_score']), 2) if stats['avg_score'] else 0,
                'total_time_minutes': round(stats['total_time'] / 60, 1),
            })
        
        total_quizzes = sum(d['total_quizzes'] for d in activity_data)
        all_students = set()
        for result in results:
            all_students.add(result.student_id)
        
        return jsonify({
            'success': True,
            'period': period,
            'activity_data': activity_data,
            'summary': {
                'total_quizzes': total_quizzes,
                'unique_students': len(all_students),
                'avg_daily_quizzes': round(total_quizzes / len(activity_data), 1) if activity_data else 0,
            },
            'timezone': user_timezone,
        })
        
    except Exception as e:
        print(f'❌ Error in activity report: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 6. تقرير تفصيلي لطالب واحد - النسخة المبسطة (API) ====================
@reports_bp.route('/api/student/<int:student_id>', methods=['GET'])
@login_required
@admin_required
def api_student_report(student_id):
    """
    API: تقرير تفصيلي عن طالب واحد (النسخة المبسطة)
    ملاحظة: هناك أيضاً /api/student/<int:student_id>/detail للنسخة الكاملة
    """
    try:
        user_timezone = get_user_timezone_from_request()
        
        student = Student.query.get(student_id)
        if not student:
            return jsonify({'success': False, 'error': 'Student not found'}), 404
        
        # ✅ استخدام StudentResult
        results = StudentResult.query.filter_by(student_id=student_id).order_by(StudentResult.created_at.desc()).all()
        
        # تحويل تاريخ التسجيل
        created_at_local = convert_utc_to_timezone(student.created_at, user_timezone) if student.created_at else None
        
        if not results:
            return jsonify({
                'success': True,
                'student': {
                    'id': student.id,
                    'name': student.name,
                    'email': student.email or '',
                    'phone': student.phone or '',
                    'school': student.school or '',
                    'grade': student.grade or '',
                    'created_at': created_at_local.isoformat() if created_at_local else None,
                    'is_active': student.is_active,
                },
                'stats': {
                    'total_quizzes': 0,
                    'avg_score': 0,
                    'message': 'لم يقم الطالب بأي اختبار بعد'
                },
                'timezone': user_timezone,
            })
        
        # حساب الإحصائيات
        scores = [r.score_percentage for r in results]
        avg_score = sum(scores) / len(scores)
        total_time = sum(r.time_spent or 0 for r in results)
        
        # تحليل الأداء عبر الوقت (آخر 20 اختبار)
        performance_timeline = []
        for result in reversed(results[-20:]):
            result_time_local = convert_utc_to_timezone(result.created_at, user_timezone)
            
            performance_timeline.append({
                'date': result_time_local.isoformat() if result_time_local else None,
                'quiz_name': result.quiz_name or 'اختبار غير معروف',  # ✅
                'score': round(result.score_percentage, 2),
                'time_spent': result.time_spent or 0,
            })
        
        # تحليل نقاط القوة والضعف
        strengths, weaknesses = analyze_topics_from_student_results(student_id)
        
        # حساب الاتجاه
        improvement_trend = calculate_improvement_trend(results)
        
        # آخر نشاط
        last_activity_local = convert_utc_to_timezone(results[0].created_at, user_timezone)
        
        # توزيع الدرجات
        score_distribution = {
            'excellent': sum(1 for s in scores if s >= 90),
            'very_good': sum(1 for s in scores if 80 <= s < 90),
            'good': sum(1 for s in scores if 70 <= s < 80),
            'acceptable': sum(1 for s in scores if 60 <= s < 70),
            'weak': sum(1 for s in scores if s < 60),
        }
        
        # ✅ آخر 10 نتائج
        recent_results = []
        for r in results[:10]:
            result_time_local = convert_utc_to_timezone(r.created_at, user_timezone)
            
            recent_results.append({
                'quiz_id': r.quiz_id or r.id,
                'quiz_name': r.quiz_name or 'اختبار غير معروف',  # ✅
                'score': round(r.score_percentage, 2),
                'correct_answers': r.correct_answers,
                'total_questions': r.total_questions,
                'time_spent': r.time_spent or 0,
                'date': result_time_local.isoformat() if result_time_local else None,
            })
        
        return jsonify({
            'success': True,
            'student': {
                'id': student.id,
                'name': student.name,
                'email': student.email or '',
                'phone': student.phone or '',
                'school': student.school or '',
                'grade': student.grade or '',
                'created_at': created_at_local.isoformat() if created_at_local else None,
                'is_active': student.is_active,
            },
            'stats': {
                'total_quizzes': len(results),
                'avg_score': round(avg_score, 2),
                'best_score': round(max(scores), 2),
                'worst_score': round(min(scores), 2),
                'total_time_spent': total_time,
                'avg_time_per_quiz': round(total_time / len(results), 1),
                'last_activity': last_activity_local.isoformat() if last_activity_local else None,
                'improvement_trend': round(improvement_trend, 2),
                'performance_level': get_performance_level(avg_score),
                'active_status': get_activity_status(results[0].created_at),
            },
            'performance_timeline': performance_timeline,
            'score_distribution': score_distribution,
            'strengths': strengths,
            'weaknesses': weaknesses,
            'recent_results': recent_results,
            'timezone': user_timezone,
        })
        
    except Exception as e:
        print(f'❌ Error in student report: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 7. تقرير تفصيلي لطالب واحد - النسخة الكاملة (API) ====================
@reports_bp.route('/api/student/<int:student_id>/detail', methods=['GET'])
@login_required
@admin_required
def api_student_detail(student_id):
    """
    API: تقرير تفصيلي عن طالب محدد
    يستخدم من التطبيق (Flutter)
    """
    try:
        user_timezone = get_user_timezone_from_request()
        
        # جلب الطالب
        student = Student.query.get_or_404(student_id)
        
        # ✅ جلب نتائج الطالب من StudentResult
        results = StudentResult.query.filter_by(student_id=student.id).order_by(StudentResult.created_at.desc()).all()
        
        if not results:
            return jsonify({
                'success': True,
                'student': {
                    'id': student.id,
                    'name': student.name,
                    'email': student.email or '',
                    'phone': student.phone or '',
                    'school': student.school or '',
                    'grade': student.grade or '',
                    'created_at': convert_utc_to_timezone(student.created_at, user_timezone).isoformat() if student.created_at else None,
                    'is_active': student.is_active,
                },
                'stats': {
                    'total_quizzes': 0,
                    'avg_score': 0,
                    'best_score': 0,
                    'worst_score': 0,
                    'total_time_spent': 0,
                    'avg_time_per_quiz': 0,
                    'last_activity': None,
                    'improvement_trend': 0,
                    'performance_level': 'لا توجد بيانات',
                    'active_status': 'غير نشط',
                },
                'performance_timeline': [],
                'score_distribution': {
                    'excellent': 0,
                    'very_good': 0,
                    'good': 0,
                    'acceptable': 0,
                    'weak': 0,
                },
                'strengths': [],
                'weaknesses': [],
                'recent_results': [],
                'timezone': user_timezone,
            })
        
        # حساب الإحصائيات
        scores = [r.score_percentage for r in results]
        avg_score = sum(scores) / len(scores)
        total_time = sum(r.time_spent or 0 for r in results)
        
        # تحويل التواريخ
        created_at_local = convert_utc_to_timezone(student.created_at, user_timezone) if student.created_at else None
        last_activity_local = convert_utc_to_timezone(results[0].created_at, user_timezone) if results else None
        
        # حساب اتجاه التحسن
        improvement_trend = calculate_improvement_trend(results)
        
        # ✅ تحليل نقاط القوة والضعف (محدث)
        strengths, weaknesses = analyze_topics_from_student_results(student.id)
        
        # خط الأداء الزمني (آخر 20 نتيجة)
        performance_timeline = []
        for idx, r in enumerate(reversed(results[-20:])):
            result_time = convert_utc_to_timezone(r.created_at, user_timezone)
            performance_timeline.append({
                'index': idx + 1,
                'score': round(r.score_percentage, 2),
                'quiz_name': r.quiz_name,
                'date': result_time.isoformat() if result_time else None,
            })
        
        # توزيع الدرجات
        score_distribution = {
            'excellent': sum(1 for s in scores if s >= 90),
            'very_good': sum(1 for s in scores if 80 <= s < 90),
            'good': sum(1 for s in scores if 70 <= s < 80),
            'acceptable': sum(1 for s in scores if 60 <= s < 70),
            'weak': sum(1 for s in scores if s < 60),
        }
        
        # ✅ آخر 10 نتائج (محدث - استخدام quiz_name مباشرة)
        recent_results = []
        for r in results[:10]:
            result_time_local = convert_utc_to_timezone(r.created_at, user_timezone)
            
            recent_results.append({
                'quiz_id': r.quiz_id or r.id,  # استخدام quiz_id إن وجد، أو id النتيجة
                'quiz_name': r.quiz_name or 'اختبار غير معروف',  # ✅ استخدام quiz_name مباشرة
                'score': round(r.score_percentage, 2),
                'correct_answers': r.correct_answers,
                'total_questions': r.total_questions,
                'time_spent': r.time_spent or 0,
                'date': result_time_local.isoformat() if result_time_local else None,
            })
        
        return jsonify({
            'success': True,
            'student': {
                'id': student.id,
                'name': student.name,
                'email': student.email or '',
                'phone': student.phone or '',
                'school': student.school or '',
                'grade': student.grade or '',
                'created_at': created_at_local.isoformat() if created_at_local else None,
                'is_active': student.is_active,
            },
            'stats': {
                'total_quizzes': len(results),
                'avg_score': round(avg_score, 2),
                'best_score': round(max(scores), 2),
                'worst_score': round(min(scores), 2),
                'total_time_spent': total_time,
                'avg_time_per_quiz': round(total_time / len(results), 1),
                'last_activity': last_activity_local.isoformat() if last_activity_local else None,
                'improvement_trend': round(improvement_trend, 2),
                'performance_level': get_performance_level(avg_score),
                'active_status': get_activity_status(results[0].created_at),
            },
            'performance_timeline': performance_timeline,
            'score_distribution': score_distribution,
            'strengths': strengths,
            'weaknesses': weaknesses,
            'recent_results': recent_results,
            'timezone': user_timezone,
        })
        
    except Exception as e:
        print(f'❌ Error in student detail: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 8. صفحة التقارير الرئيسية (Web) ====================
@reports_bp.route('/')
@login_required
@admin_required
def index():
    """صفحة التقارير الرئيسية"""
    return render_template('admin/reports.html')


# ==================== 9. صفحة تفاصيل طالب (Web) ====================
@reports_bp.route('/student/<int:student_id>')
@login_required
@admin_required
def student_detail_page(student_id):
    """صفحة تفاصيل طالب"""
    student = Student.query.get_or_404(student_id)
    return render_template('admin/student_detail.html', student=student)


# ==================== 10. صفحة الطلاب المتميزين (Web) ====================
@reports_bp.route('/top-performers')
@login_required
@admin_required
def top_performers_page():
    """صفحة الطلاب المتميزين"""
    return render_template('admin/top_performers.html')


# ==================== 11. تصدير Excel (API) ====================
@reports_bp.route('/api/export-excel', methods=['POST'])
@login_required
@admin_required
def api_export_excel():
    """تصدير تقرير الطلاب إلى Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        body = request.json or {}
        data = body.get('students', body.get('data', []))
        report_author = body.get('author', 'الإدارة')

        if not data:
            return jsonify({'success': False, 'error': 'لا توجد بيانات للتصدير'}), 400

        wb = Workbook()

        # ==================== ورقة 1: تقرير الطلاب ====================
        ws = wb.active
        ws.title = 'تقرير الطلاب'
        ws.sheet_view.rightToLeft = True

        # أنماط مشتركة
        header_font      = Font(bold=True, size=11, color='FFFFFF')
        header_fill      = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_align     = Alignment(horizontal='center', vertical='center', wrap_text=True)
        center_align     = Alignment(horizontal='center', vertical='center')
        right_align      = Alignment(horizontal='right',  vertical='center')
        border_side      = Side(style='thin', color='CCCCCC')
        border           = Border(left=border_side, right=border_side,
                                  top=border_side,  bottom=border_side)
        title_font       = Font(bold=True, size=14, color='1E3A5F')
        meta_font        = Font(size=10, color='555555')

        # ── صف العنوان (مدمج) ──
        now_str = datetime.now().strftime('%Y-%m-%d  %H:%M')
        ws.merge_cells('A1:O1')
        title_cell = ws['A1']
        title_cell.value = f'تقرير أداء الطلاب  |  أعدّه: {report_author}  |  {now_str}'
        title_cell.font      = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill      = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
        ws.row_dimensions[1].height = 30

        # ── رأس الجدول (صف 2) ──
        headers = [
            '#', 'الاسم', 'الصف', 'البريد الإلكتروني', 'الهاتف',
            'عدد الاختبارات', 'المعدل العام', 'أفضل درجة', 'أقل درجة',
            'الوقت الكلي', 'متوسط الوقت', 'آخر نشاط',
            'حالة النشاط', 'مستوى الأداء', 'اتجاه التحسن'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(2, col, header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = border
        ws.row_dimensions[2].height = 28

        # ── بيانات الطلاب (من صف 3) ──
        even_fill = PatternFill(start_color='F8FAFF', end_color='F8FAFF', fill_type='solid')

        for row_idx, student in enumerate(data, 3):
            row_fill = even_fill if row_idx % 2 == 0 else None

            def _cell(col, value, align=center_align, bold=False, bg=None):
                c = ws.cell(row_idx, col, value)
                c.alignment = align
                c.border    = border
                c.font      = Font(bold=bold, size=10)
                if bg:
                    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                elif row_fill:
                    c.fill = row_fill

            _cell(1, row_idx - 2)
            _cell(2, student.get('student_name', ''), right_align, bold=True)
            _cell(3, student.get('grade', ''))
            _cell(4, student.get('email', ''), right_align)
            _cell(5, student.get('phone', ''))
            _cell(6, student.get('total_quizzes', 0))

            # المعدل — ملوّن
            avg_score = student.get('avg_score', 0)
            avg_bg = None
            if   avg_score >= 80: avg_bg = 'C6EFCE'
            elif avg_score >= 60: avg_bg = 'FFEB9C'
            elif avg_score >  0:  avg_bg = 'FFC7CE'
            _cell(7, f"{avg_score:.1f}%", bold=True, bg=avg_bg)

            best  = student.get('best_score',  0)
            worst = student.get('worst_score', 0)
            _cell(8, f"{best:.1f}%"  if best  > 0 else '-')
            _cell(9, f"{worst:.1f}%" if worst > 0 else '-')

            total_time = student.get('total_time_spent', 0)
            _cell(10, format_time_arabic(total_time))
            _cell(11, f"{student.get('avg_time_per_quiz', 0):.1f} ث")

            last_activity = student.get('last_activity', '')
            if last_activity:
                try:
                    dt = datetime.fromisoformat(last_activity.replace('Z', '+00:00'))
                    _cell(12, dt.strftime('%Y-%m-%d %H:%M'))
                except:
                    _cell(12, 'غير متاح')
            else:
                _cell(12, 'غير متاح')

            _cell(13, student.get('active_status', ''))
            _cell(14, student.get('performance_level', ''))

            trend = student.get('improvement_trend', 0)
            trend_text = f"+{trend:.1f}%" if trend > 0 else f"{trend:.1f}%"
            trend_bg = None
            if   trend >  5: trend_bg = 'C6EFCE'
            elif trend < -5: trend_bg = 'FFC7CE'
            _cell(15, trend_text, bg=trend_bg)

        # ضبط عرض الأعمدة
        col_widths = [5, 26, 12, 28, 14, 14, 13, 12, 10, 14, 14, 20, 14, 15, 14]
        for idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        # تجميد الصفين الأولين
        ws.freeze_panes = 'A3'

        # صف المصدر (آخر صف)
        footer_row = len(data) + 3
        ws.merge_cells(f'A{footer_row}:O{footer_row}')
        footer_cell = ws[f'A{footer_row}']
        footer_cell.value = '⚗️  تم استخراج هذا التقرير من تطبيق كيم تحصيلي  |  منصة تعليمية للكيمياء  |  جميع الحقوق محفوظة ©'
        footer_cell.font      = Font(size=9, color='888888', italic=True)
        footer_cell.alignment = Alignment(horizontal='center', vertical='center')
        footer_cell.fill      = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
        ws.row_dimensions[footer_row].height = 18

        # ==================== ورقة 2: الملخص ====================
        ws2 = wb.create_sheet('الملخص')
        ws2.sheet_view.rightToLeft = True

        total   = len(data)
        active  = sum(1 for s in data if s.get('active_status') in ['نشط جداً', 'نشط'])
        scores  = [s.get('avg_score', 0) for s in data]
        avg_all = sum(scores) / total if total else 0
        quizzes = sum(s.get('total_quizzes', 0) for s in data)

        dist = {
            'ممتاز (80%+)':            sum(1 for s in scores if s >= 80),
            'جيد (60-79%)':            sum(1 for s in scores if 60 <= s < 80),
            'يحتاج انتباه (40-59%)':   sum(1 for s in scores if 40 <= s < 60),
            'حرج (أقل من 40%)':        sum(1 for s in scores if 0 < s < 40),
            'لا توجد بيانات':          sum(1 for s in scores if s == 0),
        }

        s2_title_font  = Font(bold=True, size=13, color='FFFFFF')
        s2_header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        s2_label_font  = Font(bold=True, size=11)
        s2_val_font    = Font(size=11)

        def _s2_header(row, text):
            ws2.merge_cells(f'A{row}:B{row}')
            c = ws2[f'A{row}']
            c.value     = text
            c.font      = s2_title_font
            c.fill      = s2_header_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws2.row_dimensions[row].height = 24

        def _s2_row(row, label, value, val_bg=None):
            lc = ws2.cell(row, 1, label)
            lc.font      = s2_label_font
            lc.alignment = right_align
            lc.border    = border
            vc = ws2.cell(row, 2, value)
            vc.font      = s2_val_font
            vc.alignment = center_align
            vc.border    = border
            if val_bg:
                vc.fill = PatternFill(start_color=val_bg, end_color=val_bg, fill_type='solid')

        # معلومات التقرير
        _s2_header(1, 'معلومات التقرير')
        _s2_row(2, 'أعدّه',         report_author)
        _s2_row(3, 'تاريخ الإنشاء', now_str)
        _s2_row(4, 'عدد الطلاب في التقرير', total)

        # إحصائيات عامة
        _s2_header(6, 'إحصائيات عامة')
        _s2_row(7,  'إجمالي الطلاب',  total)
        _s2_row(8,  'النشطون',         active,           val_bg='C6EFCE')
        _s2_row(9,  'غير النشطين',     total - active,   val_bg='FFC7CE')
        _s2_row(10, 'المعدل العام',    f"{avg_all:.1f}%")
        _s2_row(11, 'إجمالي الاختبارات', quizzes)

        # توزيع الأداء
        _s2_header(13, 'توزيع مستويات الأداء')
        dist_colors = ['C6EFCE', 'FFEB9C', 'FFD18C', 'FFC7CE', 'EEEEEE']
        for i, (label, count) in enumerate(dist.items()):
            pct = f"{(count/total*100):.1f}%" if total else '0%'
            _s2_row(14 + i, label, f"{count} طالب  ({pct})", val_bg=dist_colors[i])

        ws2.column_dimensions['A'].width = 28
        ws2.column_dimensions['B'].width = 22

        # صف المصدر في ورقة الملخص
        s2_footer_row = 20
        ws2.merge_cells(f'A{s2_footer_row}:B{s2_footer_row}')
        s2_footer = ws2[f'A{s2_footer_row}']
        s2_footer.value = '⚗️  تم استخراج هذا التقرير من تطبيق كيم تحصيلي  |  منصة تعليمية للكيمياء  |  جميع الحقوق محفوظة ©'
        s2_footer.font      = Font(size=9, color='888888', italic=True)
        s2_footer.alignment = Alignment(horizontal='center', vertical='center')
        s2_footer.fill      = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
        ws2.row_dimensions[s2_footer_row].height = 18

        # ==================== حفظ وإرسال ====================
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = f'تقرير_الطلاب_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f'❌ Error exporting Excel: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== Helper Functions ====================

def calculate_improvement_trend(results):
    """حساب اتجاه التحسن"""
    if len(results) < 3:
        return 0
    
    recent = [r.score_percentage for r in results[:min(5, len(results))]]
    older = [r.score_percentage for r in results[-min(5, len(results)):]]
    
    avg_recent = sum(recent) / len(recent)
    avg_older = sum(older) / len(older)
    
    return avg_recent - avg_older


def get_activity_status(last_activity_utc):
    """تحديد حالة النشاط"""
    if not last_activity_utc:
        return 'غير نشط'

    current_time_utc = get_utc_time()

    # التأكد من أن كلا التاريخين إما aware أو naive
    if last_activity_utc.tzinfo is None and current_time_utc.tzinfo is not None:
        # last_activity_utc naive و current_time_utc aware
        last_activity_utc = pytz.utc.localize(last_activity_utc)
    elif last_activity_utc.tzinfo is not None and current_time_utc.tzinfo is None:
        # last_activity_utc aware و current_time_utc naive
        current_time_utc = pytz.utc.localize(current_time_utc)

    days_since = (current_time_utc - last_activity_utc).days

    if days_since <= 1:
        return 'نشط جداً'
    elif days_since <= 7:
        return 'نشط'
    elif days_since <= 14:
        return 'خامل'
    else:
        return 'غير نشط'

def get_performance_level(avg_score):
    """تحديد مستوى الأداء"""
    if avg_score >= 90:
        return 'ممتاز'
    elif avg_score >= 80:
        return 'جيد جداً'
    elif avg_score >= 70:
        return 'جيد'
    elif avg_score >= 60:
        return 'مقبول'
    elif avg_score > 0:
        return 'ضعيف'
    else:
        return 'لا توجد بيانات'


def get_badge(rank):
    """الحصول على شارة حسب الترتيب"""
    badges = {
        1: '🏆 المركز الأول',
        2: '🥈 المركز الثاني',
        3: '🥉 المركز الثالث'
    }
    return badges.get(rank, f'⭐ المركز {rank}')


def analyze_topics_from_student_results(student_id):
    """
    ✅ تحليل نقاط القوة والضعف حسب المواضيع من StudentResult
    """
    results = StudentResult.query.filter_by(student_id=student_id).all()
    
    if not results:
        return [], []
    
    # تجميع النتائج حسب نوع الاختبار
    course_scores = {}
    
    for result in results:
        # استخدام course_id إذا كان موجوداً، وإلا استخدام quiz_type أو quiz_name
        topic_key = None
        
        if result.course_id:
            # إذا كان course_id موجود، نجلب اسم المنهج
            course = Course.query.get(result.course_id)
            topic_key = course.name if course else f"منهج {result.course_id}"
        elif result.quiz_type == 'course':
            topic_key = result.quiz_name or 'منهج غير محدد'
        elif result.quiz_type == 'unit':
            topic_key = result.quiz_name or 'وحدة غير محددة'
        elif result.quiz_type == 'lesson':
            topic_key = result.quiz_name or 'درس غير محدد'
        else:
            topic_key = result.quiz_name or 'اختبار غير محدد'
        
        if topic_key not in course_scores:
            course_scores[topic_key] = []
        course_scores[topic_key].append(result.score_percentage)
    
    # حساب المتوسطات
    course_averages = []
    for topic_name, scores in course_scores.items():
        avg = sum(scores) / len(scores)
        course_averages.append({
            'course': topic_name,
            'avg_score': round(avg, 2),
            'total_attempts': len(scores)
        })
    
    # ترتيب
    course_averages.sort(key=lambda x: x['avg_score'], reverse=True)
    
    strengths = course_averages[:3]
    weaknesses = list(reversed(course_averages[-3:]))
    
    return strengths, weaknesses


def format_time_arabic(seconds):
    """تنسيق الوقت بالعربية"""
    if seconds < 60:
        return f'{seconds} ثانية'
    elif seconds < 3600:
        minutes = seconds // 60
        return f'{minutes} دقيقة'
    else:
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        if minutes > 0:
            return f'{hours} ساعة و {minutes} دقيقة'
        return f'{hours} ساعة'

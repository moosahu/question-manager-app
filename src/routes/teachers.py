"""
إدارة المعلمين - Routes
يسمح للأدمن بإضافة وتعديل وحذف المعلمين
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
from src.extensions import db
from src.models.teacher import Teacher
from src.models.email_verification import RegistrationSettings  # ✅ جديد
from src.middleware.auth_middleware import verify_teacher_token
from functools import wraps
from datetime import datetime

teachers_bp = Blueprint('teachers', __name__, url_prefix='/teachers')


def admin_required(f):
    """التحقق من صلاحيات الأدمن"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('ليس لديك صلاحية الوصول لهذه الصفحة', 'danger')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function


# ==================== صفحة قائمة المعلمين ====================
@teachers_bp.route('/')
@login_required
@admin_required
def list_teachers():
    """عرض قائمة المعلمين"""
    search = request.args.get('search', '')
    
    if search:
        teachers = Teacher.search_teachers(search)
    else:
        teachers = Teacher.get_all_teachers()
    
    # إحصائيات
    total = Teacher.query.count()
    active = Teacher.query.filter_by(is_active=True).count()
    inactive = total - active
    
    # ✅ جديد: جلب حالة التسجيل
    settings = RegistrationSettings.get_settings()
    
    return render_template('teachers/list.html', 
                         teachers=teachers,
                         search=search,
                         total=total,
                         active=active,
                         inactive=inactive,
                         is_teacher_registration_open=settings.is_teacher_registration_open)  # ✅ جديد


# ==================== إضافة معلم جديد ====================
@teachers_bp.route('/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_teacher():
    """إضافة معلم جديد"""
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        email = request.form.get('email', '').strip() or None
        phone = request.form.get('phone', '').strip() or None
        school = request.form.get('school', '').strip() or None
        is_active = request.form.get('is_active') == 'on'
        notes = request.form.get('notes', '').strip() or None
        
        # التحقق من البيانات
        if not name or not username or not password:
            flash('الاسم واسم المستخدم وكلمة المرور مطلوبة', 'danger')
            return render_template('teachers/add.html')
        
        # التحقق من عدم تكرار اسم المستخدم
        if Teacher.query.filter_by(username=username).first():
            flash('اسم المستخدم موجود مسبقاً', 'danger')
            return render_template('teachers/add.html')
        
        # التحقق من عدم تكرار البريد
        if email and Teacher.query.filter_by(email=email).first():
            flash('البريد الإلكتروني موجود مسبقاً', 'danger')
            return render_template('teachers/add.html')
        
        # إنشاء المعلم
        teacher = Teacher(
            name=name,
            username=username,
            email=email,
            phone=phone,
            school=school,
            is_active=is_active,
            notes=notes
        )
        teacher.set_password(password)
        
        try:
            db.session.add(teacher)
            db.session.commit()
            flash(f'تم إضافة المعلم "{name}" بنجاح', 'success')
            return redirect(url_for('teachers.list_teachers'))
        except Exception as e:
            db.session.rollback()
            flash(f'خطأ في إضافة المعلم: {str(e)}', 'danger')
    
    return render_template('teachers/add.html')


# ==================== تعديل معلم ====================
@teachers_bp.route('/edit/<int:teacher_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_teacher(teacher_id):
    """تعديل بيانات معلم"""
    teacher = Teacher.query.get_or_404(teacher_id)
    
    if request.method == 'POST':
        teacher.name = request.form.get('name', '').strip()
        teacher.email = request.form.get('email', '').strip() or None
        teacher.phone = request.form.get('phone', '').strip() or None
        teacher.school = request.form.get('school', '').strip() or None
        teacher.is_active = request.form.get('is_active') == 'on'
        teacher.notes = request.form.get('notes', '').strip() or None
        
        # تغيير كلمة المرور (اختياري)
        new_password = request.form.get('password', '')
        if new_password:
            teacher.set_password(new_password)
        
        try:
            db.session.commit()
            flash(f'تم تحديث بيانات المعلم "{teacher.name}" بنجاح', 'success')
            return redirect(url_for('teachers.list_teachers'))
        except Exception as e:
            db.session.rollback()
            flash(f'خطأ في تحديث البيانات: {str(e)}', 'danger')
    
    return render_template('teachers/edit.html', teacher=teacher)


# ==================== حذف معلم ====================
@teachers_bp.route('/delete/<int:teacher_id>', methods=['POST'])
@login_required
@admin_required
def delete_teacher(teacher_id):
    """حذف معلم مع بياناته المرتبطة"""
    teacher = Teacher.query.get_or_404(teacher_id)
    name = teacher.name
    email = teacher.email

    try:
        db.session.expunge(teacher)

        # حذف البيانات المرتبطة
        for stmt, params in [
            ("DELETE FROM login_attempts WHERE user_id = :id AND user_type = 'teacher'", {'id': teacher_id}),
            ("DELETE FROM delete_account_otps WHERE user_id = :id AND user_type = 'teacher'", {'id': teacher_id}),
        ]:
            try:
                with db.session.begin_nested():
                    db.session.execute(db.text(stmt), params)
            except Exception:
                pass

        db.session.execute(
            db.text("DELETE FROM teachers WHERE id = :id"),
            {'id': teacher_id}
        )
        db.session.commit()

        # إشعار في صندوق الوارد
        try:
            from src.models.notification import Notification
            from src.models.user import User
            admin_user = User.query.filter_by(is_admin=True).first()
            if admin_user:
                notif = Notification(
                    title='🗑️ تم حذف معلم بواسطة الأدمن',
                    message=f'الاسم: {name}\nالإيميل: {email}',
                    type='admin_event',
                    user_id=admin_user.id,
                    is_read=False,
                )
                db.session.add(notif)
                db.session.commit()
        except Exception:
            pass

        flash(f'تم حذف المعلم "{name}" بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ في حذف المعلم: {str(e)}', 'danger')

    return redirect(url_for('teachers.list_teachers'))


# ==================== تبديل حالة المعلم ====================
@teachers_bp.route('/toggle/<int:teacher_id>', methods=['POST'])
@login_required
@admin_required
def toggle_teacher(teacher_id):
    """تفعيل/تعطيل حساب معلم"""
    teacher = Teacher.query.get_or_404(teacher_id)
    teacher.is_active = not teacher.is_active
    
    try:
        db.session.commit()
        status = "مفعل" if teacher.is_active else "معطل"
        flash(f'تم تغيير حالة المعلم "{teacher.name}" إلى {status}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ في تغيير الحالة: {str(e)}', 'danger')
    
    return redirect(url_for('teachers.list_teachers'))


# ==================== إعادة تعيين جهاز المعلم ====================
@teachers_bp.route('/reset-device/<int:teacher_id>', methods=['POST'])
@login_required
@admin_required
def reset_teacher_device(teacher_id):
    """إزالة ربط جهاز المعلم"""
    teacher = Teacher.query.get_or_404(teacher_id)
    
    try:
        teacher.clear_device_info()
        flash(f'تم إزالة ربط جهاز المعلم "{teacher.name}" بنجاح', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'خطأ في إزالة ربط الجهاز: {str(e)}', 'danger')
    
    return redirect(url_for('teachers.edit_teacher', teacher_id=teacher_id))


# ==================== ✅ جديد: تبديل حالة التسجيل الذاتي ====================
@teachers_bp.route('/toggle-registration', methods=['POST'])
@login_required
@admin_required
def toggle_registration():
    """تبديل حالة التسجيل الذاتي للمعلمين"""
    try:
        settings = RegistrationSettings.get_settings()
        new_status = not settings.is_teacher_registration_open
        
        RegistrationSettings.update_settings(
            is_teacher_open=new_status,
            admin_id=current_user.id
        )
        
        if new_status:
            flash('تم فتح التسجيل الذاتي للمعلمين ✓', 'success')
        else:
            flash('تم إغلاق التسجيل الذاتي للمعلمين', 'warning')
            
    except Exception as e:
        flash(f'خطأ في تغيير حالة التسجيل: {str(e)}', 'danger')
    
    return redirect(url_for('teachers.list_teachers'))


# ==================== Mobile API Endpoints ====================

@teachers_bp.route('/api/mobile/list', methods=['GET'])
@login_required
@admin_required
def api_mobile_list_teachers():
    """قائمة المعلمين للموبايل"""
    from src.models.teacher import Teacher
    search = request.args.get('search', '')
    query = Teacher.query
    if search:
        query = query.filter(
            db.or_(
                Teacher.name.ilike(f'%{search}%'),
                Teacher.username.ilike(f'%{search}%'),
            )
        )
    teachers = query.order_by(Teacher.name).all()
    return jsonify({
        'success': True,
        'teachers': [
            {
                'id': t.id,
                'name': t.name,
                'username': t.username,
                'email': t.email or '',
                'is_active': t.is_active if hasattr(t, 'is_active') else True,
                'created_at': t.created_at.isoformat() if t.created_at else '',
                'last_login': t.last_login.isoformat() if t.last_login else '',
            }
            for t in teachers
        ],
    })


@teachers_bp.route('/api/mobile/add', methods=['POST'])
@login_required
@admin_required
def api_mobile_add_teacher():
    """إضافة معلم جديد"""
    from src.models.teacher import Teacher
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    username = data.get('username', '').strip()
    password = data.get('password', '')
    email = data.get('email', '').strip() or None

    if not name or not username or not password:
        return jsonify({'success': False, 'error': 'الاسم واسم المستخدم وكلمة المرور مطلوبة'}), 400

    if Teacher.query.filter_by(username=username).first():
        return jsonify({'success': False, 'error': 'اسم المستخدم موجود مسبقاً'}), 409

    teacher = Teacher(name=name, username=username, email=email)
    teacher.set_password(password)

    try:
        db.session.add(teacher)
        db.session.commit()
        try:
            from src.models.audit_log import AuditLog
            AuditLog.log(
                action='add_teacher',
                description=f'إضافة معلم جديد: {name} (@{username})',
                admin_name=getattr(current_user, 'username', 'الادمن'),
                target_type='teacher',
                target_id=teacher.id,
            )
            db.session.commit()
        except Exception:
            pass
        return jsonify({'success': True, 'message': f'تم إضافة المعلم "{name}"', 'teacher_id': teacher.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@teachers_bp.route('/api/mobile/<int:teacher_id>/edit', methods=['POST'])
@login_required
@admin_required
def api_mobile_edit_teacher(teacher_id):
    """تعديل بيانات معلم"""
    from src.models.teacher import Teacher
    teacher = Teacher.query.get_or_404(teacher_id)
    data = request.get_json() or {}
    if 'name' in data:
        teacher.name = data['name'].strip()
    if 'email' in data:
        teacher.email = data['email'].strip() or None
    new_password = data.get('password', '')
    if new_password:
        teacher.set_password(new_password)
    try:
        db.session.commit()
        try:
            from src.models.audit_log import AuditLog
            AuditLog.log(
                action='edit_teacher',
                description=f'تعديل بيانات المعلم: {teacher.name}',
                admin_name=getattr(current_user, 'username', 'الادمن'),
                target_type='teacher',
                target_id=teacher_id,
            )
            db.session.commit()
        except Exception:
            pass
        return jsonify({'success': True, 'message': 'تم التحديث بنجاح'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@teachers_bp.route('/api/mobile/<int:teacher_id>/delete', methods=['POST'])
@login_required
@admin_required
def api_mobile_delete_teacher(teacher_id):
    """حذف معلم مع بياناته المرتبطة"""
    from src.models.teacher import Teacher
    teacher = Teacher.query.get_or_404(teacher_id)
    name = teacher.name
    email = teacher.email
    try:
        db.session.expunge(teacher)

        for stmt, params in [
            ("DELETE FROM login_attempts WHERE user_id = :id AND user_type = 'teacher'", {'id': teacher_id}),
            ("DELETE FROM delete_account_otps WHERE user_id = :id AND user_type = 'teacher'", {'id': teacher_id}),
        ]:
            try:
                with db.session.begin_nested():
                    db.session.execute(db.text(stmt), params)
            except Exception:
                pass

        db.session.execute(
            db.text("DELETE FROM teachers WHERE id = :id"),
            {'id': teacher_id}
        )
        db.session.commit()

        # إشعار في صندوق الوارد
        try:
            from src.models.notification import Notification
            from src.models.user import User
            admin_user = User.query.filter_by(is_admin=True).first()
            if admin_user:
                notif = Notification(
                    title='🗑️ تم حذف معلم بواسطة الأدمن',
                    message=f'الاسم: {name}\nالإيميل: {email}',
                    type='admin_event',
                    user_id=admin_user.id,
                    is_read=False,
                )
                db.session.add(notif)
                db.session.commit()
        except Exception:
            pass

        try:
            from src.models.audit_log import AuditLog
            AuditLog.log(
                action='delete_teacher',
                description=f'حذف المعلم: {name}',
                admin_name=getattr(current_user, 'username', 'الادمن'),
                target_type='teacher',
                target_id=teacher_id,
            )
            db.session.commit()
        except Exception:
            pass

        return jsonify({'success': True, 'message': f'تم حذف المعلم "{name}"'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@teachers_bp.route('/api/mobile/<int:teacher_id>/toggle', methods=['POST'])
@login_required
@admin_required
def api_mobile_toggle_teacher(teacher_id):
    """تفعيل/تعطيل حساب معلم عبر الموبايل"""
    from src.models.teacher import Teacher
    teacher = Teacher.query.get_or_404(teacher_id)
    teacher.is_active = not teacher.is_active
    try:
        db.session.commit()
        status = 'مفعّل' if teacher.is_active else 'معطّل'
        try:
            from src.models.audit_log import AuditLog
            AuditLog.log(
                action='toggle_teacher',
                description=f'تغيير حالة المعلم "{teacher.name}" إلى {status}',
                admin_name=getattr(current_user, 'username', 'الادمن'),
                target_type='teacher',
                target_id=teacher_id,
            )
            db.session.commit()
        except Exception:
            pass
        return jsonify({'success': True, 'message': f'تم تغيير حالة المعلم "{teacher.name}" إلى {status}',
                        'is_active': teacher.is_active})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== Teacher-Student Link Endpoints ====================

@teachers_bp.route('/api/mobile/<int:teacher_id>/class-code', methods=['GET'])
@login_required
@admin_required
def api_get_teacher_class_code(teacher_id):
    """جلب كود الربط الخاص بالمعلم (للأدمن)"""
    from src.models.teacher import Teacher
    teacher = Teacher.query.get_or_404(teacher_id)
    teacher.ensure_class_code()
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
    return jsonify({
        'success': True,
        'class_code': teacher.class_code,
        'teacher_name': teacher.name,
    })


@teachers_bp.route('/api/mobile/<int:teacher_id>/students', methods=['GET'])
@login_required
@admin_required
def api_get_teacher_students(teacher_id):
    """قائمة طلاب المعلم مع آخر نشاط ومتوسط الدرجات (للأدمن)"""
    from src.models.teacher import Teacher
    from src.models.teacher_student import TeacherStudent
    from src.models.student import Student
    from src.models.student_result import StudentResult
    from sqlalchemy import func

    teacher = Teacher.query.get_or_404(teacher_id)
    links = TeacherStudent.query.filter_by(teacher_id=teacher_id).all()

    result = []
    for link in links:
        s = link.student
        if not s:
            continue
        # متوسط الدرجات
        avg = db.session.query(func.avg(StudentResult.score_percentage))\
            .filter_by(student_id=s.id).scalar()
        # عدد الاختبارات
        count = StudentResult.query.filter_by(student_id=s.id).count()
        result.append({
            'student_id': s.id,
            'name': s.name,
            'username': s.username,
            'grade': s.grade or '',
            'school': s.school or '',
            'is_active': s.is_active,
            'last_login': s.last_login.isoformat() if s.last_login else '',
            'joined_at': link.joined_at.isoformat() if link.joined_at else '',
            'avg_score': round(avg, 1) if avg is not None else None,
            'quiz_count': count,
        })

    return jsonify({
        'success': True,
        'teacher_name': teacher.name,
        'class_code': teacher.class_code or '',
        'students': result,
    })


def _build_student_results_response(student_id):
    """دالة مساعدة: تبني response النتائج المحسّن لأي طالب"""
    from src.models.student_result import StudentResult
    from src.models.student import Student

    student = Student.query.get_or_404(student_id)
    results = StudentResult.query.filter_by(student_id=student_id)\
        .order_by(StudentResult.created_at.desc()).limit(100).all()

    # إحصائيات عامة
    scores = [r.score_percentage for r in results if r.score_percentage is not None]
    if scores:
        avg_score = round(sum(scores) / len(scores), 1)
        max_score = max(scores)
        min_score = min(scores)
        last_score = scores[0] if scores else None
    else:
        avg_score = max_score = min_score = last_score = None

    # حساب trend (آخر 5 مقارنةً بالـ 5 السابقة)
    trend = 'stable'
    if len(scores) >= 6:
        recent_avg = sum(scores[:5]) / 5
        older_avg = sum(scores[5:10]) / len(scores[5:10])
        if recent_avg > older_avg + 5:
            trend = 'improving'
        elif recent_avg < older_avg - 5:
            trend = 'declining'

    # تجميع حسب quiz_name
    topics_map = {}
    for r in results:
        name = r.quiz_name or 'غير محدد'
        if name not in topics_map:
            topics_map[name] = {'scores': [], 'last_score': None}
        if r.score_percentage is not None:
            topics_map[name]['scores'].append(r.score_percentage)
        if topics_map[name]['last_score'] is None and r.score_percentage is not None:
            topics_map[name]['last_score'] = r.score_percentage

    topics = [
        {
            'name': name,
            'count': len(data['scores']),
            'avg': round(sum(data['scores']) / len(data['scores']), 1) if data['scores'] else None,
            'last_score': data['last_score'],
        }
        for name, data in topics_map.items()
    ]

    # آخر 10 اختبارات للرسم البياني (من الأقدم للأحدث)
    chart_results = [r for r in reversed(results[:10]) if r.score_percentage is not None]
    chart_data = [
        {
            'score': r.score_percentage,
            'date': f"{r.created_at.day}/{r.created_at.month}" if r.created_at else '',
        }
        for r in chart_results
    ]

    # topics_breakdown مقسّم حسب quiz_type للمقارنة
    units_map = {}
    lessons_map = {}
    for r in results:
        name = r.quiz_name or 'غير محدد'
        if r.score_percentage is None:
            continue
        if r.quiz_type == 'lesson':
            lessons_map.setdefault(name, []).append(r.score_percentage)
        else:
            units_map.setdefault(name, []).append(r.score_percentage)

    def _to_breakdown(d):
        lst = [{'name': n, 'avg_score': round(sum(s) / len(s), 2), 'total_attempts': len(s)} for n, s in d.items()]
        return sorted(lst, key=lambda x: x['name'])

    return {
        'success': True,
        'student': {
            'id': student.id,
            'name': student.name,
            'grade': student.grade or '',
        },
        'stats': {
            'avg_score': avg_score,
            'max_score': max_score,
            'min_score': min_score,
            'last_score': last_score,
            'total_quizzes': len(results),
            'trend': trend,
            'chart_data': chart_data,
        },
        'topics': topics,
        'topics_breakdown': {'units': _to_breakdown(units_map), 'lessons': _to_breakdown(lessons_map)},
        'results': [
            {
                'id': r.id,
                'quiz_name': r.quiz_name,
                'score_percentage': r.score_percentage,
                'correct_answers': r.correct_answers,
                'total_questions': r.total_questions,
                'wrong_answers': r.wrong_answers,
                'time_spent': r.time_spent,
                'quiz_type': r.quiz_type,
                'created_at': r.created_at.isoformat() if r.created_at else '',
            }
            for r in results
        ],
    }


@teachers_bp.route('/api/mobile/<int:teacher_id>/students/<int:student_id>/results', methods=['GET'])
@login_required
@admin_required
def api_get_teacher_student_results(teacher_id, student_id):
    """نتائج طالب معين تابع لمعلم (للأدمن)"""
    from src.models.teacher_student import TeacherStudent

    # تحقق أن الطالب فعلاً مرتبط بهذا المعلم
    TeacherStudent.query.filter_by(
        teacher_id=teacher_id, student_id=student_id).first_or_404()

    return jsonify(_build_student_results_response(student_id))


@teachers_bp.route('/api/mobile/<int:teacher_id>/students/<int:student_id>/remove', methods=['POST'])
@login_required
@admin_required
def api_remove_teacher_student(teacher_id, student_id):
    """إزالة طالب من قائمة المعلم"""
    from src.models.teacher_student import TeacherStudent
    link = TeacherStudent.query.filter_by(
        teacher_id=teacher_id, student_id=student_id).first_or_404()
    try:
        db.session.delete(link)
        db.session.commit()
        return jsonify({'success': True, 'message': 'تم إزالة الطالب من قائمتك'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@teachers_bp.route('/api/mobile/assign-by-code', methods=['POST'])
@login_required
@admin_required
def api_admin_assign_by_code():
    """
    الأدمن يربط طالب بمعلم عن طريق الكود (بدون تحديد teacher_id)
    Body: { student_id, class_code }
    """
    from src.models.teacher import Teacher
    from src.models.student import Student
    from src.models.teacher_student import TeacherStudent

    data = request.get_json() or {}
    student_id = data.get('student_id')
    class_code = (data.get('class_code') or '').strip().upper()

    if not student_id or not class_code:
        return jsonify({'success': False, 'error': 'student_id والكود مطلوبان'}), 400

    student = Student.query.get(student_id)
    if not student:
        return jsonify({'success': False, 'error': 'الطالب غير موجود'}), 404

    # البحث بالكود في المعلمين أو الأدمن
    from src.models.user import User
    teacher = Teacher.query.filter_by(class_code=class_code).first()
    admin_owner = None
    if not teacher:
        admin_owner = User.query.filter_by(class_code=class_code, is_admin=True).first()
    if not teacher and not admin_owner:
        return jsonify({'success': False, 'error': 'الكود غير صحيح'}), 404

    owner_name = teacher.name if teacher else admin_owner.username

    existing = TeacherStudent.query.filter_by(student_id=student_id).first()
    if existing:
        same = (teacher and existing.teacher_id == teacher.id) or \
               (admin_owner and existing.admin_id == admin_owner.id)
        if same:
            return jsonify({'success': False, 'error': 'الطالب مرتبط بهذا الكود بالفعل'}), 409
        existing.teacher_id = teacher.id if teacher else None
        existing.admin_id = admin_owner.id if admin_owner else None
        existing.joined_at = datetime.utcnow()
        try:
            db.session.commit()
            return jsonify({'success': True,
                            'message': f'تم نقل الطالب "{student.name}" إلى "{owner_name}"'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

    link = TeacherStudent(
        teacher_id=teacher.id if teacher else None,
        admin_id=admin_owner.id if admin_owner else None,
        student_id=student_id,
    )
    try:
        db.session.add(link)
        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'تم ربط الطالب "{student.name}" بـ "{owner_name}"',
            'teacher_name': owner_name,
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@teachers_bp.route('/api/mobile/<int:teacher_id>/students/assign', methods=['POST'])
@login_required
@admin_required
def api_admin_assign_student(teacher_id):
    """
    الأدمن يربط طالب بمعلم مباشرة (بدون كود)
    Body: { student_id }
    """
    from src.models.teacher import Teacher
    from src.models.student import Student
    from src.models.teacher_student import TeacherStudent

    teacher = Teacher.query.get_or_404(teacher_id)
    data = request.get_json() or {}
    student_id = data.get('student_id')

    if not student_id:
        return jsonify({'success': False, 'error': 'student_id مطلوب'}), 400

    student = Student.query.get(student_id)
    if not student:
        return jsonify({'success': False, 'error': 'الطالب غير موجود'}), 404

    # هل مرتبط بمعلم آخر؟
    existing = TeacherStudent.query.filter_by(student_id=student_id).first()
    if existing:
        if existing.teacher_id == teacher_id:
            return jsonify({'success': False, 'error': 'الطالب مرتبط بهذا المعلم بالفعل'}), 409
        # الأدمن يقدر يغيّر المعلم مباشرة
        existing.teacher_id = teacher_id
        existing.joined_at = datetime.utcnow()
        try:
            db.session.commit()
            return jsonify({
                'success': True,
                'message': f'تم نقل الطالب "{student.name}" إلى المعلم "{teacher.name}"',
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

    link = TeacherStudent(teacher_id=teacher_id, student_id=student_id)
    try:
        db.session.add(link)
        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'تم ربط الطالب "{student.name}" بالمعلم "{teacher.name}"',
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== Admin Class Endpoints (الأدمن يتابع طلابه) ====================

@teachers_bp.route('/api/mobile/admin/class-code', methods=['GET'])
@login_required
@admin_required
def api_get_admin_class_code():
    """جلب كود الأدمن (يُولَّد تلقائياً عند أول طلب)"""
    from src.models.user import User
    admin = User.query.filter_by(is_admin=True).first_or_404()
    admin.ensure_class_code()
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
    return jsonify({
        'success': True,
        'class_code': admin.class_code,
        'admin_name': admin.username,
    })


@teachers_bp.route('/api/mobile/admin/students', methods=['GET'])
@login_required
@admin_required
def api_get_admin_students():
    """قائمة الطلاب المرتبطين بالأدمن مع إحصائياتهم"""
    from src.models.user import User
    from src.models.teacher_student import TeacherStudent
    from src.models.student_result import StudentResult
    from sqlalchemy import func

    admin = User.query.filter_by(is_admin=True).first_or_404()
    links = TeacherStudent.query.filter_by(admin_id=admin.id).all()

    result = []
    for link in links:
        s = link.student
        if not s:
            continue
        avg = db.session.query(func.avg(StudentResult.score_percentage))\
            .filter_by(student_id=s.id).scalar()
        count = StudentResult.query.filter_by(student_id=s.id).count()
        result.append({
            'student_id': s.id,
            'name': s.name,
            'username': s.username,
            'grade': s.grade or '',
            'school': s.school or '',
            'is_active': s.is_active,
            'last_login': s.last_login.isoformat() if s.last_login else '',
            'joined_at': link.joined_at.isoformat() if link.joined_at else '',
            'avg_score': round(avg, 1) if avg is not None else None,
            'quiz_count': count,
        })

    return jsonify({
        'success': True,
        'class_code': admin.class_code or '',
        'students': result,
    })


@teachers_bp.route('/api/mobile/admin/students/<int:student_id>/results', methods=['GET'])
@login_required
@admin_required
def api_get_admin_student_results(student_id):
    """نتائج طالب مرتبط بالأدمن"""
    from src.models.teacher_student import TeacherStudent
    from src.models.user import User

    admin = User.query.filter_by(is_admin=True).first_or_404()
    TeacherStudent.query.filter_by(
        admin_id=admin.id, student_id=student_id).first_or_404()

    return jsonify(_build_student_results_response(student_id))


@teachers_bp.route('/api/mobile/admin/students/<int:student_id>/remove', methods=['POST'])
@login_required
@admin_required
def api_remove_admin_student(student_id):
    """إزالة طالب من قائمة الأدمن"""
    from src.models.teacher_student import TeacherStudent
    from src.models.user import User

    admin = User.query.filter_by(is_admin=True).first_or_404()
    link = TeacherStudent.query.filter_by(
        admin_id=admin.id, student_id=student_id).first_or_404()
    try:
        db.session.delete(link)
        db.session.commit()
        return jsonify({'success': True, 'message': 'تم إزالة الطالب من قائمتك'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@teachers_bp.route('/api/mobile/admin/students/search', methods=['GET'])
@login_required
def api_search_unlinked_students():
    """
    بحث عن طلاب غير مرتبطين بأي قائمة
    GET /teachers/api/mobile/admin/students/search?q=اسم
    """
    from src.models.teacher_student import TeacherStudent

    q = (request.args.get('q') or '').strip()
    if len(q) < 2:
        return jsonify({'success': True, 'students': []})

    # IDs المرتبطين بهذا الأدمن تحديداً (لا نعرضهم — موجودون بالفعل)
    my_linked_ids = [
        ts.student_id for ts in
        TeacherStudent.query
        .with_entities(TeacherStudent.student_id)
        .filter_by(admin_id=current_user.id)
        .all()
    ]

    filters = [
        Student.is_active == True,
        db.or_(
            Student.name.ilike(f'%{q}%'),
            Student.username.ilike(f'%{q}%'),
        )
    ]
    if my_linked_ids:
        filters.append(~Student.id.in_(my_linked_ids))

    students = Student.query.filter(*filters).limit(20).all()

    # لكل طالب: هل مرتبط بمعلم آخر؟ نوضّح للأدمن
    linked_elsewhere = {
        ts.student_id for ts in
        TeacherStudent.query.with_entities(TeacherStudent.student_id).all()
    }

    return jsonify({
        'success': True,
        'students': [
            {
                'id':       s.id,
                'name':     s.name,
                'username': s.username,
                'grade':    s.grade or '',
                'school':   s.school or '',
                'linked_to_other': s.id in linked_elsewhere,
            }
            for s in students
        ],
    })


@teachers_bp.route('/api/mobile/admin/students/link', methods=['POST'])
@login_required
def api_link_student_to_admin():
    """
    الأدمن يربط طالب موجود بقائمته يدوياً
    Body: { student_id }
    """
    from src.models.teacher_student import TeacherStudent
    from src.models.user import User

    data       = request.get_json(silent=True) or {}
    student_id = data.get('student_id')

    if not student_id:
        return jsonify({'success': False, 'message': 'student_id مطلوب'}), 400

    student = Student.query.get(student_id)
    if not student or not student.is_active:
        return jsonify({'success': False, 'message': 'الطالب غير موجود أو معطّل'}), 404

    # هل مرتبط بهذا الأدمن بالفعل؟
    already_mine = TeacherStudent.query.filter_by(
        student_id=student_id, admin_id=current_user.id
    ).first()
    if already_mine:
        return jsonify({'success': False, 'message': 'الطالب في قائمتك بالفعل'}), 409

    # هل مرتبط بمعلم/أدمن آخر؟ نفكّ الربط القديم ونربطه بهذا الأدمن
    existing = TeacherStudent.query.filter_by(student_id=student_id).first()
    if existing:
        db.session.delete(existing)
        db.session.flush()

    link = TeacherStudent(admin_id=current_user.id, student_id=student_id)
    db.session.add(link)
    db.session.commit()

    return jsonify({'success': True, 'message': f'تم ربط "{student.name}" بقائمتك'})


# ==================== Teacher Self-Service Endpoints (توكن المعلم) ====================

@teachers_bp.route('/api/my/class-code', methods=['GET'])
@verify_teacher_token
def api_teacher_my_class_code():
    """المعلم يجلب كوده الخاص"""
    teacher = Teacher.query.get_or_404(request.teacher_id)
    teacher.ensure_class_code()
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
    return jsonify({
        'success': True,
        'class_code': teacher.class_code,
        'teacher_name': teacher.name,
    })


@teachers_bp.route('/api/my/students', methods=['GET'])
@verify_teacher_token
def api_teacher_my_students():
    """المعلم يجلب قائمة طلابه"""
    from src.models.teacher_student import TeacherStudent
    from src.models.student_result import StudentResult
    from sqlalchemy import func

    teacher = Teacher.query.get_or_404(request.teacher_id)
    links = TeacherStudent.query.filter_by(teacher_id=teacher.id).all()

    result = []
    for link in links:
        s = link.student
        if not s:
            continue
        avg = db.session.query(func.avg(StudentResult.score_percentage))\
            .filter_by(student_id=s.id).scalar()
        count = StudentResult.query.filter_by(student_id=s.id).count()
        result.append({
            'student_id': s.id,
            'name': s.name,
            'grade': s.grade or '',
            'school': s.school or '',
            'is_active': s.is_active,
            'last_login': s.last_login.isoformat() if s.last_login else '',
            'joined_at': link.joined_at.isoformat() if link.joined_at else '',
            'avg_score': round(avg, 1) if avg is not None else None,
            'quiz_count': count,
        })

    return jsonify({
        'success': True,
        'teacher_name': teacher.name,
        'class_code': teacher.class_code or '',
        'students': result,
    })


@teachers_bp.route('/api/my/students/<int:student_id>/results', methods=['GET'])
@verify_teacher_token
def api_teacher_my_student_results(student_id):
    """المعلم يجلب نتائج طالب من قائمته"""
    from src.models.teacher_student import TeacherStudent

    TeacherStudent.query.filter_by(
        teacher_id=request.teacher_id, student_id=student_id).first_or_404()

    return jsonify(_build_student_results_response(student_id))


@teachers_bp.route('/api/my/students/<int:student_id>/remove', methods=['POST'])
@verify_teacher_token
def api_teacher_remove_my_student(student_id):
    """المعلم يزيل طالباً من قائمته"""
    from src.models.teacher_student import TeacherStudent

    link = TeacherStudent.query.filter_by(
        teacher_id=request.teacher_id, student_id=student_id).first_or_404()
    try:
        db.session.delete(link)
        db.session.commit()
        return jsonify({'success': True, 'message': 'تم إزالة الطالب من قائمتك'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@teachers_bp.route('/api/my/export-excel', methods=['POST'])
@verify_teacher_token
def api_teacher_export_excel():
    """المعلم يصدّر قائمة طلابه كـ Excel — بدون بريد أو جوال"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        body = request.json or {}
        data = body.get('students', [])
        teacher_name = body.get('teacher_name', 'المعلم')

        if not data:
            return jsonify({'success': False, 'error': 'لا توجد بيانات للتصدير'}), 400

        wb = Workbook()

        # ==================== ورقة 1: تقرير طلابي ====================
        ws = wb.active
        ws.title = 'تقرير طلابي'
        ws.sheet_view.rightToLeft = True

        header_font  = Font(bold=True, size=11, color='FFFFFF')
        header_fill  = PatternFill(start_color='0D9488', end_color='0D9488', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        center_align = Alignment(horizontal='center', vertical='center')
        right_align  = Alignment(horizontal='right',  vertical='center')
        border_side  = Side(style='thin', color='CCCCCC')
        border       = Border(left=border_side, right=border_side,
                              top=border_side,  bottom=border_side)
        even_fill    = PatternFill(start_color='F0FDFA', end_color='F0FDFA', fill_type='solid')

        now_str = datetime.now().strftime('%Y-%m-%d  %H:%M')

        # صف العنوان
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value     = f'تقرير طلابي  |  {teacher_name}  |  {now_str}'
        title_cell.font      = Font(bold=True, size=13, color='134E4A')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill      = PatternFill(start_color='CCFBF1', end_color='CCFBF1', fill_type='solid')
        ws.row_dimensions[1].height = 28

        # رأس الجدول (صف 2) — 7 أعمدة بدون بريد وجوال
        headers = ['#', 'الطالب', 'الصف', 'المدرسة', 'المعدل', 'الاختبارات', 'آخر دخول', 'الحالة']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(2, col, h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = border
        ws.row_dimensions[2].height = 26

        # البيانات
        for row_idx, student in enumerate(data, 3):
            row_fill = even_fill if row_idx % 2 == 0 else None

            def _c(col, value, align=center_align, bold=False, bg=None):
                c = ws.cell(row_idx, col, value)
                c.alignment = align
                c.border    = border
                c.font      = Font(bold=bold, size=10)
                if bg:
                    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                elif row_fill:
                    c.fill = row_fill

            _c(1, row_idx - 2)
            _c(2, student.get('name', ''), right_align, bold=True)
            _c(3, student.get('grade', ''))
            _c(4, student.get('school', ''), right_align)

            avg = student.get('avg_score')
            avg_bg = None
            if avg is not None:
                if   avg >= 80: avg_bg = 'C6EFCE'
                elif avg >= 60: avg_bg = 'FFEB9C'
                else:           avg_bg = 'FFC7CE'
            _c(5, f"{avg:.1f}%" if avg is not None else 'لا نتائج', bold=True, bg=avg_bg)

            _c(6, student.get('quiz_count', 0))

            last_login = student.get('last_login', '')
            if last_login:
                try:
                    dt = datetime.fromisoformat(last_login.replace('Z', '+00:00'))
                    _c(7, dt.strftime('%Y-%m-%d'))
                except:
                    _c(7, 'غير متاح')
            else:
                _c(7, 'غير متاح')

            is_active = student.get('is_active', False)
            status_text = 'نشط' if is_active else 'غير نشط'
            status_bg   = 'C6EFCE' if is_active else 'FFC7CE'
            _c(8, status_text, bg=status_bg)

        # عرض الأعمدة
        for idx, w in enumerate([5, 26, 14, 22, 12, 13, 14, 12], 1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        ws.freeze_panes = 'A3'

        # صف المصدر
        footer_row = len(data) + 3
        ws.merge_cells(f'A{footer_row}:H{footer_row}')
        fc = ws[f'A{footer_row}']
        fc.value     = f'⚗️  تم استخراج هذا التقرير من تطبيق كيم تحصيلي  |  منصة تعليمية للكيمياء  |  جميع الحقوق محفوظة © {datetime.now().year}'
        fc.font      = Font(size=9, color='888888', italic=True)
        fc.alignment = Alignment(horizontal='center', vertical='center')
        fc.fill      = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
        ws.row_dimensions[footer_row].height = 18

        # ==================== ورقة 2: الملخص ====================
        ws2 = wb.create_sheet('الملخص')
        ws2.sheet_view.rightToLeft = True

        total       = len(data)
        active      = sum(1 for s in data if s.get('is_active'))
        with_results= sum(1 for s in data if (s.get('quiz_count') or 0) > 0)
        scores      = [s['avg_score'] for s in data if s.get('avg_score') is not None]
        avg_all     = sum(scores) / len(scores) if scores else 0

        dist = {
            'ممتاز (80%+)':           sum(1 for s in scores if s >= 80),
            'جيد (60-79%)':           sum(1 for s in scores if 60 <= s < 80),
            'يحتاج انتباه (40-59%)':  sum(1 for s in scores if 40 <= s < 60),
            'حرج (أقل من 40%)':       sum(1 for s in scores if 0 < s < 40),
            'لا توجد نتائج':          sum(1 for s in data if not s.get('avg_score')),
        }

        s2_hfont = Font(bold=True, size=11, color='FFFFFF')
        s2_hfill = PatternFill(start_color='0D9488', end_color='0D9488', fill_type='solid')
        s2_lf    = Font(bold=True, size=11)
        s2_vf    = Font(size=11)

        def _h2(row, text):
            ws2.merge_cells(f'A{row}:B{row}')
            c = ws2[f'A{row}']
            c.value = text; c.font = s2_hfont; c.fill = s2_hfill
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws2.row_dimensions[row].height = 22

        def _r2(row, label, value, bg=None):
            lc = ws2.cell(row, 1, label)
            lc.font = s2_lf; lc.alignment = right_align; lc.border = border
            vc = ws2.cell(row, 2, value)
            vc.font = s2_vf; vc.alignment = center_align; vc.border = border
            if bg:
                vc.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')

        _h2(1, 'معلومات التقرير')
        _r2(2, 'المعلم',          teacher_name)
        _r2(3, 'تاريخ الإنشاء',   now_str)
        _r2(4, 'عدد الطلاب',      total)

        _h2(6, 'إحصائيات عامة')
        _r2(7,  'إجمالي الطلاب',   total)
        _r2(8,  'النشطون',          active,       bg='C6EFCE')
        _r2(9,  'غير النشطين',      total-active, bg='FFC7CE')
        _r2(10, 'لهم نتائج',        with_results)
        _r2(11, 'المعدل العام',      f"{avg_all:.1f}%" if scores else 'لا نتائج')

        _h2(13, 'توزيع مستويات الأداء')
        dist_colors = ['C6EFCE', 'FFEB9C', 'FFD18C', 'FFC7CE', 'EEEEEE']
        for i, (label, count) in enumerate(dist.items()):
            pct = f"{(count/total*100):.1f}%" if total else '0%'
            _r2(14+i, label, f"{count} طالب  ({pct})", bg=dist_colors[i])

        ws2.column_dimensions['A'].width = 28
        ws2.column_dimensions['B'].width = 22

        # صف المصدر في الملخص
        ws2.merge_cells('A20:B20')
        sf = ws2['A20']
        sf.value     = f'⚗️  تم استخراج هذا التقرير من تطبيق كيم تحصيلي  |  منصة تعليمية للكيمياء  |  جميع الحقوق محفوظة © {datetime.now().year}'
        sf.font      = Font(size=9, color='888888', italic=True)
        sf.alignment = Alignment(horizontal='center', vertical='center')
        sf.fill      = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
        ws2.row_dimensions[20].height = 18

        # ==================== حفظ وإرسال ====================
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = f'تقرير_طلابي_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@teachers_bp.route('/api/mobile/admin/export-excel', methods=['POST'])
@login_required
@admin_required
def api_admin_export_excel():
    """الأدمن يصدّر قائمة الطلاب كـ Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        body = request.json or {}
        data = body.get('students', [])
        admin_name = body.get('admin_name', 'الإدارة')

        if not data:
            return jsonify({'success': False, 'error': 'لا توجد بيانات للتصدير'}), 400

        wb = Workbook()

        # ==================== ورقة 1: تقرير الطلاب ====================
        ws = wb.active
        ws.title = 'تقرير الطلاب'
        ws.sheet_view.rightToLeft = True

        header_font  = Font(bold=True, size=11, color='FFFFFF')
        header_fill  = PatternFill(start_color='1e3a8a', end_color='2563EB', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        center_align = Alignment(horizontal='center', vertical='center')
        right_align  = Alignment(horizontal='right',  vertical='center')
        border_side  = Side(style='thin', color='CCCCCC')
        border       = Border(left=border_side, right=border_side,
                              top=border_side,  bottom=border_side)
        even_fill    = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')

        now_str = datetime.now().strftime('%Y-%m-%d  %H:%M')

        # صف العنوان
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value     = f'تقرير الطلاب  |  {admin_name}  |  {now_str}'
        title_cell.font      = Font(bold=True, size=13, color='1e3a8a')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill      = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
        ws.row_dimensions[1].height = 28

        # رأس الجدول (صف 2)
        headers = ['#', 'الطالب', 'الصف', 'المدرسة', 'المعدل', 'الاختبارات', 'آخر دخول', 'الحالة']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(2, col, h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = border
        ws.row_dimensions[2].height = 26

        # البيانات
        for row_idx, student in enumerate(data, 3):
            row_fill = even_fill if row_idx % 2 == 0 else None

            def _c(col, value, align=center_align, bold=False, bg=None):
                c = ws.cell(row_idx, col, value)
                c.alignment = align
                c.border    = border
                c.font      = Font(bold=bold, size=10)
                if bg:
                    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                elif row_fill:
                    c.fill = row_fill

            _c(1, row_idx - 2)
            _c(2, student.get('name', ''), right_align, bold=True)
            _c(3, student.get('grade', ''))
            _c(4, student.get('school', ''), right_align)

            avg = student.get('avg_score')
            avg_bg = None
            if avg is not None:
                if   avg >= 80: avg_bg = 'C6EFCE'
                elif avg >= 60: avg_bg = 'FFEB9C'
                else:           avg_bg = 'FFC7CE'
            _c(5, f"{avg:.1f}%" if avg is not None else 'لا نتائج', bold=True, bg=avg_bg)

            _c(6, student.get('quiz_count', 0))

            last_login = student.get('last_login', '')
            if last_login:
                try:
                    dt = datetime.fromisoformat(last_login.replace('Z', '+00:00'))
                    _c(7, dt.strftime('%Y-%m-%d'))
                except:
                    _c(7, 'غير متاح')
            else:
                _c(7, 'غير متاح')

            is_active = student.get('is_active', False)
            status_text = 'نشط' if is_active else 'غير نشط'
            status_bg   = 'C6EFCE' if is_active else 'FFC7CE'
            _c(8, status_text, bg=status_bg)

        # عرض الأعمدة
        for idx, w in enumerate([5, 26, 14, 22, 12, 13, 14, 12], 1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        ws.freeze_panes = 'A3'

        # صف المصدر
        footer_row = len(data) + 3
        ws.merge_cells(f'A{footer_row}:H{footer_row}')
        fc = ws[f'A{footer_row}']
        fc.value     = f'⚗️  تم استخراج هذا التقرير من تطبيق كيم تحصيلي  |  منصة تعليمية للكيمياء  |  جميع الحقوق محفوظة © {datetime.now().year}'
        fc.font      = Font(size=9, color='888888', italic=True)
        fc.alignment = Alignment(horizontal='center', vertical='center')
        fc.fill      = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
        ws.row_dimensions[footer_row].height = 18

        # ==================== ورقة 2: الملخص ====================
        ws2 = wb.create_sheet('الملخص')
        ws2.sheet_view.rightToLeft = True

        total       = len(data)
        active      = sum(1 for s in data if s.get('is_active'))
        with_results= sum(1 for s in data if (s.get('quiz_count') or 0) > 0)
        scores      = [s['avg_score'] for s in data if s.get('avg_score') is not None]
        avg_all     = sum(scores) / len(scores) if scores else 0

        dist = {
            'ممتاز (80%+)':           sum(1 for s in scores if s >= 80),
            'جيد (60-79%)':           sum(1 for s in scores if 60 <= s < 80),
            'يحتاج انتباه (40-59%)':  sum(1 for s in scores if 40 <= s < 60),
            'حرج (أقل من 40%)':       sum(1 for s in scores if 0 < s < 40),
            'لا توجد نتائج':          sum(1 for s in data if not s.get('avg_score')),
        }

        s2_hfont = Font(bold=True, size=11, color='FFFFFF')
        s2_hfill = PatternFill(start_color='1e3a8a', end_color='2563EB', fill_type='solid')
        s2_lf    = Font(bold=True, size=11)
        s2_vf    = Font(size=11)

        def _h2(row, text):
            ws2.merge_cells(f'A{row}:B{row}')
            c = ws2[f'A{row}']
            c.value = text; c.font = s2_hfont; c.fill = s2_hfill
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws2.row_dimensions[row].height = 22

        def _r2(row, label, value, bg=None):
            lc = ws2.cell(row, 1, label)
            lc.font = s2_lf; lc.alignment = right_align; lc.border = border
            vc = ws2.cell(row, 2, value)
            vc.font = s2_vf; vc.alignment = center_align; vc.border = border
            if bg:
                vc.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')

        _h2(1, 'معلومات التقرير')
        _r2(2, 'الإدارة',          admin_name)
        _r2(3, 'تاريخ الإنشاء',   now_str)
        _r2(4, 'عدد الطلاب',      total)

        _h2(6, 'إحصائيات عامة')
        _r2(7,  'إجمالي الطلاب',   total)
        _r2(8,  'النشطون',          active,       bg='C6EFCE')
        _r2(9,  'غير النشطين',      total-active, bg='FFC7CE')
        _r2(10, 'لهم نتائج',        with_results)
        _r2(11, 'المعدل العام',      f"{avg_all:.1f}%" if scores else 'لا نتائج')

        _h2(13, 'توزيع مستويات الأداء')
        dist_colors = ['C6EFCE', 'FFEB9C', 'FFD18C', 'FFC7CE', 'EEEEEE']
        for i, (label, count) in enumerate(dist.items()):
            pct = f"{(count/total*100):.1f}%" if total else '0%'
            _r2(14+i, label, f"{count} طالب  ({pct})", bg=dist_colors[i])

        ws2.column_dimensions['A'].width = 28
        ws2.column_dimensions['B'].width = 22

        ws2.merge_cells('A20:B20')
        sf = ws2['A20']
        sf.value     = f'⚗️  تم استخراج هذا التقرير من تطبيق كيم تحصيلي  |  منصة تعليمية للكيمياء  |  جميع الحقوق محفوظة © {datetime.now().year}'
        sf.font      = Font(size=9, color='888888', italic=True)
        sf.alignment = Alignment(horizontal='center', vertical='center')
        sf.fill      = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
        ws2.row_dimensions[20].height = 18

        # ==================== حفظ وإرسال ====================
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = f'تقرير_الطلاب_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@teachers_bp.route('/api/my/fcm-token', methods=['POST'])
@verify_teacher_token
def api_teacher_save_fcm_token():
    """المعلم يحفظ FCM Token للإشعارات"""
    data = request.get_json() or {}
    fcm_token = data.get('fcm_token', '').strip()
    if not fcm_token:
        return jsonify({'success': False, 'error': 'fcm_token مطلوب'}), 400

    teacher = Teacher.query.get_or_404(request.teacher_id)
    try:
        teacher.update_fcm_token(fcm_token)
        db.session.commit()
        return jsonify({'success': True, 'message': 'تم حفظ FCM Token'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@teachers_bp.route('/api/my/fcm-token', methods=['DELETE'])
@verify_teacher_token
def api_teacher_delete_fcm_token():
    """المعلم يحذف FCM Token عند تسجيل الخروج"""
    teacher = Teacher.query.get_or_404(request.teacher_id)
    try:
        teacher.clear_fcm_token()
        db.session.commit()
        return jsonify({'success': True, 'message': 'تم حذف FCM Token'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== Teacher Notifications Endpoints ====================

@teachers_bp.route('/api/my/notifications', methods=['GET'])
@verify_teacher_token
def api_teacher_my_notifications():
    """المعلم يجلب إشعاراته مع عدد غير المقروء"""
    from src.models.teacher_notification import TeacherNotification

    teacher_id = request.teacher_id
    notifications = (
        TeacherNotification.query
        .filter_by(teacher_id=teacher_id)
        .order_by(TeacherNotification.created_at.desc())
        .limit(50)
        .all()
    )
    unread_count = (
        TeacherNotification.query
        .filter_by(teacher_id=teacher_id, is_read=False)
        .count()
    )
    return jsonify({
        'success': True,
        'notifications': [n.to_dict() for n in notifications],
        'unread_count': unread_count,
    })


@teachers_bp.route('/api/my/notifications/read', methods=['POST'])
@verify_teacher_token
def api_teacher_mark_notifications_read():
    """تعليم كل إشعارات المعلم كمقروءة"""
    from src.models.teacher_notification import TeacherNotification

    try:
        TeacherNotification.query.filter_by(
            teacher_id=request.teacher_id, is_read=False
        ).update({'is_read': True})
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@teachers_bp.route('/api/my/notifications/<int:notif_id>', methods=['DELETE'])
@verify_teacher_token
def api_teacher_delete_notification(notif_id):
    """حذف إشعار واحد (يخص المعلم الحالي فقط)"""
    from src.models.teacher_notification import TeacherNotification

    notif = TeacherNotification.query.filter_by(
        id=notif_id, teacher_id=request.teacher_id
    ).first_or_404()
    try:
        db.session.delete(notif)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== Admin → Teacher Notify Endpoints ====================

@teachers_bp.route('/api/admin/notify/<int:teacher_id>', methods=['POST'])
@login_required
@admin_required
def api_admin_notify_teacher(teacher_id):
    """الأدمن يرسل إشعاراً لمعلم معين"""
    from src.models.teacher_notification import TeacherNotification

    teacher = Teacher.query.get_or_404(teacher_id)
    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    message = (data.get('message') or '').strip()

    if not title or not message:
        return jsonify({'success': False, 'error': 'title و message مطلوبان'}), 400

    notif = TeacherNotification.create(
        teacher_id=teacher.id,
        title=title,
        message=message,
        type='system',
    )
    if notif is None:
        return jsonify({'success': False, 'error': 'فشل في إرسال الإشعار'}), 500

    try:
        from src.models.audit_log import AuditLog
        AuditLog.log(
            action='send_notification',
            description=f'إرسال إشعار للمعلم "{teacher.name}": {title}',
            admin_name=getattr(current_user, 'username', 'الادمن'),
            target_type='teacher',
            target_id=teacher.id,
        )
        db.session.commit()
    except Exception as log_err:
        print(f'⚠️ AuditLog error: {log_err}')

    return jsonify({'success': True})


@teachers_bp.route('/api/admin/notify-all', methods=['POST'])
@login_required
@admin_required
def api_admin_notify_all_teachers():
    """الأدمن يرسل إشعاراً لجميع المعلمين"""
    from src.models.teacher_notification import TeacherNotification

    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    message = (data.get('message') or '').strip()

    if not title or not message:
        return jsonify({'success': False, 'error': 'title و message مطلوبان'}), 400

    teachers = Teacher.query.filter_by(is_active=True).all()
    count = 0
    try:
        for t in teachers:
            notif = TeacherNotification(
                teacher_id=t.id,
                title=title,
                message=message,
                type='system',
            )
            db.session.add(notif)
            count += 1
        db.session.commit()
        try:
            from src.models.audit_log import AuditLog
            AuditLog.log(
                action='send_notification',
                description=f'إرسال إشعار لجميع المعلمين ({count} معلم): {title}',
                admin_name=getattr(current_user, 'username', 'الادمن'),
                target_type='teachers_all',
            )
            db.session.commit()
        except Exception as log_err:
            print(f'⚠️ AuditLog error: {log_err}')
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

    return jsonify({'success': True, 'count': count})

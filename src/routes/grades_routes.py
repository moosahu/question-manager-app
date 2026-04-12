"""
نظام الدرجات والحضور اليومي
- admin web: إدارة grade_config (الفترات والفئات)
- teacher API: إدخال درجات، حضور، إرسال درجات
- student API: عرض ما أُرسل من درجات
"""
from flask import Blueprint, request, jsonify, render_template, redirect, url_for, flash
from flask_login import login_required
from sqlalchemy import func
from datetime import datetime, date

from src.extensions import db
from src.middleware.auth_middleware import verify_teacher_token, verify_student_token
from src.models.grade_period import GradePeriod
from src.models.grade_config import GradeConfig
from src.models.grade_entry import GradeEntry
from src.models.student_attendance import StudentAttendance
from src.models.student_grade_release import StudentGradeRelease
from src.models.teacher_student import TeacherStudent
from src.models.student import Student

grades_bp = Blueprint('grades', __name__)


# ──────────────────────────────────────────────────────
#  دوال مساعدة
# ──────────────────────────────────────────────────────

def _sync_to_old_system(student_id, category_id, teacher_id=None, admin_id=None):
    """مزامنة: لما يُدخَل GradeEntry في فئة اختبار → يُحدَّث StudentSemesterGrade تلقائياً"""
    try:
        from src.models.student_semester_grade import StudentSemesterGrade
        cat = GradeConfig.query.get(category_id)
        if not cat:
            return
        period = GradePeriod.query.get(cat.period_id)
        if not period or not period.semester_number:
            return

        # فقط نزامن إذا كانت هذه الفئة هي الأعلى درجة في الفترة (= فئة الاختبار)
        top_cat = GradeConfig.query.filter_by(
            period_id=cat.period_id, is_active=True
        ).order_by(GradeConfig.max_score.desc()).first()
        if not top_cat or top_cat.id != cat.id:
            return

        # متوسط جميع إدخالات الطالب في هذه الفئة
        q = GradeEntry.query.filter_by(student_id=student_id, category_id=category_id)
        if teacher_id:
            q = q.filter_by(teacher_id=teacher_id)
        elif admin_id:
            q = q.filter_by(admin_id=admin_id)
        entries = q.all()
        if not entries:
            return

        avg_ratio = sum(e.score for e in entries) / len(entries)
        grade     = round(avg_ratio * cat.max_score, 2)

        existing = StudentSemesterGrade.get_grade(student_id, period.semester_number)
        if existing:
            existing.grade     = grade
            existing.max_grade = cat.max_score
            existing.updated_at = datetime.utcnow()
        else:
            db.session.add(StudentSemesterGrade(
                student_id=student_id,
                period=period.semester_number,
                grade=grade,
                max_grade=cat.max_score,
            ))
    except Exception as e:
        print(f"⚠️ _sync_to_old_system error: {e}")


def _calc_category_score(entries):
    """متوسط الإدخالات (score بين 0-1) × max_score"""
    if not entries:
        return None
    avg = sum(e.score for e in entries) / len(entries)
    return round(avg * entries[0].category.max_score, 2)


def _build_student_grades(student_id, teacher_id=None, admin_id=None):
    """يبني هيكل درجات الطالب — يدعم teacher_id أو admin_id"""
    periods = GradePeriod.query.filter_by(is_active=True).order_by(GradePeriod.sort_order).all()
    result = {}
    for period in periods:
        cats = GradeConfig.query.filter_by(period_id=period.id, is_active=True)\
                                .order_by(GradeConfig.sort_order).all()
        period_data = {'period_name': period.period_name, 'categories': {}}
        for cat in cats:
            q = GradeEntry.query.filter_by(student_id=student_id, category_id=cat.id)
            if teacher_id:
                q = q.filter_by(teacher_id=teacher_id)
            elif admin_id:
                q = q.filter_by(admin_id=admin_id)
            entries = q.order_by(GradeEntry.entry_date).all()
            avg = _calc_category_score(entries) if entries else None
            period_data['categories'][cat.id] = {
                'category_name': cat.category_name,
                'max_score':     cat.max_score,
                'entries':       [e.to_dict() for e in entries],
                'avg_score':     avg,
            }
        result[period.id] = period_data
    return result


# ══════════════════════════════════════════════════════
#  ADMIN WEB: إدارة grade_config
# ══════════════════════════════════════════════════════

@grades_bp.route('/admin/grade-config', methods=['GET'])
@login_required
def admin_grade_config():
    """صفحة إدارة التقسيمة"""
    periods = GradePeriod.query.order_by(GradePeriod.sort_order).all()
    configs = GradeConfig.query.order_by(GradeConfig.period_id, GradeConfig.sort_order).all()
    return render_template('grades/grade_config.html', periods=periods, configs=configs)


@grades_bp.route('/admin/grade-config/add-category', methods=['POST'])
@login_required
def admin_add_category():
    period_id     = request.form.get('period_id', type=int)
    category_name = request.form.get('category_name', '').strip()
    max_score     = request.form.get('max_score', type=float, default=10)
    sort_order    = request.form.get('sort_order', type=int, default=0)

    if not period_id or not category_name:
        flash('الفترة واسم الفئة مطلوبان', 'danger')
        return redirect(url_for('grades.admin_grade_config'))

    cat = GradeConfig(
        period_id=period_id,
        category_name=category_name,
        max_score=max_score,
        sort_order=sort_order,
        is_active=True,
    )
    db.session.add(cat)
    db.session.commit()
    flash(f'تمت إضافة الفئة "{category_name}"', 'success')
    return redirect(url_for('grades.admin_grade_config'))


@grades_bp.route('/admin/grade-config/<int:cat_id>/edit', methods=['POST'])
@login_required
def admin_edit_category(cat_id):
    cat = GradeConfig.query.get_or_404(cat_id)
    cat.category_name = request.form.get('category_name', cat.category_name).strip()
    cat.max_score     = request.form.get('max_score', type=float, default=cat.max_score)
    cat.sort_order    = request.form.get('sort_order', type=int, default=cat.sort_order)
    db.session.commit()
    flash('تم تحديث الفئة', 'success')
    return redirect(url_for('grades.admin_grade_config'))


@grades_bp.route('/admin/grade-config/<int:cat_id>/toggle', methods=['POST'])
@login_required
def admin_toggle_category(cat_id):
    cat = GradeConfig.query.get_or_404(cat_id)
    cat.is_active = not cat.is_active
    db.session.commit()
    return redirect(url_for('grades.admin_grade_config'))


@grades_bp.route('/admin/grade-config/<int:cat_id>/delete', methods=['POST'])
@login_required
def admin_delete_category(cat_id):
    cat = GradeConfig.query.get_or_404(cat_id)
    db.session.delete(cat)
    db.session.commit()
    flash('تم حذف الفئة', 'success')
    return redirect(url_for('grades.admin_grade_config'))


@grades_bp.route('/admin/grade-config/add-period', methods=['POST'])
@login_required
def admin_add_period():
    period_name = request.form.get('period_name', '').strip()
    sort_order  = request.form.get('sort_order', type=int, default=0)
    if not period_name:
        flash('اسم الفترة مطلوب', 'danger')
        return redirect(url_for('grades.admin_grade_config'))
    p = GradePeriod(period_name=period_name, sort_order=sort_order, is_active=True)
    db.session.add(p)
    db.session.commit()
    flash(f'تمت إضافة الفترة "{period_name}"', 'success')
    return redirect(url_for('grades.admin_grade_config'))


# ══════════════════════════════════════════════════════
#  TEACHER API: التقسيمة
# ══════════════════════════════════════════════════════

@grades_bp.route('/api/teacher/grade-config', methods=['GET'])
@verify_teacher_token
def api_teacher_grade_config():
    """المعلم يجلب التقسيمة الكاملة"""
    periods = GradePeriod.query.filter_by(is_active=True).order_by(GradePeriod.sort_order).all()
    data = []
    for p in periods:
        cats = GradeConfig.query.filter_by(period_id=p.id, is_active=True)\
                                .order_by(GradeConfig.sort_order).all()
        data.append({
            'period_id':   p.id,
            'period_name': p.period_name,
            'categories':  [c.to_dict() for c in cats],
        })
    return jsonify({'success': True, 'periods': data})


@grades_bp.route('/api/admin/grade-config', methods=['GET'])
@login_required
def api_admin_grade_config():
    """الأدمن يجلب التقسيمة — session cookie"""
    periods = GradePeriod.query.filter_by(is_active=True).order_by(GradePeriod.sort_order).all()
    data = []
    for p in periods:
        cats = GradeConfig.query.filter_by(period_id=p.id, is_active=True)\
                                .order_by(GradeConfig.sort_order).all()
        data.append({
            'period_id':   p.id,
            'period_name': p.period_name,
            'categories':  [c.to_dict() for c in cats],
        })
    return jsonify({'success': True, 'periods': data})


# ══════════════════════════════════════════════════════
#  TEACHER API: إدخال درجات
# ══════════════════════════════════════════════════════

@grades_bp.route('/api/teacher/grade-entries', methods=['POST'])
@verify_teacher_token
def api_add_grade_entry():
    """إضافة إدخال درجة لطالب في فئة معينة"""
    teacher_id = request.teacher_id
    data = request.get_json() or {}

    student_id  = data.get('student_id')
    category_id = data.get('category_id')
    entry_label = data.get('entry_label', '').strip()
    score       = data.get('score')          # 0 – max_score

    if not all([student_id, category_id, entry_label, score is not None]):
        return jsonify({'success': False, 'error': 'البيانات ناقصة'}), 400

    cat = GradeConfig.query.get(category_id)
    if not cat:
        return jsonify({'success': False, 'error': 'الفئة غير موجودة'}), 404

    # التحقق أن الطالب تابع لهذا المعلم
    link = TeacherStudent.query.filter_by(teacher_id=teacher_id, student_id=student_id).first()
    if not link:
        return jsonify({'success': False, 'error': 'الطالب غير مرتبط بحسابك'}), 403

    # score يُخزّن كنسبة 0-1
    score_ratio = float(score) / cat.max_score
    score_ratio = max(0.0, min(1.0, score_ratio))

    entry = GradeEntry(
        student_id=student_id,
        teacher_id=teacher_id,
        category_id=category_id,
        entry_label=entry_label,
        score=score_ratio,
        entry_date=date.today(),
    )
    db.session.add(entry)
    db.session.flush()
    _sync_to_old_system(student_id, category_id, teacher_id=teacher_id)
    db.session.commit()
    return jsonify({'success': True, 'entry': entry.to_dict()})


@grades_bp.route('/api/teacher/grade-entries/bulk', methods=['POST'])
@verify_teacher_token
def api_bulk_add_grade_entries():
    """إدخال درجة واجب/اختبار لعدة طلاب دفعة واحدة"""
    teacher_id = request.teacher_id
    data = request.get_json() or {}

    category_id = data.get('category_id')
    entry_label = data.get('entry_label', '').strip()
    entries     = data.get('entries', [])   # [{student_id, score}, ...]

    if not all([category_id, entry_label, entries]):
        return jsonify({'success': False, 'error': 'البيانات ناقصة'}), 400

    cat = GradeConfig.query.get(category_id)
    if not cat:
        return jsonify({'success': False, 'error': 'الفئة غير موجودة'}), 404

    saved, failed = 0, 0
    for item in entries:
        sid   = item.get('student_id')
        score = item.get('score')
        if sid is None or score is None:
            failed += 1
            continue
        link = TeacherStudent.query.filter_by(teacher_id=teacher_id, student_id=sid).first()
        if not link:
            failed += 1
            continue
        ratio = max(0.0, min(1.0, float(score) / cat.max_score))
        db.session.add(GradeEntry(
            student_id=sid, teacher_id=teacher_id,
            category_id=category_id, entry_label=entry_label,
            score=ratio, entry_date=date.today(),
        ))
        saved += 1

    db.session.flush()
    for item in entries:
        sid = item.get('student_id')
        if sid:
            _sync_to_old_system(sid, category_id, teacher_id=teacher_id)
    db.session.commit()
    return jsonify({'success': True, 'saved': saved, 'failed': failed})


@grades_bp.route('/api/teacher/grade-entries/<int:entry_id>', methods=['DELETE'])
@verify_teacher_token
def api_delete_grade_entry(entry_id):
    teacher_id = request.teacher_id
    entry = GradeEntry.query.get_or_404(entry_id)
    if entry.teacher_id != teacher_id:
        return jsonify({'success': False, 'error': 'غير مصرح'}), 403
    db.session.delete(entry)
    db.session.commit()
    return jsonify({'success': True})


@grades_bp.route('/api/teacher/students/<int:student_id>/grades', methods=['GET'])
@verify_teacher_token
def api_teacher_student_grades(student_id):
    """المعلم يجلب درجات طالب بالتفصيل"""
    teacher_id = request.teacher_id
    link = TeacherStudent.query.filter_by(teacher_id=teacher_id, student_id=student_id).first()
    if not link:
        return jsonify({'success': False, 'error': 'الطالب غير مرتبط بحسابك'}), 403

    grades = _build_student_grades(student_id, teacher_id)
    return jsonify({'success': True, 'grades': grades})


# ══════════════════════════════════════════════════════
#  TEACHER API: الحضور
# ══════════════════════════════════════════════════════

@grades_bp.route('/api/teacher/attendance', methods=['POST'])
@verify_teacher_token
def api_record_attendance():
    """تسجيل / تحديث حضور طالب"""
    teacher_id = request.teacher_id
    data = request.get_json() or {}

    student_id      = data.get('student_id')
    attendance_date = data.get('attendance_date')  # YYYY-MM-DD
    period_number   = data.get('period_number', 1)
    status          = data.get('status', 'present')  # present/absent/late/sleeping
    notes           = data.get('notes', '')

    if not all([student_id, attendance_date]):
        return jsonify({'success': False, 'error': 'student_id و attendance_date مطلوبان'}), 400

    if status not in ('present', 'absent', 'late', 'sleeping'):
        return jsonify({'success': False, 'error': 'status غير صحيح'}), 400

    try:
        att_date = datetime.strptime(attendance_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'error': 'تنسيق التاريخ غير صحيح (YYYY-MM-DD)'}), 400

    link = TeacherStudent.query.filter_by(teacher_id=teacher_id, student_id=student_id).first()
    if not link:
        return jsonify({'success': False, 'error': 'الطالب غير مرتبط بحسابك'}), 403

    existing = StudentAttendance.query.filter_by(
        student_id=student_id,
        teacher_id=teacher_id,
        attendance_date=att_date,
        period_number=period_number,
    ).first()

    if existing:
        existing.status = status
        existing.notes  = notes
        att = existing
    else:
        att = StudentAttendance(
            student_id=student_id,
            teacher_id=teacher_id,
            attendance_date=att_date,
            period_number=period_number,
            status=status,
            notes=notes,
        )
        db.session.add(att)

    db.session.commit()
    return jsonify({'success': True, 'attendance': att.to_dict()})


@grades_bp.route('/api/teacher/attendance', methods=['GET'])
@verify_teacher_token
def api_get_attendance():
    """سجل الحضور للمعلم — فلترة بالتاريخ أو الطالب"""
    teacher_id = request.teacher_id
    student_id = request.args.get('student_id', type=int)
    date_str   = request.args.get('date')  # YYYY-MM-DD

    q = StudentAttendance.query.filter_by(teacher_id=teacher_id)
    if student_id:
        q = q.filter_by(student_id=student_id)
    if date_str:
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
            q = q.filter_by(attendance_date=d)
        except ValueError:
            pass

    records = q.order_by(StudentAttendance.attendance_date.desc(),
                         StudentAttendance.period_number).all()
    return jsonify({'success': True, 'attendance': [r.to_dict() for r in records]})


# ══════════════════════════════════════════════════════
#  TEACHER API: إرسال درجات للطالب
# ══════════════════════════════════════════════════════

@grades_bp.route('/api/teacher/send-grades', methods=['POST'])
@verify_teacher_token
def api_send_grades():
    """المعلم يرسل درجات فترة معينة لطالب"""
    teacher_id = request.teacher_id
    data = request.get_json() or {}

    student_id     = data.get('student_id')
    period_id      = data.get('period_id')
    release_type   = data.get('release_type', 'both')   # test/yearly/both
    custom_message = data.get('custom_message', '')

    if not all([student_id, period_id]):
        return jsonify({'success': False, 'error': 'student_id و period_id مطلوبان'}), 400

    if release_type not in ('test', 'yearly', 'both'):
        return jsonify({'success': False, 'error': 'release_type غير صحيح'}), 400

    link = TeacherStudent.query.filter_by(teacher_id=teacher_id, student_id=student_id).first()
    if not link:
        return jsonify({'success': False, 'error': 'الطالب غير مرتبط بحسابك'}), 403

    # إنشاء أو تحديث سجل الإرسال
    existing = StudentGradeRelease.query.filter_by(
        student_id=student_id,
        teacher_id=teacher_id,
        period_id=period_id,
    ).first()

    if existing:
        existing.release_type   = release_type
        existing.custom_message = custom_message
        existing.released_at    = datetime.utcnow()
        rel = existing
    else:
        rel = StudentGradeRelease(
            student_id=student_id,
            teacher_id=teacher_id,
            period_id=period_id,
            release_type=release_type,
            custom_message=custom_message,
        )
        db.session.add(rel)

    db.session.commit()

    # ── إشعار FCM للطالب ──────────────────────────────
    try:
        from src.services.notification_service import NotificationService
        student = Student.query.get(student_id)
        period  = GradePeriod.query.get(period_id)
        if student and student.fcm_token:
            type_label = {'test': 'الاختبار', 'yearly': 'أعمال السنة', 'both': 'الدرجات'}.get(release_type, 'الدرجات')
            period_name = period.period_name if period else ''
            notif_body  = f'أرسل معلمك {type_label} - {period_name}'
            if custom_message:
                notif_body += f'\n{custom_message}'
            NotificationService.send_fcm_notification(
                student.fcm_token,
                '📊 درجاتك الجديدة',
                notif_body,
                {'type': 'grades_released', 'period_id': str(period_id)},
            )
    except Exception as e:
        print(f"⚠️ grades FCM error: {e}")

    return jsonify({'success': True, 'release': rel.to_dict()})


# ══════════════════════════════════════════════════════
#  TEACHER API: إرسال جماعي
# ══════════════════════════════════════════════════════

@grades_bp.route('/api/teacher/send-grades/bulk', methods=['POST'])
@verify_teacher_token
def api_bulk_send_grades():
    """المعلم يرسل الدرجات لمجموعة طلاب دفعة واحدة"""
    teacher_id = request.teacher_id
    data = request.get_json() or {}

    period_id      = data.get('period_id')
    release_type   = data.get('release_type', 'both')
    student_ids    = data.get('student_ids', [])
    custom_message = data.get('custom_message', '')

    if not period_id or not student_ids:
        return jsonify({'success': False, 'error': 'period_id و student_ids مطلوبان'}), 400
    if release_type not in ('test', 'yearly', 'both'):
        return jsonify({'success': False, 'error': 'release_type غير صحيح'}), 400

    period = GradePeriod.query.get(period_id)
    saved, failed = 0, 0

    for sid in student_ids:
        link = TeacherStudent.query.filter_by(teacher_id=teacher_id, student_id=sid).first()
        if not link:
            failed += 1
            continue
        existing = StudentGradeRelease.query.filter_by(
            student_id=sid, teacher_id=teacher_id, period_id=period_id).first()
        if existing:
            existing.release_type   = release_type
            existing.custom_message = custom_message
            existing.released_at    = datetime.utcnow()
        else:
            db.session.add(StudentGradeRelease(
                student_id=sid, teacher_id=teacher_id,
                period_id=period_id, release_type=release_type,
                custom_message=custom_message,
            ))
        saved += 1

        # FCM
        try:
            from src.services.notification_service import NotificationService
            student = Student.query.get(sid)
            if student and student.fcm_token:
                type_label  = {'test': 'الاختبار', 'yearly': 'أعمال السنة', 'both': 'الدرجات'}.get(release_type, 'الدرجات')
                period_name = period.period_name if period else ''
                body = f'أرسل معلمك {type_label} - {period_name}'
                if custom_message:
                    body += f'\n{custom_message}'
                NotificationService.send_fcm_notification(
                    student.fcm_token, '📊 درجاتك الجديدة', body,
                    {'type': 'grades_released', 'period_id': str(period_id)})
        except Exception as e:
            print(f"⚠️ bulk FCM error: {e}")

    db.session.commit()
    return jsonify({'success': True, 'sent': saved, 'failed': failed})


@grades_bp.route('/api/admin/send-grades/bulk', methods=['POST'])
@login_required
def api_admin_bulk_send_grades():
    """الأدمن يرسل الدرجات لمجموعة طلاب دفعة واحدة"""
    from flask_login import current_user
    data = request.get_json() or {}

    period_id      = data.get('period_id')
    release_type   = data.get('release_type', 'both')
    student_ids    = data.get('student_ids', [])
    custom_message = data.get('custom_message', '')

    if not period_id or not student_ids:
        return jsonify({'success': False, 'error': 'بيانات ناقصة'}), 400
    if release_type not in ('test', 'yearly', 'both'):
        return jsonify({'success': False, 'error': 'release_type غير صحيح'}), 400

    period = GradePeriod.query.get(period_id)
    saved, failed = 0, 0

    for sid in student_ids:
        existing = StudentGradeRelease.query.filter_by(
            student_id=sid, admin_id=current_user.id, period_id=period_id).first()
        if existing:
            existing.release_type   = release_type
            existing.custom_message = custom_message
            existing.released_at    = datetime.utcnow()
        else:
            db.session.add(StudentGradeRelease(
                student_id=sid, admin_id=current_user.id,
                period_id=period_id, release_type=release_type,
                custom_message=custom_message,
            ))
        saved += 1

        try:
            from src.services.notification_service import NotificationService
            student = Student.query.get(sid)
            if student and student.fcm_token:
                type_label  = {'test': 'الاختبار', 'yearly': 'أعمال السنة', 'both': 'الدرجات'}.get(release_type, 'الدرجات')
                period_name = period.period_name if period else ''
                body = f'أُرسلت لك {type_label} - {period_name}'
                if custom_message:
                    body += f'\n{custom_message}'
                NotificationService.send_fcm_notification(
                    student.fcm_token, '📊 درجاتك الجديدة', body,
                    {'type': 'grades_released', 'period_id': str(period_id)})
        except Exception as e:
            print(f"⚠️ bulk FCM error: {e}")

    db.session.commit()
    return jsonify({'success': True, 'sent': saved, 'failed': failed})


# ══════════════════════════════════════════════════════
#  STUDENT API: عرض الدرجات المُرسلة
# ══════════════════════════════════════════════════════

@grades_bp.route('/api/student/my-grades', methods=['GET'])
@verify_student_token
def api_student_my_grades():
    """الطالب يجلب الدرجات التي أرسلها المعلم له"""
    student_id = request.student_id

    # جلب كل سجلات الإرسال لهذا الطالب
    releases = StudentGradeRelease.query.filter_by(student_id=student_id).all()
    if not releases:
        return jsonify({'success': True, 'data': []})

    output = []
    for rel in releases:
        teacher_id    = rel.teacher_id
        period_id     = rel.period_id
        release_type  = rel.release_type

        # جلب الفئات المطلوبة فقط حسب release_type
        period = GradePeriod.query.get(period_id)
        cats   = GradeConfig.query.filter_by(period_id=period_id, is_active=True)\
                                  .order_by(GradeConfig.sort_order).all()

        categories_data = []
        for cat in cats:
            # فلترة حسب نوع الإرسال (بسيط: test = يحتوي "اختبار"، yearly = غيره)
            is_test_cat = 'اختبار' in cat.category_name
            if release_type == 'test' and not is_test_cat:
                continue
            if release_type == 'yearly' and is_test_cat:
                continue

            entries = GradeEntry.query.filter_by(
                student_id=student_id,
                teacher_id=teacher_id,
                category_id=cat.id,
            ).order_by(GradeEntry.entry_date).all()

            avg = _calc_category_score(entries) if entries else None
            categories_data.append({
                'category_name': cat.category_name,
                'max_score':     cat.max_score,
                'avg_score':     avg,
                'entries':       [{'label': e.entry_label, 'score': round(e.score * cat.max_score, 2)} for e in entries],
            })

        output.append({
            'period_id':      period_id,
            'period_name':    period.period_name if period else '',
            'release_type':   release_type,
            'custom_message': rel.custom_message or '',
            'released_at':    rel.to_dict()['released_at'],
            'categories':     categories_data,
        })

    return jsonify({'success': True, 'data': output})


# ══════════════════════════════════════════════════════
#  ADMIN API: إدارة درجات وحضور (session cookie)
#  نفس منطق المعلم لكن teacher_id يأتي من الـ body/params
# ══════════════════════════════════════════════════════

@grades_bp.route('/api/admin/students/<int:student_id>/grades', methods=['GET'])
@login_required
def api_admin_student_grades(student_id):
    """الأدمن يجلب درجات طالبه — يستخدم admin_id من الجلسة"""
    from flask_login import current_user
    try:
        grades = _build_student_grades(student_id, admin_id=current_user.id)
        # تحويل المفاتيح الرقمية إلى نصوص لضمان JSON صحيح
        grades_str = {str(k): v for k, v in grades.items()}
        for pdata in grades_str.values():
            pdata['categories'] = {str(k): v for k, v in pdata['categories'].items()}
        return jsonify({'success': True, 'grades': grades_str})
    except Exception as e:
        import traceback
        print(f"❌ api_admin_student_grades error: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500


@grades_bp.route('/api/admin/grade-entries', methods=['POST'])
@login_required
def api_admin_add_grade_entry():
    from flask_login import current_user
    data        = request.get_json() or {}
    student_id  = data.get('student_id')
    category_id = data.get('category_id')
    entry_label = data.get('entry_label', '').strip()
    score       = data.get('score')

    if not all([student_id, category_id, entry_label, score is not None]):
        return jsonify({'success': False, 'error': 'البيانات ناقصة'}), 400

    cat = GradeConfig.query.get(category_id)
    if not cat:
        return jsonify({'success': False, 'error': 'الفئة غير موجودة'}), 404

    score_ratio = max(0.0, min(1.0, float(score) / cat.max_score))
    entry = GradeEntry(
        student_id=student_id, admin_id=current_user.id,
        category_id=category_id, entry_label=entry_label,
        score=score_ratio, entry_date=date.today(),
    )
    db.session.add(entry)
    db.session.flush()
    _sync_to_old_system(student_id, category_id, admin_id=current_user.id)
    db.session.commit()
    return jsonify({'success': True, 'entry': entry.to_dict()})


@grades_bp.route('/api/admin/grade-entries/bulk', methods=['POST'])
@login_required
def api_admin_bulk_add_grade_entries():
    """إدخال درجات جماعية — الأدمن"""
    from flask_login import current_user
    data = request.get_json() or {}

    category_id = data.get('category_id')
    entry_label = data.get('entry_label', '').strip()
    entries     = data.get('entries', [])

    if not all([category_id, entry_label, entries]):
        return jsonify({'success': False, 'error': 'البيانات ناقصة'}), 400

    cat = GradeConfig.query.get(category_id)
    if not cat:
        return jsonify({'success': False, 'error': 'الفئة غير موجودة'}), 404

    saved, failed = 0, 0
    for item in entries:
        sid   = item.get('student_id')
        score = item.get('score')
        if sid is None or score is None:
            failed += 1
            continue
        ratio = max(0.0, min(1.0, float(score) / cat.max_score))
        db.session.add(GradeEntry(
            student_id=sid, admin_id=current_user.id,
            category_id=category_id, entry_label=entry_label,
            score=ratio, entry_date=date.today(),
        ))
        saved += 1

    db.session.flush()
    for item in entries:
        sid = item.get('student_id')
        if sid:
            _sync_to_old_system(sid, category_id, admin_id=current_user.id)
    db.session.commit()
    return jsonify({'success': True, 'saved': saved, 'failed': failed})


@grades_bp.route('/api/admin/sync-grades', methods=['POST'])
@login_required
def api_admin_sync_grades():
    """إعادة مزامنة كل إدخالات الاختبار (أعلى max_score) مع النظام القديم"""
    from flask_login import current_user
    synced = 0
    entries = GradeEntry.query.filter_by(admin_id=current_user.id).all()
    for entry in entries:
        try:
            _sync_to_old_system(entry.student_id, entry.category_id,
                                admin_id=current_user.id)
            synced += 1
        except Exception:
            pass
    db.session.commit()
    return jsonify({'success': True, 'synced': synced})


@grades_bp.route('/api/admin/grade-entries/<int:entry_id>', methods=['DELETE'])
@login_required
def api_admin_delete_grade_entry(entry_id):
    entry = GradeEntry.query.get_or_404(entry_id)
    db.session.delete(entry)
    db.session.commit()
    return jsonify({'success': True})


@grades_bp.route('/api/admin/attendance', methods=['POST'])
@login_required
def api_admin_record_attendance():
    from flask_login import current_user
    data            = request.get_json() or {}
    student_id      = data.get('student_id')
    attendance_date = data.get('attendance_date')
    period_number   = data.get('period_number', 1)
    status          = data.get('status', 'present')
    notes           = data.get('notes', '')

    if not all([student_id, attendance_date]):
        return jsonify({'success': False, 'error': 'بيانات ناقصة'}), 400
    if status not in ('present', 'absent', 'late', 'sleeping'):
        return jsonify({'success': False, 'error': 'status غير صحيح'}), 400
    try:
        att_date = datetime.strptime(attendance_date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'error': 'تنسيق التاريخ غير صحيح'}), 400

    existing = StudentAttendance.query.filter_by(
        student_id=student_id, admin_id=current_user.id,
        attendance_date=att_date, period_number=period_number,
    ).first()
    if existing:
        existing.status = status
        existing.notes  = notes
        att = existing
    else:
        att = StudentAttendance(
            student_id=student_id, admin_id=current_user.id,
            attendance_date=att_date, period_number=period_number,
            status=status, notes=notes,
        )
        db.session.add(att)
    db.session.commit()
    return jsonify({'success': True, 'attendance': att.to_dict()})


@grades_bp.route('/api/admin/attendance', methods=['GET'])
@login_required
def api_admin_get_attendance():
    from flask_login import current_user
    student_id = request.args.get('student_id', type=int)
    date_str   = request.args.get('date')
    q = StudentAttendance.query.filter_by(admin_id=current_user.id)
    if student_id:
        q = q.filter_by(student_id=student_id)
    if date_str:
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
            q = q.filter_by(attendance_date=d)
        except ValueError:
            pass
    records = q.order_by(StudentAttendance.attendance_date.desc(),
                         StudentAttendance.period_number).all()
    return jsonify({'success': True, 'attendance': [r.to_dict() for r in records]})


@grades_bp.route('/api/admin/send-grades', methods=['POST'])
@login_required
def api_admin_send_grades():
    from flask_login import current_user
    data           = request.get_json() or {}
    student_id     = data.get('student_id')
    period_id      = data.get('period_id')
    release_type   = data.get('release_type', 'both')
    custom_message = data.get('custom_message', '')

    if not all([student_id, period_id]):
        return jsonify({'success': False, 'error': 'بيانات ناقصة'}), 400
    if release_type not in ('test', 'yearly', 'both'):
        return jsonify({'success': False, 'error': 'release_type غير صحيح'}), 400

    existing = StudentGradeRelease.query.filter_by(
        student_id=student_id, admin_id=current_user.id, period_id=period_id,
    ).first()
    if existing:
        existing.release_type   = release_type
        existing.custom_message = custom_message
        existing.released_at    = datetime.utcnow()
        rel = existing
    else:
        rel = StudentGradeRelease(
            student_id=student_id, admin_id=current_user.id,
            period_id=period_id, release_type=release_type,
            custom_message=custom_message,
        )
        db.session.add(rel)
    db.session.commit()

    # إشعار FCM
    try:
        from src.services.notification_service import NotificationService
        student = Student.query.get(student_id)
        period  = GradePeriod.query.get(period_id)
        if student and student.fcm_token:
            type_label  = {'test': 'الاختبار', 'yearly': 'أعمال السنة', 'both': 'الدرجات'}.get(release_type, 'الدرجات')
            period_name = period.period_name if period else ''
            notif_body  = f'أُرسلت لك {type_label} - {period_name}'
            if custom_message:
                notif_body += f'\n{custom_message}'
            NotificationService.send_fcm_notification(
                student.fcm_token, '📊 درجاتك الجديدة', notif_body,
                {'type': 'grades_released', 'period_id': str(period_id)},
            )
    except Exception as e:
        print(f"⚠️ grades FCM error: {e}")

    return jsonify({'success': True, 'release': rel.to_dict()})


# ══════════════════════════════════════════════════════
#  ADMIN: تحميل قالب الدرجات (Excel)
# ══════════════════════════════════════════════════════

@grades_bp.route('/api/admin/grade-template', methods=['GET'])
@login_required
def api_admin_grade_template():
    from flask_login import current_user
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from flask import send_file

    period_id = request.args.get('period_id', type=int)
    if not period_id:
        return jsonify({'success': False, 'error': 'period_id مطلوب'}), 400

    period = GradePeriod.query.get(period_id)
    if not period:
        return jsonify({'success': False, 'error': 'الفترة غير موجودة'}), 404

    cats = GradeConfig.query.filter_by(period_id=period_id, is_active=True)\
                            .order_by(GradeConfig.sort_order).all()
    if not cats:
        return jsonify({'success': False, 'error': 'لا توجد فئات لهذه الفترة'}), 404

    students = Student.query.filter(Student.is_active != False)\
                            .order_by(Student.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = period.period_name
    ws.sheet_view.rightToLeft = True

    blue_fill   = PatternFill('solid', fgColor='1e3a8a')
    gray_fill   = PatternFill('solid', fgColor='F1F5F9')
    yellow_fill = PatternFill('solid', fgColor='FEF9C3')
    thin   = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    total_cols = 2 + len(cats)

    # الصف 1: عنوان
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    tc = ws.cell(1, 1, value=f'قالب درجات — {period.period_name}  |  يُرفع في التطبيق بعد التعبئة')
    tc.font      = Font(bold=True, color='FFFFFF', size=13)
    tc.fill      = blue_fill
    tc.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
    ws.row_dimensions[1].height = 28

    # الصف 2: رؤوس
    headers = ['student_id', 'اسم الطالب'] + \
              [f"{c.category_name}\n/{int(c.max_score)}" for c in cats]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(2, col, value=h)
        cell.font      = Font(bold=True, color='FFFFFF', size=11)
        cell.fill      = blue_fill
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True, readingOrder=2)
        cell.border = border
    ws.row_dimensions[2].height = 40
    ws.column_dimensions['A'].hidden = True

    # بيانات الطلاب
    for r, student in enumerate(students, 3):
        ws.cell(r, 1, value=student.id)
        nc = ws.cell(r, 2, value=student.name)
        nc.fill      = gray_fill
        nc.font      = Font(bold=True, size=11)
        nc.alignment = Alignment(horizontal='right', vertical='center', readingOrder=2)
        nc.border    = border
        for col in range(3, total_cols + 1):
            cell = ws.cell(r, col)
            cell.fill      = yellow_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = border
        ws.row_dimensions[r].height = 22

    ws.column_dimensions['B'].width = 28
    for col in range(3, total_cols + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

    # صفحة تعليمات
    ws2 = wb.create_sheet('تعليمات')
    ws2.sheet_view.rightToLeft = True
    ws2['A1'] = 'تعليمات الاستخدام'
    ws2['A1'].font = Font(bold=True, size=14, color='1e3a8a')
    for i, txt in enumerate([
        '١. لا تغيّر أسماء الطلاب أو ترتيبهم',
        '٢. لا تغيّر رؤوس الأعمدة',
        '٣. أدخل الدرجة في الخلايا الصفراء فقط',
        '٤. الخلية الفارغة = لا يُضاف شيء لهذا الطالب',
        '٥. بعد التعبئة، ارفع الملف من زر "استيراد درجات" في التطبيق',
        f'٦. هذا القالب خاص بـ: {period.period_name}',
    ], 2):
        ws2[f'A{i}'] = txt
        ws2[f'A{i}'].font = Font(size=12)
    ws2.column_dimensions['A'].width = 55

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'درجات_{period.period_name}.xlsx')


# ══════════════════════════════════════════════════════
#  ADMIN: استيراد درجات من Excel
# ══════════════════════════════════════════════════════

@grades_bp.route('/api/admin/import-grades-excel', methods=['POST'])
@login_required
def api_admin_import_grades_excel():
    from flask_login import current_user
    from io import BytesIO
    import openpyxl

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'لم يُرسَل ملف'}), 400

    period_id = request.form.get('period_id', type=int)
    if not period_id:
        return jsonify({'success': False, 'error': 'period_id مطلوب'}), 400

    period = GradePeriod.query.get(period_id)
    if not period:
        return jsonify({'success': False, 'error': 'الفترة غير موجودة'}), 404

    cats = GradeConfig.query.filter_by(period_id=period_id, is_active=True)\
                            .order_by(GradeConfig.sort_order).all()
    if not cats:
        return jsonify({'success': False, 'error': 'لا توجد فئات'}), 404

    file = request.files['file']
    try:
        wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
        ws = wb.worksheets[0]
    except Exception:
        return jsonify({'success': False, 'error': 'الملف تالف أو غير صالح'}), 400

    saved, errors = 0, []

    for row in ws.iter_rows(min_row=3, values_only=True):
        student_id = row[0]
        if not student_id:
            continue
        student = Student.query.get(int(student_id))
        if not student:
            errors.append(f'طالب ID {student_id} غير موجود')
            continue

        for i, cat in enumerate(cats):
            raw = row[2 + i] if (2 + i) < len(row) else None
            if raw is None or str(raw).strip() == '':
                continue
            try:
                score_val = float(raw)
            except (ValueError, TypeError):
                errors.append(f'{student.name} — {cat.category_name}: قيمة غير صالحة ({raw})')
                continue
            if score_val < 0 or score_val > cat.max_score:
                errors.append(f'{student.name} — {cat.category_name}: خارج النطاق ({score_val}/{cat.max_score})')
                continue

            label = f'استيراد {period.period_name}'
            GradeEntry.query.filter_by(
                student_id=student.id, category_id=cat.id, entry_label=label).delete()
            db.session.add(GradeEntry(
                student_id=student.id,
                category_id=cat.id,
                admin_id=current_user.id,
                entry_label=label,
                score=score_val / cat.max_score,
            ))
            db.session.flush()
            _sync_to_old_system(student.id, cat.id, admin_id=current_user.id)
            saved += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'خطأ في الحفظ: {str(e)}'}), 500

    return jsonify({
        'success': True,
        'saved':   saved,
        'errors':  errors[:20],
        'message': f'تم حفظ {saved} درجة' + (f' — {len(errors)} تحذير' if errors else ''),
    })
